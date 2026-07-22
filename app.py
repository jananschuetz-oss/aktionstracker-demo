from flask import Flask, render_template, request, redirect, url_for, session, send_file, g, flash, jsonify, abort
import sqlite3
import os
import math
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass
import uuid
import secrets
import shutil
import base64
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
import zipfile
from datetime import datetime, date, timedelta
import io
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
from flask_wtf import CSRFProtect
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from PIL import Image
import urllib.request
import urllib.parse
import time as _time
import threading
import functools
import holidays as _holidays_lib
import re

app = Flask(__name__)
_SECRET_KEY = os.environ.get('SECRET_KEY')
if not _SECRET_KEY:
    # Kein Fallback: ein fest im Code stehender Schlüssel wäre für jeden mit
    # Repo-Zugriff einsehbar und würde erlauben, Session-Cookies (inkl. Admin-
    # Rolle) selbst zu signieren. Lieber der Start scheitert, als dass die App
    # mit einem kompromittierten Schlüssel läuft.
    raise RuntimeError(
        "SECRET_KEY fehlt in der Umgebung – App wird aus Sicherheitsgründen nicht gestartet. "
        "Lokal: .env mit SECRET_KEY=<zufälliger Wert> anlegen. Produktion: Railway-Variable setzen."
    )
app.secret_key = _SECRET_KEY

# CSRF-Schutz (Bugreport 2026-07-21): global für alle POST/PUT/PATCH/DELETE-Routen.
# HTML-Formulare bekommen das Token über {{ csrf_token() }} (siehe Templates), Fetch-
# Requests aus dem Frontend über den global gepatchten window.fetch in base.html, der
# das Token als X-CSRFToken-Header anhängt (siehe dortiger Kommentar).
csrf = CSRFProtect(app)

# IP-basiertes Rate-Limiting (Bugreport 2026-07-21, Mittel): zusätzliche Schutzschicht
# neben dem konto-basierten Login-Lockout. In-Memory-Storage genügt für das aktuelle
# Deployment (wenige Gunicorn-Worker, keine Multi-Instanz-Skalierung).
limiter = Limiter(get_remote_address, app=app, default_limits=['300 per hour'], storage_uri='memory://')

# Flask setzt den Logger in Produktion (debug=False) standardmäßig auf WARNING – dadurch
# wären sämtliche app.logger.info()-Meldungen (Wochenbericht/Monatsbericht/Export-Status
# usw.) in den Railway-Logs unsichtbar. Explizit auf INFO heben.
import logging as _logging
app.logger.setLevel(_logging.INFO)

# ── Session-Sicherheit ────────────────────────────────────────────────────────
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=8)
app.config['SESSION_COOKIE_HTTPONLY']    = True
app.config['SESSION_COOKIE_SAMESITE']   = 'Lax'
# Explizite FORCE_HTTPS-Variable statt ausschließlich der Railway-spezifischen Heuristik
# (Bugreport 2026-07-21): funktioniert weiterhin ohne Konfiguration auf Railway, gibt aber
# zusätzlich einen dokumentierten, plattformunabhängigen Schalter für andere Hosts/On-Premise.
app.config['SESSION_COOKIE_SECURE']     = (
    os.environ.get('FORCE_HTTPS', '').strip().lower() in ('1', 'true', 'yes')
    or os.environ.get('RAILWAY_ENVIRONMENT') is not None
)

# HTTP-Security-Header (Bugreport 2026-07-21): keine strikte Content-Security-Policy, da
# die App durchgängig Inline-<script>-Blöcke nutzt (kein Nonce-System) – das würde ohne
# aufwändigen Umbau alle Seiten brechen. Diese Header schützen trotzdem gegen Clickjacking
# (X-Frame-Options) und MIME-Sniffing-Angriffe (X-Content-Type-Options), ohne Risiko für
# bestehende Funktionalität.
@app.after_request
def set_security_headers(response):
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-Frame-Options'] = 'DENY'
    response.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'
    if app.config['SESSION_COOKIE_SECURE']:
        response.headers['Strict-Transport-Security'] = 'max-age=31536000; includeSubDomains'
    # Content-Security-Policy (Bugreport 2026-07-21, Mittel): bewusst kein striktes
    # script-src ohne 'unsafe-inline' (viele Inline-Scripts ohne Nonce-System). Blockt
    # trotzdem das Nachladen von Skripten/Objekten von FREMDEN Domains. connect-src
    # braucht cdn.jsdelivr.net zusätzlich zu 'self': der Service Worker (static/sw.js)
    # cached Bootstrap/Chart.js beim Install-Event per cache.addAll() – fetch()-Requests
    # fallen unter connect-src, nicht unter script-src.
    response.headers['Content-Security-Policy'] = (
        "default-src 'self'; "
        "script-src 'self' 'unsafe-inline' https://cdn.jsdelivr.net; "
        "style-src 'self' 'unsafe-inline' https://cdn.jsdelivr.net; "
        "font-src 'self' https://cdn.jsdelivr.net; "
        "img-src 'self' data: https://*.tile.openstreetmap.org; "
        "connect-src 'self' https://cdn.jsdelivr.net; "
        "object-src 'none'; base-uri 'self'; frame-ancestors 'none'"
    )
    return response

DATABASE = os.environ.get('DATABASE_PATH', 'brewery.db')
LOGO_VERSION = '3'  # cache-bust

# ── Branding (pro-Kunde via Railway ENV Variables anpassbar) ─────────────────
# Railway: Settings → Variables → diese Variablen setzen
COMPANY_NAME   = os.environ.get('COMPANY_NAME',   'Ihre Firma GmbH')
COMPANY_SHORT  = os.environ.get('COMPANY_SHORT',  'Demo')
LOGO_URL       = os.environ.get('LOGO_URL',       '')    # externe Bild-URL oder leer → lokale Datei
ADMIN_PASSWORD = os.environ.get('ADMIN_PASSWORD')
if not ADMIN_PASSWORD:
    # Kein Fallback aus demselben Grund wie bei SECRET_KEY: ein im Code stehendes
    # Standardpasswort wäre bei jedem vergessenen Deployment sofort erratbar.
    raise RuntimeError(
        "ADMIN_PASSWORD fehlt in der Umgebung – App wird aus Sicherheitsgründen nicht gestartet. "
        "Lokal: .env mit ADMIN_PASSWORD=<eigenes Passwort> anlegen. Produktion: Railway-Variable setzen."
    )
EXPORT_EMAIL   = os.environ.get('EXPORT_EMAIL',   '')        # E-Mail für automatischen 4-Wochen-Export
KARTE_MODUS    = os.environ.get('KARTE_MODUS',   'basis')   # 'aus' | 'basis' | 'heatmap'
TOUREN_MODUS   = os.getenv('TOUREN_MODUS', 'aus')             # 'aus' | 'an'
ARBEITSZEIT_MODUS = os.getenv('ARBEITSZEIT_MODUS', 'aus') == 'an'  # Zusatzmodul, standardmäßig aus (Add-on)
TANKEN_MODUS   = os.getenv('TANKEN_MODUS', 'aus') == 'an'  # Firmenwagen-Tanken-Sonderkategorie, standardmäßig aus (Add-on)
GRATISWARE_MODUS = os.getenv('GRATISWARE_MODUS', 'aus') == 'an'  # Gratisware-Report, standardmäßig aus (Add-on)
UNIT_LABEL       = os.environ.get('UNIT_LABEL',      'Einheiten')  # Mengenbezeichnung z.B. 'Kisten', 'Kartons', 'Paletten'
MAX_MITARBEITER  = int(os.environ.get('MAX_MITARBEITER', 0))  # 0 = kein Limit (nicht konfiguriert)
DEFAULT_PASSWORD = os.environ.get('DEFAULT_PASSWORD', 'demo123')  # Standard-Passwort für neue Mitarbeiter
DEMO_MODUS = os.environ.get('INIT_DEMO_USERS', 'true').lower() == 'true'

UPLOAD_FOLDER = os.getenv('UPLOAD_FOLDER', os.path.join(os.path.dirname(__file__), 'static', 'uploads'))
ALLOWED_EXTENSIONS = {'jpg', 'jpeg', 'png', 'gif', 'webp', 'heic', 'heif'}
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16 MB max upload

BACKUP_FOLDER = os.path.join(os.path.dirname(__file__), 'backups')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(BACKUP_FOLDER, exist_ok=True)

# ── E-Mail-Konfiguration (via Umgebungsvariablen setzen) ──────────────────────
# Railway: Settings → Variables → diese Variablen eintragen
MAIL_SERVER   = os.environ.get('MAIL_SERVER',   '')          # z.B. smtp.gmail.com
MAIL_PORT     = int(os.environ.get('MAIL_PORT',  587))
MAIL_USE_TLS  = os.environ.get('MAIL_USE_TLS',  'true').lower() == 'true'
MAIL_USERNAME = os.environ.get('MAIL_USERNAME', '')
MAIL_PASSWORD = os.environ.get('MAIL_PASSWORD', '')          # App-Passwort bei Gmail
MAIL_FROM     = os.environ.get('MAIL_FROM',     MAIL_USERNAME)
APP_BASE_URL  = os.environ.get('APP_BASE_URL',  '')          # z.B. https://mein-tool.up.railway.app

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# Saisonaler Verteilungsschlüssel für Produktverkäufe (Jan–Dez, Summe = 1.0)
# Sommer-Peak Juli/August, Herbst-Spike September/Oktober
SONNENSCHLUESSEL = [0.05, 0.05, 0.07, 0.08, 0.10, 0.12, 0.13, 0.12, 0.09, 0.08, 0.06, 0.05]
M_NAMEN = ['Jan', 'Feb', 'Mär', 'Apr', 'Mai', 'Jun', 'Jul', 'Aug', 'Sep', 'Okt', 'Nov', 'Dez']

@app.template_filter('from_json')
def from_json_filter(s):
    import json
    return json.loads(s)

@app.template_filter('todatetime')
def todatetime_filter(s):
    """Konvertiert 'YYYY-MM-DD' String zu date-Objekt für Datumsvergleiche im Template."""
    try:
        return date.fromisoformat(str(s))
    except Exception:
        return date.today()

@app.template_filter('isoweek')
def isoweek_filter(s):
    """Gibt die ISO-Kalenderwoche für ein 'YYYY-MM-DD' Datum zurück."""
    try:
        return date.fromisoformat(str(s)).isocalendar()[1]
    except Exception:
        return ''

@app.template_filter('weekday_de')
def weekday_de_filter(s):
    """Gibt den deutschen Wochentagsnamen für ein 'YYYY-MM-DD' Datum zurück."""
    _namen = ['Montag','Dienstag','Mittwoch','Donnerstag','Freitag','Samstag','Sonntag']
    try:
        return _namen[date.fromisoformat(str(s)).weekday()]
    except Exception:
        return ''

@app.before_request
def check_session_lifetime():
    """Session-Timer bei jedem Request erneuern (Sliding Window, 8h Inaktivität)."""
    pass

@app.context_processor
def inject_now():
    ctx = {
        'now':           datetime.now(),
        'company_name':  COMPANY_NAME,
        'company_short': COMPANY_SHORT,
        'logo_url':      LOGO_URL or '/static/logo.png',
        'karte_modus':      KARTE_MODUS,
        'touren_modus':     TOUREN_MODUS,
        'arbeitszeit_modus': ARBEITSZEIT_MODUS,
        'tanken_modus':     TANKEN_MODUS,
        'gratisware_modus': GRATISWARE_MODUS,
        'unit_label':       UNIT_LABEL,
        'max_mitarbeiter':  MAX_MITARBEITER,
        'default_password': DEFAULT_PASSWORD,
        'demo_modus':       DEMO_MODUS,
        'meine_vertretungen': [],
        'alle_kollegen':      [],
        'offene_urlaubsantraege': 0,
        'offene_vs_hinweise': 0,
        'offene_aufbauten_freigaben': 0,
        'abgelehnte_aufbauten': [],
        'mein_email': '',
        'mein_signatur': None,
        'karte_benachrichtigung': None,
        'urlaub_konto': None,
    }
    if session.get('user_id'):
        try:
            _benachr = query("SELECT karte_benachrichtigung FROM mitarbeiter WHERE id=?", (session['user_id'],), one=True)
            ctx['karte_benachrichtigung'] = (_benachr['karte_benachrichtigung'] if _benachr else None)
            ctx['meine_vertretungen'] = query(
                '''SELECT v.id, v.von, v.bis, v.status, v.typ, m.name AS vertreter_name
                   FROM vertretung v
                   LEFT JOIN mitarbeiter m ON m.id = v.vertreter_id
                   WHERE v.abwesender_id = ?
                   ORDER BY v.von DESC''',
                (session['user_id'],)
            )
            ctx['urlaub_konto'] = _urlaub_konto(session['user_id'], date.today().year)
            ctx['alle_kollegen'] = query(
                "SELECT id, name FROM mitarbeiter WHERE rolle IN ('rep','verkaufsleiter') AND id != ? ORDER BY name",
                (session['user_id'],)
            )
            _me = query("SELECT email, signatur_bild FROM mitarbeiter WHERE id=?", (session['user_id'],), one=True)
            ctx['mein_email'] = (_me['email'] or '') if _me else ''
            ctx['mein_signatur'] = (_me['signatur_bild'] or None) if _me else None
            # Zähler offener Urlaubsanträge (für Navbar-Markierung der Manager)
            if session.get('rolle') in ('admin', 'verkaufsleiter'):
                _tc, _tp = _team_m_clause('m')
                _cnt = query(
                    f'''SELECT COUNT(*) AS n FROM vertretung v
                        JOIN mitarbeiter m ON m.id = v.abwesender_id
                        WHERE v.status = 'angefragt'{_tc}''',
                    _tp, one=True)
                ctx['offene_urlaubsantraege'] = _cnt['n'] if _cnt else 0
            # Zähler offener Verkaufsstellen-Hinweise (nur Admin – nur der pflegt Stammdaten)
            if session.get('rolle') == 'admin':
                _hcnt = query("SELECT COUNT(*) AS n FROM vs_hinweis_meldung WHERE status = 'offen'", one=True)
                ctx['offene_vs_hinweise'] = _hcnt['n'] if _hcnt else 0
            # Zähler offener Aufbauten-Freigaben (für Navbar-Markierung der Manager)
            if session.get('rolle') in ('admin', 'verkaufsleiter'):
                _tc, _tp = _freigabe_scope_clause('a')
                _fcnt = query(f'''
                    SELECT COUNT(*) AS n FROM displayposition dp JOIN aktivitaet a ON a.id = dp.aktivitaet_id
                    WHERE dp.status = 'offen'{_tc}
                ''', _tp, one=True)
                ctx['offene_aufbauten_freigaben'] = _fcnt['n'] if _fcnt else 0
            # Abgelehnte Aufbauten, die der Mitarbeiter noch nicht gesehen hat – als
            # Hinweis-Banner auf dem Dashboard (s. /aufbau/<id>/gesehen zum Bestätigen).
            if session.get('rolle') in ('rep', 'verkaufsleiter'):
                ctx['abgelehnte_aufbauten'] = query('''
                    SELECT dp.id AS dp_id, dp.anzahl, dp.ablehnungsgrund,
                           ds.name AS ds_name, a.datum, v.name AS vs_name
                    FROM displayposition dp
                    JOIN displaysorte ds ON ds.id = dp.displaysorte_id
                    JOIN aktivitaet a ON a.id = dp.aktivitaet_id
                    JOIN verkaufsstelle v ON v.id = a.verkaufsstelle_id
                    WHERE dp.status = 'abgelehnt' AND dp.mitarbeiter_gesehen = 0
                          AND a.mitarbeiter_id = ?
                    ORDER BY a.datum DESC
                ''', (session['user_id'],))
        except Exception:
            pass
    return ctx


# ─── DB Helpers ───────────────────────────────────────────────────────────────

def get_db():
    db = getattr(g, '_database', None)
    if db is None:
        db = g._database = sqlite3.connect(DATABASE)
        db.row_factory = sqlite3.Row
        db.execute("PRAGMA foreign_keys = ON")
    return db

@app.teardown_appcontext
def close_connection(exception):
    db = getattr(g, '_database', None)
    if db is not None:
        db.close()

def query(sql, args=(), one=False):
    cur = get_db().execute(sql, args)
    rv = cur.fetchall()
    return (rv[0] if rv else None) if one else rv

def execute(sql, args=()):
    db = get_db()
    cur = db.execute(sql, args)
    db.commit()
    return cur.lastrowid


FOTO_AUFBEWAHRUNG_WOCHEN = 6   # Fotos werden nach 6 Wochen gelöscht (Export am 1. um 08:00, Cleanup um 09:00)

# ── E-Mail versenden ──────────────────────────────────────────────────────────

_smtp_last_error = ''   # Letzten E-Mail-Fehler für Diagnose merken
RESEND_API_KEY = os.environ.get('RESEND_API_KEY', '')

def send_email(to: str, subject: str, body_html: str) -> bool:
    """Sendet eine HTML-E-Mail via Resend (bevorzugt) oder SMTP (Fallback)."""
    global _smtp_last_error
    _smtp_last_error = ''

    # ── Resend API (funktioniert auf Railway, da SMTP-Ports oft blockiert sind) ──
    if RESEND_API_KEY:
        try:
            import resend as _resend
            _resend.api_key = RESEND_API_KEY
            from_addr = MAIL_FROM or f'Aktionstracker <{MAIL_USERNAME}>'
            _resend.Emails.send({
                'from':    from_addr,
                'to':      [to],
                'subject': subject,
                'html':    body_html,
            })
            app.logger.info(f"E-Mail via Resend gesendet an {to}")
            return True
        except Exception as e:
            _smtp_last_error = f'Resend: {e}'
            app.logger.error(f"Resend-Fehler: {e}")
            return False

    # ── SMTP-Fallback (lokale Entwicklung / andere Server) ────────────────────
    if not MAIL_SERVER or not MAIL_USERNAME:
        _smtp_last_error = 'MAIL_SERVER oder MAIL_USERNAME nicht gesetzt'
        app.logger.warning("E-Mail nicht konfiguriert (MAIL_SERVER / MAIL_USERNAME fehlen).")
        return False
    try:
        msg = MIMEMultipart('alternative')
        msg['Subject'] = subject
        msg['From']    = MAIL_FROM
        msg['To']      = to
        msg.attach(MIMEText(body_html, 'html', 'utf-8'))
        with smtplib.SMTP(MAIL_SERVER, MAIL_PORT, timeout=10) as smtp:
            if MAIL_USE_TLS:
                smtp.starttls()
            smtp.login(MAIL_USERNAME, MAIL_PASSWORD)
            smtp.send_message(msg)
        return True
    except Exception as e:
        _smtp_last_error = str(e)
        app.logger.error(f"E-Mail-Fehler (SMTP): {e}")
        return False


def send_email_with_attachments(to: str, subject: str, body_html: str,
                                attachments: list) -> bool:
    """Sendet eine HTML-E-Mail mit Dateianhängen.
    attachments: Liste von (dateiname, bytes_daten, content_type) Tupeln."""

    # ── Resend API ────────────────────────────────────────────────────────────
    if RESEND_API_KEY:
        try:
            import base64 as _b64
            import resend as _resend
            _resend.api_key = RESEND_API_KEY
            from_addr = MAIL_FROM or f'Aktionstracker <{MAIL_USERNAME}>'
            resend_attachments = [
                {'filename': name, 'content': _b64.b64encode(data).decode('ascii')}
                for name, data, _ in attachments
            ]
            _resend.Emails.send({
                'from':        from_addr,
                'to':          [to],
                'subject':     subject,
                'html':        body_html,
                'attachments': resend_attachments,
            })
            app.logger.info(f"E-Mail+Anhang via Resend gesendet an {to}")
            return True
        except Exception as e:
            app.logger.error(f"Resend-Fehler (Anhang): {e}")
            return False

    # ── SMTP-Fallback ─────────────────────────────────────────────────────────
    if not MAIL_SERVER or not MAIL_USERNAME:
        app.logger.warning("E-Mail nicht konfiguriert – Versand mit Anhang nicht möglich.")
        return False
    try:
        msg = MIMEMultipart('mixed')
        msg['Subject'] = subject
        msg['From']    = MAIL_FROM
        msg['To']      = to
        msg.attach(MIMEText(body_html, 'html', 'utf-8'))
        for dateiname, daten, content_type in attachments:
            haupttyp, untertyp = content_type.split('/', 1)
            part = MIMEBase(haupttyp, untertyp)
            part.set_payload(daten)
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', 'attachment', filename=dateiname)
            msg.attach(part)
        with smtplib.SMTP(MAIL_SERVER, MAIL_PORT, timeout=60) as smtp:
            if MAIL_USE_TLS:
                smtp.starttls()
            smtp.login(MAIL_USERNAME, MAIL_PASSWORD)
            smtp.send_message(msg)
        return True
    except Exception as e:
        app.logger.error(f"E-Mail-Fehler (Anhang): {e}")
        return False


def _html_to_pdf(html: str) -> bytes | None:
    """Konvertiert einen HTML-String zu PDF-Bytes via xhtml2pdf."""
    import traceback as _tb
    try:
        from xhtml2pdf import pisa
        # Spacer-TDs (class="kpi-spc") und @media-Style entfernen,
        # da xhtml2pdf sonst negative verfügbare Breite berechnet.
        clean = re.sub(r'<td[^>]*class=["\']kpi-spc["\'][^>]*>.*?</td>', '', html, flags=re.DOTALL)
        clean = re.sub(r'<style[^>]*>.*?</style>', '', clean, flags=re.DOTALL)
        buf = io.BytesIO()
        status = pisa.CreatePDF(clean.encode('utf-8'), dest=buf, encoding='utf-8')
        data = buf.getvalue()
        if not status.err and data:
            app.logger.info(f"PDF generiert: {len(data)} Bytes")
            return data
        app.logger.error(f"xhtml2pdf Fehler={status.err}, Bytes={len(data)}")
        return None
    except Exception as e:
        app.logger.error(f"PDF-Fehler: {e}\n{_tb.format_exc()}")
        return None


# ── Datenbank-Backup ──────────────────────────────────────────────────────────

def backup_db():
    """Erstellt ein tägliches Backup der DB. Behält die letzten 7 Tage."""
    if not os.path.exists(DATABASE):
        return
    heute       = date.today().isoformat()
    backup_pfad = os.path.join(BACKUP_FOLDER, f'brewery_{heute}.db')
    if not os.path.exists(backup_pfad):
        shutil.copy2(DATABASE, backup_pfad)
        # Nur die letzten 7 Backups behalten
        alle = sorted([
            f for f in os.listdir(BACKUP_FOLDER)
            if f.startswith('brewery_') and f.endswith('.db')
        ])
        for alt in alle[:-7]:
            try:
                os.remove(os.path.join(BACKUP_FOLDER, alt))
            except Exception:
                pass


def cleanup_alte_fotos():
    """Löscht Foto-Dateien die älter als FOTO_AUFBEWAHRUNG_WOCHEN Wochen sind
    und setzt foto_pfad in der DB auf NULL. Gibt Anzahl gelöschter Fotos zurück."""
    from datetime import timedelta
    grenzwert = (date.today() - timedelta(weeks=FOTO_AUFBEWAHRUNG_WOCHEN)).isoformat()
    db = get_db()
    alte_akte = db.execute(
        "SELECT id, foto_pfad, foto_pfad_2, foto_pfad_3 FROM aktivitaet "
        "WHERE ((foto_pfad IS NOT NULL AND foto_pfad != '') "
        "    OR (foto_pfad_2 IS NOT NULL AND foto_pfad_2 != '') "
        "    OR (foto_pfad_3 IS NOT NULL AND foto_pfad_3 != '')) "
        "AND datum < ?",
        (grenzwert,)
    ).fetchall()
    count = 0
    for akt in alte_akte:
        for spalte in ('foto_pfad', 'foto_pfad_2', 'foto_pfad_3'):
            if not akt[spalte]:
                continue
            pfad = os.path.join(UPLOAD_FOLDER, akt[spalte])
            if os.path.exists(pfad):
                os.remove(pfad)
                count += 1
        db.execute("UPDATE aktivitaet SET foto_pfad = NULL, foto_pfad_2 = NULL, foto_pfad_3 = NULL WHERE id = ?", (akt['id'],))
    if alte_akte:
        db.commit()
    return count


def _az_netto_minuten(beginn, ende, pause_minuten=0):
    """Berechnet die Netto-Arbeitszeit in Minuten aus Beginn/Ende (HH:MM) minus Pause.
    Gibt None zurück, wenn Beginn/Ende fehlen oder Ende vor Beginn liegt.
    Setzt IMMER mindestens die gesetzliche Pflichtpause an (>6 Std → 30 Min,
    >9 Std → 45 Min) – unabhängig davon, was in pause_minuten gespeichert ist.
    Die Speicher-Routen korrigieren pause_minuten zwar schon beim Sichern nach
    oben, aber bei einem Race zwischen zwei schnellen Feld-Speicherungen
    (Beginn→Ende kurz hintereinander) oder bei Altdaten von vor dieser Regel
    kann der gespeicherte Wert trotzdem 0 bleiben – die Anzeige muss sich
    daher selbst absichern, statt dem gespeicherten Wert blind zu vertrauen."""
    if not beginn or not ende:
        return None
    try:
        h1, m1 = (int(x) for x in beginn.split(':'))
        h2, m2 = (int(x) for x in ende.split(':'))
    except (ValueError, AttributeError):
        return None
    minuten = (h2 * 60 + m2) - (h1 * 60 + m1)
    if minuten < 0:
        return None
    effektive_pause = max(pause_minuten or 0, _az_pflichtpause_minuten(minuten))
    return max(0, minuten - effektive_pause)


def _az_brutto_minuten(beginn, ende):
    """Bruttoarbeitszeit in Minuten aus Beginn/Ende (HH:MM), ohne Pausenabzug.
    Gibt None zurück, wenn Beginn/Ende fehlen oder Ende vor Beginn liegt."""
    if not beginn or not ende:
        return None
    try:
        h1, m1 = (int(x) for x in beginn.split(':'))
        h2, m2 = (int(x) for x in ende.split(':'))
    except (ValueError, AttributeError):
        return None
    minuten = (h2 * 60 + m2) - (h1 * 60 + m1)
    return minuten if minuten >= 0 else None


def _az_pflichtpause_minuten(brutto_minuten):
    """Gesetzliche Mindestpause: >9 Std → 45 Min, >6 Std → 30 Min, sonst 0."""
    if brutto_minuten is None:
        return 0
    if brutto_minuten > 9 * 60:
        return 45
    if brutto_minuten > 6 * 60:
        return 30
    return 0


def _az_fmt_std(minuten):
    """Formatiert Minuten als 'Hh MMmin', z.B. 7h 30min."""
    if minuten is None:
        return '–'
    return f"{minuten // 60}h {minuten % 60:02d}min"


def komprimiere_foto(quelle, ziel_pfad: str, max_px: int = 1200, qualitaet: int = 75):
    """Öffnet Foto aus Datei-Objekt oder bytes-Puffer, skaliert auf max_px (längste Seite)
    und speichert als JPEG mit gegebener Qualität. Gibt Dateipfad zurück."""
    img = Image.open(quelle)
    img = img.convert("RGB")          # HEIC/PNG/WEBP → JPEG-kompatibel
    img.thumbnail((max_px, max_px), Image.LANCZOS)
    img.save(ziel_pfad, format="JPEG", quality=qualitaet, optimize=True)
    return ziel_pfad


def init_db():
    with app.app_context():
        db = get_db()
        db.executescript('''
            CREATE TABLE IF NOT EXISTS mitarbeiter (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                kuerzel TEXT NOT NULL UNIQUE,
                rolle TEXT DEFAULT 'rep',
                passwort TEXT DEFAULT 'start123'
            );

            CREATE TABLE IF NOT EXISTS verkaufsstelle (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                ort TEXT,
                typ TEXT,
                aktiv INTEGER DEFAULT 1
            );

            CREATE TABLE IF NOT EXISTS biersorte (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                einheit TEXT DEFAULT 'Kiste (20x0.5L)',
                aktiv INTEGER DEFAULT 1
            );

            CREATE TABLE IF NOT EXISTS aktivitaet (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                datum DATE NOT NULL,
                mitarbeiter_id INTEGER NOT NULL,
                verkaufsstelle_id INTEGER NOT NULL,
                anzahl_displays INTEGER DEFAULT 0,
                notizen TEXT,
                foto_pfad TEXT,
                erstellt_am TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (mitarbeiter_id) REFERENCES mitarbeiter(id),
                FOREIGN KEY (verkaufsstelle_id) REFERENCES verkaufsstelle(id)
            );

            CREATE TABLE IF NOT EXISTS bestellposition (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                aktivitaet_id INTEGER NOT NULL,
                biersorte_id INTEGER NOT NULL,
                kisten_anzahl INTEGER NOT NULL,
                FOREIGN KEY (aktivitaet_id) REFERENCES aktivitaet(id) ON DELETE CASCADE,
                FOREIGN KEY (biersorte_id) REFERENCES biersorte(id)
            );

            CREATE TABLE IF NOT EXISTS zielzahlen (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                mitarbeiter_id INTEGER,
                jahr INTEGER NOT NULL,
                displays_ziel INTEGER DEFAULT 0,
                kisten_ziel INTEGER DEFAULT 0,
                UNIQUE(mitarbeiter_id, jahr),
                FOREIGN KEY (mitarbeiter_id) REFERENCES mitarbeiter(id)
            );

            CREATE TABLE IF NOT EXISTS displaysorte (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL UNIQUE,
                aktiv INTEGER DEFAULT 1
            );

            CREATE TABLE IF NOT EXISTS displayposition (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                aktivitaet_id INTEGER NOT NULL,
                displaysorte_id INTEGER NOT NULL,
                anzahl INTEGER NOT NULL DEFAULT 0,
                FOREIGN KEY (aktivitaet_id) REFERENCES aktivitaet(id) ON DELETE CASCADE,
                FOREIGN KEY (displaysorte_id) REFERENCES displaysorte(id)
            );

            CREATE TABLE IF NOT EXISTS mitarbeiter_verkaufsstelle (
                mitarbeiter_id    INTEGER NOT NULL,
                verkaufsstelle_id INTEGER NOT NULL,
                PRIMARY KEY (mitarbeiter_id, verkaufsstelle_id),
                FOREIGN KEY (mitarbeiter_id)    REFERENCES mitarbeiter(id)    ON DELETE CASCADE,
                FOREIGN KEY (verkaufsstelle_id) REFERENCES verkaufsstelle(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS vertretung (
                id            INTEGER PRIMARY KEY AUTOINCREMENT,
                abwesender_id INTEGER NOT NULL,
                vertreter_id  INTEGER,
                von           DATE NOT NULL,
                bis           DATE NOT NULL,
                status        TEXT DEFAULT 'bestätigt',  -- angefragt|bestätigt|abgelehnt
                FOREIGN KEY (abwesender_id) REFERENCES mitarbeiter(id) ON DELETE CASCADE,
                FOREIGN KEY (vertreter_id)  REFERENCES mitarbeiter(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS wochenbericht_config (
                id             INTEGER PRIMARY KEY CHECK (id = 1),
                aktiv          INTEGER DEFAULT 0,
                empfaenger_2   TEXT    DEFAULT '',
                empfaenger_3   TEXT    DEFAULT '',
                zuletzt_gesendet TEXT  DEFAULT ''
            );
            INSERT OR IGNORE INTO wochenbericht_config (id) VALUES (1);

            CREATE TABLE IF NOT EXISTS tagesplan (
                id                INTEGER PRIMARY KEY AUTOINCREMENT,
                mitarbeiter_id    INTEGER NOT NULL REFERENCES mitarbeiter(id) ON DELETE CASCADE,
                verkaufsstelle_id INTEGER NOT NULL REFERENCES verkaufsstelle(id) ON DELETE CASCADE,
                datum             TEXT NOT NULL,
                reihenfolge       INTEGER DEFAULT 0,
                notiz             TEXT,
                erledigt          INTEGER DEFAULT 0,
                geloescht         INTEGER DEFAULT 0,
                geloescht_am      TEXT,
                aktivitaet_id     INTEGER REFERENCES aktivitaet(id) ON DELETE SET NULL,
                erstellt_von      INTEGER REFERENCES mitarbeiter(id) ON DELETE SET NULL,
                erstellt_am       TEXT DEFAULT (datetime('now'))
            );
        ''')

        # Freigabe-Workflow für Aufbauten: vor der ALTER-Schleife prüfen, ob die
        # displayposition.status-Spalte neu ist – nur dann gelten Bestandsdaten als bereits
        # genehmigt (s. Update unten nach der Schleife).
        try:
            _dp_cols_vorher = {c['name'] for c in db.execute("PRAGMA table_info(displayposition)").fetchall()}
            _dp_status_neu = 'status' not in _dp_cols_vorher
        except Exception:
            _dp_status_neu = False

        # Migrationen für bestehende DBs
        for migration in [
            "ALTER TABLE aktivitaet    ADD COLUMN foto_pfad          TEXT",
            # KONZEPT-V2: Aktivitätstyp (Aufbau/Bestellung/Besuch). Bestand → 'Aufbau'.
            "ALTER TABLE aktivitaet    ADD COLUMN aktionstyp         TEXT DEFAULT 'Aufbau'",
            # KONZEPT-V2 Phase 2: Lebenszyklus offener Bestellungen
            "ALTER TABLE aktivitaet    ADD COLUMN bestell_status     TEXT",   # offen|aufgebaut|storniert (nur Bestellung)
            "ALTER TABLE aktivitaet    ADD COLUMN storno_grund       TEXT",
            "ALTER TABLE aktivitaet    ADD COLUMN realisiert_am      TEXT",   # Phase 3: wann Bestellung aufgebaut/storniert wurde
            # Bestand: bereits vorhandene Bestellungen als 'offen' markieren
            "UPDATE aktivitaet SET bestell_status='offen' WHERE aktionstyp='Bestellung' AND bestell_status IS NULL",
            "ALTER TABLE mitarbeiter   ADD COLUMN email               TEXT",
            "ALTER TABLE mitarbeiter   ADD COLUMN reset_token         TEXT",
            "ALTER TABLE mitarbeiter   ADD COLUMN reset_token_ablauf  DATETIME",
            "ALTER TABLE mitarbeiter   ADD COLUMN muss_passwort_aendern INTEGER DEFAULT 0",
            "ALTER TABLE verkaufsstelle ADD COLUMN strasse             TEXT",
            "ALTER TABLE verkaufsstelle ADD COLUMN ansprechpartner    TEXT",
            "ALTER TABLE verkaufsstelle ADD COLUMN lat                REAL",
            "ALTER TABLE verkaufsstelle ADD COLUMN lng                REAL",
            "ALTER TABLE mitarbeiter   ADD COLUMN karte_benachrichtigung TEXT",
            # Wochenbericht-Config – nachrüsten falls DB vor diesem Feature erstellt wurde
            """CREATE TABLE IF NOT EXISTS wochenbericht_config (
                id             INTEGER PRIMARY KEY CHECK (id = 1),
                aktiv          INTEGER DEFAULT 0,
                empfaenger_2   TEXT    DEFAULT '',
                empfaenger_3   TEXT    DEFAULT '',
                zuletzt_gesendet TEXT  DEFAULT ''
            )""",
            "INSERT OR IGNORE INTO wochenbericht_config (id) VALUES (1)",
            # Multi-Team-Feature
            """CREATE TABLE IF NOT EXISTS team (
                id   INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL
            )""",
            "ALTER TABLE mitarbeiter ADD COLUMN team_id INTEGER REFERENCES team(id) ON DELETE SET NULL",
            "ALTER TABLE wochenbericht_config ADD COLUMN zuletzt_gesendet_monat TEXT DEFAULT ''",
            "ALTER TABLE wochenbericht_config ADD COLUMN urlaubsmail_empfaenger TEXT DEFAULT ''",
            "ALTER TABLE wochenbericht_config ADD COLUMN neue_vs_empfaenger TEXT DEFAULT ''",
            "ALTER TABLE tagesplan ADD COLUMN geloescht INTEGER DEFAULT 0",
            "ALTER TABLE tagesplan ADD COLUMN geloescht_am TEXT",
            "ALTER TABLE aktivitaet    ADD COLUMN von_uhrzeit       TEXT",
            "ALTER TABLE aktivitaet    ADD COLUMN bis_uhrzeit       TEXT",
            "ALTER TABLE verkaufsstelle ADD COLUMN plz              TEXT",
            "ALTER TABLE verkaufsstelle ADD COLUMN landkreis        TEXT",
            "ALTER TABLE mitarbeiter   ADD COLUMN aktiv             INTEGER DEFAULT 1",
            "ALTER TABLE verkaufsstelle ADD COLUMN geocode_quelle   TEXT",
            """CREATE TABLE IF NOT EXISTS plz_zentrum (
                plz  TEXT PRIMARY KEY,
                lat  REAL NOT NULL,
                lng  REAL NOT NULL
            )""",
            "ALTER TABLE aktivitaet    ADD COLUMN foto_pfad_2        TEXT",
            "ALTER TABLE aktivitaet    ADD COLUMN foto_pfad_3        TEXT",
            "ALTER TABLE verkaufsstelle ADD COLUMN hinweis             TEXT",
            "ALTER TABLE verkaufsstelle ADD COLUMN lieferant           TEXT",
            "ALTER TABLE verkaufsstelle ADD COLUMN kundennummer        TEXT",
            "ALTER TABLE displaysorte ADD COLUMN zaehlt_zur_zielerreichung INTEGER DEFAULT 1",
            """CREATE TABLE IF NOT EXISTS arbeitszeit (
                id             INTEGER PRIMARY KEY AUTOINCREMENT,
                mitarbeiter_id INTEGER NOT NULL REFERENCES mitarbeiter(id) ON DELETE CASCADE,
                datum          TEXT NOT NULL,
                beginn         TEXT,
                ende           TEXT,
                erstellt_am    TEXT DEFAULT (datetime('now','localtime')),
                pause_minuten  INTEGER DEFAULT 0,
                UNIQUE(mitarbeiter_id, datum)
            )""",
            """CREATE TABLE IF NOT EXISTS vs_hinweis_meldung (
                id                INTEGER PRIMARY KEY AUTOINCREMENT,
                verkaufsstelle_id INTEGER NOT NULL,
                mitarbeiter_id    INTEGER NOT NULL,
                aktivitaet_id     INTEGER,
                text              TEXT NOT NULL,
                status            TEXT DEFAULT 'offen',
                erstellt_am       TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                erledigt_am       TIMESTAMP,
                erledigt_von_id   INTEGER,
                FOREIGN KEY (verkaufsstelle_id) REFERENCES verkaufsstelle(id) ON DELETE CASCADE,
                FOREIGN KEY (mitarbeiter_id)    REFERENCES mitarbeiter(id),
                FOREIGN KEY (aktivitaet_id)     REFERENCES aktivitaet(id) ON DELETE SET NULL
            )""",
            # Performance: Indizes für die häufigsten Filter/Joins (v.a. Dashboard,
            # Aktivitäten-Liste, context_processor bei JEDEM Request) – ohne diese
            # macht SQLite bei wachsender Datenmenge Full-Table-Scans auf jeder Seite.
            "CREATE INDEX IF NOT EXISTS idx_aktivitaet_datum ON aktivitaet(datum)",
            "CREATE INDEX IF NOT EXISTS idx_aktivitaet_mitarbeiter ON aktivitaet(mitarbeiter_id)",
            "CREATE INDEX IF NOT EXISTS idx_aktivitaet_verkaufsstelle ON aktivitaet(verkaufsstelle_id)",
            # Covering-Index für die BP-Subquery im Dashboard (SUM(kisten_anzahl)
            # GROUP BY aktivitaet_id, über die gesamte Historie, bei jedem Aufruf).
            # Ersetzt den schmaleren idx_bestellposition_aktivitaet(aktivitaet_id) –
            # jede Abfrage, die der schmale Index bedienen konnte, kann genauso gut
            # das führende Präfix dieses Composite-Index nutzen, ohne den doppelten
            # Schreib-Overhead bei jedem INSERT auf bestellposition.
            "CREATE INDEX IF NOT EXISTS idx_bestellposition_akt_kisten ON bestellposition(aktivitaet_id, kisten_anzahl)",
            "DROP INDEX IF EXISTS idx_bestellposition_aktivitaet",
            "CREATE INDEX IF NOT EXISTS idx_displayposition_aktivitaet ON displayposition(aktivitaet_id)",
            "CREATE INDEX IF NOT EXISTS idx_mitarbeiter_team ON mitarbeiter(team_id)",
            "CREATE INDEX IF NOT EXISTS idx_mitarbeiter_verkaufsstelle_ma ON mitarbeiter_verkaufsstelle(mitarbeiter_id)",
            "CREATE INDEX IF NOT EXISTS idx_mitarbeiter_verkaufsstelle_vs ON mitarbeiter_verkaufsstelle(verkaufsstelle_id)",
            "CREATE INDEX IF NOT EXISTS idx_tagesplan_ma_datum ON tagesplan(mitarbeiter_id, datum)",
            "CREATE INDEX IF NOT EXISTS idx_arbeitszeit_ma_datum ON arbeitszeit(mitarbeiter_id, datum)",
            # Expression-Index: deckt alle Stellen ab, die per strftime('%Y', datum) = ?
            # nach Jahr filtern (Dashboard, Aktivitäten, Exporte, ...).
            "CREATE INDEX IF NOT EXISTS idx_aktivitaet_jahr ON aktivitaet(strftime('%Y', datum))",
            # Heatmap-Performance: api_karte_heatmap() joint pro Verkaufsstelle gegen
            # aktivitaet im Jahreszeitraum. Mit strftime('%Y',datum)=? wählt SQLite oft den
            # falschen Index (idx_aktivitaet_jahr) und scannt je VS über alle Aktivitäten des
            # Jahres statt direkt zur VS zu springen – bei vielen VS/Aktivitäten langsam. Mit
            # Datumsbereich (a.datum >= ? AND a.datum < ?) statt strftime kann dieser
            # zusammengesetzte Index als Covering-Index genutzt werden.
            "CREATE INDEX IF NOT EXISTS idx_aktivitaet_vs_datum ON aktivitaet(verkaufsstelle_id, datum)",
            # vertretung wird bei JEDEM Request im context_processor abgefragt
            # (Urlaubsantrag-Badge, meine_vertretungen) – ohne Index bisher ein
            # Full-Table-Scan auf jeder einzelnen Seite.
            "CREATE INDEX IF NOT EXISTS idx_vertretung_abwesender ON vertretung(abwesender_id)",
            "CREATE INDEX IF NOT EXISTS idx_vertretung_status ON vertretung(status)",
            # Urlaubskonto: Jahresanspruch + Übertrag Vorjahr je Mitarbeiter, Bundesland für
            # die gesetzlichen Feiertage bei der Werktage-Berechnung (siehe _werktage_zaehlen).
            "ALTER TABLE mitarbeiter ADD COLUMN urlaubsanspruch_jahr INTEGER DEFAULT 30",
            "ALTER TABLE mitarbeiter ADD COLUMN urlaub_uebertrag_vorjahr INTEGER DEFAULT 0",
            "ALTER TABLE mitarbeiter ADD COLUMN bundesland TEXT DEFAULT 'BY'",
            # Manuell erfasste, bereits genommene Urlaubstage des laufenden Jahres, die NICHT
            # über das Tool liefen (z.B. Alturlaub vor Einführung des Urlaubskontos) – fließt
            # zusätzlich zu den über 'vertretung' erfassten Tagen in "genommen" ein.
            "ALTER TABLE mitarbeiter ADD COLUMN urlaub_manuell_genommen INTEGER DEFAULT 0",
            # Abwesenheitstyp – nur 'urlaub' zieht vom Urlaubskonto ab, der Rest bleibt reine
            # Abwesenheitsdokumentation (Bestand vor diesem Feature: alles war faktisch Urlaub).
            "ALTER TABLE vertretung ADD COLUMN typ TEXT DEFAULT 'urlaub'",
            # Homeoffice: Wohnort je Mitarbeiter + Verweis auf die automatisch angelegte
            # Homeoffice-„Verkaufsstelle" (siehe verkaufsstelle.homeoffice_mitarbeiter_id
            # unten), damit Homeoffice als normaler Besuchsplanungs-Stopp funktioniert,
            # ohne dass Tagesplan/Aktivität-Grundgerüst angefasst werden muss.
            "ALTER TABLE mitarbeiter ADD COLUMN wohnort_strasse TEXT",
            "ALTER TABLE mitarbeiter ADD COLUMN wohnort_plz TEXT",
            "ALTER TABLE mitarbeiter ADD COLUMN wohnort_ort TEXT",
            "ALTER TABLE mitarbeiter ADD COLUMN homeoffice_vs_id INTEGER REFERENCES verkaufsstelle(id) ON DELETE SET NULL",
            "ALTER TABLE verkaufsstelle ADD COLUMN homeoffice_mitarbeiter_id INTEGER REFERENCES mitarbeiter(id) ON DELETE CASCADE",
            # Freigabe-Workflow für Aufbauten: der VKL muss jeden neu erfassten zielrelevanten
            # Aufbautyp einzeln prüfen (Details/Fotos ansehen) und freigeben oder ablehnen (mit
            # Begründung), bevor er in die Zielzahlen einfließt – siehe /aufbauten/freigabe.
            "ALTER TABLE displayposition ADD COLUMN status TEXT DEFAULT 'offen'",
            "ALTER TABLE displayposition ADD COLUMN ablehnungsgrund TEXT",
            "ALTER TABLE displayposition ADD COLUMN entschieden_von INTEGER",
            "ALTER TABLE displayposition ADD COLUMN entschieden_am TEXT",
            "ALTER TABLE displayposition ADD COLUMN mitarbeiter_gesehen INTEGER DEFAULT 1",
            # Firmenwagen-Tanken (Kernfunktion, standardmäßig per TANKEN_MODUS-Env-Var
            # deaktiviert): Pflichtfelder aus dem Tankbeleg, damit die spätere Erfassung
            # nicht mehr im Nachhinein per Fotoauswertung nachgetragen werden muss.
            "ALTER TABLE aktivitaet ADD COLUMN tank_datum TEXT",
            "ALTER TABLE aktivitaet ADD COLUMN tank_kennzeichen TEXT",
            "ALTER TABLE aktivitaet ADD COLUMN tank_kraftstoffsorte TEXT",
            "ALTER TABLE aktivitaet ADD COLUMN tank_liter REAL",
            "ALTER TABLE aktivitaet ADD COLUMN tank_km_stand INTEGER",
            # Gratisware (Kernfunktion, standardmäßig per GRATISWARE_MODUS-Env-Var
            # deaktiviert): eigene Kisten-Produkte (Verleger/Kofferraum), gesondert von der
            # regulären Bestellmenge im Gratisware-Report auswertbar.
            "ALTER TABLE biersorte ADD COLUMN ist_gratisware INTEGER DEFAULT 0",
            # Urlaubsantrag-PDF: Antragsdatum wird für den "Datum"-Vermerk des Mitarbeiters
            # auf dem erzeugten PDF benötigt. Kein Default-Ausdruck möglich (SQLite erlaubt
            # bei ALTER TABLE ADD COLUMN keinen nicht-konstanten DEFAULT) – wird stattdessen
            # explizit bei den beiden nutzerseitigen INSERT INTO vertretung gesetzt.
            "ALTER TABLE vertretung ADD COLUMN erstellt_am TEXT",
            # Login-Sperre nach Fehlversuchen (2026-07-21): fehlgeschlagene_logins zählt
            # falsche Passworteingaben in Folge, wird bei jedem erfolgreichen Login auf 0
            # zurückgesetzt. konto_gesperrt greift nach 3 Fehlversuchen für normale
            # Mitarbeiter-Accounts und bleibt bestehen, bis Admin/VKL das Passwort neu setzt
            # (admin_mitarbeiter_passwort) – bewusst kein Selbst-Reset über "Passwort
            # vergessen", damit ein gesperrtes Konto wirklich über die Leitung läuft.
            # gesperrt_bis ist nur für den ADMIN-Meta-Account relevant (zeitbasierte
            # Sperre statt "wer sperrt den Admin frei" – siehe login()).
            "ALTER TABLE mitarbeiter ADD COLUMN fehlgeschlagene_logins INTEGER DEFAULT 0",
            "ALTER TABLE mitarbeiter ADD COLUMN konto_gesperrt INTEGER DEFAULT 0",
            "ALTER TABLE mitarbeiter ADD COLUMN gesperrt_bis TEXT",
            # Digitale Signatur (2026-07-21): einmalig hinterlegte Unterschrift pro Mitarbeiter
            # (als PNG-Data-URI), wird beim Urlaubsantrag sowohl für den Antragsteller als
            # auch für den genehmigenden VKL/Admin ins PDF eingebettet – kein erneutes
            # Unterschreiben je Antrag nötig. bestaetigt_von_id hält fest, wer genehmigt hat,
            # damit dessen Signatur (nicht die eines beliebigen anderen Managers) verwendet wird.
            "ALTER TABLE mitarbeiter ADD COLUMN signatur_bild TEXT",
            "ALTER TABLE vertretung ADD COLUMN bestaetigt_von_id INTEGER REFERENCES mitarbeiter(id) ON DELETE SET NULL",
        ]:
            try:
                db.execute(migration)
                db.commit()
            except Exception:
                pass  # Spalte existiert bereits

        # Bestehende Displaypositionen (vor Einführung des Freigabe-Workflows) gelten als
        # bereits genehmigt, damit sich anzahl_displays/Zielzahlen nicht rückwirkend ändern.
        if _dp_status_neu:
            try:
                db.execute("UPDATE displayposition SET status='freigegeben'")
                db.commit()
            except Exception:
                pass

        # Migration: vertreter_id nullable machen (Urlaub ohne Vertretung).
        # SQLite kann NOT NULL nicht per ALTER entfernen → Tabelle neu aufbauen.
        try:
            _cols = db.execute("PRAGMA table_info(vertretung)").fetchall()
            _vid  = next((c for c in _cols if c['name'] == 'vertreter_id'), None)
            if _vid and _vid['notnull'] == 1:
                db.executescript('''
                    PRAGMA foreign_keys=off;
                    CREATE TABLE vertretung_neu (
                        id            INTEGER PRIMARY KEY AUTOINCREMENT,
                        abwesender_id INTEGER NOT NULL,
                        vertreter_id  INTEGER,
                        von           DATE NOT NULL,
                        bis           DATE NOT NULL,
                        FOREIGN KEY (abwesender_id) REFERENCES mitarbeiter(id) ON DELETE CASCADE,
                        FOREIGN KEY (vertreter_id)  REFERENCES mitarbeiter(id) ON DELETE CASCADE
                    );
                    INSERT INTO vertretung_neu (id, abwesender_id, vertreter_id, von, bis)
                        SELECT id, abwesender_id, vertreter_id, von, bis FROM vertretung;
                    DROP TABLE vertretung;
                    ALTER TABLE vertretung_neu RENAME TO vertretung;
                    PRAGMA foreign_keys=on;
                ''')
                db.commit()
        except Exception:
            pass

        # Migration: Status-Spalte für Urlaubs-Genehmigung (angefragt/bestätigt/abgelehnt).
        # Bestandseinträge gelten als 'bestätigt', damit nichts kippt.
        try:
            db.execute("ALTER TABLE vertretung ADD COLUMN status TEXT DEFAULT 'bestätigt'")
            db.commit()
        except Exception:
            pass

        # Migration: Freitext-Grund für Abwesenheiten (v.a. "Frei/Sonderurlaub").
        try:
            db.execute("ALTER TABLE vertretung ADD COLUMN grund TEXT")
            db.commit()
        except Exception:
            pass

        # Idempotent (läuft bei jedem Start, aber nach der ersten Korrektur ohne
        # Effekt): Pflichtpause rückwirkend auf bestehende Arbeitszeit-Einträge
        # anwenden, bei denen sie noch nicht korrekt gespeichert war (z.B. durch
        # das Race zwischen zwei schnellen Feld-Speicherungen Beginn→Ende). Die
        # Anzeige (_az_netto_minuten) setzt die Pflichtpause zwar selbst schon
        # immer an, aber die gespeicherte Pause-Spalte soll denselben Wert
        # zeigen statt "0", damit sie nicht wie ein Fehler aussieht.
        try:
            _az_rows = db.execute(
                "SELECT id, beginn, ende, pause_minuten FROM arbeitszeit WHERE beginn IS NOT NULL AND ende IS NOT NULL"
            ).fetchall()
            for _r in _az_rows:
                _pflicht = _az_pflichtpause_minuten(_az_brutto_minuten(_r['beginn'], _r['ende']))
                if _pflicht > (_r['pause_minuten'] or 0):
                    db.execute("UPDATE arbeitszeit SET pause_minuten=? WHERE id=?", (_pflicht, _r['id']))
            db.commit()
        except Exception:
            pass

        # Admin + Demo-Leitung + Verkaufsleiter (Passwörter via ENV konfigurierbar). Nur
        # (neu) hashen, wenn nötig – sonst würde jeder Neustart ein evtl. schon migriertes
        # Hash-Passwort wieder mit dem ENV-Klartext überschreiben (siehe login()).
        def _seed_passwort(kuerzel, name, rolle, klartext_pw):
            row = db.execute("SELECT passwort FROM mitarbeiter WHERE kuerzel=?", (kuerzel,)).fetchone()
            if not row or not _ist_passwort_hash(row['passwort']) or not check_password_hash(row['passwort'], klartext_pw):
                pw_hash = generate_password_hash(klartext_pw)
                db.execute("INSERT OR IGNORE INTO mitarbeiter (name, kuerzel, rolle, passwort) VALUES (?,?,?,?)", (name, kuerzel, rolle, pw_hash))
                db.execute("UPDATE mitarbeiter SET passwort=? WHERE kuerzel=?", (pw_hash, kuerzel))
        _seed_passwort('ADMIN', 'Administrator', 'admin', ADMIN_PASSWORD)
        # Demo Leitung (Login: Demo) – einziger Demo-GF-Zugang für Interessenten
        _seed_passwort('Demo', 'Demo Leitung', 'admin', os.environ.get('DEMO_PASSWORT', 'demo2026'))
        _seed_passwort('VKL', 'Verkaufsleiter', 'verkaufsleiter', DEFAULT_PASSWORD)

        # Beispiel-Mitarbeiter (nur bei INIT_DEMO_USERS=true) – 4 Reps + 1 VKL
        if os.environ.get('INIT_DEMO_USERS', 'true').lower() == 'true':
            reps = [
                ('Max Müller',    'MM', DEFAULT_PASSWORD),
                ('Anna Schmidt',  'AS', DEFAULT_PASSWORD),
                ('Thomas Weber',  'TW', DEFAULT_PASSWORD),
                ('Lisa Fischer',  'LF', DEFAULT_PASSWORD),
            ]
            for name, kuerzel, pw in reps:
                _seed_passwort(kuerzel, name, 'rep', pw)
            # KH deaktivieren falls aus Altbestand vorhanden
            db.execute("UPDATE mitarbeiter SET aktiv=0 WHERE kuerzel='KH'")

        # Displaysorten – nur einfügen wenn Tabelle leer
        if not db.execute("SELECT 1 FROM displaysorte LIMIT 1").fetchone():
            for ds_name in ['Regal-Display', 'Eingangs-Display',
                            'Counter-Display', 'Schaufenster', 'Außenwerbung']:
                db.execute("INSERT OR IGNORE INTO displaysorte (name) VALUES (?)", (ds_name,))

        # Produkte – nur einfügen wenn Tabelle leer
        if not db.execute("SELECT 1 FROM biersorte LIMIT 1").fetchone():
            produkte = [
                ('Produkt A',  'Karton (12 Stück)'),
                ('Produkt B',  'Karton (12 Stück)'),
                ('Produkt C',  'Karton (24 Stück)'),
                ('Produkt D',  'Karton (24 Stück)'),
                ('Produkt E',  'Palette (100 Stück)'),
                ('Produkt F',  'Palette (100 Stück)'),
            ]
            for name, einheit in produkte:
                db.execute("INSERT INTO biersorte (name, einheit) VALUES (?, ?)", (name, einheit))

        # Beispiel-Kunden – nur einfügen wenn Tabelle leer
        if not db.execute("SELECT 1 FROM verkaufsstelle LIMIT 1").fetchone():
            stellen = [
                ('Supermarkt Mitte',       'Berlin',   'Einzelhandel'),
                ('Fachmarkt Nord',         'Hamburg',  'Einzelhandel'),
                ('Restaurant Zur Post',    'München',  'Gastronomie'),
                ('Hotel Stadtblick',       'Frankfurt','Hotel'),
                ('Großhandel Meyer',       'Köln',     'Getränkehandel'),
                ('Kiosk am Bahnhof',       'Düsseldorf','Kiosk'),
                ('Sportverein 1902',       'Stuttgart','Verein'),
                ('Café Central',           'Leipzig',  'Gastronomie'),
            ]
            for name, ort, typ in stellen:
                db.execute("INSERT INTO verkaufsstelle (name, ort, typ) VALUES (?, ?, ?)", (name, ort, typ))

        db.commit()

        # Gratisware-Produkte (Verleger/Kofferraum) – eigene Kisten-Posten, gesondert von
        # der regulären Kistenware ausgewertet (ist_gratisware=1). Idempotent. Muss NACH der
        # obigen Demo-Produkte-Seedung laufen, sonst würde die "Tabelle leer?"-Prüfung dort
        # durch diese beiden Zeilen fälschlich als "schon befüllt" erkannt und die
        # Demo-Produkte (A–F) würden nie angelegt.
        for _gw_name in ('Gratisware Verleger', 'Gratisware Kofferraum'):
            if not db.execute("SELECT 1 FROM biersorte WHERE name=?", (_gw_name,)).fetchone():
                db.execute(
                    "INSERT INTO biersorte (name, einheit, ist_gratisware) VALUES (?, 'Kiste', 1)",
                    (_gw_name,)
                )
        db.commit()

        # Beispieldaten einfügen wenn DB noch leer
        if not db.execute("SELECT 1 FROM aktivitaet LIMIT 1").fetchone():
            seed_demo_data_relativ(db)
            seed_demo_besuchsplan(db)

        # Zielzahlen immer aktualisieren (unabhängig ob Seed gelaufen)
        _ziele_2026 = {'MM':(80,2900),'AS':(120,4600),'TW':(80,2600),'LF':(90,3900),'KH':(80,2500)}
        _reps_all = db.execute("SELECT id, kuerzel FROM mitarbeiter WHERE rolle='rep'").fetchall()
        for _r in _reps_all:
            if _r['kuerzel'] in _ziele_2026:
                _d, _k = _ziele_2026[_r['kuerzel']]
                db.execute('''INSERT INTO zielzahlen (mitarbeiter_id,jahr,displays_ziel,kisten_ziel)
                    VALUES (?,2026,?,?) ON CONFLICT(mitarbeiter_id,jahr) DO UPDATE SET
                    displays_ziel=excluded.displays_ziel, kisten_ziel=excluded.kisten_ziel''',
                    (_r['id'], _d, _k))
        db.execute('''INSERT INTO zielzahlen (mitarbeiter_id,jahr,displays_ziel,kisten_ziel)
            VALUES (NULL,2026,450,16500) ON CONFLICT(mitarbeiter_id,jahr) DO UPDATE SET
            displays_ziel=excluded.displays_ziel, kisten_ziel=excluded.kisten_ziel''')
        db.commit()

        # Bestellpositionen nachfüllen wenn Tabelle leer aber Bestellungen existieren
        if not db.execute("SELECT 1 FROM bestellposition LIMIT 1").fetchone():
            import random as _rnd
            _biere = db.execute("SELECT id FROM biersorte WHERE aktiv=1 AND COALESCE(ist_gratisware,0)=0").fetchall()
            _bier_ids = [b['id'] for b in _biere]
            if _bier_ids:
                _best_akt = db.execute(
                    "SELECT id FROM aktivitaet WHERE aktionstyp='Bestellung' AND bestell_status IN ('offen','aufgebaut')"
                ).fetchall()
                for _a in _best_akt:
                    for _bid in _rnd.sample(_bier_ids, k=min(_rnd.randint(2,4), len(_bier_ids))):
                        db.execute(
                            "INSERT OR IGNORE INTO bestellposition (aktivitaet_id,biersorte_id,kisten_anzahl) VALUES (?,?,?)",
                            (_a['id'], _bid, _rnd.randint(5, 45))
                        )
                db.commit()
                app.logger.info("Bestellpositionen nachgefüllt für %d Bestellungen.", len(_best_akt))

        # Beispielfotos einmalig zuweisen – NACH seed_demo_data (Aktivitäten müssen existieren)
        fotos_in_db = db.execute(
            "SELECT COUNT(*) FROM aktivitaet WHERE foto_pfad IS NOT NULL AND foto_pfad != ''"
        ).fetchone()[0]
        if fotos_in_db == 0 and os.path.isdir(UPLOAD_FOLDER):
            upload_files = sorted([
                f for f in os.listdir(UPLOAD_FOLDER)
                if f.startswith('akt_') and f.endswith('.jpg')
            ])
            if upload_files:
                mai_akte = db.execute("""
                    SELECT id FROM aktivitaet
                    WHERE strftime('%Y-%m', datum) = '2026-05'
                      AND (foto_pfad IS NULL OR foto_pfad = '')
                    ORDER BY datum, id
                    LIMIT ?
                """, (len(upload_files),)).fetchall()
                for akt_row, dateiname in zip(mai_akte, upload_files):
                    db.execute("UPDATE aktivitaet SET foto_pfad = ? WHERE id = ?",
                               (dateiname, akt_row['id']))
                db.commit()

        # Stationszuordnung: Demo-Reps geografisch nach Region – idempotent
        # Trigger: MM hat Stationen außerhalb Bayerns ODER es gibt unzugeordnete Stationen
        _mm_row = db.execute("SELECT id FROM mitarbeiter WHERE kuerzel='MM'").fetchone()
        _mm_hat_falsch = bool(_mm_row and db.execute("""
            SELECT COUNT(*) FROM mitarbeiter_verkaufsstelle mv
            JOIN verkaufsstelle v ON v.id = mv.verkaufsstelle_id
            WHERE mv.mitarbeiter_id=? AND v.ort NOT IN ('München','Nürnberg')
        """, (_mm_row['id'],)).fetchone()[0])
        _unzugeordnet = db.execute("""
            SELECT COUNT(*) FROM verkaufsstelle v
            LEFT JOIN mitarbeiter_verkaufsstelle mv ON mv.verkaufsstelle_id = v.id
            WHERE v.aktiv = 1 AND mv.mitarbeiter_id IS NULL
        """).fetchone()[0]
        if _unzugeordnet > 0 or _mm_hat_falsch:
            import random as _rnd_assign
            _rnd_assign.seed(42)
            # Rep-Zuordnungen löschen und geografisch neu vergeben
            db.execute("DELETE FROM mitarbeiter_verkaufsstelle WHERE mitarbeiter_id IN "
                       "(SELECT id FROM mitarbeiter WHERE rolle='rep')")
            DEMO_GEO = {
                'MM': ('München', 'Nürnberg'),
                'AS': ('Hamburg', 'Hannover', 'Bremen', 'Berlin'),
                'TW': ('Frankfurt', 'Wiesbaden', 'Leipzig'),
                'LF': ('Köln', 'Düsseldorf', 'Dortmund', 'Essen', 'Bonn'),
                'KH': ('Stuttgart', 'Freiburg', 'Mannheim'),
            }
            for _kz, _staedte in DEMO_GEO.items():
                _r = db.execute("SELECT id FROM mitarbeiter WHERE kuerzel=?", (_kz,)).fetchone()
                if not _r:
                    continue
                for _s in db.execute(
                    "SELECT id FROM verkaufsstelle WHERE ort IN ({}) AND aktiv=1".format(
                        ','.join('?' * len(_staedte))
                    ), _staedte
                ).fetchall():
                    db.execute(
                        "INSERT OR IGNORE INTO mitarbeiter_verkaufsstelle (mitarbeiter_id, verkaufsstelle_id) VALUES (?,?)",
                        (_r['id'], _s['id'])
                    )
            # Catch-all: verbleibende unzugeordnete Stationen gleichmäßig auf Reps verteilen
            _rest_reps = db.execute("SELECT id FROM mitarbeiter WHERE rolle='rep'").fetchall()
            _rest_vs = db.execute("""
                SELECT v.id FROM verkaufsstelle v
                WHERE v.aktiv=1 AND v.id NOT IN (
                    SELECT verkaufsstelle_id FROM mitarbeiter_verkaufsstelle
                    WHERE mitarbeiter_id IN (SELECT id FROM mitarbeiter WHERE rolle='rep')
                )
            """).fetchall()
            if _rest_vs and _rest_reps:
                _pool_r = [s['id'] for s in _rest_vs]
                _rnd_assign.shuffle(_pool_r)
                for _i, _sid in enumerate(_pool_r):
                    db.execute(
                        "INSERT OR IGNORE INTO mitarbeiter_verkaufsstelle (mitarbeiter_id, verkaufsstelle_id) VALUES (?,?)",
                        (_rest_reps[_i % len(_rest_reps)]['id'], _sid)
                    )
            # VKL bekommt bewusst keine Stationen zugewiesen (sieht alles über die Gesamtansicht)
            db.commit()
            app.logger.info("Demo: Stationszuordnung geografisch nach Regionen verteilt.")

        # Demo-Koordinaten: Stationen ohne lat/lng anhand der Stadt direkt setzen
        ohne_coords = db.execute(
            "SELECT id, ort FROM verkaufsstelle WHERE aktiv=1 AND (lat IS NULL OR lng IS NULL)"
        ).fetchall()
        if ohne_coords:
            import random as _rnd_geo
            _rnd_geo.seed(77)
            STADT_COORDS = {
                'Berlin':     (52.5200, 13.4050),
                'Hamburg':    (53.5753, 10.0153),
                'München':    (48.1374, 11.5755),
                'Frankfurt':  (50.1109,  8.6821),
                'Köln':       (50.9333,  6.9500),
                'Düsseldorf': (51.2217,  6.7762),
                'Stuttgart':  (48.7758,  9.1829),
                'Leipzig':    (51.3397, 12.3731),
                'Nürnberg':   (49.4521, 11.0767),
                'Hannover':   (52.3759,  9.7320),
                'Mannheim':   (49.4875,  8.4660),
                'Dortmund':   (51.5135,  7.4653),
                'Bremen':     (53.0793,  8.8017),
                'Wiesbaden':  (50.0800,  8.2400),
                'Bonn':       (50.7374,  7.0982),
                'Freiburg':   (47.9990,  7.8421),
                'Essen':      (51.4508,  7.0131),
                'Augsburg':   (48.3717, 10.8983),
                'Starnberg':  (47.9986, 11.3381),
                'Dachau':     (48.2604, 11.4335),
            }
            gesetzt = 0
            for vs in ohne_coords:
                base = STADT_COORDS.get(vs['ort'])
                if base:
                    lat = base[0] + _rnd_geo.uniform(-0.05, 0.05)
                    lng = base[1] + _rnd_geo.uniform(-0.07, 0.07)
                    db.execute("UPDATE verkaufsstelle SET lat=?, lng=? WHERE id=?", (lat, lng, vs['id']))
                    gesetzt += 1
            db.commit()
            app.logger.info(f"Demo-Koordinaten: {gesetzt}/{len(ohne_coords)} Stationen mit Stadtkoordinaten gesetzt.")

        # Demo-Landkreis + PLZ: Stationen anhand Ort befüllen (idempotent)
        _demo_lk = {
            'München': ('München (Stadt)', '80331'),
            'Nürnberg': ('Nürnberg (Stadt)', '90403'),
            'Hamburg': ('Hamburg (Stadt)', '20095'),
            'Frankfurt': ('Frankfurt am Main (Stadt)', '60311'),
            'Köln': ('Köln (Stadt)', '50667'),
            'Düsseldorf': ('Düsseldorf (Stadt)', '40213'),
            'Stuttgart': ('Stuttgart (Stadt)', '70173'),
            'Leipzig': ('Leipzig (Stadt)', '04109'),
            'Hannover': ('Hannover (Region)', '30159'),
            'Mannheim': ('Mannheim (Stadt)', '68161'),
            'Dortmund': ('Dortmund (Stadt)', '44135'),
            'Bremen': ('Bremen (Stadt)', '28195'),
            'Wiesbaden': ('Wiesbaden (Stadt)', '65183'),
            'Bonn': ('Rhein-Sieg-Kreis', '53111'),
            'Freiburg': ('Freiburg im Breisgau (Stadt)', '79098'),
            'Essen': ('Essen (Stadt)', '45127'),
            'Augsburg': ('Augsburg (Stadt)', '86150'),
            'Starnberg': ('Landkreis Starnberg', '82319'),
            'Dachau': ('Landkreis Dachau', '85221'),
        }
        for _ort, (_lk, _plz) in _demo_lk.items():
            db.execute("UPDATE verkaufsstelle SET landkreis=? WHERE ort=? AND landkreis IS NULL", (_lk, _ort))
            db.execute("UPDATE verkaufsstelle SET plz=? WHERE ort=? AND plz IS NULL", (_plz, _ort))
        db.commit()

        # Migration: Aktivitäten dem geografisch zugeordneten Mitarbeiter zuweisen (idempotent)
        _falsch_n = db.execute("""
            SELECT COUNT(*) FROM aktivitaet a
            JOIN mitarbeiter_verkaufsstelle mv ON mv.verkaufsstelle_id = a.verkaufsstelle_id
            WHERE a.mitarbeiter_id != mv.mitarbeiter_id
        """).fetchone()[0]
        if _falsch_n > 0:
            db.execute("""
                UPDATE aktivitaet
                SET mitarbeiter_id = (
                    SELECT mv.mitarbeiter_id
                    FROM mitarbeiter_verkaufsstelle mv
                    WHERE mv.verkaufsstelle_id = aktivitaet.verkaufsstelle_id
                )
                WHERE EXISTS (
                    SELECT 1 FROM mitarbeiter_verkaufsstelle mv
                    WHERE mv.verkaufsstelle_id = aktivitaet.verkaufsstelle_id
                      AND mv.mitarbeiter_id != aktivitaet.mitarbeiter_id
                )
            """)
            db.commit()
            app.logger.info(f"Demo-Migration: {_falsch_n} Aktivitaeten dem richtigen Mitarbeiter zugeordnet.")

        # Migration: Aktivitäten für 02.–06. Juni 2026 nachfüllen (einmalig, KW 23 fehlt vor Sonntags-Job)
        _juni_n = db.execute(
            "SELECT COUNT(*) FROM aktivitaet WHERE datum BETWEEN '2026-06-02' AND '2026-06-06'"
        ).fetchone()[0]
        if _juni_n == 0:
            import random as _rnd_juni
            _rnd_juni.seed(20260601)
            _reps_juni = db.execute("SELECT id FROM mitarbeiter WHERE rolle='rep'").fetchall()
            _biere_juni = [r['id'] for r in db.execute("SELECT id FROM biersorte WHERE aktiv=1 AND COALESCE(ist_gratisware,0)=0").fetchall()]
            _NOTIZEN_J = ['', '', '', '',
                          'Sonderaktion vereinbart', 'Kunde sehr zufrieden',
                          'Neues Kuehlregal besprochen', 'Stammkunde, laeuft sehr gut',
                          'Termin fuer Herbstaktion vereinbart', 'Probierpaket mitgenommen']
            _WERKTAGE_J = ['2026-06-02', '2026-06-03', '2026-06-04', '2026-06-05', '2026-06-06']
            _gesamt_juni = 0
            for _rep_j in _reps_juni:
                _zugewiesen_j = db.execute("""
                    SELECT v.id FROM verkaufsstelle v
                    JOIN mitarbeiter_verkaufsstelle mv ON mv.verkaufsstelle_id = v.id
                    WHERE mv.mitarbeiter_id = ? AND v.aktiv = 1
                """, (_rep_j['id'],)).fetchall()
                if not _zugewiesen_j:
                    continue
                _tage_j   = _rnd_juni.sample(_WERKTAGE_J, k=5)
                _stellen_j = _rnd_juni.sample(list(_zugewiesen_j), k=min(5, len(_zugewiesen_j)))
                for _i_j, _datum_j in enumerate(_tage_j):
                    _vs_j    = _stellen_j[_i_j % len(_stellen_j)]
                    _disp_j  = _rnd_juni.choices([0,1,2,3,4], weights=[25,30,25,15,5])[0]
                    _cur_j   = db.execute(
                        "INSERT INTO aktivitaet (datum,mitarbeiter_id,verkaufsstelle_id,anzahl_displays,notizen) "
                        "VALUES (?,?,?,?,?)",
                        (_datum_j, _rep_j['id'], _vs_j['id'], _disp_j, _rnd_juni.choice(_NOTIZEN_J))
                    )
                    if _biere_juni and _rnd_juni.random() > 0.35:
                        db.execute(
                            "INSERT INTO bestellposition (aktivitaet_id,biersorte_id,kisten_anzahl) VALUES (?,?,?)",
                            (_cur_j.lastrowid, _rnd_juni.choice(_biere_juni), _rnd_juni.randint(1, 10))
                        )
                    _gesamt_juni += 1
            db.commit()
            app.logger.info(f"Demo-Migration: {_gesamt_juni} Aktivitaeten fuer KW23 (02.-06. Juni 2026) eingefuegt.")

        # Migration: 2025-Daten entfernen (kein weiterer Dateneingang erwartet)
        _n_2025 = db.execute(
            "SELECT COUNT(*) FROM aktivitaet WHERE strftime('%Y', datum) = '2025'"
        ).fetchone()[0]
        if _n_2025 > 0:
            db.execute("DELETE FROM aktivitaet WHERE strftime('%Y', datum) = '2025'")
            db.commit()
            app.logger.info(f"Migration: {_n_2025} Aktivitaeten aus 2025 geloescht.")

        # Migration: Überfällige Demo-Bestellungen auf max. 5 begrenzen
        _ue_ids = db.execute(
            "SELECT id FROM aktivitaet "
            "WHERE aktionstyp='Bestellung' AND COALESCE(bestell_status,'offen')='offen' "
            "AND julianday('now') - julianday(datum) > 28 "
            "ORDER BY datum ASC"
        ).fetchall()
        if len(_ue_ids) > 5:
            _close_ids = [r[0] for r in _ue_ids[:-5]]
            db.execute(
                f"UPDATE aktivitaet SET bestell_status='aufgebaut' WHERE id IN ({','.join('?'*len(_close_ids))})",
                _close_ids
            )
            db.commit()
            app.logger.info(f"Migration: {len(_close_ids)} ueberfaellige Demo-Bestellungen geschlossen.")

        # Migration: Überfällige Demo-Bestellungen sicherstellen (für Wochenbericht-Demo)
        # Wenn keine offene Bestellung >28 Tage existiert, 3 steckengebliebene einfügen
        _ue_count = db.execute(
            "SELECT COUNT(*) FROM aktivitaet "
            "WHERE aktionstyp='Bestellung' AND COALESCE(bestell_status,'offen')='offen' "
            "AND julianday('now') - julianday(datum) > 28"
        ).fetchone()[0]
        if _ue_count == 0:
            from datetime import date as _d, timedelta as _td
            _ue_reps    = db.execute("SELECT id FROM mitarbeiter WHERE rolle='rep' ORDER BY id LIMIT 3").fetchall()
            _ue_stellen = db.execute("SELECT id FROM verkaufsstelle WHERE aktiv=1 ORDER BY id LIMIT 3").fetchall()
            _ue_notizen = [
                'Herbst-Aktion – Lieferung noch ausstehend',
                'Bestellung vom Kunden noch nicht abgeholt',
                'Nachbestellung – bisher keine Rückmeldung',
            ]
            for _i, _rep in enumerate(_ue_reps):
                _ue_datum = (_d.today() - _td(days=45 - _i * 6)).isoformat()
                _ue_vs    = _ue_stellen[_i % len(_ue_stellen)]['id'] if _ue_stellen else None
                if _ue_vs:
                    db.execute(
                        "INSERT INTO aktivitaet (datum,mitarbeiter_id,verkaufsstelle_id,anzahl_displays,notizen,aktionstyp,bestell_status) "
                        "VALUES (?,?,?,0,?,'Bestellung','offen')",
                        (_ue_datum, _rep['id'], _ue_vs, _ue_notizen[_i])
                    )
            db.commit()
            app.logger.info("Migration: 3 ueberfaellige Demo-Bestellungen eingefuegt.")

        # Migration: Straßenadressen für Demo-Verkaufsstellen eintragen (name-basiert)
        _vs_adressen = [
            ('Supermarkt Mitte',           'Alexanderplatz 1'),
            ('Fachmarkt Nord',             'Moekenbeergstr. 7'),
            ('Restaurant Zur Post',        'Marienplatz 8'),
            ('Hotel Stadtblick',           'Zeil 15'),
            ('Großhandel Meyer',           'Schildergasse 22'),
            ('Kiosk am Bahnhof',           'Bahnhofstrasse 3'),
            ('Sportverein 1902',           'Schillerplatz 5'),
            ('Café Central',               'Augustusplatz 9'),
            ('Restaurant Zum Marktplatz',  'Hauptmarkt 14'),
            ('Bistro Central',             'Kroepke 6'),
            ('Ristorante Bella Vista',     'Wasserturmplatz 4'),
            ('Steakhouse Westend',         'Bockenheimer Landstr. 18'),
            ('Café Metropol',              'Westenhellweg 12'),
            ('Pizzeria Napoli',            'Am Markt 9'),
            ('Imbiss Am Stadtpark',        'Ruettenscheider Str. 3'),
            ('Gasthaus Lindenhof',         'Wilhelmstrasse 11'),
            ('Stadthotel am Ring',         'Muensterplatz 2'),
            ('Pension Garni Sonnenhof',    'Muensterplatz 7'),
            ('Supermarkt Stadtmitte',      'Koenigstrasse 26'),
            ('Verbrauchermarkt Nord',      'List 5'),
            ('Discountmarkt Westend',      'Planken 8'),
            ('Großhandel Fischer',         'Grossmarkthalle 3'),
            ('Cash & Carry Zentrum',       'Unionstrasse 11'),
            ('Handelskontor Weber',        'Schlachte 17'),
            ('Sportverein Blau-Weiß',      'Vereinsweg 4'),
            ('Schützengesellschaft 1888',  'Schuetzenstrasse 6'),
            ('TSG Vereinsheim',            'Vereinsstrasse 12'),
            ('Stadionkiosk SV Mitte',      'Schwarzwaldstrasse 20'),
        ]
        _updated = 0
        for _vs_name, _strasse in _vs_adressen:
            _r = db.execute(
                "UPDATE verkaufsstelle SET strasse=? WHERE name=? AND (strasse IS NULL OR strasse='')",
                (_strasse, _vs_name)
            ).rowcount
            _updated += _r
        if _updated:
            db.commit()
            app.logger.info(f"Migration: {_updated} Verkaufsstellen-Adressen eingetragen.")

        # Migration: Demo-Konten zusammenführen → ein "Demo Leitung" (kuerzel='Demo')
        _old_dl = db.execute(
            "SELECT id FROM mitarbeiter WHERE name='Demo Leitung' AND kuerzel='DL'"
        ).fetchone()
        if _old_dl:
            db.execute("DELETE FROM aktivitaet WHERE mitarbeiter_id=?", (_old_dl[0],))
            db.execute("DELETE FROM mitarbeiter_verkaufsstelle WHERE mitarbeiter_id=?", (_old_dl[0],))
            db.execute("DELETE FROM zielzahlen WHERE mitarbeiter_id=?", (_old_dl[0],))
            db.execute("DELETE FROM mitarbeiter WHERE id=?", (_old_dl[0],))
            db.commit()
            app.logger.info("Migration: Altes Demo Leitung (DL) entfernt.")
        _demo_zugang = db.execute(
            "SELECT id FROM mitarbeiter WHERE name='Demo-Zugang'"
        ).fetchone()
        if _demo_zugang:
            db.execute("UPDATE mitarbeiter SET name='Demo Leitung' WHERE id=?", (_demo_zugang[0],))
            db.commit()
            app.logger.info("Migration: Demo-Zugang zu 'Demo Leitung' umbenannt.")

        # Alte Fotos beim Start bereinigen
        cleanup_alte_fotos()

        # Tägliches DB-Backup
        backup_db()


def seed_demo_data(db):
    """Füllt die DB mit realistischen Beispieldaten für KW 01–22, 2026."""
    import random as rnd
    from datetime import date, timedelta
    rnd.seed(42)

    # Zusätzliche Verkaufsstellen
    extra = [
        ('Restaurant Zum Marktplatz',    'Nürnberg',          'Gastronomie'),
        ('Bistro Central',               'Hannover',          'Gastronomie'),
        ('Ristorante Bella Vista',       'Mannheim',          'Gastronomie'),
        ('Steakhouse Westend',           'Frankfurt',         'Gastronomie'),
        ('Café Metropol',                'Dortmund',          'Gastronomie'),
        ('Pizzeria Napoli',              'Bremen',            'Gastronomie'),
        ('Imbiss Am Stadtpark',          'Essen',             'Gastronomie'),
        ('Gasthaus Lindenhof',           'Wiesbaden',         'Gastronomie'),
        ('Stadthotel am Ring',           'Bonn',              'Hotel'),
        ('Pension Garni Sonnenhof',      'Freiburg',          'Hotel'),
        ('Supermarkt Stadtmitte',        'Nürnberg',          'Einzelhandel'),
        ('Verbrauchermarkt Nord',        'Hannover',          'Einzelhandel'),
        ('Discountmarkt Westend',        'Mannheim',          'Einzelhandel'),
        ('Großhandel Fischer',           'Frankfurt',         'Getränkehandel'),
        ('Cash & Carry Zentrum',         'Dortmund',          'Getränkehandel'),
        ('Handelskontor Weber',          'Bremen',            'Getränkehandel'),
        ('Sportverein Blau-Weiß',        'Essen',             'Verein'),
        ('Schützengesellschaft 1888',    'Wiesbaden',         'Verein'),
        ('TSG Vereinsheim',              'Bonn',              'Verein'),
        ('Stadionkiosk SV Mitte',        'Freiburg',          'Verein'),
    ]
    for name, ort, typ in extra:
        if not db.execute("SELECT 1 FROM verkaufsstelle WHERE name=?", (name,)).fetchone():
            db.execute("INSERT INTO verkaufsstelle (name,ort,typ) VALUES (?,?,?)", (name, ort, typ))

    reps    = db.execute("SELECT id, kuerzel FROM mitarbeiter WHERE rolle='rep'").fetchall()
    stellen = db.execute("SELECT id, typ FROM verkaufsstelle WHERE aktiv=1").fetchall()
    biere   = db.execute("SELECT id, name FROM biersorte WHERE aktiv=1").fetchall()
    bier_by = {b['name']: b['id'] for b in biere}

    PREF = {
        'Gastronomie':    ['Produkt A', 'Produkt B', 'Produkt C'],
        'Einzelhandel':   ['Produkt A', 'Produkt B', 'Produkt C', 'Produkt D'],
        'Getränkehandel': ['Produkt A', 'Produkt B', 'Produkt C', 'Produkt D', 'Produkt E', 'Produkt F'],
        'Hotel':          ['Produkt A', 'Produkt B', 'Produkt C'],
        'Verein':         ['Produkt A', 'Produkt C', 'Produkt D'],
        'Kiosk':          ['Produkt A', 'Produkt B'],
    }
    NOTIZEN = [
        '', '', '', '',
        'Sonderaktion vereinbart', 'Kunde sehr zufrieden',
        'Neues Kühlregal besprochen', 'Probierpaket mitgenommen',
        'Konkurrenzprodukte gesichtet', 'Rückgabe 3 leere Displays',
        'Termin für Herbstaktion vereinbart', 'Preiserhöhung kommuniziert',
        'Stammkunde, läuft sehr gut', 'Beschwerden über Lieferzeit',
    ]

    def positionen(typ):
        namen = PREF.get(typ, [b['name'] for b in biere])
        auswahl = rnd.sample(namen, k=rnd.randint(2, min(5, len(namen))))
        return [(bier_by[n], rnd.randint(3, 50)) for n in auswahl if n in bier_by]

    for kw in range(1, 23):
        monday = date.fromisocalendar(2026, kw, 1)
        for rep in reps:
            for tag in sorted(rnd.sample(range(5), k=rnd.randint(2, 4))):
                vs = rnd.choice(stellen)
                displays = rnd.choices([0,1,2,3,4,5], weights=[30,25,20,12,8,5])[0]
                cur = db.execute(
                    "INSERT INTO aktivitaet (datum,mitarbeiter_id,verkaufsstelle_id,anzahl_displays,notizen) "
                    "VALUES (?,?,?,?,?)",
                    ((monday + timedelta(days=tag)).isoformat(),
                     rep['id'], vs['id'], displays, rnd.choice(NOTIZEN))
                )
                for bier_id, menge in positionen(vs['typ']):
                    db.execute(
                        "INSERT INTO bestellposition (aktivitaet_id,biersorte_id,kisten_anzahl) VALUES (?,?,?)",
                        (cur.lastrowid, bier_id, menge)
                    )

    # Zielzahlen 2026 – realistisches Bild: Stern / Grün / Gelb / Rot / Sehr Rot
    # Berechnet auf Basis der IST-Werte per Ende Juni (SONNENSCHLÜSSEL Jan-Jun = 47 %)
    # MM: ~120 % Saisonstatus | TW: ~104 % | AS: ~80 % | LF: ~45 % | KH: 0 %
    ZIELE = {'MM':(80,2900),'AS':(120,4600),'TW':(80,2600),'LF':(90,3900),'KH':(80,2500)}
    for rep in reps:
        if rep['kuerzel'] in ZIELE:
            d, k = ZIELE[rep['kuerzel']]
            db.execute('''INSERT INTO zielzahlen (mitarbeiter_id,jahr,displays_ziel,kisten_ziel)
                VALUES (?,2026,?,?) ON CONFLICT(mitarbeiter_id,jahr) DO UPDATE SET
                displays_ziel=excluded.displays_ziel, kisten_ziel=excluded.kisten_ziel''',
                (rep['id'], d, k))
    db.execute('''INSERT INTO zielzahlen (mitarbeiter_id,jahr,displays_ziel,kisten_ziel)
        VALUES (NULL,2026,450,16500) ON CONFLICT(mitarbeiter_id,jahr) DO UPDATE SET
        displays_ziel=excluded.displays_ziel, kisten_ziel=excluded.kisten_ziel''')

    # Verkaufsstellen gleichmäßig auf Reps verteilen (für Karten-Demo)
    alle_stellen = db.execute("SELECT id FROM verkaufsstelle WHERE aktiv=1 ORDER BY id").fetchall()
    stellen_ids  = [s['id'] for s in alle_stellen]
    rnd.shuffle(stellen_ids)
    for idx, stelle_id in enumerate(stellen_ids):
        rep = reps[idx % len(reps)]
        db.execute(
            "INSERT OR IGNORE INTO mitarbeiter_verkaufsstelle (mitarbeiter_id, verkaufsstelle_id) VALUES (?,?)",
            (rep['id'], stelle_id)
        )

    # Fest verdrahtete "steckengebliebene" Bestellungen für Überfällig-Demo
    # Idempotent: nur einfügen wenn noch keine solche Bestellung von diesem Rep an diesem Datum
    _ue_stellen = db.execute("SELECT id FROM verkaufsstelle WHERE aktiv=1 ORDER BY id LIMIT 3").fetchall()
    _ue_notizen = [
        'Herbst-Aktion – Lieferung noch ausstehend',
        'Bestellung vom Kunden noch nicht abgeholt',
        'Nachbestellung – bisher keine Rückmeldung',
    ]
    for i, rep in enumerate(reps[:3]):
        _tage_alt = 45 - i * 6  # 45, 39, 33 Tage alt
        _ue_datum = (date.today() - timedelta(days=_tage_alt)).isoformat()
        _ue_vs    = _ue_stellen[i % len(_ue_stellen)]['id']
        already = db.execute(
            "SELECT 1 FROM aktivitaet WHERE mitarbeiter_id=? AND datum=? AND aktionstyp='Bestellung' AND bestell_status='offen'",
            (rep['id'], _ue_datum)
        ).fetchone()
        if not already:
            db.execute(
                "INSERT INTO aktivitaet (datum,mitarbeiter_id,verkaufsstelle_id,anzahl_displays,notizen,aktionstyp,bestell_status) "
                "VALUES (?,?,?,0,?,'Bestellung','offen')",
                (_ue_datum, rep['id'], _ue_vs, _ue_notizen[i])
            )

    db.commit()


def seed_demo_data_relativ(db):
    """Seeded mit Daten der letzten 6 Wochen (relativ zu heute). Ersetzt seed_demo_data beim Daily-Reset."""
    import random as rnd
    from datetime import date, timedelta

    today = date.today()
    montag_aktuell = today - timedelta(days=today.weekday())
    montag_start   = montag_aktuell - timedelta(weeks=5)  # 6 Wochen inkl. laufender Woche

    rnd.seed(42)

    reps    = db.execute("SELECT id, kuerzel FROM mitarbeiter WHERE rolle='rep' AND aktiv=1").fetchall()
    stellen = db.execute("SELECT id, typ FROM verkaufsstelle WHERE aktiv=1").fetchall()
    biere   = db.execute("SELECT id FROM biersorte WHERE aktiv=1 AND COALESCE(ist_gratisware,0)=0").fetchall()
    bier_ids = [b['id'] for b in biere]

    # Unterschiedliche Performance-Profile für Ranking-Demo
    PROFIL = {'MM': 12, 'AS': 10, 'TW': 9, 'LF': 7}

    NOTIZEN = [
        '', '', '', '', 'Sonderaktion vereinbart', 'Kunde sehr zufrieden',
        'Neues Kühlregal besprochen', 'Probierpaket mitgenommen',
        'Konkurrenzprodukte gesichtet', 'Rückgabe 3 leere Displays',
        'Termin für Herbstaktion vereinbart', 'Stammkunde, läuft sehr gut',
        'Bestellung für nächste Lieferung', 'Neues Sortiment vorgestellt',
        'Kein Bedarf aktuell, Wiedervorlage in 2 Wochen', 'Feedback eingeholt – positiv',
    ]

    for kw_offset in range(6):
        montag  = montag_start + timedelta(weeks=kw_offset)
        freitag = montag + timedelta(days=4)
        ende    = min(freitag, today - timedelta(days=1))
        if ende < montag:
            continue  # Aktuelle Woche noch nicht gestartet

        verf_tage = (ende - montag).days + 1

        for rep in reps:
            n_week = PROFIL.get(rep['kuerzel'], 9)
            if kw_offset == 5:  # Laufende Woche anteilig kürzen
                n_week = max(1, round(n_week * verf_tage / 5))

            # Mix: 45% Aufbau, 35% Bestellung, 20% Besuch
            n_aufbau = max(1, round(n_week * 0.45))
            n_best   = max(1, round(n_week * 0.35))
            n_besuch = max(0, n_week - n_aufbau - n_best)
            typen = ['Aufbau'] * n_aufbau + ['Bestellung'] * n_best + ['Besuch'] * n_besuch
            rnd.shuffle(typen)

            zugewiesen = db.execute("""
                SELECT v.id, v.typ FROM verkaufsstelle v
                JOIN mitarbeiter_verkaufsstelle mv ON mv.verkaufsstelle_id = v.id
                WHERE mv.mitarbeiter_id = ? AND v.aktiv = 1
            """, (rep['id'],)).fetchall()
            rep_stellen = list(zugewiesen) if len(zugewiesen) >= 2 else list(stellen)

            tage = sorted(rnd.choices(range(min(5, verf_tage)), k=len(typen)))

            for i, (tag, typ) in enumerate(zip(tage, typen)):
                if montag + timedelta(days=tag) > ende:
                    continue
                datum    = (montag + timedelta(days=tag)).isoformat()
                vs       = rnd.choice(rep_stellen)
                displays = rnd.choices([0,1,2,3,4], weights=[25,25,25,15,10])[0] if typ == 'Aufbau' else 0
                bestell_status = 'offen' if typ == 'Bestellung' else None
                cur = db.execute(
                    "INSERT INTO aktivitaet "
                    "(datum,mitarbeiter_id,verkaufsstelle_id,anzahl_displays,notizen,aktionstyp,bestell_status) "
                    "VALUES (?,?,?,?,?,?,?)",
                    (datum, rep['id'], vs['id'], displays, rnd.choice(NOTIZEN), typ, bestell_status)
                )
                if typ == 'Bestellung':
                    for bid in rnd.sample(bier_ids, k=min(rnd.randint(2, 4), len(bier_ids))):
                        db.execute(
                            "INSERT INTO bestellposition (aktivitaet_id,biersorte_id,kisten_anzahl) VALUES (?,?,?)",
                            (cur.lastrowid, bid, rnd.randint(5, 45))
                        )

    # Zielzahlen – realistisches Bild: Stern / Grün / Gelb / Rot / Sehr Rot
    ZIELE = {'MM': (80, 2900), 'AS': (120, 4600), 'TW': (80, 2600), 'LF': (90, 3900), 'KH': (80, 2500)}
    for rep in reps:
        if rep['kuerzel'] in ZIELE:
            d, k = ZIELE[rep['kuerzel']]
            db.execute('''INSERT INTO zielzahlen (mitarbeiter_id,jahr,displays_ziel,kisten_ziel)
                VALUES (?,?,?,?) ON CONFLICT(mitarbeiter_id,jahr) DO UPDATE SET
                displays_ziel=excluded.displays_ziel, kisten_ziel=excluded.kisten_ziel''',
                (rep['id'], today.year, d, k))
    db.execute('''INSERT INTO zielzahlen (mitarbeiter_id,jahr,displays_ziel,kisten_ziel)
        VALUES (NULL,?,450,16500) ON CONFLICT(mitarbeiter_id,jahr) DO UPDATE SET
        displays_ziel=excluded.displays_ziel, kisten_ziel=excluded.kisten_ziel''',
        (today.year,))

    # Verkaufsstellen auf Reps verteilen – nur beim ersten Mal
    if not db.execute("SELECT 1 FROM mitarbeiter_verkaufsstelle LIMIT 1").fetchone():
        alle_stellen = [s['id'] for s in db.execute("SELECT id FROM verkaufsstelle WHERE aktiv=1 ORDER BY id").fetchall()]
        rnd.shuffle(alle_stellen)
        for idx, sid in enumerate(alle_stellen):
            rep = reps[idx % len(reps)]
            db.execute(
                "INSERT OR IGNORE INTO mitarbeiter_verkaufsstelle (mitarbeiter_id, verkaufsstelle_id) VALUES (?,?)",
                (rep['id'], sid)
            )

    db.commit()


def seed_demo_besuchsplan(db):
    """Füllt Besuchsplan (tagesplan) für heute + die nächsten 3 Werktage je Rep."""
    import random as rnd
    from datetime import date, timedelta

    today = date.today()
    rnd.seed(today.toordinal())

    reps = db.execute("SELECT id FROM mitarbeiter WHERE rolle='rep' AND aktiv=1").fetchall()

    # Nächste 4 Werktage (inkl. heute)
    tage, d = [], today
    while len(tage) < 4:
        if d.weekday() < 5:
            tage.append(d)
        d += timedelta(days=1)

    for rep in reps:
        stellen = db.execute("""
            SELECT v.id FROM verkaufsstelle v
            JOIN mitarbeiter_verkaufsstelle mv ON mv.verkaufsstelle_id = v.id
            WHERE mv.mitarbeiter_id = ? AND v.aktiv = 1
        """, (rep['id'],)).fetchall()
        if not stellen:
            stellen = db.execute("SELECT id FROM verkaufsstelle WHERE aktiv=1 LIMIT 12").fetchall()
        ids = [s['id'] for s in stellen]

        for tag in tage:
            n_stops = rnd.randint(2, 3)
            gewaehlte = rnd.sample(ids, k=min(n_stops, len(ids)))
            for reihenfolge, vs_id in enumerate(gewaehlte, start=1):
                db.execute(
                    "INSERT INTO tagesplan (mitarbeiter_id, verkaufsstelle_id, datum, reihenfolge) VALUES (?,?,?,?)",
                    (rep['id'], vs_id, tag.isoformat(), reihenfolge)
                )

    db.commit()


def _do_demo_daily_reset():
    """Täglicher Reset um 03:00: alle Aktivitäts- und Plandaten löschen und neu seeden."""
    if not DEMO_MODUS:
        return
    db = get_db()
    try:
        db.execute("DELETE FROM bestellposition")
        db.execute("DELETE FROM aktivitaet")
        db.execute("DELETE FROM tagesplan")
        db.execute("DELETE FROM zielzahlen")
        db.commit()
        seed_demo_data_relativ(db)
        seed_demo_besuchsplan(db)
        _demo_tagesplan_fortschritt()
        app.logger.info("Demo-Daily-Reset abgeschlossen.")
    except Exception as e:
        app.logger.error(f"Demo-Reset Fehler: {e}", exc_info=True)


def demo_daily_reset():
    """Wrapper für APScheduler."""
    with app.app_context():
        _do_demo_daily_reset()


def _demo_tagesplan_fortschritt():
    """Markiert heutige Tagesplan-Einträge als erledigt — lässt je Rep 1–2 zufällig offen.
    Seed ist tagesabhängig: welche Einträge offen bleiben, wechselt täglich und variiert pro Rep."""
    if not DEMO_MODUS:
        return
    import random as rnd
    from datetime import date
    today = date.today()
    rnd.seed(today.toordinal())  # deterministisch pro Tag, aber täglich anders

    db  = get_db()
    reps = db.execute("SELECT id FROM mitarbeiter WHERE rolle='rep' AND aktiv=1").fetchall()

    for rep in reps:
        eintraege = db.execute(
            "SELECT id FROM tagesplan WHERE mitarbeiter_id=? AND datum=? AND COALESCE(geloescht,0)=0 ORDER BY reihenfolge",
            (rep['id'], today.isoformat())
        ).fetchall()
        if not eintraege:
            continue
        # 1–2 Einträge zufällig offen lassen (variiert pro Rep und Tag)
        n_offen  = rnd.randint(1, min(2, len(eintraege)))
        offen_ids = {row['id'] for row in rnd.sample(eintraege, k=n_offen)}
        for row in eintraege:
            if row['id'] not in offen_ids:
                db.execute("UPDATE tagesplan SET erledigt=1 WHERE id=? AND erledigt=0", (row['id'],))

    db.commit()


def demo_tagesplan_fortschritt():
    """Wrapper für APScheduler."""
    with app.app_context():
        _demo_tagesplan_fortschritt()


# ─── Auth ─────────────────────────────────────────────────────────────────────

def login_required(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        if session.get('muss_passwort_aendern'):
            return redirect(url_for('erstes_passwort'))
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        # Bugreport 2026-07-21 (Mittel): muss_passwort_aendern wurde bisher nur von
        # login_required erzwungen.
        if session.get('muss_passwort_aendern'):
            return redirect(url_for('erstes_passwort'))
        if session.get('rolle') != 'admin':
            flash('Zugriff verweigert – nur für die Leitung.', 'danger')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated

def manager_required(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        if session.get('muss_passwort_aendern'):
            return redirect(url_for('erstes_passwort'))
        if session.get('rolle') not in ('admin', 'verkaufsleiter'):
            flash('Zugriff verweigert – nur für Leitung und Verkaufsleiter.', 'danger')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated


# ─── Team-Filter-Hilfsfunktionen ──────────────────────────────────────────────

def _team_ma_clause(alias='a'):
    """Gibt (sql_fragment, params) zurück um Aktivitäts-Queries auf das VKL-Team einzugrenzen.
    Nur aktiv wenn VKL eingeloggt ist UND ein Team zugeordnet hat. Admin und Rep: kein Filter.
    alias: Tabellen-Alias der aktivitaet-Tabelle in der Query."""
    if session.get('rolle') == 'verkaufsleiter':
        tid = session.get('team_id')
        if tid:
            return (
                f' AND {alias}.mitarbeiter_id IN (SELECT id FROM mitarbeiter WHERE team_id = ?)',
                (tid,)
            )
    return '', ()

def _team_m_clause(alias='m'):
    """Gibt (sql_fragment, params) zurück um mitarbeiter-Queries auf das VKL-Team einzugrenzen.
    alias: Tabellen-Alias der mitarbeiter-Tabelle."""
    if session.get('rolle') == 'verkaufsleiter':
        tid = session.get('team_id')
        if tid:
            return f' AND {alias}.team_id = ?', (tid,)
    return '', ()


def _freigabe_scope_clause(alias='a'):
    """Gibt (sql_fragment, params) für die Aufbauten-Freigabe zurück. Ein VKL mit
    Team-Zuordnung sieht sein Team plus die eigenen Aktivitäten (VKL mit eigenem Gebiet,
    der selbst Aktivitäten für seine zugewiesenen Verkaufsstellen erfasst). Ein VKL OHNE
    Team-Zuordnung gilt als gleichberechtigter Freigeber für alle VKLs ohne Team (z.B.
    mehrere VKLs mit eigenem Gebiet, die sich gegenseitig vertreten/unterstützen) – sieht
    daher wie Admin alle offenen Freigaben, nicht nur die eigenen. Admin: kein Filter."""
    if session.get('rolle') == 'verkaufsleiter':
        tid = session.get('team_id')
        if tid:
            uid = session['user_id']
            return (
                f' AND ({alias}.mitarbeiter_id IN (SELECT id FROM mitarbeiter WHERE team_id = ?) OR {alias}.mitarbeiter_id = ?)',
                (tid, uid)
            )
    return '', ()


def _urlaub_daten(ma_ids, start_iso, end_iso):
    """Dict {(mitarbeiter_id, datum): {'typ':..., 'grund':...}} für alle Tage im Bereich
    [start_iso, end_iso], an denen der jeweilige Mitarbeiter laut bestätigter Abwesenheit
    (vertretung) abwesend ist. Verhält sich für Mitgliedschaftsprüfungen ('in urlaub_tag')
    wie die frühere Set-Variante, liefert zusätzlich typ/grund – genutzt, um in der
    Besuchsplanung „Urlaub" bzw. „Frei/Sonderurlaub" (inkl. Grund) statt „Kein Plan" anzuzeigen."""
    if not ma_ids:
        return {}
    ph = ','.join('?' * len(ma_ids))
    rows = query(
        f"SELECT abwesender_id, von, bis, typ, grund FROM vertretung "
        f"WHERE status='bestätigt' AND abwesender_id IN ({ph}) AND von <= ? AND bis >= ?",
        tuple(ma_ids) + (end_iso, start_iso)
    )
    start_d = date.fromisoformat(start_iso)
    end_d   = date.fromisoformat(end_iso)
    ergebnis = {}
    for r in rows:
        d = max(date.fromisoformat(r['von']), start_d)
        bis = min(date.fromisoformat(r['bis']), end_d)
        while d <= bis:
            ergebnis[(r['abwesender_id'], d.isoformat())] = {'typ': r['typ'], 'grund': r['grund']}
            d += timedelta(days=1)
    return ergebnis


_BUNDESLAENDER = {
    'BW': 'Baden-Württemberg', 'BY': 'Bayern', 'BE': 'Berlin', 'BB': 'Brandenburg',
    'HB': 'Bremen', 'HH': 'Hamburg', 'HE': 'Hessen', 'MV': 'Mecklenburg-Vorpommern',
    'NI': 'Niedersachsen', 'NW': 'Nordrhein-Westfalen', 'RP': 'Rheinland-Pfalz',
    'SL': 'Saarland', 'SN': 'Sachsen', 'ST': 'Sachsen-Anhalt', 'SH': 'Schleswig-Holstein',
    'TH': 'Thüringen',
}


@functools.lru_cache(maxsize=256)
def _feiertage_set(jahr, bundesland):
    """Gesetzliche Feiertage eines Bundeslands/Jahres als Set von 'YYYY-MM-DD'-Strings.
    Gecacht, da bei jedem Seitenaufruf (context_processor → _urlaub_konto) neu berechnet
    würde, obwohl sich Feiertage für ein gegebenes (Jahr, Bundesland)-Paar nie ändern."""
    try:
        return {d.isoformat() for d in _holidays_lib.Germany(subdiv=bundesland or 'BY', years=jahr)}
    except Exception:
        return set()


def _werktage_zaehlen(von_iso, bis_iso, bundesland):
    """Zählt Werktage (Mo–Fr) zwischen von/bis (inklusive), abzüglich gesetzlicher
    Feiertage des angegebenen Bundeslands. Grundlage für die Urlaubskonto-Berechnung."""
    von = date.fromisoformat(von_iso)
    bis = date.fromisoformat(bis_iso)
    if bis < von:
        return 0
    feiertage = set()
    for jahr in range(von.year, bis.year + 1):
        feiertage |= _feiertage_set(jahr, bundesland)
    n = 0
    d = von
    while d <= bis:
        if d.weekday() < 5 and d.isoformat() not in feiertage:
            n += 1
        d += timedelta(days=1)
    return n


def _urlaub_konto(mitarbeiter_id, jahr):
    """Urlaubskonto eines Mitarbeiters für ein Kalenderjahr: Anspruch + Übertrag aus dem
    Vorjahr, abzüglich bereits genehmigter Urlaubstage (nur typ='urlaub', status='bestätigt').
    Beantragte-aber-noch-nicht-genehmigte Tage werden separat ausgewiesen (reservieren das
    Kontingent noch nicht, damit eine Ablehnung nichts zurückbuchen muss)."""
    ma = query(
        "SELECT urlaubsanspruch_jahr, urlaub_uebertrag_vorjahr, bundesland, urlaub_manuell_genommen FROM mitarbeiter WHERE id=?",
        (mitarbeiter_id,), one=True
    )
    if not ma:
        return None
    anspruch    = ma['urlaubsanspruch_jahr'] if ma['urlaubsanspruch_jahr'] is not None else 30
    uebertrag   = ma['urlaub_uebertrag_vorjahr'] or 0
    bundesland  = ma['bundesland'] or 'BY'
    jahr_start  = f"{jahr}-01-01"
    jahr_ende   = f"{jahr}-12-31"
    rows = query(
        "SELECT von, bis, status FROM vertretung WHERE abwesender_id=? AND typ='urlaub' "
        "AND von <= ? AND bis >= ?",
        (mitarbeiter_id, jahr_ende, jahr_start)
    )
    # Manuell erfasster Alturlaub gilt nur fürs laufende Jahr – die Spalte ist nicht
    # jahresgebunden, da sie ausschließlich zum einmaligen Nachtragen von vor Tool-Einführung
    # bereits genommenem Urlaub des aktuellen Jahres gedacht ist.
    genommen  = (ma['urlaub_manuell_genommen'] or 0) if jahr == date.today().year else 0
    beantragt = 0
    for r in rows:
        von = max(r['von'], jahr_start)
        bis = min(r['bis'], jahr_ende)
        tage = _werktage_zaehlen(von, bis, bundesland)
        if r['status'] == 'bestätigt':
            genommen += tage
        elif r['status'] == 'angefragt':
            beantragt += tage
    return {
        'anspruch':   anspruch,
        'uebertrag':  uebertrag,
        'genommen':   genommen,
        'beantragt':  beantragt,
        'verfuegbar': anspruch + uebertrag - genommen,
    }


def _urlaubsantrag_pdf_bytes(vtr_id: int) -> bytes | None:
    """Erzeugt einen Urlaubsantrag als PDF, befüllt mit den Daten aus dem Urlaubskonto.
    Nur für typ='urlaub' gedacht – andere Vertretungs-Typen (Krankheit, Sonderurlaub,
    Frei/Sonderurlaub, unbezahlt) haben kein PDF-Pendant und werden hier bewusst nicht
    bedient (kein Papierformular, keine Lohnbüro-Relevanz). Ohne Firmen-Briefkopf, da
    Demo mandantenunabhängig ist. Schlägt nie hart fehl, gibt bei fehlenden Daten nur
    None zurück (Aufrufer verschickt die Mail dann ohne Anhang). Trägt seit 2026-07-21 die
    einmalig hinterlegten Signaturen von Mitarbeiter und Genehmiger ein, sofern beide ihre
    Unterschrift im Profil hinterlegt haben (siehe profil_signatur) – fehlt eine davon, bleibt
    der jeweilige Bereich wie bisher unterschriftslos, aber weiterhin digital gültig."""
    vtr = query(
        '''SELECT v.abwesender_id, v.von, v.bis, v.typ, v.erstellt_am, v.bestaetigt_von_id,
                  m.name AS mitarbeiter_name, m.wohnort_ort, m.bundesland, m.signatur_bild AS ma_signatur,
                  g.name AS genehmiger_name, g.signatur_bild AS genehmiger_signatur
           FROM vertretung v JOIN mitarbeiter m ON m.id = v.abwesender_id
           LEFT JOIN mitarbeiter g ON g.id = v.bestaetigt_von_id
           WHERE v.id=?''',
        (vtr_id,), one=True
    )
    if not vtr or vtr['typ'] != 'urlaub':
        return None
    konto = _urlaub_konto(vtr['abwesender_id'], date.fromisoformat(vtr['von']).year)
    if not konto:
        return None
    tage = _werktage_zaehlen(vtr['von'], vtr['bis'], vtr['bundesland'] or 'BY')
    # v.status ist zu diesem Zeitpunkt bereits auf 'bestätigt' gesetzt (Aufrufer aktualisiert
    # vor dem Mail-/PDF-Versand) – konto['verfuegbar'] spiegelt also schon den Saldo NACH
    # Abzug dieses Antrags. "Anspruch zum Zeitpunkt des Antrags" ist der Saldo VOR Abzug.
    anspruch_zum_antrag = konto['verfuegbar'] + tage
    verbleibender_resturlaub = konto['verfuegbar']

    def _fmt(iso_datum):
        return date.fromisoformat(iso_datum).strftime('%d.%m.%Y')

    antrags_datum = _fmt(vtr['erstellt_am']) if vtr['erstellt_am'] else _fmt(date.today().isoformat())
    heute         = date.today().strftime('%d.%m.%Y')
    wohnort       = vtr['wohnort_ort'] or '–'

    def _signatur_html(data_uri):
        if not data_uri or not data_uri.startswith('data:image/png;base64,'):
            return ''
        return f'<img src="{data_uri}" style="height:44px;display:block;margin-top:4px">'

    ma_signatur_html         = _signatur_html(vtr['ma_signatur'])
    genehmiger_signatur_html = _signatur_html(vtr['genehmiger_signatur'])

    html = f'''<!DOCTYPE html>
<html><head><meta charset="utf-8"></head>
<body style="margin:0;padding:0;font-family:Arial,sans-serif;color:#222">
<div style="max-width:680px;margin:0 auto;padding:30px 36px">
  <div style="font-size:15px;font-weight:bold;color:#1a3a5c;margin:0 0 18px">{COMPANY_NAME}</div>

  <div style="font-size:22px;font-weight:bold;color:#1a3a5c;margin:0 0 18px;border-bottom:2px solid #1a3a5c;padding-bottom:6px">Urlaubsantrag</div>

  <table width="100%" cellpadding="0" cellspacing="0" style="font-size:13px">
    <tr><td style="padding:7px 0;color:#555;width:58%">Name des Mitarbeiters</td>
        <td style="padding:7px 0;font-weight:bold">{vtr['mitarbeiter_name']}</td></tr>
    <tr><td style="padding:7px 0;color:#555;border-top:1px solid #eee">Resturlaub Vorjahr</td>
        <td style="padding:7px 0;border-top:1px solid #eee">{konto['uebertrag']} Tage</td></tr>
    <tr><td style="padding:7px 0;color:#555;border-top:1px solid #eee">Urlaubsanspruch laufendes Jahr</td>
        <td style="padding:7px 0;border-top:1px solid #eee">{konto['anspruch']} Tage</td></tr>
    <tr><td style="padding:7px 0;color:#555;border-top:1px solid #eee">Anspruch zum Zeitpunkt des Antrags</td>
        <td style="padding:7px 0;border-top:1px solid #eee">{anspruch_zum_antrag} Tage</td></tr>
    <tr><td style="padding:7px 0;color:#555;border-top:1px solid #eee">Beantragter Urlaub</td>
        <td style="padding:7px 0;border-top:1px solid #eee">{_fmt(vtr['von'])} bis {_fmt(vtr['bis'])} &nbsp;({tage} Tage)</td></tr>
    <tr><td style="padding:7px 0;color:#555;border-top:1px solid #eee;font-weight:bold">Verbleibender Resturlaub</td>
        <td style="padding:7px 0;border-top:1px solid #eee;font-weight:bold">{verbleibender_resturlaub} Tage</td></tr>
  </table>

  <div style="margin-top:28px;font-size:12px;color:#555">{wohnort}, {antrags_datum}</div>
  {ma_signatur_html}
  <div style="font-size:10px;color:#999;margin-top:2px">Ort, Datum, Unterschrift (Antragstellung)</div>

  <div style="margin-top:26px;padding-top:16px;border-top:1px solid #ccc;font-size:13px">
    <span style="color:#2d8a4e;font-weight:bold">&#10003;</span> Der Antrag auf Urlaub wird befürwortet / genehmigt.
  </div>

  <div style="margin-top:22px;font-size:12px;color:#555">{heute}</div>
  {genehmiger_signatur_html}
  <div style="font-size:10px;color:#999;margin-top:2px">Datum, Unterschrift (Genehmigung{', ' + vtr['genehmiger_name'] if vtr['genehmiger_name'] else ''})</div>

  <div style="margin-top:32px;padding-top:10px;border-top:1px solid #eee;font-size:10px;color:#999">
    Automatisch erzeugt vom Aktions Tracker am {heute} &middot; digital genehmigt{' mit hinterlegter Unterschrift' if (ma_signatur_html or genehmiger_signatur_html) else ', ohne Unterschrift gültig'}.
  </div>
</div>
</body></html>'''
    return _html_to_pdf(html)


# ─── PWA Manifest (dynamisch mit COMPANY_NAME aus ENV) ───────────────────────

@app.route('/manifest.json')
def manifest():
    from flask import jsonify
    data = {
        "name": f"Aktions Tracker – {COMPANY_NAME}",
        "short_name": "Aktions Tracker",
        "description": f"Außendienst-Aktivitäten und Bestellungen – {COMPANY_NAME}",
        "start_url": "/dashboard",
        "scope": "/",
        "display": "standalone",
        "background_color": "#1a3a5c",
        "theme_color": "#1a3a5c",
        "orientation": "portrait-primary",
        "icons": [
            {"src": "/static/icon-192.png", "sizes": "192x192", "type": "image/png", "purpose": "any"},
            {"src": "/static/icon-192.png", "sizes": "192x192", "type": "image/png", "purpose": "maskable"},
            {"src": "/static/icon-512.png", "sizes": "512x512", "type": "image/png", "purpose": "any"},
            {"src": "/static/icon-512.png", "sizes": "512x512", "type": "image/png", "purpose": "maskable"},
        ]
    }
    resp = jsonify(data)
    resp.headers['Content-Type'] = 'application/manifest+json'
    return resp


# ─── Health Check (Railway Deploy-Gate, kein Login nötig) ────────────────────

@app.route('/health')
def health():
    try:
        query("SELECT 1")
    except Exception as exc:
        app.logger.error(f"Health-Check fehlgeschlagen: {exc}")
        return {'ok': False, 'error': str(exc)}, 503
    return {'ok': True}, 200


# ─── Routes: Auth ─────────────────────────────────────────────────────────────

_PASSWORT_HASH_PRAEFIXE = ('pbkdf2:', 'scrypt:')

def _ist_passwort_hash(wert):
    return bool(wert) and wert.startswith(_PASSWORT_HASH_PRAEFIXE)

def _password_ok(gespeichert, eingabe):
    """True wenn eingabe zum gespeicherten Passwort passt. Erkennt sowohl
    gehashte (werkzeug, ab 2026-07-21) als auch alte Klartext-Passwörter
    (Bestand vor diesem Feature) – Klartext wird bei erfolgreichem Login
    andernorts automatisch nachträglich gehasht (Lazy-Migration)."""
    if _ist_passwort_hash(gespeichert):
        return check_password_hash(gespeichert, eingabe)
    return gespeichert == eingabe

LOGIN_MAX_VERSUCHE = 3
ADMIN_SPERRE_MINUTEN = 15
# Bugreport 2026-07-21 (Niedrig): fester Dummy-Hash, gegen den bei unbekanntem Login-
# Namen trotzdem ein check_password_hash() ausgeführt wird, um Timing-basierte
# Konto-Enumeration zu erschweren (Login funktioniert auch per Kürzel/Klarname).
_DUMMY_PASSWORT_HASH = generate_password_hash(secrets.token_urlsafe(32))

@app.route('/', methods=['GET', 'POST'])
@limiter.limit('15 per minute', methods=['POST'])
def login():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))

    if request.method == 'POST':
        email_input = request.form.get('email', '').strip()
        passwort    = request.form.get('passwort', '').strip()

        # ADMIN-Direktlogin: Passwort aus ENV, DB-unabhängig. Der Account wird trotzdem
        # als mitarbeiter-Zeile geführt (für session/Anzeige) – init_db() legt sie beim
        # Start an, daher existiert sie hier immer schon.
        _admin_pw = ADMIN_PASSWORD.strip()
        if email_input.upper() == 'ADMIN':
            db = get_db()
            admin = db.execute("SELECT * FROM mitarbeiter WHERE kuerzel='ADMIN'").fetchone()
            _admin_gesperrt_bis = admin['gesperrt_bis'] if admin and 'gesperrt_bis' in admin.keys() else None
            if _admin_gesperrt_bis and datetime.now().isoformat() < _admin_gesperrt_bis:
                _rest_min = max(1, int((datetime.fromisoformat(_admin_gesperrt_bis) - datetime.now()).total_seconds() // 60) + 1)
                flash(f'Zu viele Fehlversuche. Admin-Login ist noch ca. {_rest_min} Minute(n) gesperrt.', 'danger')
                return render_template('login.html')
            if secrets.compare_digest(passwort, _admin_pw):
                # Passwort-Hash nur neu berechnen, wenn nötig (Hashing ist bewusst langsam –
                # nicht bei jedem Login unnötig wiederholen).
                if not admin or not _ist_passwort_hash(admin['passwort']) or not check_password_hash(admin['passwort'], _admin_pw):
                    _admin_pw_hash = generate_password_hash(_admin_pw)
                    db.execute("INSERT OR IGNORE INTO mitarbeiter (name,kuerzel,rolle,passwort) VALUES ('Administrator','ADMIN','admin',?)", (_admin_pw_hash,))
                    db.execute("UPDATE mitarbeiter SET passwort=? WHERE kuerzel='ADMIN'", (_admin_pw_hash,))
                    db.commit()
                    admin = db.execute("SELECT * FROM mitarbeiter WHERE kuerzel='ADMIN'").fetchone()
                if admin:
                    db.execute("UPDATE mitarbeiter SET fehlgeschlagene_logins=0, gesperrt_bis=NULL WHERE id=?", (admin['id'],))
                    db.commit()
                    session.permanent = True
                    session['user_id'] = admin['id']
                    session['name']    = admin['name']
                    session['kuerzel'] = admin['kuerzel']
                    session['rolle']   = admin['rolle']
                    return redirect(url_for('dashboard'))
            elif admin:
                # Bugreport 2026-07-21 (Mittel): atomares SQL-Increment statt Read-Modify-
                # Write, verhindert Race Condition bei parallelen Requests.
                db.execute("UPDATE mitarbeiter SET fehlgeschlagene_logins = fehlgeschlagene_logins + 1 WHERE id=?", (admin['id'],))
                db.commit()
                _versuche = db.execute("SELECT fehlgeschlagene_logins FROM mitarbeiter WHERE id=?", (admin['id'],)).fetchone()['fehlgeschlagene_logins']
                if _versuche >= LOGIN_MAX_VERSUCHE:
                    _bis = (datetime.now() + timedelta(minutes=ADMIN_SPERRE_MINUTEN)).isoformat()
                    db.execute("UPDATE mitarbeiter SET gesperrt_bis=? WHERE id=?", (_bis, admin['id']))
                    db.commit()
                    flash(f'Zu viele Fehlversuche. Admin-Login ist {ADMIN_SPERRE_MINUTEN} Minuten gesperrt.', 'danger')
                    return render_template('login.html')

        # Normale Login-Logik für alle anderen (E-Mail, Kürzel oder vollständiger Name)
        user = query("SELECT * FROM mitarbeiter WHERE LOWER(email) = LOWER(?)", (email_input,), one=True)
        if not user:
            user = query("SELECT * FROM mitarbeiter WHERE UPPER(kuerzel) = UPPER(?)", (email_input,), one=True)
        if not user:
            user = query("SELECT * FROM mitarbeiter WHERE LOWER(name) = LOWER(?)", (email_input,), one=True)

        # ADMIN wurde bereits oben vollständig behandelt (Erfolg/Sperre/Fehlversuch-Zähler
        # via ENV-Passwort) – hier nur noch die normalen Mitarbeiter-Accounts.
        if user and user['kuerzel'] != 'ADMIN':
            # konto_gesperrt bleibt eine rein manuelle Admin-Sperre – der automatische
            # Fehlversuchs-Lockout ist seit Bugreport 2026-07-21 (Hoch) zeitbasiert statt
            # permanent: vorher konnte jeder, der nur ein Kürzel/einen Namen kennt (beides
            # gültige Login-IDs), ein fremdes Konto mit 3 Fehlversuchen DAUERHAFT
            # lahmlegen, bis ein Admin manuell entsperrt.
            if user['konto_gesperrt']:
                flash('Dieses Konto ist gesperrt. Bitte an die Verkaufsleitung/Admin wenden – das Passwort muss dort neu gesetzt werden.', 'danger')
                return render_template('login.html')
            _user_gesperrt_bis = user['gesperrt_bis'] if 'gesperrt_bis' in user.keys() else None
            if _user_gesperrt_bis and datetime.now().isoformat() < _user_gesperrt_bis:
                _rest_min = max(1, int((datetime.fromisoformat(_user_gesperrt_bis) - datetime.now()).total_seconds() // 60) + 1)
                flash(f'Zu viele Fehlversuche. Konto ist noch ca. {_rest_min} Minute(n) gesperrt.', 'danger')
                return render_template('login.html')

            if _password_ok(user['passwort'], passwort):
                # Lazy-Migration: Klartext-Passwort beim ersten erfolgreichen Login hashen.
                if not _ist_passwort_hash(user['passwort']):
                    execute("UPDATE mitarbeiter SET passwort=? WHERE id=?", (generate_password_hash(passwort), user['id']))
                execute("UPDATE mitarbeiter SET fehlgeschlagene_logins=0, gesperrt_bis=NULL WHERE id=?", (user['id'],))
                session.permanent  = True          # läuft nach PERMANENT_SESSION_LIFETIME ab
                session['user_id'] = user['id']
                session['name']    = user['name']
                session['kuerzel'] = user['kuerzel']
                session['rolle']   = user['rolle']
                session['team_id'] = user['team_id'] if 'team_id' in user.keys() else None
                session['muss_passwort_aendern'] = bool(user['muss_passwort_aendern'] if 'muss_passwort_aendern' in user.keys() else 0)
                # Karte-Benachrichtigungen werden NICHT in die Session geschrieben (kann beliebig
                # groß werden und sprengt sonst die Session-Cookie über das 4KB-Browser-Limit –
                # die Anzeige erfolgt stattdessen live aus der DB via context_processor, siehe
                # inject_now()).
                if session['muss_passwort_aendern']:
                    return redirect(url_for('erstes_passwort'))
                return redirect(url_for('dashboard'))

            # Atomares SQL-Increment (siehe Kommentar beim Admin-Zweig oben).
            execute("UPDATE mitarbeiter SET fehlgeschlagene_logins = fehlgeschlagene_logins + 1 WHERE id=?", (user['id'],))
            _versuche = query("SELECT fehlgeschlagene_logins FROM mitarbeiter WHERE id=?", (user['id'],), one=True)['fehlgeschlagene_logins']
            if _versuche >= LOGIN_MAX_VERSUCHE:
                _bis = (datetime.now() + timedelta(minutes=ADMIN_SPERRE_MINUTEN)).isoformat()
                execute("UPDATE mitarbeiter SET gesperrt_bis=? WHERE id=?", (_bis, user['id']))
                flash(f'Zu viele Fehlversuche. Konto ist {ADMIN_SPERRE_MINUTEN} Minuten gesperrt.', 'danger')
                return render_template('login.html')
        elif not user:
            # Timing-Angleichung gegen Konto-Enumeration (Login funktioniert auch per
            # Kürzel/Klarname) – Ergebnis wird verworfen.
            check_password_hash(_DUMMY_PASSWORT_HASH, passwort)

        flash('Ungültige E-Mail-Adresse oder falsches Passwort.', 'danger')

    return render_template('login.html')


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))


# ─── Passwort-Reset per E-Mail ────────────────────────────────────────────────

@app.route('/passwort-vergessen', methods=['GET', 'POST'])
@limiter.limit('5 per minute', methods=['POST'])
def passwort_vergessen():
    mail_konfiguriert = bool(MAIL_SERVER and MAIL_USERNAME)
    if request.method == 'POST':
        eingabe = request.form.get('eingabe', '').strip()
        ma = query(
            "SELECT * FROM mitarbeiter WHERE email=? AND rolle!='admin'",
            (eingabe,), one=True
        )
        if ma and ma['email']:
            token  = secrets.token_urlsafe(32)
            ablauf = (datetime.now() + timedelta(hours=1)).isoformat()
            execute("UPDATE mitarbeiter SET reset_token=?, reset_token_ablauf=? WHERE id=?",
                    (token, ablauf, ma['id']))
            base = APP_BASE_URL or request.host_url.rstrip('/')
            reset_url = f"{base}{url_for('passwort_reset', token=token)}"
            html = f"""
            <div style="font-family:Arial,sans-serif;max-width:500px;margin:0 auto">
              <div style="background:#1a3a5c;padding:20px 30px;border-radius:8px 8px 0 0">
                <h2 style="color:#fff;margin:0">📊 Aktions Tracker</h2>
              </div>
              <div style="background:#f9f9f9;padding:30px;border:1px solid #ddd;border-radius:0 0 8px 8px">
                <p>Hallo <strong>{ma['name']}</strong>,</p>
                <p>Sie haben eine Passwort-Zurücksetzung angefordert. Klicken Sie auf den Button:</p>
                <p style="text-align:center;margin:30px 0">
                  <a href="{reset_url}"
                     style="background:#1a3a5c;color:#fff;padding:14px 32px;border-radius:6px;
                            text-decoration:none;font-weight:bold;font-size:16px">
                    Passwort zurücksetzen
                  </a>
                </p>
                <p style="color:#888;font-size:13px">
                  Der Link ist <strong>1 Stunde</strong> gültig.<br>
                  Falls Sie diese Anfrage nicht gestellt haben, ignorieren Sie diese E-Mail.
                </p>
                <hr style="border:none;border-top:1px solid #ddd;margin:20px 0">
                <p style="color:#aaa;font-size:11px">
                  Aktions Tracker &mdash; automatisch generiert
                </p>
              </div>
            </div>"""
            send_email(ma['email'], f'Passwort zurücksetzen – {COMPANY_NAME}', html)
        # Immer dieselbe Meldung (Sicherheit: kein Hinweis ob Konto existiert)
        flash('Falls ein Konto mit diesen Daten existiert, wurde eine E-Mail gesendet.', 'info')
        return redirect(url_for('login'))
    return render_template('passwort_vergessen.html', mail_konfiguriert=mail_konfiguriert)


@app.route('/passwort-reset/<token>', methods=['GET', 'POST'])
def passwort_reset(token):
    ma = query(
        "SELECT * FROM mitarbeiter WHERE reset_token=? AND reset_token_ablauf > ?",
        (token, datetime.now().isoformat()), one=True
    )
    if not ma:
        flash('Der Reset-Link ist ungültig oder abgelaufen. Bitte neu anfordern.', 'danger')
        return redirect(url_for('login'))
    if request.method == 'POST':
        neues_pw = request.form.get('passwort', '').strip()
        bestaet  = request.form.get('passwort2', '').strip()
        if len(neues_pw) < 8:
            flash('Passwort muss mindestens 8 Zeichen haben.', 'danger')
            return render_template('passwort_reset.html', token=token, name=ma['name'])
        if neues_pw != bestaet:
            flash('Passwörter stimmen nicht überein.', 'danger')
            return render_template('passwort_reset.html', token=token, name=ma['name'])
        execute("UPDATE mitarbeiter SET passwort=?, reset_token=NULL, reset_token_ablauf=NULL WHERE id=?",
                (generate_password_hash(neues_pw), ma['id']))
        flash('Passwort erfolgreich geändert! Bitte jetzt anmelden.', 'success')
        return redirect(url_for('login'))
    return render_template('passwort_reset.html', token=token, name=ma['name'])


# --- Erstlogin: Passwort-AEnderung erzwingen ---

@app.route('/erstes-passwort', methods=['GET', 'POST'])
def erstes_passwort():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    if not session.get('muss_passwort_aendern'):
        return redirect(url_for('dashboard'))
    if request.method == 'POST':
        neues_pw  = request.form.get('passwort', '').strip()
        bestaet   = request.form.get('passwort2', '').strip()
        if len(neues_pw) < 8:
            flash('Passwort muss mindestens 8 Zeichen haben.', 'danger')
            return render_template('erstes_passwort.html')
        if neues_pw != bestaet:
            flash('Passwoerter stimmen nicht ueberein.', 'danger')
            return render_template('erstes_passwort.html')
        execute(
            'UPDATE mitarbeiter SET passwort=?, muss_passwort_aendern=0 WHERE id=?',
            (generate_password_hash(neues_pw), session['user_id'])
        )
        session['muss_passwort_aendern'] = False
        flash('Passwort erfolgreich gesetzt! Willkommen.', 'success')
        return redirect(url_for('dashboard'))
    return render_template('erstes_passwort.html')


# ─── Rechtliche Seiten ────────────────────────────────────────────────────────

@app.route('/impressum')
def impressum():
    return render_template('impressum.html')

@app.route('/datenschutz')
def datenschutz():
    return render_template('datenschutz.html')

@app.route('/agb')
def agb():
    return render_template('agb.html')

@app.route('/avv')
def avv():
    return render_template('avv.html')


# ─── Logo direkt ausliefern (umgeht Static-Cache) ────────────────────────────

@app.route('/logo.png')
def serve_logo():
    if LOGO_URL:
        from flask import redirect as _redir
        return _redir(LOGO_URL, code=302)
    logo_path = os.path.join(os.path.dirname(__file__), 'static', 'logo.png')
    return send_file(logo_path, mimetype='image/png',
                     max_age=3600, conditional=True)


# ─── Admin: DB-Backup herunterladen ──────────────────────────────────────────

@app.route('/admin/backup/herunterladen')
@login_required
def admin_backup_herunterladen():
    if session.get('rolle') != 'admin':
        flash('Keine Berechtigung.', 'danger')
        return redirect(url_for('dashboard'))
    if not os.path.exists(DATABASE):
        flash('Datenbank nicht gefunden.', 'danger')
        return redirect(url_for('admin'))
    heute = date.today().isoformat()
    app.logger.info(f"Audit: DB-Backup heruntergeladen von {session.get('kuerzel')} (Mitarbeiter-ID {session.get('user_id')})")
    return send_file(DATABASE, as_attachment=True,
                     download_name=f'brewery_backup_{heute}.db',
                     mimetype='application/x-sqlite3')


@app.route('/admin/export/jetzt-senden', methods=['POST'])
@login_required
def admin_export_jetzt_senden():
    if session.get('rolle') != 'admin':
        flash('Keine Berechtigung.', 'danger')
        return redirect(url_for('dashboard'))
    if not EXPORT_EMAIL:
        flash('EXPORT_EMAIL ist nicht gesetzt – Export kann nicht versendet werden.', 'danger')
        return redirect(url_for('admin'))
    try:
        app.logger.info(f"Audit: Manueller Foto-/Excel-Export ausgelöst von {session.get('kuerzel')} (Mitarbeiter-ID {session.get('user_id')}), Versand an {EXPORT_EMAIL}")
        auto_export_job()
        flash(f'Export wurde ausgelöst und an {EXPORT_EMAIL} gesendet.', 'success')
    except Exception as e:
        app.logger.error(f"Manueller Export Fehler: {e}", exc_info=True)
        flash('Export fehlgeschlagen – siehe Logs.', 'danger')
    return redirect(url_for('admin'))


# ─── Routes: Dashboard ────────────────────────────────────────────────────────

@app.route('/dashboard')
@login_required
def dashboard():
    jahr      = request.args.get('jahr', date.today().year, type=int)
    is_admin  = session.get('rolle') == 'admin'
    is_manager = session.get('rolle') in ('admin', 'verkaufsleiter')
    ma_filter = request.args.get('ma', '', type=str)
    # VKL mit eigenem Gebiet sehen standardmäßig ihre eigene (Mitarbeiter-)Ansicht,
    # solange sie nicht per Dropdown gezielt einen anderen Mitarbeiter ansehen (ma_filter gesetzt).
    zeige_eigene_ansicht = (not is_admin) and (not ma_filter)
    ma_clause = "AND a.mitarbeiter_id = ?" if ma_filter else ""
    ma_params = (ma_filter,) if ma_filter else ()

    # KW-Daten (Wochenübersicht)
    # Subquery: Kisten pro Aktivität voraggregieren → verhindert Duplikation von anzahl_displays
    BP = "(SELECT aktivitaet_id, SUM(kisten_anzahl) AS kisten_total FROM bestellposition GROUP BY aktivitaet_id)"

    # Displays zählen nur bei Aufbau (inkl. Altdaten/NULL); Kisten nur bei Bestellung
    _AUF     = "COALESCE(a.aktionstyp,'Aufbau')='Aufbau'"
    DISP_IST = f"SUM(CASE WHEN {_AUF} THEN a.anzahl_displays ELSE 0 END)"
    KIST_IST = "COALESCE(SUM(CASE WHEN a.aktionstyp='Bestellung' THEN b.kisten_total ELSE 0 END), 0)"

    # Team-Filter (VKL mit zugewiesenem Team sieht nur eigene Team-Mitglieder)
    t_ma_sql, t_ma_p = _team_ma_clause('a')

    if is_manager:
        kw_data = query(f'''
            SELECT strftime('%W', a.datum) AS kw,
                   CAST(strftime('%W', a.datum) AS INTEGER) AS kw_int,
                   {DISP_IST} AS displays,
                   {KIST_IST} AS kisten,
                   COUNT(a.id) AS besuche
            FROM aktivitaet a
            LEFT JOIN {BP} b ON b.aktivitaet_id = a.id
            WHERE strftime('%Y', a.datum) = ? {ma_clause}{t_ma_sql}
            GROUP BY kw
            ORDER BY kw
        ''', (str(jahr),) + ma_params + t_ma_p)
    else:
        kw_data = query(f'''
            SELECT strftime('%W', a.datum) AS kw,
                   CAST(strftime('%W', a.datum) AS INTEGER) AS kw_int,
                   {DISP_IST} AS displays,
                   {KIST_IST} AS kisten,
                   COUNT(a.id) AS besuche
            FROM aktivitaet a
            LEFT JOIN {BP} b ON b.aktivitaet_id = a.id
            WHERE strftime('%Y', a.datum) = ? AND a.mitarbeiter_id = ?
            GROUP BY kw
            ORDER BY kw
        ''', (str(jahr), session['user_id']))

    # Jahresgesamtwerte
    if is_manager:
        jahres = query(f'''
            SELECT {DISP_IST} AS displays,
                   {KIST_IST} AS kisten,
                   COUNT(a.id) AS besuche,
                   COUNT(DISTINCT a.mitarbeiter_id) AS mitarbeiter_aktiv
            FROM aktivitaet a
            LEFT JOIN {BP} b ON b.aktivitaet_id = a.id
            WHERE strftime('%Y', a.datum) = ? {ma_clause}{t_ma_sql}
        ''', (str(jahr),) + ma_params + t_ma_p, one=True)
    else:
        jahres = query(f'''
            SELECT {DISP_IST} AS displays,
                   {KIST_IST} AS kisten,
                   COUNT(a.id) AS besuche
            FROM aktivitaet a
            LEFT JOIN {BP} b ON b.aktivitaet_id = a.id
            WHERE strftime('%Y', a.datum) = ? AND a.mitarbeiter_id = ?
        ''', (str(jahr), session['user_id']), one=True)

    # KONZEPT-V2: Pipeline-Kennzahlen (Manager: Teaser mit Link; Rep: nur eigene offene)
    p_aufgebaut = p_storniert = p_ueberfaellig = 0
    offene_rep_liste = []
    if is_manager:
        vorgemerkt, p_aufgebaut, p_storniert, p_ueberfaellig = _bestell_kennzahlen()
    else:
        vorgemerkt = query(
            "SELECT COUNT(*) AS n FROM aktivitaet a WHERE a.aktionstyp='Bestellung' AND COALESCE(a.bestell_status,'offen')='offen' AND a.mitarbeiter_id=?",
            (session['user_id'],), one=True)['n']
        offene_rep_liste = query(
            """SELECT a.id, a.verkaufsstelle_id AS vs_id,
                      v.name AS station, v.strasse, v.ort, a.datum, a.notizen,
                      COALESCE(a.anzahl_displays, 0) AS displays,
                      COALESCE((SELECT SUM(kisten_anzahl) FROM bestellposition WHERE aktivitaet_id=a.id), 0) AS kisten,
                      CAST((julianday('now') - julianday(a.datum)) AS INTEGER) AS alter_tage
               FROM aktivitaet a JOIN verkaufsstelle v ON v.id = a.verkaufsstelle_id
               WHERE a.aktionstyp='Bestellung' AND COALESCE(a.bestell_status,'offen')='offen'
                 AND (a.mitarbeiter_id=?
                      OR a.verkaufsstelle_id IN (
                          SELECT verkaufsstelle_id FROM mitarbeiter_verkaufsstelle WHERE mitarbeiter_id=?
                      ))
               ORDER BY a.datum ASC""",
            (session['user_id'], session['user_id'])
        )

    # Top Biersorten – direkt über bestellposition, kein Display-Problem hier
    if is_manager:
        top_bier = query(f'''
            SELECT bs.name, SUM(bp.kisten_anzahl) AS kisten
            FROM bestellposition bp
            JOIN biersorte bs ON bs.id = bp.biersorte_id
            JOIN aktivitaet a ON a.id = bp.aktivitaet_id
            WHERE strftime('%Y', a.datum) = ? AND a.aktionstyp='Bestellung' {ma_clause}{t_ma_sql}
            GROUP BY bs.id ORDER BY kisten DESC LIMIT 6
        ''', (str(jahr),) + ma_params + t_ma_p)
    else:
        top_bier = query('''
            SELECT bs.name, SUM(bp.kisten_anzahl) AS kisten
            FROM bestellposition bp
            JOIN biersorte bs ON bs.id = bp.biersorte_id
            JOIN aktivitaet a ON a.id = bp.aktivitaet_id
            WHERE strftime('%Y', a.datum) = ? AND a.mitarbeiter_id = ? AND a.aktionstyp='Bestellung'
            GROUP BY bs.id ORDER BY kisten DESC LIMIT 6
        ''', (str(jahr), session['user_id']))

    # Mitarbeiter-Ranking (Manager-Sicht, nur ohne Einzelfilter)
    t_m_sql, t_m_p = _team_m_clause('m')
    rep_stats = []
    if is_manager and not ma_filter:
        rep_stats = query(f'''
            SELECT m.name, m.kuerzel,
                   {DISP_IST} AS displays,
                   {KIST_IST} AS kisten,
                   COUNT(a.id) AS besuche
            FROM mitarbeiter m
            JOIN aktivitaet a ON a.mitarbeiter_id = m.id
            LEFT JOIN {BP} b ON b.aktivitaet_id = a.id
            WHERE strftime('%Y', a.datum) = ?{t_m_sql}
            GROUP BY m.id ORDER BY kisten DESC
        ''', (str(jahr),) + t_m_p)

    # Letzte Aktivitäten
    if not zeige_eigene_ansicht:
        letzte = query(f'''
            SELECT a.id, a.datum, m.name AS mitarbeiter, v.name AS verkaufsstelle,
                   a.anzahl_displays, COALESCE(SUM(b.kisten_anzahl), 0) AS kisten
            FROM aktivitaet a
            JOIN mitarbeiter m ON m.id = a.mitarbeiter_id
            JOIN verkaufsstelle v ON v.id = a.verkaufsstelle_id
            LEFT JOIN bestellposition b ON b.aktivitaet_id = a.id
            WHERE strftime('%Y', a.datum) = ? {ma_clause}{t_ma_sql}
            GROUP BY a.id ORDER BY a.datum DESC, a.erstellt_am DESC LIMIT 10
        ''', (str(jahr),) + ma_params + t_ma_p)
    else:
        letzte = query('''
            SELECT a.id, a.datum, m.name AS mitarbeiter, v.name AS verkaufsstelle,
                   a.anzahl_displays, COALESCE(SUM(b.kisten_anzahl), 0) AS kisten
            FROM aktivitaet a
            JOIN mitarbeiter m ON m.id = a.mitarbeiter_id
            JOIN verkaufsstelle v ON v.id = a.verkaufsstelle_id
            LEFT JOIN bestellposition b ON b.aktivitaet_id = a.id
            WHERE strftime('%Y', a.datum) = ? AND a.mitarbeiter_id = ?
            GROUP BY a.id ORDER BY a.datum DESC, a.erstellt_am DESC LIMIT 10
        ''', (str(jahr), session['user_id']))

    _tm_sql, _tm_p = _team_m_clause('m')
    alle_ma = query(
        f"SELECT id, name FROM mitarbeiter WHERE rolle IN ('rep','verkaufsleiter'){_tm_sql} ORDER BY name",
        _tm_p
    ) if is_manager else []

    verfuegbare_jahre = [r[0] for r in query(
        "SELECT DISTINCT CAST(strftime('%Y', datum) AS INTEGER) FROM aktivitaet ORDER BY 1 DESC"
    )]
    if not verfuegbare_jahre:
        verfuegbare_jahre = [date.today().year]
    if date.today().year not in verfuegbare_jahre:
        verfuegbare_jahre.insert(0, date.today().year)

    chart_kw   = [f"KW {int(r['kw_int']):02d}" for r in kw_data]
    chart_disp = [r['displays'] or 0 for r in kw_data]
    chart_kist = [r['kisten'] or 0 for r in kw_data]
    bier_namen = [r['name'] for r in top_bier]
    bier_kist  = [r['kisten'] for r in top_bier]

    # Zielzahlen für Fortschrittsanzeige im Dashboard
    if is_manager and not ma_filter:
        ziel = query(
            "SELECT displays_ziel, kisten_ziel FROM zielzahlen WHERE mitarbeiter_id IS NULL AND jahr=?",
            (str(jahr),), one=True
        )
    elif is_manager and ma_filter:
        ziel = query(
            "SELECT displays_ziel, kisten_ziel FROM zielzahlen WHERE mitarbeiter_id=? AND jahr=?",
            (int(ma_filter), str(jahr)), one=True
        )
    else:
        ziel = query(
            "SELECT displays_ziel, kisten_ziel FROM zielzahlen WHERE mitarbeiter_id=? AND jahr=?",
            (session['user_id'], str(jahr)), one=True
        )

    # Inaktivitäts-Warnung: Reps ohne Aktivität diese Woche (ab Mittwoch sichtbar)
    # Reps mit aktiver Vertretung werden ausgeschlossen
    inaktiv_reps = []
    if is_manager and not ma_filter:
        _heute = date.today()
        if _heute.weekday() >= 2:  # erst ab Mittwoch sinnvoll (Mo=0, Di=1, Mi=2)
            _mo_kw = _heute - timedelta(days=_heute.weekday())  # Montag dieser Woche
            _t_sql, _t_p = _team_m_clause('m')
            inaktiv_reps = query(
                f"""SELECT m.id, m.name, m.kuerzel
                    FROM mitarbeiter m
                    WHERE m.rolle = 'rep' {_t_sql}
                    AND m.id NOT IN (
                        SELECT DISTINCT mitarbeiter_id FROM aktivitaet
                        WHERE datum >= ?
                    )
                    AND m.id NOT IN (
                        SELECT abwesender_id FROM vertretung
                        WHERE von <= ? AND bis >= ? AND status = 'bestätigt'
                    )
                    ORDER BY m.name""",
                _t_p + (_mo_kw.isoformat(), _heute.isoformat(), _heute.isoformat())
            )

    # Offene Urlaubsanträge (Manager: zum Bestätigen/Ablehnen direkt im Dashboard)
    urlaubsantraege = []
    if is_manager and not ma_filter:
        _ta_sql, _ta_p = _team_m_clause('m')
        urlaubsantraege = query(
            f"""SELECT v.id, v.von, v.bis, v.status,
                       m.name AS abwesender, r.name AS vertreter
                FROM vertretung v
                JOIN mitarbeiter m ON m.id = v.abwesender_id
                LEFT JOIN mitarbeiter r ON r.id = v.vertreter_id
                WHERE v.status = 'angefragt' {_ta_sql}
                ORDER BY v.von""",
            _ta_p
        )

    # Rep-Dashboard: Tages-/Wochen-/Monatszahlen
    heute_stats = diese_woche_stats = vorwoche_stats = dieser_monat_stats = None
    kw_aktuell = date.today().isocalendar()[1]
    _monat_namen = ['Januar','Februar','März','April','Mai','Juni',
                    'Juli','August','September','Oktober','November','Dezember']
    monat_name = _monat_namen[date.today().month - 1]
    # Persönliche Tages-/Wochen-/Monatszahlen: eigene Ansicht (Rep, oder VKL ohne ma_filter)
    if zeige_eigene_ansicht:
        _uid = (session['user_id'],)
        heute_stats = query(f'''
            SELECT {DISP_IST} AS displays, {KIST_IST} AS kisten, COUNT(a.id) AS besuche
            FROM aktivitaet a LEFT JOIN {BP} b ON b.aktivitaet_id = a.id
            WHERE a.datum = date('now','localtime') AND a.mitarbeiter_id = ?
        ''', _uid, one=True)
        diese_woche_stats = query(f'''
            SELECT {DISP_IST} AS displays, {KIST_IST} AS kisten, COUNT(a.id) AS besuche
            FROM aktivitaet a LEFT JOIN {BP} b ON b.aktivitaet_id = a.id
            WHERE strftime('%Y-%W', a.datum) = strftime('%Y-%W', date('now','localtime'))
            AND a.mitarbeiter_id = ?
        ''', _uid, one=True)
        vorwoche_stats = query(f'''
            SELECT {DISP_IST} AS displays, {KIST_IST} AS kisten, COUNT(a.id) AS besuche
            FROM aktivitaet a LEFT JOIN {BP} b ON b.aktivitaet_id = a.id
            WHERE strftime('%Y-%W', a.datum) = strftime('%Y-%W', date('now','localtime','-7 days'))
            AND a.mitarbeiter_id = ?
        ''', _uid, one=True)
        dieser_monat_stats = query(f'''
            SELECT {DISP_IST} AS displays, {KIST_IST} AS kisten, COUNT(a.id) AS besuche
            FROM aktivitaet a LEFT JOIN {BP} b ON b.aktivitaet_id = a.id
            WHERE strftime('%Y-%m', a.datum) = strftime('%Y-%m', date('now','localtime'))
            AND a.mitarbeiter_id = ?
        ''', _uid, one=True)

    # Tagesplan für Rep: Montag–Sonntag der gewählten (oder aktuellen) Woche
    tagesplan_rep = []
    alle_verkaufsstellen_rep = []
    _today = date.today()
    _tp_woche_str = request.args.get('tp_woche', None)
    if _tp_woche_str:
        try:
            _tp_w = date.fromisoformat(_tp_woche_str)
        except ValueError:
            _tp_w = _today
        tp_woche_montag = _tp_w - timedelta(days=_tp_w.weekday())
    else:
        tp_woche_montag = _today - timedelta(days=_today.weekday())
    tp_woche_sonntag = tp_woche_montag + timedelta(days=6)
    tp_kw            = tp_woche_montag.isocalendar()[1]
    tp_prev_kw       = tp_kw - 1 if tp_kw > 1 else 52
    tp_next_kw       = tp_kw + 1 if tp_kw < 52 else 1
    tp_prev_woche    = (tp_woche_montag - timedelta(days=7)).isoformat()
    tp_next_woche    = (tp_woche_montag + timedelta(days=7)).isoformat()
    datum_woche_rep  = [(tp_woche_montag + timedelta(days=i)).isoformat() for i in range(7)]
    urlaub_woche_rep = {
        d: info for (_mid, d), info in _urlaub_daten([session['user_id']], tp_woche_montag.isoformat(), tp_woche_sonntag.isoformat()).items()
    }
    # Feiertage in der Besuchsplanung anzeigen statt "Kein Plan" – Bundesland des Reps,
    # damit Feiertage korrekt regional zugeordnet sind (analog zur Urlaubskonto-Logik).
    _bundesland_rep = query("SELECT bundesland FROM mitarbeiter WHERE id=?", (session['user_id'],), one=True)
    _bundesland_rep = (_bundesland_rep['bundesland'] if _bundesland_rep else None) or 'BY'
    _feiertage_rep = set()
    for _j in {tp_woche_montag.year, tp_woche_sonntag.year}:
        _feiertage_rep |= _feiertage_set(_j, _bundesland_rep)
    feiertage_woche_rep = {d for d in datum_woche_rep if d in _feiertage_rep}
    # Eigener Besuchsplan: eigene Ansicht (Rep, oder VKL ohne ma_filter)
    if zeige_eigene_ansicht:
        tagesplan_rep = query('''
            SELECT tp.id, tp.datum, tp.reihenfolge, tp.notiz, tp.erledigt,
                   v.name AS station, v.strasse, v.plz, v.ort, v.id AS vs_id,
                   v.lieferant, v.ansprechpartner, v.hinweis,
                   (SELECT a.von_uhrzeit FROM aktivitaet a
                    WHERE a.mitarbeiter_id = tp.mitarbeiter_id
                      AND a.verkaufsstelle_id = tp.verkaufsstelle_id
                      AND a.datum = tp.datum ORDER BY a.erstellt_am DESC LIMIT 1) AS von_uhrzeit,
                   (SELECT a.bis_uhrzeit FROM aktivitaet a
                    WHERE a.mitarbeiter_id = tp.mitarbeiter_id
                      AND a.verkaufsstelle_id = tp.verkaufsstelle_id
                      AND a.datum = tp.datum ORDER BY a.erstellt_am DESC LIMIT 1) AS bis_uhrzeit
            FROM tagesplan tp
            JOIN verkaufsstelle v ON v.id = tp.verkaufsstelle_id
            WHERE tp.mitarbeiter_id = ?
              AND tp.datum >= ?
              AND tp.datum <= ?
              AND COALESCE(tp.geloescht, 0) = 0
            ORDER BY tp.datum, (von_uhrzeit IS NULL), von_uhrzeit, tp.reihenfolge, tp.id
        ''', (session['user_id'], tp_woche_montag.isoformat(), tp_woche_sonntag.isoformat()))
        # Stationsliste für Self-Service-Formular. Gedeckelt (Performance) – bei
        # sehr vielen aktiven Stationen würde das feste Einbetten ALLER Stationen
        # als <option> die Seite massiv aufblähen (siehe VS_DASHBOARD_SEITENGROESSE);
        # die freie Suche im Dropdown (JS) fragt für Treffer außerhalb dieser ersten
        # Seite zusätzlich /api/verkaufsstellen nach.
        assigned = query(
            "SELECT verkaufsstelle_id FROM mitarbeiter_verkaufsstelle WHERE mitarbeiter_id=?",
            (session['user_id'],)
        )
        if assigned:
            _vs_ids = [r['verkaufsstelle_id'] for r in assigned]
            _ph = ','.join('?' * len(_vs_ids))
            alle_verkaufsstellen_rep = query(
                f"SELECT id, name, plz, ort, strasse, typ, landkreis FROM verkaufsstelle WHERE aktiv=1 AND id IN ({_ph}) ORDER BY name",
                _vs_ids
            )
        else:
            alle_verkaufsstellen_rep = query(
                "SELECT id, name, plz, ort, strasse, typ, landkreis FROM verkaufsstelle "
                "WHERE aktiv=1 AND (homeoffice_mitarbeiter_id IS NULL OR homeoffice_mitarbeiter_id=?) "
                "ORDER BY name LIMIT ?",
                (session['user_id'], VS_DASHBOARD_SEITENGROESSE)
            )

    return render_template('dashboard.html',
        jahr=jahr, kw_data=kw_data, jahres=jahres,
        rep_stats=rep_stats, letzte=letzte,
        verfuegbare_jahre=verfuegbare_jahre,
        chart_kw=json.dumps(chart_kw),
        chart_disp=json.dumps(chart_disp),
        chart_kist=json.dumps(chart_kist),
        bier_namen=json.dumps(bier_namen),
        bier_kist=json.dumps(bier_kist),
        is_admin=is_admin,
        is_manager=is_manager,
        zeige_eigene_ansicht=zeige_eigene_ansicht,
        ma_filter=ma_filter,
        alle_ma=alle_ma,
        vorgemerkt=vorgemerkt,
        offene_rep_liste=offene_rep_liste,
        p_aufgebaut=p_aufgebaut,
        p_storniert=p_storniert,
        p_ueberfaellig=p_ueberfaellig,
        ziel=ziel,
        inaktiv_reps=inaktiv_reps,
        urlaubsantraege=urlaubsantraege,
        heute_stats=heute_stats,
        diese_woche_stats=diese_woche_stats,
        vorwoche_stats=vorwoche_stats,
        dieser_monat_stats=dieser_monat_stats,
        kw_aktuell=kw_aktuell,
        monat_name=monat_name,
        tagesplan_rep=tagesplan_rep,
        alle_verkaufsstellen_rep=alle_verkaufsstellen_rep,
        datum_woche_rep=datum_woche_rep,
        urlaub_woche_rep=urlaub_woche_rep,
        feiertage_woche_rep=feiertage_woche_rep,
        tp_kw=tp_kw,
        tp_prev_kw=tp_prev_kw,
        tp_next_kw=tp_next_kw,
        tp_prev_woche=tp_prev_woche,
        tp_next_woche=tp_next_woche,
        today_str=_today.isoformat(),
        tomorrow_str=(_today + timedelta(days=1)).isoformat(),
    )


# ─── Tourenplanung ───────────────────────────────────────────────────────────

@app.route('/api/tourenplanung/mitarbeiter/<int:ma_id>/verkaufsstellen')
@manager_required
def api_tourenplanung_mitarbeiter_verkaufsstellen(ma_id):
    """Verkaufsstellen eines Mitarbeiters für das 'Neuer Stopp'-Formular auf der
    Tourenplanung-Seite (gleiche Logik wie das Rep-Selbstservice-Widget im
    Dashboard, nur für VKL/Admin nutzbar zur Planung für sich selbst oder das Team)."""
    # Bugreport 2026-07-21 (Niedrig-mittel): ma_id kam bisher ungeprüft rein – ein VKL
    # konnte durch Hochzählen die zugewiesenen Verkaufsstellen (inkl. Homeoffice-/
    # Wohnadresse) fremder Mitarbeiter außerhalb des eigenen Teams einsehen.
    if not _mitarbeiter_im_eigenen_team(ma_id):
        abort(403)
    assigned = query(
        "SELECT verkaufsstelle_id FROM mitarbeiter_verkaufsstelle WHERE mitarbeiter_id=?", (ma_id,)
    )
    vs_ids = [r['verkaufsstelle_id'] for r in assigned] if assigned else []
    if vs_ids:
        ph = ','.join('?' * len(vs_ids))
        vs_rows = query(
            f"SELECT id, name, plz, ort FROM verkaufsstelle WHERE aktiv=1 AND id IN ({ph}) "
            "AND (homeoffice_mitarbeiter_id IS NULL OR homeoffice_mitarbeiter_id=?) ORDER BY name",
            vs_ids + [ma_id]
        )
    else:
        vs_rows = query(
            "SELECT id, name, plz, ort FROM verkaufsstelle WHERE aktiv=1 "
            "AND (homeoffice_mitarbeiter_id IS NULL OR homeoffice_mitarbeiter_id=?) ORDER BY name",
            (ma_id,)
        )
    return jsonify([{'id': v['id'], 'name': v['name'], 'plz': v['plz'], 'ort': v['ort']} for v in vs_rows])


@app.route('/api/arbeitszeit/heute', methods=['GET'])
@login_required
def api_arbeitszeit_heute():
    if not ARBEITSZEIT_MODUS:
        return jsonify({'ok': False, 'error': 'Arbeitszeiterfassung ist nicht aktiviert.'}), 403
    row = query(
        "SELECT beginn, ende, pause_minuten FROM arbeitszeit WHERE mitarbeiter_id=? AND datum=?",
        (session['user_id'], date.today().isoformat()), one=True
    )
    return jsonify({
        'ok': True,
        'beginn': row['beginn'] if row else None,
        'ende': row['ende'] if row else None,
        'pause_minuten': (row['pause_minuten'] or 0) if row else 0,
    })


@app.route('/api/arbeitszeit/speichern', methods=['POST'])
@login_required
def api_arbeitszeit_speichern():
    if not ARBEITSZEIT_MODUS:
        return jsonify({'ok': False, 'error': 'Arbeitszeiterfassung ist nicht aktiviert.'}), 403
    data = request.get_json(force=True, silent=True) or {}
    feld = data.get('feld')
    if feld in ('beginn', 'ende'):
        wert = (data.get('uhrzeit') or '').strip()
        if not wert:
            return jsonify({'ok': False, 'error': 'Ungültige Eingabe'}), 400
    elif feld == 'pause':
        try:
            wert = max(0, int(data.get('minuten', 0)))
        except (TypeError, ValueError):
            return jsonify({'ok': False, 'error': 'Ungültige Eingabe'}), 400
        feld = 'pause_minuten'
    else:
        return jsonify({'ok': False, 'error': 'Ungültige Eingabe'}), 400

    heute = date.today().isoformat()
    row = query(
        "SELECT id, beginn, ende, pause_minuten FROM arbeitszeit WHERE mitarbeiter_id=? AND datum=?",
        (session['user_id'], heute), one=True
    )

    # Bei Beginn/Ende-Änderung: gesetzliche Mindestpause automatisch nach oben korrigieren
    # (>6 Std → mind. 30 Min, >9 Std → mind. 45 Min). Manuelle Pausen-Eingabe bleibt unangetastet.
    pause_bump = None
    if feld in ('beginn', 'ende'):
        neu_beginn = wert if feld == 'beginn' else (row['beginn'] if row else None)
        neu_ende   = wert if feld == 'ende'   else (row['ende'] if row else None)
        pflicht = _az_pflichtpause_minuten(_az_brutto_minuten(neu_beginn, neu_ende))
        vorhandene_pause = (row['pause_minuten'] or 0) if row else 0
        if pflicht > vorhandene_pause:
            pause_bump = pflicht

    if row:
        execute(f"UPDATE arbeitszeit SET {feld}=? WHERE id=?", (wert, row['id']))
        if pause_bump is not None:
            execute("UPDATE arbeitszeit SET pause_minuten=? WHERE id=?", (pause_bump, row['id']))
    else:
        execute(f"INSERT INTO arbeitszeit (mitarbeiter_id, datum, {feld}) VALUES (?,?,?)",
                (session['user_id'], heute, wert))
        if pause_bump is not None:
            execute(
                "UPDATE arbeitszeit SET pause_minuten=? WHERE mitarbeiter_id=? AND datum=?",
                (pause_bump, session['user_id'], heute)
            )
    return jsonify({'ok': True, 'pause_minuten': pause_bump if pause_bump is not None else ((row['pause_minuten'] or 0) if row else 0)})


@app.route('/api/arbeitszeit/admin-speichern', methods=['POST'])
@manager_required
def api_arbeitszeit_admin_speichern():
    """Admin kann Arbeitszeit für jeden Mitarbeiter und jedes Datum – auch in der
    Vergangenheit – korrigieren. VKL genauso, aber nur für die eigenen Teammitglieder."""
    if not ARBEITSZEIT_MODUS:
        return jsonify({'ok': False, 'error': 'Arbeitszeiterfassung ist nicht aktiviert.'}), 403
    data = request.get_json(force=True, silent=True) or {}
    try:
        ma_id = int(data.get('mitarbeiter_id'))
    except (TypeError, ValueError):
        return jsonify({'ok': False, 'error': 'Ungültiger Mitarbeiter'}), 400
    datum = (data.get('datum') or '').strip()
    try:
        date.fromisoformat(datum)
    except ValueError:
        return jsonify({'ok': False, 'error': 'Ungültiges Datum'}), 400

    beginn = (data.get('beginn') or '').strip() or None
    ende   = (data.get('ende') or '').strip() or None
    try:
        pause_minuten = max(0, int(data.get('pause_minuten') or 0))
    except (TypeError, ValueError):
        pause_minuten = 0

    # Gesetzliche Mindestpause auch hier nur nach oben korrigieren, nie nach unten
    pflicht = _az_pflichtpause_minuten(_az_brutto_minuten(beginn, ende))
    if pflicht > pause_minuten:
        pause_minuten = pflicht

    ma = query("SELECT id, team_id FROM mitarbeiter WHERE id=?", (ma_id,), one=True)
    if not ma:
        return jsonify({'ok': False, 'error': 'Mitarbeiter nicht gefunden'}), 404
    if session.get('rolle') == 'verkaufsleiter' and ma['team_id'] != session.get('team_id'):
        return jsonify({'ok': False, 'error': 'Kein Zugriff'}), 403

    row = query(
        "SELECT id FROM arbeitszeit WHERE mitarbeiter_id=? AND datum=?",
        (ma_id, datum), one=True
    )
    if row:
        execute(
            "UPDATE arbeitszeit SET beginn=?, ende=?, pause_minuten=? WHERE id=?",
            (beginn, ende, pause_minuten, row['id'])
        )
    else:
        execute(
            "INSERT INTO arbeitszeit (mitarbeiter_id, datum, beginn, ende, pause_minuten) VALUES (?,?,?,?,?)",
            (ma_id, datum, beginn, ende, pause_minuten)
        )
    return jsonify({'ok': True})


@app.route('/arbeitszeit')
@login_required
def arbeitszeit_uebersicht():
    if not ARBEITSZEIT_MODUS:
        flash('Die Arbeitszeiterfassung ist in Ihrem aktuellen Paket nicht verfügbar.', 'warning')
        return redirect(url_for('dashboard'))
    is_admin   = session.get('rolle') == 'admin'
    is_manager = session.get('rolle') in ('admin', 'verkaufsleiter')
    modus = request.args.get('modus', 'woche')
    if modus not in ('tag', 'woche', 'monat'):
        modus = 'woche'
    today = date.today()
    today_str = today.isoformat()

    mitarbeiter_liste = []
    ma_id = session['user_id']
    if is_manager:
        _tm_sql, _tm_p = _team_m_clause('m')
        mitarbeiter_liste = query(
            f"SELECT id, name, kuerzel FROM mitarbeiter m WHERE rolle IN ('rep','verkaufsleiter') AND aktiv=1 {_tm_sql} ORDER BY name",
            _tm_p
        )
        ma_param = request.args.get('ma', '')
        ma_id = int(ma_param) if ma_param.isdigit() else None
        # Bugreport 2026-07-21 (Mittel): ma-Parameter kam bisher ungeprüft rein – ein VKL
        # konnte per ?ma=<fremde-ID> Arbeitszeitdaten außerhalb des eigenen Teams einsehen.
        if ma_id and not _mitarbeiter_im_eigenen_team(ma_id):
            ma_id = session['user_id']

    # Eigene Ansicht (Mitarbeiter/VKL auf ihre eigenen Daten geschaut) darf für
    # den heutigen Tag selbst bearbeitet werden; Admin darf immer & überall.
    eigene_ansicht = (ma_id == session['user_id'])

    # ── Tag ──
    tag_datum_str = request.args.get('datum', today_str)
    try:
        tag_d = date.fromisoformat(tag_datum_str)
    except ValueError:
        tag_d = today
    tag_datum_str = tag_d.isoformat()
    tag_prev = (tag_d - timedelta(days=1)).isoformat()
    tag_next = (tag_d + timedelta(days=1)).isoformat()

    tag_eintrag = None
    tag_team = []
    if modus == 'tag':
        if ma_id:
            row = query(
                "SELECT beginn, ende, pause_minuten FROM arbeitszeit WHERE mitarbeiter_id=? AND datum=?",
                (ma_id, tag_datum_str), one=True
            )
            netto = _az_netto_minuten(row['beginn'], row['ende'], row['pause_minuten']) if row else None
            tag_eintrag = {
                'beginn': row['beginn'] if row else None,
                'ende': row['ende'] if row else None,
                'pause_minuten': (row['pause_minuten'] or 0) if row else 0,
                'netto_fmt': _az_fmt_std(netto),
            }
        elif is_manager:
            _tm_sql, _tm_p = _team_m_clause('m')
            rows = query(f'''
                SELECT m.id, m.name, m.kuerzel, az.beginn, az.ende, az.pause_minuten
                FROM mitarbeiter m
                LEFT JOIN arbeitszeit az ON az.mitarbeiter_id = m.id AND az.datum = ?
                WHERE m.rolle IN ('rep','verkaufsleiter') AND m.aktiv=1 {_tm_sql}
                ORDER BY m.name
            ''', (tag_datum_str,) + _tm_p)
            tag_team = [{
                'id': r['id'], 'name': r['name'], 'kuerzel': r['kuerzel'],
                'beginn': r['beginn'], 'ende': r['ende'], 'pause_minuten': r['pause_minuten'] or 0,
                'netto_fmt': _az_fmt_std(_az_netto_minuten(r['beginn'], r['ende'], r['pause_minuten'])),
            } for r in rows]

    # ── Woche ──
    woche_start_str = request.args.get('woche')
    try:
        woche_start = date.fromisoformat(woche_start_str) if woche_start_str else today
    except ValueError:
        woche_start = today
    woche_start = woche_start - timedelta(days=woche_start.weekday())
    woche_ende  = woche_start + timedelta(days=6)
    woche_kw    = woche_start.isocalendar()[1]

    # ── Monat ──
    monat_str = request.args.get('monat')
    try:
        monat_start = date.fromisoformat(monat_str + '-01') if monat_str else today.replace(day=1)
    except ValueError:
        monat_start = today.replace(day=1)
    if monat_start.month == 12:
        monat_ende = date(monat_start.year, 12, 31)
    else:
        monat_ende = date(monat_start.year, monat_start.month + 1, 1) - timedelta(days=1)
    prev_monat = (monat_start - timedelta(days=1)).replace(day=1)
    next_monat = (monat_ende + timedelta(days=1)).replace(day=1)
    _monat_namen = ['Januar','Februar','März','April','Mai','Juni',
                    'Juli','August','September','Oktober','November','Dezember']
    monat_label = f"{_monat_namen[monat_start.month - 1]} {monat_start.year}"

    zeitraum_start = woche_start if modus == 'woche' else monat_start
    zeitraum_ende  = woche_ende  if modus == 'woche' else monat_ende

    tage_de = ['Montag','Dienstag','Mittwoch','Donnerstag','Freitag','Samstag','Sonntag']
    eigene_tage = []
    eigene_summe = 0
    team_summary = []

    if modus in ('woche', 'monat'):
        if ma_id:
            rows = query(
                "SELECT datum, beginn, ende, pause_minuten FROM arbeitszeit "
                "WHERE mitarbeiter_id=? AND datum BETWEEN ? AND ? ORDER BY datum",
                (ma_id, zeitraum_start.isoformat(), zeitraum_ende.isoformat())
            )
            by_datum = {r['datum']: r for r in rows}
            if modus == 'woche':
                for i in range(7):
                    d = woche_start + timedelta(days=i)
                    r = by_datum.get(d.isoformat())
                    netto = _az_netto_minuten(r['beginn'], r['ende'], r['pause_minuten']) if r else None
                    if netto:
                        eigene_summe += netto
                    eigene_tage.append({
                        'datum': d.isoformat(), 'tag': tage_de[i],
                        'beginn': r['beginn'] if r else None,
                        'ende': r['ende'] if r else None,
                        'pause_minuten': (r['pause_minuten'] or 0) if r else 0,
                        'netto': netto, 'netto_fmt': _az_fmt_std(netto),
                    })
            else:
                # Monat: pro Kalenderwoche gruppiert
                wochen = {}
                for r in rows:
                    d = date.fromisoformat(r['datum'])
                    kw = d.isocalendar()[1]
                    netto = _az_netto_minuten(r['beginn'], r['ende'], r['pause_minuten'])
                    if kw not in wochen:
                        wochen[kw] = {'kw': kw, 'summe': 0, 'tage': 0}
                    if netto:
                        wochen[kw]['summe'] += netto
                        wochen[kw]['tage']  += 1
                    eigene_summe += netto or 0
                eigene_tage = [
                    {'kw': w['kw'], 'summe': w['summe'], 'summe_fmt': _az_fmt_std(w['summe']), 'tage': w['tage']}
                    for w in sorted(wochen.values(), key=lambda w: w['kw'])
                ]

        if is_manager and not ma_id:
            _tm_sql, _tm_p = _team_m_clause('m')
            rows = query(
                f"SELECT az.mitarbeiter_id, az.datum, az.beginn, az.ende, az.pause_minuten "
                f"FROM arbeitszeit az JOIN mitarbeiter m ON m.id = az.mitarbeiter_id "
                f"WHERE az.datum BETWEEN ? AND ? {_tm_sql}",
                (zeitraum_start.isoformat(), zeitraum_ende.isoformat()) + _tm_p
            )
            by_ma = {}
            for r in rows:
                by_ma.setdefault(r['mitarbeiter_id'], {})[r['datum']] = r
            team_summary = []
            for m in mitarbeiter_liste:
                ma_rows = by_ma.get(m['id'], {})
                summe = 0
                tage_list = []
                if modus == 'woche':
                    for i in range(7):
                        d = woche_start + timedelta(days=i)
                        r = ma_rows.get(d.isoformat())
                        netto = _az_netto_minuten(r['beginn'], r['ende'], r['pause_minuten']) if r else None
                        if netto:
                            summe += netto
                        tage_list.append({
                            'datum': d.isoformat(), 'tag': tage_de[i],
                            'beginn': r['beginn'] if r else None,
                            'ende': r['ende'] if r else None,
                            'pause_minuten': (r['pause_minuten'] or 0) if r else 0,
                            'netto_fmt': _az_fmt_std(netto),
                        })
                else:
                    for r in ma_rows.values():
                        summe += _az_netto_minuten(r['beginn'], r['ende'], r['pause_minuten']) or 0
                team_summary.append({
                    'id': m['id'], 'name': m['name'], 'kuerzel': m['kuerzel'],
                    'summe': summe, 'summe_fmt': _az_fmt_std(summe), 'tage': tage_list,
                })
            team_summary.sort(key=lambda x: x['name'])

    return render_template('arbeitszeit.html',
        is_admin=is_admin, is_manager=is_manager, mitarbeiter_liste=mitarbeiter_liste, ma_id=ma_id,
        modus=modus, eigene_ansicht=eigene_ansicht, today_str=today_str,
        tag_datum=tag_datum_str, tag_prev=tag_prev, tag_next=tag_next,
        tag_eintrag=tag_eintrag, tag_team=tag_team,
        woche_start=woche_start.isoformat(), woche_ende=woche_ende.isoformat(), woche_kw=woche_kw,
        prev_woche=(woche_start - timedelta(days=7)).isoformat(),
        next_woche=(woche_start + timedelta(days=7)).isoformat(),
        monat_str=monat_start.strftime('%Y-%m'), monat_label=monat_label,
        prev_monat=prev_monat.strftime('%Y-%m'), next_monat=next_monat.strftime('%Y-%m'),
        eigene_tage=eigene_tage, eigene_summe=eigene_summe, eigene_summe_fmt=_az_fmt_std(eigene_summe),
        team_summary=team_summary,
    )


@app.route('/tourenplanung')
@login_required
def tourenplanung():
    if session.get('rolle') not in ('admin', 'verkaufsleiter'):
        return redirect(url_for('dashboard'))
    modus = request.args.get('modus', 'tag')
    today = date.today()
    _tm_sql, _tm_p = _team_m_clause('m')
    reps = query(
        f"SELECT id, name, kuerzel, bundesland FROM mitarbeiter m WHERE rolle='rep' AND aktiv=1 {_tm_sql} ORDER BY name",
        _tm_p
    )
    rep_ids = [r['id'] for r in reps]

    # Tag-Modus
    datum = request.args.get('datum', today.isoformat())
    try:
        _datum_d = date.fromisoformat(datum)
    except ValueError:
        _datum_d = today
    _datum_montag  = _datum_d - timedelta(days=_datum_d.weekday())
    datum_woche    = [(_datum_montag + timedelta(days=i)).isoformat() for i in range(7)]
    tag_kw         = _datum_montag.isocalendar()[1]
    tag_prev_kw    = tag_kw - 1 if tag_kw > 1 else 52
    tag_next_kw    = tag_kw + 1 if tag_kw < 52 else 1
    tag_prev_datum = (_datum_d - timedelta(days=7)).isoformat()
    tag_next_datum = (_datum_d + timedelta(days=7)).isoformat()
    urlaub_tag = _urlaub_daten(rep_ids, datum, datum) if modus == 'tag' else set()
    # Feiertag statt "Kein Plan" anzeigen – Bundesland ist je Mitarbeiter individuell.
    feiertag_tag = {r['id'] for r in reps if datum in _feiertage_set(_datum_d.year, r['bundesland'] or 'BY')}
    plan_tag = query(f'''
        SELECT tp.id, tp.datum, tp.reihenfolge, tp.notiz, tp.erledigt,
               COALESCE(tp.geloescht, 0) AS geloescht, tp.geloescht_am,
               v.name AS station, v.plz, v.ort, v.id AS vs_id,
               v.lieferant, v.ansprechpartner, v.hinweis,
               m.name AS mitarbeiter, m.kuerzel, m.id AS ma_id,
               (SELECT a.von_uhrzeit FROM aktivitaet a
                WHERE a.mitarbeiter_id = tp.mitarbeiter_id
                  AND a.verkaufsstelle_id = tp.verkaufsstelle_id
                  AND a.datum = tp.datum ORDER BY a.erstellt_am DESC LIMIT 1) AS von_uhrzeit,
               (SELECT a.bis_uhrzeit FROM aktivitaet a
                WHERE a.mitarbeiter_id = tp.mitarbeiter_id
                  AND a.verkaufsstelle_id = tp.verkaufsstelle_id
                  AND a.datum = tp.datum ORDER BY a.erstellt_am DESC LIMIT 1) AS bis_uhrzeit
        FROM tagesplan tp
        JOIN verkaufsstelle v ON v.id = tp.verkaufsstelle_id
        JOIN mitarbeiter m ON m.id = tp.mitarbeiter_id
        WHERE tp.datum = ? {_tm_sql.replace('AND', 'AND', 1)}
        ORDER BY m.name, (von_uhrzeit IS NULL), von_uhrzeit, tp.reihenfolge, tp.id
    ''', (datum,) + _tm_p) if modus == 'tag' else []

    # Wochen-Modus
    woche_start_str = request.args.get('woche', None)
    if woche_start_str:
        try:
            woche_start = date.fromisoformat(woche_start_str)
        except ValueError:
            woche_start = today - timedelta(days=today.weekday())
    else:
        woche_start = today - timedelta(days=today.weekday())
    woche_start = woche_start - timedelta(days=woche_start.weekday())  # ensure Monday
    woche_ende = woche_start + timedelta(days=6)
    woche_tage = [(woche_start + timedelta(days=i)).isoformat() for i in range(7)]
    plan_woche = query(f'''
        SELECT tp.id, tp.datum, tp.reihenfolge, tp.notiz, tp.erledigt,
               COALESCE(tp.geloescht, 0) AS geloescht, tp.geloescht_am,
               v.name AS station, v.plz, v.ort, v.id AS vs_id,
               v.lieferant, v.ansprechpartner, v.hinweis,
               m.name AS mitarbeiter, m.kuerzel, m.id AS ma_id,
               (SELECT a.von_uhrzeit FROM aktivitaet a
                WHERE a.mitarbeiter_id = tp.mitarbeiter_id
                  AND a.verkaufsstelle_id = tp.verkaufsstelle_id
                  AND a.datum = tp.datum ORDER BY a.erstellt_am DESC LIMIT 1) AS von_uhrzeit,
               (SELECT a.bis_uhrzeit FROM aktivitaet a
                WHERE a.mitarbeiter_id = tp.mitarbeiter_id
                  AND a.verkaufsstelle_id = tp.verkaufsstelle_id
                  AND a.datum = tp.datum ORDER BY a.erstellt_am DESC LIMIT 1) AS bis_uhrzeit
        FROM tagesplan tp
        JOIN verkaufsstelle v ON v.id = tp.verkaufsstelle_id
        JOIN mitarbeiter m ON m.id = tp.mitarbeiter_id
        WHERE tp.datum >= ? AND tp.datum <= ? {_tm_sql}
        ORDER BY m.name, tp.datum, (von_uhrzeit IS NULL), von_uhrzeit, tp.reihenfolge, tp.id
    ''', (woche_start.isoformat(), woche_ende.isoformat()) + _tm_p) if modus == 'woche' else []
    urlaub_woche = _urlaub_daten(rep_ids, woche_start.isoformat(), woche_ende.isoformat()) if modus == 'woche' else set()
    urlaub_ganze_woche = {
        ma_id for ma_id in rep_ids
        if all((ma_id, tag) in urlaub_woche for tag in woche_tage)
    } if modus == 'woche' else set()
    # Feiertag statt "kein Plan" je Tag/Mitarbeiter im Wochen-Modus – Bundesland individuell.
    feiertag_woche = {
        (r['id'], tag) for r in reps for tag in woche_tage
        if tag in _feiertage_set(date.fromisoformat(tag).year, r['bundesland'] or 'BY')
    } if modus == 'woche' else set()

    return render_template('tourenplanung.html',
        reps=reps,
        modus=modus,
        # Tag
        datum=datum,
        datum_woche=datum_woche,
        tag_kw=tag_kw,
        tag_prev_kw=tag_prev_kw,
        tag_next_kw=tag_next_kw,
        tag_prev_datum=tag_prev_datum,
        tag_next_datum=tag_next_datum,
        plan_tag=plan_tag,
        urlaub_tag=urlaub_tag,
        feiertag_tag=feiertag_tag,
        today_str=today.isoformat(),
        tomorrow_str=(today + timedelta(days=1)).isoformat(),
        # Woche
        woche_start=woche_start.isoformat(),
        woche_ende=woche_ende.isoformat(),
        woche_tage=woche_tage,
        plan_woche=plan_woche,
        urlaub_woche=urlaub_woche,
        urlaub_ganze_woche=urlaub_ganze_woche,
        feiertag_woche=feiertag_woche,
        prev_woche=(woche_start - timedelta(days=7)).isoformat(),
        next_woche=(woche_start + timedelta(days=7)).isoformat(),
    )


@app.route('/tourenplanung/neu', methods=['POST'])
@login_required
def tourenplanung_neu():
    is_manager = session.get('rolle') in ('admin', 'verkaufsleiter')
    is_rep = session.get('rolle') == 'rep'
    if not is_manager and not is_rep:
        abort(403)
    ma_id  = request.form.get('mitarbeiter_id', type=int)
    vs_ids = request.form.getlist('verkaufsstelle_id')
    datum  = request.form.get('datum', date.today().isoformat()).strip()
    notiz  = request.form.get('notiz', '').strip()
    # Reps dürfen nur für sich selbst planen
    if is_rep and ma_id != session['user_id']:
        abort(403)
    # VKL/Admin, die für sich SELBST planen (eigene Dashboard-Ansicht), landen wie
    # ein Mitarbeiter zurück im Dashboard statt auf der Team-Tourenplanung-Seite.
    eigene_planung = ma_id == session['user_id'] and session.get('rolle') != 'admin'
    if not ma_id or not vs_ids or not datum:
        flash('Bitte alle Pflichtfelder ausfüllen.', 'warning')
        if is_manager and not eigene_planung:
            return redirect(url_for('tourenplanung', datum=datum))
        return redirect(url_for('dashboard'))
    # VKL darf nur für Mitarbeiter des eigenen Teams planen (Bugreport 2026-07-21: bislang
    # ohne Team-Check).
    if session.get('rolle') == 'verkaufsleiter' and session.get('team_id'):
        _ma_team = query("SELECT team_id FROM mitarbeiter WHERE id=?", (ma_id,), one=True)
        if not _ma_team or _ma_team['team_id'] != session.get('team_id'):
            abort(403)
    for vs_id in vs_ids:
        max_r = query(
            "SELECT COALESCE(MAX(reihenfolge), 0) AS m FROM tagesplan WHERE mitarbeiter_id=? AND datum=?",
            (ma_id, datum), one=True
        )['m']
        execute(
            "INSERT INTO tagesplan (mitarbeiter_id, verkaufsstelle_id, datum, reihenfolge, notiz, erstellt_von) VALUES (?,?,?,?,?,?)",
            (ma_id, int(vs_id), datum, max_r + 1, notiz or None, session['user_id'])
        )
    if is_manager and not eigene_planung:
        return redirect(url_for('tourenplanung', datum=datum, ma=ma_id))
    return redirect(url_for('dashboard') + '#tab-tagesplan-btn')


@app.route('/api/tagesplan/stopp/neu', methods=['POST'])
@login_required
def api_tagesplan_stopp_neu():
    if session.get('rolle') not in ('rep', 'verkaufsleiter', 'admin'):
        return jsonify({'ok': False, 'error': 'Kein Zugriff'}), 403
    data  = request.get_json(silent=True) or {}
    vs_id = data.get('vs_id')
    datum = data.get('datum', date.today().isoformat())
    if not vs_id:
        return jsonify({'ok': False, 'error': 'Keine Verkaufsstelle'})
    ma_id = session['user_id']
    max_r = query(
        "SELECT COALESCE(MAX(reihenfolge), 0) AS m FROM tagesplan WHERE mitarbeiter_id=? AND datum=?",
        (ma_id, datum), one=True
    )['m']
    execute(
        "INSERT INTO tagesplan (mitarbeiter_id, verkaufsstelle_id, datum, reihenfolge, erstellt_von) VALUES (?,?,?,?,?)",
        (ma_id, int(vs_id), datum, max_r + 1, ma_id)
    )
    return jsonify({'ok': True})


@app.route('/api/tagesplan/stopp/<int:tp_id>/details', methods=['GET'])
@login_required
def api_tagesplan_stopp_details(tp_id):
    row = query(
        "SELECT mitarbeiter_id, verkaufsstelle_id, datum FROM tagesplan WHERE id=?",
        (tp_id,), one=True
    )
    if not row:
        return jsonify({'ok': False, 'error': 'Nicht gefunden'}), 404
    if not _mitarbeiter_im_eigenen_team(row['mitarbeiter_id']):
        return jsonify({'ok': False, 'error': 'Kein Zugriff'}), 403
    akt = query(
        "SELECT id, COALESCE(aktionstyp,'Aufbau') AS aktionstyp, notizen, anzahl_displays, "
        "foto_pfad, foto_pfad_2, foto_pfad_3 FROM aktivitaet "
        "WHERE mitarbeiter_id=? AND verkaufsstelle_id=? AND datum=? ORDER BY erstellt_am DESC LIMIT 1",
        (row['mitarbeiter_id'], row['verkaufsstelle_id'], row['datum']), one=True
    )
    if not akt:
        return jsonify({'ok': True, 'gefunden': False})
    bestellungen = query(
        "SELECT b.name, bp.kisten_anzahl FROM bestellposition bp "
        "JOIN biersorte b ON b.id=bp.biersorte_id WHERE bp.aktivitaet_id=? ORDER BY b.name",
        (akt['id'],)
    )
    displays = query(
        "SELECT d.name, dp.anzahl FROM displayposition dp "
        "JOIN displaysorte d ON d.id=dp.displaysorte_id WHERE dp.aktivitaet_id=? AND dp.anzahl>0 ORDER BY d.name",
        (akt['id'],)
    )
    return jsonify({
        'ok': True, 'gefunden': True,
        'aktionstyp': akt['aktionstyp'],
        'notizen': akt['notizen'] or '',
        'anzahl_displays': akt['anzahl_displays'] or 0,
        'bestellungen': [{'name': b['name'], 'kisten': b['kisten_anzahl']} for b in bestellungen],
        'displays':     [{'name': d['name'], 'anzahl': d['anzahl']}        for d in displays],
        'foto_pfad':    akt['foto_pfad'] or '',
        'fotos':        [p for p in (akt['foto_pfad'], akt['foto_pfad_2'], akt['foto_pfad_3']) if p],
    })


@app.route('/tourenplanung/<int:tp_id>/loeschen', methods=['POST'])
@login_required
def tourenplanung_loeschen(tp_id):
    row = query("SELECT datum, mitarbeiter_id, erledigt FROM tagesplan WHERE id=?", (tp_id,), one=True)
    if not row:
        abort(404)
    is_manager = session.get('rolle') in ('admin', 'verkaufsleiter')
    if not _mitarbeiter_im_eigenen_team(row['mitarbeiter_id']):
        abort(403)
    if not is_manager and (row['erledigt'] or row['datum'] < date.today().isoformat()):
        abort(403)
    execute(
        "UPDATE tagesplan SET geloescht=1, geloescht_am=datetime('now','localtime') WHERE id=?",
        (tp_id,)
    )
    if is_manager:
        return redirect(url_for('tourenplanung', datum=row['datum']))
    return redirect(url_for('dashboard') + '#tab-tagesplan-btn')


@app.route('/tourenplanung/<int:tp_id>/erledigt', methods=['POST'])
@login_required
def tourenplanung_erledigt(tp_id):
    row = query("SELECT mitarbeiter_id, erledigt FROM tagesplan WHERE id=?", (tp_id,), one=True)
    if not row:
        abort(404)
    if not _mitarbeiter_im_eigenen_team(row['mitarbeiter_id']):
        abort(403)
    execute("UPDATE tagesplan SET erledigt=? WHERE id=?", (0 if row['erledigt'] else 1, tp_id))
    return redirect(request.referrer or url_for('dashboard'))


@app.route('/tourenplanung/<int:tp_id>/reihenfolge', methods=['POST'])
@login_required
def tourenplanung_reihenfolge(tp_id):
    data     = request.get_json(force=True, silent=True) or {}
    richtung = data.get('richtung')
    if richtung not in ('hoch', 'runter'):
        return jsonify({'ok': False})

    cur = query(
        "SELECT id, mitarbeiter_id, datum, reihenfolge FROM tagesplan WHERE id=? AND COALESCE(geloescht,0)=0",
        (tp_id,), one=True
    )
    if not cur:
        return jsonify({'ok': False})

    if not _mitarbeiter_im_eigenen_team(cur['mitarbeiter_id']):
        return jsonify({'ok': False})

    if richtung == 'hoch':
        nb = query('''
            SELECT id, reihenfolge FROM tagesplan
            WHERE mitarbeiter_id=? AND datum=? AND reihenfolge < ? AND COALESCE(geloescht,0)=0
            ORDER BY reihenfolge DESC LIMIT 1
        ''', (cur['mitarbeiter_id'], cur['datum'], cur['reihenfolge']), one=True)
    else:
        nb = query('''
            SELECT id, reihenfolge FROM tagesplan
            WHERE mitarbeiter_id=? AND datum=? AND reihenfolge > ? AND COALESCE(geloescht,0)=0
            ORDER BY reihenfolge ASC LIMIT 1
        ''', (cur['mitarbeiter_id'], cur['datum'], cur['reihenfolge']), one=True)

    if not nb:
        return jsonify({'ok': False})

    db = get_db()
    db.execute("UPDATE tagesplan SET reihenfolge=? WHERE id=?", (nb['reihenfolge'], tp_id))
    db.execute("UPDATE tagesplan SET reihenfolge=? WHERE id=?", (cur['reihenfolge'], nb['id']))
    db.commit()
    return jsonify({'ok': True})


# ─── Routenoptimierung (Tagesplan-Reihenfolgevorschlag) ───────────────────────

_DE_LAT = (47.2, 55.2)
_DE_LNG = (5.8,  15.2)


def _geocode_freitext(text, timeout=5):
    """Geocodiert einen beliebigen Freitext (z.B. Start-/Zieladresse für die
    Routenoptimierung) via Photon, ohne die strikte Orts-Validierung von
    _geocode_adresse (die ist auf strukturierte VS-Adressfelder zugeschnitten).
    Gibt (lat, lng) oder (None, None) zurück."""
    text = (text or '').strip()
    if not text:
        return None, None
    url = ('https://photon.komoot.io/api/?q='
           + urllib.parse.quote(text)
           + '&limit=1&lang=de&bbox=5.8,47.2,15.2,55.2')
    try:
        req = urllib.request.Request(
            url, headers={'User-Agent': 'AktionsTracker/1.0'}
        )
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            data = json.loads(resp.read().decode())
        features = data.get('features', [])
        if features:
            coords = features[0]['geometry']['coordinates']
            lat, lng = float(coords[1]), float(coords[0])  # GeoJSON: [lng, lat]
            if _DE_LAT[0] <= lat <= _DE_LAT[1] and _DE_LNG[0] <= lng <= _DE_LNG[1]:
                return lat, lng
    except Exception as exc:
        app.logger.warning(f"Geocode Freitext '{text}': {exc}")
    return None, None


def _haversine_km(lat1, lng1, lat2, lng2):
    """Luftlinien-Distanz zwischen zwei Koordinaten in km."""
    R = 6371.0
    phi1, phi2 = math.radians(lat1), math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlambda = math.radians(lng2 - lng1)
    a = math.sin(dphi / 2) ** 2 + math.cos(phi1) * math.cos(phi2) * math.sin(dlambda / 2) ** 2
    return 2 * R * math.asin(math.sqrt(a))


def _route_optimieren(stop_coords, start_coord=None, ende_coord=None):
    """Nearest-Neighbor + 2-opt auf Luftlinien-Distanz. stop_coords: Liste von
    (lat,lng) der zu sortierenden Stopps. start_coord/ende_coord: optionale
    Fixpunkte, die NICHT Teil des Ergebnisses sind, aber die Distanzberechnung
    an den Rändern beeinflussen (fehlender Fixpunkt = freies Ende). Gibt eine
    Liste von Indizes (bezogen auf stop_coords) in optimierter Reihenfolge
    zurück. Für die üblichen 3-15 Tagesstopps läuft das in Millisekunden."""
    n = len(stop_coords)
    if n <= 1:
        return list(range(n))

    def dist(p1, p2):
        return _haversine_km(p1[0], p1[1], p2[0], p2[1])

    dmat = [[dist(stop_coords[i], stop_coords[j]) for j in range(n)] for i in range(n)]

    start_idx = 0
    if start_coord:
        start_idx = min(range(n), key=lambda i: dist(start_coord, stop_coords[i]))
    unbesucht = set(range(n)) - {start_idx}
    route = [start_idx]
    aktuell = start_idx
    while unbesucht:
        naechster = min(unbesucht, key=lambda j: dmat[aktuell][j])
        route.append(naechster)
        unbesucht.remove(naechster)
        aktuell = naechster

    def routen_laenge(r):
        total = dist(start_coord, stop_coords[r[0]]) if start_coord else 0.0
        for i in range(len(r) - 1):
            total += dmat[r[i]][r[i + 1]]
        if ende_coord:
            total += dist(stop_coords[r[-1]], ende_coord)
        return total

    aktuelle_laenge = routen_laenge(route)
    verbessert = True
    while verbessert:
        verbessert = False
        for i in range(n - 1):
            for j in range(i + 1, n):
                neu = route[:i] + route[i:j + 1][::-1] + route[j + 1:]
                neue_laenge = routen_laenge(neu)
                if neue_laenge < aktuelle_laenge - 1e-9:
                    route = neu
                    aktuelle_laenge = neue_laenge
                    verbessert = True
    return route


MAPBOX_MAX_PUNKTE = 12  # Mapbox Optimized Trips API v1: max. 12 Koordinaten inkl. Start/Ziel


def _route_optimieren_mapbox(stop_coords, start_coord=None, ende_coord=None, timeout=6):
    """Straßenbasierter Routenvorschlag über die Mapbox Optimized Trips API
    (v1, max. MAPBOX_MAX_PUNKTE Koordinaten inkl. Start/Ziel). Gibt eine Liste
    von Indizes (bezogen auf stop_coords) zurück, oder None bei fehlendem
    Token/Limit-Überschreitung/jedem Fehler – der Aufrufer fällt dann auf
    _route_optimieren() (Luftlinie) zurück, damit das Feature nie hart
    fehlschlägt, nur weil Mapbox gerade nicht erreichbar ist."""
    token = os.environ.get('MAPBOX_ACCESS_TOKEN', '').strip()
    if not token:
        return None

    n = len(stop_coords)
    if n == 0:
        return []
    if n == 1:
        return [0]

    # Die Mapbox Optimized-Trips-API unterstützt bei roundtrip=false NUR die
    # Kombination source=first + destination=last (beide gesetzt) – ein
    # einseitiger Anker (nur Start ODER nur Ziel, freies anderes Ende) liefert
    # "NotImplemented". Für diesen selteneren Fall daher bewusst kein Mapbox-
    # Aufruf, sondern direkter Rückfall auf die Luftlinien-Berechnung.
    if bool(start_coord) != bool(ende_coord):
        return None

    punkte = []
    if start_coord:
        punkte.append(start_coord)
    punkte += list(stop_coords)
    ende_separat = bool(ende_coord) and ende_coord != start_coord
    if ende_separat:
        punkte.append(ende_coord)

    if len(punkte) > MAPBOX_MAX_PUNKTE:
        return None

    koord_str = ';'.join(f"{lng},{lat}" for lat, lng in punkte)
    params = {'access_token': token, 'overview': 'false'}
    if start_coord and ende_separat:
        params['source'] = 'first'
        params['destination'] = 'last'
        params['roundtrip'] = 'false'
    elif start_coord and ende_coord:
        # Gleicher Start-/Zielpunkt (z.B. Homeoffice beide Richtungen) -> Rundtour
        params['source'] = 'first'
        params['roundtrip'] = 'true'
    else:
        # Weder Start noch Ziel vorgegeben -> freie Optimierung ohne Anker
        params['roundtrip'] = 'true'

    url = ('https://api.mapbox.com/optimized-trips/v1/mapbox/driving/'
           + koord_str + '?' + urllib.parse.urlencode(params))
    try:
        req = urllib.request.Request(url, headers={'User-Agent': 'AktionsTracker/1.0'})
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            data = json.loads(resp.read().decode())
        if data.get('code') != 'Ok' or not data.get('trips'):
            return None
        waypoints = data.get('waypoints', [])
        eingabe_offset = 1 if start_coord else 0
        stop_eintraege = []
        for punkt_idx, wp in enumerate(waypoints):
            if punkt_idx < eingabe_offset:
                continue
            if ende_separat and punkt_idx == len(punkte) - 1:
                continue
            stop_idx = punkt_idx - eingabe_offset
            stop_eintraege.append((wp['waypoint_index'], stop_idx))
        stop_eintraege.sort(key=lambda t: t[0])
        return [stop_idx for _, stop_idx in stop_eintraege]
    except Exception as exc:
        app.logger.warning(f"Mapbox Routenoptimierung fehlgeschlagen: {exc}")
        return None


def _homeoffice_punkt(ma_id):
    """Koordinaten der Homeoffice-VS eines Mitarbeiters, falls hinterlegt und
    bereits geocodiert. Gibt (lat, lng) oder None zurück."""
    ma = query("SELECT homeoffice_vs_id FROM mitarbeiter WHERE id=?", (ma_id,), one=True)
    if not ma or not ma['homeoffice_vs_id']:
        return None
    vs = query("SELECT lat, lng FROM verkaufsstelle WHERE id=?", (ma['homeoffice_vs_id'],), one=True)
    if vs and vs['lat'] is not None and vs['lng'] is not None:
        return (vs['lat'], vs['lng'])
    return None


@app.route('/api/tourenplanung/route-vorschlag', methods=['POST'])
@login_required
def api_route_vorschlag():
    """Berechnet einen Reihenfolge-Vorschlag für die Tagesplanung eines Tages.
    Ändert nichts an der DB – reine Berechnung, Anwenden erfolgt separat über
    api_route_anwenden()."""
    data  = request.get_json(silent=True) or {}
    datum = (data.get('datum') or '').strip()
    try:
        ma_id = int(data.get('mitarbeiter_id') or session['user_id'])
    except (TypeError, ValueError):
        return jsonify({'ok': False, 'error': 'Ungültige Mitarbeiter-ID'}), 400
    if ma_id != session['user_id'] and not _mitarbeiter_im_eigenen_team(ma_id):
        return jsonify({'ok': False, 'error': 'Keine Berechtigung'}), 403
    if not datum:
        return jsonify({'ok': False, 'error': 'Datum fehlt'}), 400

    stops = query(
        "SELECT tp.id AS tp_id, v.name, v.strasse, v.plz, v.ort, v.lat, v.lng FROM tagesplan tp "
        "JOIN verkaufsstelle v ON v.id = tp.verkaufsstelle_id "
        "WHERE tp.mitarbeiter_id=? AND tp.datum=? AND COALESCE(tp.geloescht,0)=0 "
        "ORDER BY tp.reihenfolge, tp.id",
        (ma_id, datum)
    )
    ohne_koordinaten = [s['name'] for s in stops if s['lat'] is None or s['lng'] is None]
    stops = [s for s in stops if s['lat'] is not None and s['lng'] is not None]
    if len(stops) < 2:
        return jsonify({'ok': False, 'error': 'Mindestens 2 Stopps mit Koordinaten nötig für einen Vorschlag.',
                        'ohne_koordinaten': ohne_koordinaten})

    start_text = (data.get('start') or '').strip()
    ziel_text  = (data.get('ziel') or '').strip()
    start_punkt = _geocode_freitext(start_text) if start_text else _homeoffice_punkt(ma_id)
    if start_punkt == (None, None):
        start_punkt = None
    ziel_punkt = _geocode_freitext(ziel_text) if ziel_text else _homeoffice_punkt(ma_id)
    if ziel_punkt == (None, None):
        ziel_punkt = None

    coords = [(s['lat'], s['lng']) for s in stops]

    # Mapbox-Straßenrouting versuchen (genauer als Luftlinie), bei fehlendem
    # Token/zu vielen Punkten/jedem API-Fehler automatisch auf die Luftlinien-
    # Berechnung zurückfallen – das Feature darf nie hart fehlschlagen.
    mapbox_token_vorhanden = bool(os.environ.get('MAPBOX_ACCESS_TOKEN', '').strip())
    anzahl_punkte = len(coords) + (1 if start_punkt else 0) + (1 if (ziel_punkt and ziel_punkt != start_punkt) else 0)
    order = None
    verfahren = 'luftlinie'
    zu_viele_fuer_mapbox = mapbox_token_vorhanden and anzahl_punkte > MAPBOX_MAX_PUNKTE
    if mapbox_token_vorhanden and not zu_viele_fuer_mapbox:
        order = _route_optimieren_mapbox(coords, start_coord=start_punkt, ende_coord=ziel_punkt)
        if order is not None:
            verfahren = 'strasse'
    if order is None:
        order = _route_optimieren(coords, start_coord=start_punkt, ende_coord=ziel_punkt)
        verfahren = 'luftlinie'

    def _adresse(s):
        teile = [t for t in [s['strasse'], (f"{s['plz'] or ''} {s['ort'] or ''}").strip()] if t]
        return ', '.join(teile)

    return jsonify({
        'ok': True,
        'reihenfolge': [stops[i]['tp_id'] for i in order],
        'namen': [stops[i]['name'] for i in order],
        'adressen': [_adresse(stops[i]) for i in order],
        'start_gefunden': start_punkt is not None,
        'ziel_gefunden': ziel_punkt is not None,
        'ohne_koordinaten': ohne_koordinaten,
        'verfahren': verfahren,
        'zu_viele_fuer_strassenrouting': zu_viele_fuer_mapbox,
    })


@app.route('/api/tourenplanung/route-anwenden', methods=['POST'])
@login_required
def api_route_anwenden():
    """Übernimmt eine vorgeschlagene Reihenfolge – schreibt dieselben
    tagesplan.reihenfolge-Werte, die auch die Auf/Ab-Pfeile setzen."""
    data = request.get_json(silent=True) or {}
    tp_ids = data.get('reihenfolge')
    if not tp_ids or not isinstance(tp_ids, list):
        return jsonify({'ok': False, 'error': 'Keine Reihenfolge übergeben'}), 400
    try:
        tp_ids = [int(x) for x in tp_ids]
    except (TypeError, ValueError):
        return jsonify({'ok': False, 'error': 'Ungültige IDs'}), 400

    ph = ','.join('?' * len(tp_ids))
    rows = query(f"SELECT id, mitarbeiter_id FROM tagesplan WHERE id IN ({ph})", tp_ids)
    gefunden = {r['id'] for r in rows}
    if gefunden != set(tp_ids):
        return jsonify({'ok': False, 'error': 'Ungültige oder nicht mehr vorhandene Stopps'}), 400
    eigentuemer = {r['mitarbeiter_id'] for r in rows}
    fremde = {m for m in eigentuemer if m != session['user_id'] and not _mitarbeiter_im_eigenen_team(m)}
    if fremde:
        return jsonify({'ok': False, 'error': 'Keine Berechtigung'}), 403

    for idx, tp_id in enumerate(tp_ids, start=1):
        execute("UPDATE tagesplan SET reihenfolge=? WHERE id=?", (idx, tp_id))
    return jsonify({'ok': True})


@app.route('/api/tourenplanung/letzter-stopp-vortag')
@login_required
def api_letzter_stopp_vortag():
    """Für den 'Wie gestern'-Button: letzter Stopp des Vortages als Textadresse,
    damit sie ins Start-Feld übernommen werden kann (wird beim Berechnen wie
    jede andere Freitext-Adresse geocodiert)."""
    datum = (request.args.get('datum') or '').strip()
    try:
        ma_id = int(request.args.get('mitarbeiter_id') or session['user_id'])
    except (TypeError, ValueError):
        return jsonify({'ok': False, 'error': 'Ungültige Mitarbeiter-ID'}), 400
    if ma_id != session['user_id'] and not _mitarbeiter_im_eigenen_team(ma_id):
        return jsonify({'ok': False, 'error': 'Keine Berechtigung'}), 403
    try:
        d = date.fromisoformat(datum)
    except ValueError:
        return jsonify({'ok': False, 'error': 'Ungültiges Datum'}), 400

    vortag = (d - timedelta(days=1)).isoformat()
    row = query(
        "SELECT v.name, v.strasse, v.plz, v.ort FROM tagesplan tp "
        "JOIN verkaufsstelle v ON v.id = tp.verkaufsstelle_id "
        "WHERE tp.mitarbeiter_id=? AND tp.datum=? AND COALESCE(tp.geloescht,0)=0 "
        "ORDER BY tp.reihenfolge DESC, tp.id DESC LIMIT 1",
        (ma_id, vortag), one=True
    )
    if not row:
        return jsonify({'ok': False, 'error': f'Kein Stopp am {vortag} gefunden.'})
    adresse = ', '.join(filter(None, [row['strasse'], ' '.join(filter(None, [row['plz'], row['ort']]))]))
    return jsonify({'ok': True, 'name': row['name'], 'adresse': adresse})


@app.route('/api/verkaufsstelle/<int:vs_id>/aktivitaeten')
@login_required
def api_vs_aktivitaeten(vs_id):
    rolle  = session.get('rolle')
    ma_id  = session.get('user_id')
    if rolle == 'rep':
        ok = query(
            "SELECT 1 FROM mitarbeiter_verkaufsstelle WHERE mitarbeiter_id=? AND verkaufsstelle_id=?",
            (ma_id, vs_id), one=True
        )
        if not ok:
            return jsonify({'ok': False, 'error': 'Kein Zugriff'})
    elif rolle not in ('admin', 'verkaufsleiter'):
        return jsonify({'ok': False, 'error': 'Kein Zugriff'})

    vs = query("SELECT name, ort FROM verkaufsstelle WHERE id=? AND aktiv=1", (vs_id,), one=True)
    if not vs:
        return jsonify({'ok': False, 'error': 'Nicht gefunden'})

    rows = query('''
        SELECT a.datum, m.name AS mitarbeiter, m.kuerzel,
               COALESCE(a.aktionstyp, 'Besuch') AS aktionstyp,
               a.anzahl_displays, a.notizen,
               COALESCE(
                   (SELECT GROUP_CONCAT(bs.name||' '||bp.kisten_anzahl, ', ')
                    FROM bestellposition bp
                    JOIN biersorte bs ON bs.id = bp.biersorte_id
                    WHERE bp.aktivitaet_id = a.id),
                   ''
               ) AS bestellungen
        FROM aktivitaet a
        JOIN mitarbeiter m ON m.id = a.mitarbeiter_id
        WHERE a.verkaufsstelle_id = ?
        ORDER BY a.datum DESC, a.erstellt_am DESC
        LIMIT 30
    ''', (vs_id,))

    return jsonify({
        'ok':   True,
        'name': vs['name'],
        'ort':  vs['ort'] or '',
        'data': [dict(r) for r in rows]
    })


# ─── API: Letzter Besuch ─────────────────────────────────────────────────────

@app.route('/api/letzter-besuch/<int:vs_id>')
@login_required
def api_letzter_besuch(vs_id):
    """Gibt die letzten 3 Besuche bei einer Verkaufsstelle zurück.
    Manager sehen alle Reps, normale Reps nur ihre eigenen."""
    ma_id = session['user_id']
    is_manager = session.get('rolle') in ('admin', 'verkaufsleiter')
    if is_manager:
        rows = query('''
            SELECT a.id, a.datum, a.anzahl_displays, a.notizen,
                   m.name AS mitarbeiter,
                   COALESCE(a.aktionstyp, 'Aufbau') AS aktionstyp,
                   COALESCE((SELECT SUM(bp.kisten_anzahl) FROM bestellposition bp
                             WHERE bp.aktivitaet_id = a.id), 0) AS kisten_gesamt
            FROM aktivitaet a
            JOIN mitarbeiter m ON m.id = a.mitarbeiter_id
            WHERE a.verkaufsstelle_id = ?
            ORDER BY a.datum DESC, a.id DESC LIMIT 3
        ''', (vs_id,))
    else:
        rows = query('''
            SELECT a.id, a.datum, a.anzahl_displays, a.notizen,
                   NULL AS mitarbeiter,
                   COALESCE(a.aktionstyp, 'Aufbau') AS aktionstyp,
                   COALESCE((SELECT SUM(bp.kisten_anzahl) FROM bestellposition bp
                             WHERE bp.aktivitaet_id = a.id), 0) AS kisten_gesamt
            FROM aktivitaet a
            WHERE a.verkaufsstelle_id = ? AND a.mitarbeiter_id = ?
            ORDER BY a.datum DESC, a.id DESC LIMIT 3
        ''', (vs_id, ma_id))
    if not rows:
        return jsonify({'besuche': []})

    # Einzelpositionen des neuesten Besuchs für "Letzte Bestellung übernehmen"
    neueste_id = rows[0]['id']
    bier_pos = query(
        "SELECT biersorte_id, kisten_anzahl FROM bestellposition WHERE aktivitaet_id = ?",
        (neueste_id,)
    )
    disp_pos = query(
        "SELECT displaysorte_id, anzahl FROM displayposition WHERE aktivitaet_id = ?",
        (neueste_id,)
    )
    letzte_bestellung = {
        'bier':     {str(r['biersorte_id']):   r['kisten_anzahl'] for r in bier_pos},
        'displays': {str(r['displaysorte_id']): r['anzahl']        for r in disp_pos},
    }

    besuche = []
    for i, row in enumerate(rows):
        try:
            tage = (date.today() - date.fromisoformat(row['datum'])).days
        except Exception:
            tage = '?'
        try:
            d = date.fromisoformat(row['datum'])
            datum_fmt = f"{d.day:02d}.{d.month:02d}.{d.year}"
        except Exception:
            datum_fmt = row['datum']
        notizen = (row['notizen'] or '').strip()
        if len(notizen) > 60:
            notizen = notizen[:60] + '…'
        entry = {
            'datum':       datum_fmt,
            'tage_ago':    tage,
            'displays':    row['anzahl_displays'] or 0,
            'kisten':      row['kisten_gesamt']   or 0,
            'notizen':     notizen,
            'mitarbeiter': row['mitarbeiter'] or None,
            'aktionstyp':  row['aktionstyp'] or 'Aufbau',
        }
        if i == 0:
            entry['letzte_bestellung'] = letzte_bestellung
        besuche.append(entry)
    return jsonify({'besuche': besuche})


# ─── KONZEPT-V2: Offene Bestellungen einer Station (Pipeline) ──────────────────

@app.route('/api/offene-bestellungen/<int:vs_id>')
@login_required
def api_offene_bestellungen(vs_id):
    """Offene (noch nicht aufgebaute/stornierte) Bestellungen dieser Station. Zugriff nur
    fürs eigene Gebiet (Rep) bzw. Team (VKL) – hatte bisher gar keine Prüfung (Bugreport
    2026-07-20, von Arco portiert: ließ sich für jede beliebige Verkaufsstelle firmenweit
    abrufen)."""
    if not _verkaufsstelle_im_eigenen_gebiet(vs_id):
        return jsonify({'error': 'Kein Zugriff'}), 403
    rows = query('''
        SELECT a.id, a.datum, m.name AS mitarbeiter
        FROM aktivitaet a
        JOIN mitarbeiter m ON m.id = a.mitarbeiter_id
        WHERE a.verkaufsstelle_id = ?
          AND a.aktionstyp = 'Bestellung'
          AND COALESCE(a.bestell_status, 'offen') = 'offen'
        ORDER BY a.datum ASC, a.id ASC
    ''', (vs_id,))
    ergebnis = []
    for r in rows:
        bier = {str(b['biersorte_id']): b['kisten_anzahl'] for b in query(
            "SELECT biersorte_id, kisten_anzahl FROM bestellposition WHERE aktivitaet_id=?", (r['id'],))}
        disp = {str(d['displaysorte_id']): d['anzahl'] for d in query(
            "SELECT displaysorte_id, anzahl FROM displayposition WHERE aktivitaet_id=?", (r['id'],))}
        d_sum, k_sum = sum(disp.values()), sum(bier.values())
        teile = []
        if d_sum: teile.append(f"{d_sum} Display" + ("s" if d_sum != 1 else ""))
        if k_sum: teile.append(f"{k_sum} {UNIT_LABEL}")
        try:
            tage = (date.today() - datetime.strptime(r['datum'], '%Y-%m-%d').date()).days
        except Exception:
            tage = 0
        ergebnis.append({
            'id': r['id'], 'datum': r['datum'], 'tage_ago': tage,
            'mitarbeiter': r['mitarbeiter'], 'bier': bier, 'displays': disp,
            'zusammenfassung': ' · '.join(teile) if teile else 'ohne Mengen',
            'ueberfaellig': tage > 28,
        })
    return jsonify({'bestellungen': ergebnis})


@app.route('/aktivitaet/<int:akt_id>/stornieren', methods=['POST'])
@login_required
def aktivitaet_stornieren(akt_id):
    """Soft-Close: offene Bestellung ohne Aufbau schließen, mit Grund. Rep darf nur
    eigene Bestellungen stornieren, VKL nur die seines Teams (Bugreport 2026-07-21:
    bislang konnte jeder eingeloggte Nutzer jede Bestellung im System stornieren)."""
    grund = request.form.get('grund', '').strip()
    if grund not in ('Nicht/falsch geliefert', 'Fehleingabe', 'Kunde abgesprungen'):
        return jsonify({'ok': False, 'error': 'Ungültiger Grund'}), 400
    if session.get('rolle') == 'rep':
        scope_sql, scope_p = ' AND a.mitarbeiter_id = ?', (session['user_id'],)
    else:
        scope_sql, scope_p = _team_ma_clause('a')
    best = query(
        f"SELECT a.id FROM aktivitaet a WHERE a.id=? AND a.aktionstyp='Bestellung' "
        f"AND COALESCE(a.bestell_status,'offen')='offen'{scope_sql}", (akt_id,) + scope_p, one=True)
    if not best:
        return jsonify({'ok': False, 'error': 'Offene Bestellung nicht gefunden'}), 404
    execute("UPDATE aktivitaet SET bestell_status='storniert', storno_grund=?, realisiert_am=datetime('now','localtime') WHERE id=?", (grund, akt_id))
    return jsonify({'ok': True})


# ─── KONZEPT-V2 Phase 3: Pipeline-Übersicht für VKL/Leitung ───────────────────

def _bestell_kennzahlen():
    """Pipeline-Kennzahlen team-scoped (VKL: eigenes Team, Leitung: alles)."""
    t_sql, t_p = _team_ma_clause('a')
    base = f"FROM aktivitaet a WHERE a.aktionstyp='Bestellung'{t_sql}"
    offen     = query(f"SELECT COUNT(*) AS n {base} AND COALESCE(a.bestell_status,'offen')='offen'", t_p, one=True)['n']
    aufgebaut = query(f"SELECT COUNT(*) AS n {base} AND a.bestell_status='aufgebaut'", t_p, one=True)['n']
    storniert = query(f"SELECT COUNT(*) AS n {base} AND a.bestell_status='storniert'", t_p, one=True)['n']
    ueberfaellig = query(
        f"SELECT COUNT(*) AS n {base} AND COALESCE(a.bestell_status,'offen')='offen' "
        f"AND julianday('now') - julianday(a.datum) > 28", t_p, one=True)['n']
    return offen, aufgebaut, storniert, ueberfaellig


@app.route('/bestellungen')
@login_required
def bestellungen_uebersicht():
    if session.get('rolle') not in ('admin', 'verkaufsleiter'):
        return redirect(url_for('dashboard'))
    t_sql, t_p = _team_ma_clause('a')
    base = f"FROM aktivitaet a WHERE a.aktionstyp='Bestellung'{t_sql}"

    offen, aufgebaut, storniert, ueberfaellig_gesamt = _bestell_kennzahlen()
    gesamt = offen + aufgebaut + storniert
    quote  = round(aufgebaut / gesamt * 100) if gesamt else 0

    dl = query(
        f"SELECT AVG(julianday(a.realisiert_am) - julianday(a.datum)) AS d {base} "
        f"AND a.bestell_status='aufgebaut' AND a.realisiert_am IS NOT NULL", t_p, one=True)['d']
    durchlauf = round(dl) if dl is not None else None

    gruende = query(
        f"SELECT a.storno_grund AS grund, COUNT(*) AS n {base} "
        f"AND a.bestell_status='storniert' AND a.storno_grund IS NOT NULL "
        f"GROUP BY a.storno_grund ORDER BY n DESC", t_p)
    storno_max = max([g['n'] for g in gruende], default=0)

    rows = query(f'''
        SELECT a.id, a.datum, v.name AS station, m.name AS rep, a.anzahl_displays,
               CAST(julianday('now') - julianday(a.datum) AS INTEGER) AS tage,
               COALESCE((SELECT SUM(kisten_anzahl) FROM bestellposition WHERE aktivitaet_id=a.id),0) AS kisten
        FROM aktivitaet a
        JOIN verkaufsstelle v ON v.id = a.verkaufsstelle_id
        JOIN mitarbeiter m ON m.id = a.mitarbeiter_id
        WHERE a.aktionstyp='Bestellung' AND COALESCE(a.bestell_status,'offen')='offen'
          AND julianday('now') - julianday(a.datum) > 28{t_sql}
        ORDER BY tage DESC
        LIMIT 200
    ''', t_p)
    ueberfaellig = []
    for u in rows:
        teile = []
        if u['anzahl_displays']: teile.append(f"{u['anzahl_displays']} Displays")
        if u['kisten']: teile.append(f"{u['kisten']} {UNIT_LABEL}")
        ueberfaellig.append({'station': u['station'], 'rep': u['rep'], 'tage': u['tage'],
                             'menge': ' · '.join(teile) if teile else '–'})

    def _menge(d, k):
        t = []
        if d: t.append(f"{d} Displays")
        if k: t.append(f"{k} {UNIT_LABEL}")
        return ' · '.join(t) if t else '–'

    raw = query(f'''
        SELECT a.id, a.datum, v.name AS station, m.name AS rep, a.anzahl_displays,
               CAST(julianday('now')-julianday(a.datum) AS INTEGER) AS tage,
               COALESCE((SELECT SUM(kisten_anzahl) FROM bestellposition WHERE aktivitaet_id=a.id),0) AS kisten
        FROM aktivitaet a JOIN verkaufsstelle v ON v.id=a.verkaufsstelle_id
        JOIN mitarbeiter m ON m.id=a.mitarbeiter_id
        WHERE a.aktionstyp='Bestellung' AND COALESCE(a.bestell_status,'offen')='offen'{t_sql}
        ORDER BY a.datum ASC
        LIMIT 200
    ''', t_p)
    liste_offen = [{'id': r['id'], 'datum': r['datum'], 'station': r['station'],
                    'rep': r['rep'], 'tage': r['tage'],
                    'menge': _menge(r['anzahl_displays'], r['kisten'])} for r in raw]

    raw = query(f'''
        SELECT a.id, a.datum, a.realisiert_am, v.name AS station, m.name AS rep,
               a.anzahl_displays,
               CAST(julianday(COALESCE(a.realisiert_am,datetime('now')))-julianday(a.datum) AS INTEGER) AS durchlauf_tage,
               COALESCE((SELECT SUM(kisten_anzahl) FROM bestellposition WHERE aktivitaet_id=a.id),0) AS kisten
        FROM aktivitaet a JOIN verkaufsstelle v ON v.id=a.verkaufsstelle_id
        JOIN mitarbeiter m ON m.id=a.mitarbeiter_id
        WHERE a.aktionstyp='Bestellung' AND a.bestell_status='aufgebaut'{t_sql}
        ORDER BY a.realisiert_am DESC
        LIMIT 200
    ''', t_p)
    liste_aufgebaut = [{'id': r['id'], 'datum': r['datum'], 'realisiert_am': r['realisiert_am'],
                        'station': r['station'], 'rep': r['rep'],
                        'durchlauf_tage': r['durchlauf_tage'],
                        'menge': _menge(r['anzahl_displays'], r['kisten'])} for r in raw]

    raw = query(f'''
        SELECT a.id, a.datum, v.name AS station, m.name AS rep, a.anzahl_displays,
               a.storno_grund,
               COALESCE((SELECT SUM(kisten_anzahl) FROM bestellposition WHERE aktivitaet_id=a.id),0) AS kisten
        FROM aktivitaet a JOIN verkaufsstelle v ON v.id=a.verkaufsstelle_id
        JOIN mitarbeiter m ON m.id=a.mitarbeiter_id
        WHERE a.aktionstyp='Bestellung' AND a.bestell_status='storniert'{t_sql}
        ORDER BY a.datum DESC
        LIMIT 200
    ''', t_p)
    liste_storniert = [{'id': r['id'], 'datum': r['datum'], 'station': r['station'],
                        'rep': r['rep'], 'storno_grund': r['storno_grund'],
                        'menge': _menge(r['anzahl_displays'], r['kisten'])} for r in raw]

    return render_template('bestellungen.html',
        offen=offen, aufgebaut=aufgebaut, storniert=storniert, gesamt=gesamt,
        quote=quote, durchlauf=durchlauf, gruende=gruende, storno_max=storno_max,
        ueberfaellig=ueberfaellig, ueberfaellig_gesamt=ueberfaellig_gesamt,
        liste_offen=liste_offen, liste_aufgebaut=liste_aufgebaut,
        liste_storniert=liste_storniert)


@app.route('/api/aktivitaet/offline-sync', methods=['POST'])
@csrf.exempt  # Offline-Warteschlange kann lange nach dem Laden der Seite (Token-Alter
              # unklar bei langer Offline-Zeit) nachsenden; login_required bleibt die
              # eigentliche Zugriffskontrolle, CSRF-Risiko hier gering (nur eigene Aktivität).
@login_required
def api_aktivitaet_offline_sync():
    """Nimmt eine offline gespeicherte Aktivität als JSON (base64-Foto) entgegen."""
    data = request.get_json(force=True, silent=True)
    if not data:
        return jsonify({'ok': False, 'error': 'Kein JSON'}), 400

    datum   = data.get('datum', '').strip()
    vs_id   = data.get('verkaufsstelle_id', '')
    notizen = data.get('notizen', '')
    fotos_b64   = data.get('fotos') or ([data['foto']] if data.get('foto') else [])  # ['data:image/jpeg;base64,...']
    fotos_b64   = fotos_b64[:3]
    displays    = data.get('displays', {})  # {ds_id: menge}
    bier_map    = data.get('bier', {})      # {bier_id: kisten}
    von_uhrzeit = data.get('von_uhrzeit', '') or None
    bis_uhrzeit = data.get('bis_uhrzeit', '') or None

    if not datum or not vs_id:
        return jsonify({'ok': False, 'error': 'Datum und Verkaufsstelle fehlen'}), 400

    # Bugreport 2026-07-21 (Hoch): dieser Endpunkt umging bisher die Pflichtfeld-Regeln
    # von neue_aktivitaet() (Foto-Pflicht, "nur aktuelle Woche"-Datumslimit für Reps) –
    # ein direkter API-Call (curl/Postman) konnte damit z.B. eine Aktivität ganz ohne
    # Foto-Nachweis anlegen. Offline-Sync kennt keinen aktionstyp-Parameter (Default ist
    # 'Aufbau', das striktere Foto-Pflicht hat) – daher hier: Foto Pflicht außer bei
    # Homeoffice-Verkaufsstellen (einzige Sonderkategorie, die der Offline-Pfad kennt).
    _vs_check = query(
        "SELECT homeoffice_mitarbeiter_id FROM verkaufsstelle WHERE id=?", (vs_id,), one=True
    ) if str(vs_id).isdigit() else None
    if not _vs_check:
        return jsonify({'ok': False, 'error': 'Ungültige Verkaufsstelle'}), 400
    _is_homeoffice = bool(_vs_check['homeoffice_mitarbeiter_id'])

    if session.get('rolle') == 'rep':
        _heute = date.today()
        _min_datum = (_heute - timedelta(days=_heute.weekday())).isoformat()
        if datum < _min_datum:
            return jsonify({'ok': False, 'error': 'Nur die aktuelle Woche kann eingetragen werden.'}), 400

    if not fotos_b64 and not _is_homeoffice:
        return jsonify({'ok': False, 'error': 'Foto ist bei dieser Aktivität Pflicht.'}), 400

    # Fotos dekodieren, komprimieren und speichern (max. 3)
    foto_pfade = [None, None, None]
    for i, foto_b64 in enumerate(fotos_b64):
        if not (foto_b64 and ',' in foto_b64):
            continue
        try:
            _, b64data = foto_b64.split(',', 1)
            foto_bytes = base64.b64decode(b64data)
            dateiname  = f"akt_{uuid.uuid4().hex}.jpg"
            ziel       = os.path.join(UPLOAD_FOLDER, dateiname)
            komprimiere_foto(io.BytesIO(foto_bytes), ziel)
            foto_pfade[i] = dateiname
        except Exception as exc:
            app.logger.warning(f"Offline-Sync Foto-Fehler: {exc}")
    foto_pfad, foto_pfad_2, foto_pfad_3 = foto_pfade

    # Freigabe-Workflow: Tier-1-Typen (zaehlt_zur_zielerreichung=1) starten als "offen" und
    # fließen erst nach VKL-Freigabe in anzahl_displays ein – muss identisch zum Online-Pfad
    # (neue_aktivitaet) sein, sonst unterscheidet sich die Zielerreichung je nachdem ob
    # online oder offline erfasst. anzahl_displays startet daher hier bewusst bei 0.
    _tier1_ids = {str(r['id']) for r in query(
        "SELECT id FROM displaysorte WHERE zaehlt_zur_zielerreichung=1"
    )}
    anzahl_displays = 0

    akt_id = execute(
        "INSERT INTO aktivitaet (datum, mitarbeiter_id, verkaufsstelle_id, "
        "anzahl_displays, notizen, foto_pfad, foto_pfad_2, foto_pfad_3, von_uhrzeit, bis_uhrzeit) VALUES (?,?,?,?,?,?,?,?,?,?)",
        (datum, session['user_id'], vs_id, anzahl_displays, notizen, foto_pfad, foto_pfad_2, foto_pfad_3, von_uhrzeit, bis_uhrzeit)
    )

    for ds_id, menge in displays.items():
        menge = int(menge)
        if menge > 0:
            _status = 'offen' if ds_id in _tier1_ids else 'freigegeben'
            execute("INSERT INTO displayposition (aktivitaet_id, displaysorte_id, anzahl, status)"
                    " VALUES (?,?,?,?)", (akt_id, int(ds_id), menge, _status))

    for bier_id, kisten in bier_map.items():
        kisten = int(kisten)
        if kisten > 0:
            execute("INSERT INTO bestellposition (aktivitaet_id, biersorte_id, kisten_anzahl)"
                    " VALUES (?,?,?)", (akt_id, int(bier_id), kisten))

    # Bugreport 2026-07-22: Offline-Sync fasste den Tagesplan bisher gar nicht an – ein
    # geplanter Stopp blieb offen und ein ungeplanter Stopp wurde nicht nachgetragen,
    # obwohl die Aktivität selbst korrekt gespeichert wurde. Identische Logik wie im
    # Online-Pfad (neue_aktivitaet(), siehe dort). Analog zum selben Fix in Arcobräu.
    execute(
        "UPDATE tagesplan SET erledigt=1 WHERE mitarbeiter_id=? AND verkaufsstelle_id=? AND datum=? AND erledigt=0 AND COALESCE(geloescht,0)=0",
        (session['user_id'], vs_id, datum)
    )
    if datum == date.today().isoformat():
        existing = query(
            "SELECT id FROM tagesplan WHERE mitarbeiter_id=? AND verkaufsstelle_id=? AND datum=? AND COALESCE(geloescht,0)=0",
            (session['user_id'], vs_id, datum), one=True
        )
        if not existing:
            execute(
                "INSERT INTO tagesplan (mitarbeiter_id, verkaufsstelle_id, datum, erledigt, erstellt_von) VALUES (?,?,?,1,?)",
                (session['user_id'], vs_id, datum, session['user_id'])
            )

    app.logger.info(f"Offline-Sync: Aktivität {akt_id} für User {session['user_id']} gespeichert")
    return jsonify({'ok': True, 'akt_id': akt_id})


# ─── Routes: Aktivitäten ──────────────────────────────────────────────────────

@app.route('/aktivitaet/neu', methods=['GET', 'POST'])
@login_required
def neue_aktivitaet():
    today = date.today().isoformat()

    # Datum-Mindestgrenze: Mitarbeiter dürfen nur aktuelle Woche eintragen
    is_rep = session.get('rolle') == 'rep'
    if is_rep:
        _d = date.today()
        min_datum = (_d - timedelta(days=_d.weekday())).isoformat()
    else:
        min_datum = None

    # Homeoffice-VS, die das reduzierte Formular auslösen (kein Bestellung/Aufbau-Typ,
    # Foto optional). Rollenbasiert: Admin/VKL erfassen ggf. für andere Mitarbeiter nach,
    # brauchen also mehr als nur ihren eigenen Homeoffice-Stopp.
    _rolle = session.get('rolle')
    if _rolle == 'admin' or (_rolle == 'verkaufsleiter' and not session.get('team_id')):
        homeoffice_vs_ids = [r['id'] for r in query(
            "SELECT id FROM verkaufsstelle WHERE homeoffice_mitarbeiter_id IS NOT NULL"
        )]
    elif _rolle == 'verkaufsleiter':
        homeoffice_vs_ids = [r['id'] for r in query(
            "SELECT id FROM verkaufsstelle WHERE homeoffice_mitarbeiter_id IN "
            "(SELECT id FROM mitarbeiter WHERE team_id=?)", (session['team_id'],)
        )]
    else:
        homeoffice_vs_ids = [r['id'] for r in query(
            "SELECT id FROM verkaufsstelle WHERE homeoffice_mitarbeiter_id=?", (session['user_id'],)
        )]

    # Firmenwagen-Tanken: geteilte Sonderkategorie (nicht personengebunden wie Homeoffice) –
    # unconditional, da für alle Mitarbeiter unabhängig vom eigenen Gebiet sichtbar.
    tanken_vs_ids = [
        r['id'] for r in query("SELECT id FROM verkaufsstelle WHERE typ='Firmenwagen-Tanken' AND aktiv=1")
    ]

    # Mitarbeiter und VKL sehen nur ihre zugeordneten Verkaufsstellen (wenn Zuordnung gesetzt)
    if session.get('rolle') in ('rep', 'verkaufsleiter'):
        assigned = query(
            "SELECT verkaufsstelle_id FROM mitarbeiter_verkaufsstelle WHERE mitarbeiter_id=?",
            (session['user_id'],)
        )
        if assigned:
            vs_ids = [r['verkaufsstelle_id'] for r in assigned]
            ph = ','.join('?' * len(vs_ids))
            verkaufsstellen = query(
                f"SELECT * FROM verkaufsstelle WHERE aktiv=1 AND id IN ({ph}) ORDER BY name",
                vs_ids
            )
        else:
            verkaufsstellen = query("SELECT * FROM verkaufsstelle WHERE aktiv=1 ORDER BY name")
    else:
        verkaufsstellen = query("SELECT * FROM verkaufsstelle WHERE aktiv=1 ORDER BY name")

    # Aktive Vertretungen: zusätzliche VS des Abwesenden einblenden
    vertretungs_gruppen = []  # [{name, vs: [...]}, ...]
    aktive_vtr = query(
        '''SELECT v.id, v.abwesender_id, m.name AS abwesender_name
           FROM vertretung v
           JOIN mitarbeiter m ON m.id = v.abwesender_id
           WHERE v.vertreter_id = ? AND v.von <= ? AND v.bis >= ? AND v.status = 'bestätigt' ''',
        (session['user_id'], today, today)
    )
    eigene_vs_ids = {vs['id'] for vs in verkaufsstellen}
    for vtr in aktive_vtr:
        vtr_assigned = query(
            "SELECT verkaufsstelle_id FROM mitarbeiter_verkaufsstelle WHERE mitarbeiter_id=?",
            (vtr['abwesender_id'],)
        )
        if vtr_assigned:
            vtr_ids = [r['verkaufsstelle_id'] for r in vtr_assigned]
            ph = ','.join('?' * len(vtr_ids))
            vtr_vs = query(
                f"SELECT * FROM verkaufsstelle WHERE aktiv=1 AND id IN ({ph}) ORDER BY name",
                vtr_ids
            )
        else:
            vtr_vs = []
        # Nur VS hinzufügen, die nicht schon im eigenen Gebiet sind
        extra = [v for v in vtr_vs if v['id'] not in eigene_vs_ids]
        if extra:
            vertretungs_gruppen.append({'name': vtr['abwesender_name'], 'vs': extra})

    biersorten      = query("SELECT * FROM biersorte      WHERE aktiv=1 ORDER BY name")
    if not GRATISWARE_MODUS:
        biersorten = [b for b in biersorten if not b['ist_gratisware']]
    displaysorte    = query("SELECT * FROM displaysorte   WHERE aktiv=1 ORDER BY name")

    if request.method == 'POST':
        datum   = request.form.get('datum')
        vs_id   = request.form.get('verkaufsstelle_id')
        notizen = request.form.get('notizen', '')
        aktionstyp = request.form.get('aktionstyp', 'Aufbau')
        if aktionstyp not in ('Aufbau', 'Bestellung', 'Besuch'):
            aktionstyp = 'Aufbau'

        # Homeoffice-Stopp: kein Bestellung/Aufbau, aktionstyp wird server-seitig auf
        # 'Besuch' erzwungen – das macht (über die bestehende Foto-Pflicht-Regel unten,
        # die nur beim Typ 'Aufbau' greift) automatisch auch das Foto optional.
        _vs_homeoffice_check = query(
            "SELECT homeoffice_mitarbeiter_id FROM verkaufsstelle WHERE id=?", (vs_id,), one=True
        ) if vs_id and vs_id.isdigit() else None
        is_homeoffice = bool(_vs_homeoffice_check and _vs_homeoffice_check['homeoffice_mitarbeiter_id'])
        # Firmenwagen-Tanken: wie Homeoffice ein reduziertes Formular (kein Bestellung/
        # Aufbau-Typ), aber Foto bleibt Pflicht (Tankbeleg).
        _vs_typ_check = query(
            "SELECT typ FROM verkaufsstelle WHERE id=?", (vs_id,), one=True
        ) if vs_id and vs_id.isdigit() else None
        is_tanken = bool(_vs_typ_check and _vs_typ_check['typ'] == 'Firmenwagen-Tanken')
        if is_homeoffice or is_tanken:
            aktionstyp = 'Besuch'

        von_uhrzeit = request.form.get('von_uhrzeit', '').strip()
        bis_uhrzeit = request.form.get('bis_uhrzeit', '').strip()
        foto_files = [f for f in request.files.getlist('fotos') if f and f.filename][:3]
        # KONZEPT-V2: Foto ist nur beim Aufbau Pflicht – außer bei Firmenwagen-Tanken,
        # da dort trotz aktionstyp='Besuch' bewusst weiter ein Foto (Tankbeleg) nötig ist.
        if (aktionstyp == 'Aufbau' or is_tanken) and not foto_files:
            _foto_meldung = 'Bitte den Tankbeleg als Foto hochladen.' if is_tanken \
                else 'Bitte ein Foto hochladen – beim Aufbau ist das Foto Pflicht.'
            flash(_foto_meldung, 'danger')
            return render_template('neue_aktivitaet.html',
                verkaufsstellen=verkaufsstellen, biersorten=biersorten,
                displaysorte=displaysorte, vertretungs_gruppen=vertretungs_gruppen,
                heute=date.today().isoformat(), min_datum=min_datum,
                homeoffice_vs_ids=homeoffice_vs_ids, tanken_vs_ids=tanken_vs_ids)

        # Firmenwagen-Tanken: Pflichtfelder aus dem Tankbeleg (Tankdatum, Kennzeichen,
        # Kraftstoffsorte, Menge, Kilometerstand) – ermöglichen den Tanken-Report ohne
        # manuelles Nachschlagen im Foto.
        tank_datum           = request.form.get('tank_datum', '').strip()
        tank_kennzeichen     = request.form.get('tank_kennzeichen', '').strip()
        tank_kraftstoffsorte = request.form.get('tank_kraftstoffsorte', '').strip()
        tank_liter_str       = request.form.get('tank_liter', '').strip()
        tank_km_stand_str    = request.form.get('tank_km_stand', '').strip()
        tank_liter    = None
        tank_km_stand = None
        tank_felder_fehlen = False
        if is_tanken:
            if not (tank_datum and tank_kennzeichen and tank_kraftstoffsorte in ('Diesel', 'Benzin', 'Super')):
                tank_felder_fehlen = True
            try:
                tank_liter = float(tank_liter_str.replace(',', '.'))
                if tank_liter <= 0:
                    tank_felder_fehlen = True
            except ValueError:
                tank_felder_fehlen = True
            try:
                tank_km_stand = int(tank_km_stand_str)
                if tank_km_stand < 0:
                    tank_felder_fehlen = True
            except ValueError:
                tank_felder_fehlen = True
        if tank_felder_fehlen:
            flash('Bitte alle Tankbeleg-Daten vollständig ausfüllen (Tankdatum, Kennzeichen, Kraftstoffsorte, Menge, Kilometerstand).', 'danger')
            return render_template('neue_aktivitaet.html',
                verkaufsstellen=verkaufsstellen, biersorten=biersorten,
                displaysorte=displaysorte, vertretungs_gruppen=vertretungs_gruppen,
                heute=date.today().isoformat(), min_datum=min_datum,
                homeoffice_vs_ids=homeoffice_vs_ids, tanken_vs_ids=tanken_vs_ids)

        if not datum or not vs_id:
            flash('Datum und Verkaufsstelle sind Pflichtfelder.', 'danger')
            return render_template('neue_aktivitaet.html',
                verkaufsstellen=verkaufsstellen, biersorten=biersorten,
                displaysorte=displaysorte, heute=date.today().isoformat(),
                min_datum=min_datum, homeoffice_vs_ids=homeoffice_vs_ids, tanken_vs_ids=tanken_vs_ids)

        if is_rep and min_datum and datum < min_datum:
            flash('Aktivitäten können nur für die aktuelle Woche eingetragen werden.', 'danger')
            return render_template('neue_aktivitaet.html',
                verkaufsstellen=verkaufsstellen, biersorten=biersorten,
                displaysorte=displaysorte, heute=date.today().isoformat(),
                min_datum=min_datum, homeoffice_vs_ids=homeoffice_vs_ids, tanken_vs_ids=tanken_vs_ids)

        # Displaypositionen sammeln. Freigabe-Workflow: zielrelevante (Tier-1) Typen starten
        # als "offen" und fließen erst nach VKL-Freigabe in anzahl_displays ein (s.
        # /aufbauten/freigabe, aufbau_freigeben() erhöht anzahl_displays dann nachträglich) –
        # anzahl_displays startet daher hier bewusst bei 0. Auffüllen-Typen (Tier-2) zählten
        # ohnehin nie zur Zielerreichung und werden automatisch freigegeben.
        anzahl_displays  = 0
        disp_positionen  = []
        for ds in displaysorte:
            menge_str = request.form.get(f'disp_{ds["id"]}', '').strip()
            if menge_str and menge_str.isdigit() and int(menge_str) > 0:
                menge = int(menge_str)
                disp_positionen.append((ds['id'], menge, ds['zaehlt_zur_zielerreichung']))

        # Fotos verarbeiten + komprimieren (max. 3)
        foto_pfade = [None, None, None]
        for i, foto_file in enumerate(foto_files):
            if not (foto_file and foto_file.filename and allowed_file(foto_file.filename)):
                continue
            dateiname = f"akt_{uuid.uuid4().hex}.jpg"
            ziel = os.path.join(UPLOAD_FOLDER, dateiname)
            try:
                komprimiere_foto(foto_file, ziel)
                foto_pfade[i] = dateiname
            except Exception as exc:
                # Kein Fallback auf ungeprüftes Speichern mehr (Bugreport 2026-07-21):
                # Image.open() in komprimiere_foto() ist die einzige echte Content-Prüfung,
                # ob die Datei tatsächlich ein Bild ist – schlägt sie fehl, wird der Upload
                # übersprungen statt die rohen Bytes unter einer Bild-Endung zu speichern.
                app.logger.warning(f"Foto-Upload abgelehnt (keine gültige Bilddatei): {exc}")
        foto_pfad, foto_pfad_2, foto_pfad_3 = foto_pfade

        bestell_status = 'offen' if aktionstyp == 'Bestellung' else None
        akt_id = execute(
            "INSERT INTO aktivitaet (datum, mitarbeiter_id, verkaufsstelle_id, anzahl_displays, notizen, foto_pfad, foto_pfad_2, foto_pfad_3, aktionstyp, bestell_status, von_uhrzeit, bis_uhrzeit, tank_datum, tank_kennzeichen, tank_kraftstoffsorte, tank_liter, tank_km_stand) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (datum, session['user_id'], vs_id, anzahl_displays, notizen, foto_pfad, foto_pfad_2, foto_pfad_3, aktionstyp, bestell_status, von_uhrzeit or None, bis_uhrzeit or None,
             tank_datum or None, tank_kennzeichen or None, tank_kraftstoffsorte or None, tank_liter, tank_km_stand)
        )

        # Displaypositionen speichern – zielrelevante (Tier-1) Typen starten als "offen" und
        # müssen erst vom VKL freigegeben werden, bevor sie in die Zielzahlen einfließen
        # (Freigabe-Workflow, s. /aufbauten/freigabe). Auffüllen-Typen zählen ohnehin nicht
        # zur Zielerreichung und werden automatisch freigegeben (keine Prüfung nötig).
        for ds_id, menge, zaehlt in disp_positionen:
            _status = 'offen' if zaehlt else 'freigegeben'
            execute(
                "INSERT INTO displayposition (aktivitaet_id, displaysorte_id, anzahl, status) VALUES (?,?,?,?)",
                (akt_id, ds_id, menge, _status)
            )

        # Bestellpositionen speichern
        for bier in biersorten:
            menge = request.form.get(f'bier_{bier["id"]}', '').strip()
            if menge and menge.isdigit() and int(menge) > 0:
                execute(
                    "INSERT INTO bestellposition (aktivitaet_id, biersorte_id, kisten_anzahl) VALUES (?,?,?)",
                    (akt_id, bier['id'], int(menge))
                )

        # Optionaler Hinweis an den Admin (z.B. Adress-/Namensänderung der Verkaufsstelle)
        admin_hinweis = request.form.get('admin_hinweis', '').strip()
        if admin_hinweis:
            execute(
                "INSERT INTO vs_hinweis_meldung (verkaufsstelle_id, mitarbeiter_id, aktivitaet_id, text) VALUES (?,?,?,?)",
                (vs_id, session['user_id'], akt_id, admin_hinweis)
            )

        # KONZEPT-V2: aus offenen Bestellungen aufgebaut → diese Bestellungen schließen
        if aktionstyp == 'Aufbau':
            erledigt = request.form.get('erledigt_bestellung_ids', '')
            for bid in [x.strip() for x in erledigt.split(',') if x.strip().isdigit()]:
                execute(
                    "UPDATE aktivitaet SET bestell_status='aufgebaut', realisiert_am=datetime('now','localtime') "
                    "WHERE id=? AND aktionstyp='Bestellung' AND COALESCE(bestell_status,'offen')='offen'",
                    (bid,)
                )

        if foto_pfad:
            cleanup_alte_fotos()

        # Passenden Tagesplan-Stop automatisch als erledigt markieren
        execute(
            "UPDATE tagesplan SET erledigt=1 WHERE mitarbeiter_id=? AND verkaufsstelle_id=? AND datum=? AND erledigt=0 AND COALESCE(geloescht,0)=0",
            (session['user_id'], vs_id, datum)
        )

        # Ungeplante Verkaufsstelle für heute automatisch in den Tagesplan aufnehmen
        if datum == date.today().isoformat():
            existing = query(
                "SELECT id FROM tagesplan WHERE mitarbeiter_id=? AND verkaufsstelle_id=? AND datum=? AND COALESCE(geloescht,0)=0",
                (session['user_id'], vs_id, datum), one=True
            )
            if not existing:
                execute(
                    "INSERT INTO tagesplan (mitarbeiter_id, verkaufsstelle_id, datum, erledigt, erstellt_von) VALUES (?,?,?,1,?)",
                    (session['user_id'], vs_id, datum, session['user_id'])
                )

        flash('Aktivität erfolgreich gespeichert!', 'success')
        return redirect(url_for('neue_aktivitaet'))

    preselect_vs  = request.args.get('vs_id', '', type=str)
    preselect_typ = request.args.get('typ', '', type=str)
    bestellung_id = request.args.get('bestellung_id', '', type=str)

    bestellung_info = None
    if bestellung_id.isdigit():
        _b = query(
            '''SELECT a.id, a.datum, a.notizen, a.verkaufsstelle_id,
                      v.name AS station, v.plz, v.ort
               FROM aktivitaet a JOIN verkaufsstelle v ON v.id=a.verkaufsstelle_id
               WHERE a.id=? AND a.aktionstyp='Bestellung'
                 AND COALESCE(a.bestell_status,'offen')='offen' ''',
            (int(bestellung_id),), one=True
        )
        if _b:
            _pos = query(
                '''SELECT bs.name, bp.kisten_anzahl FROM bestellposition bp
                   JOIN biersorte bs ON bs.id=bp.biersorte_id
                   WHERE bp.aktivitaet_id=? ORDER BY bp.kisten_anzahl DESC''',
                (int(bestellung_id),)
            )
            bestellung_info = dict(_b)
            bestellung_info['positionen'] = _pos
            if not preselect_vs:
                preselect_vs = str(_b['verkaufsstelle_id'])

    tagesplan_heute = []
    if is_rep:
        tagesplan_heute = query('''
            SELECT tp.verkaufsstelle_id AS vs_id, v.name AS station, v.plz, v.ort
            FROM tagesplan tp
            JOIN verkaufsstelle v ON v.id = tp.verkaufsstelle_id
            WHERE tp.mitarbeiter_id = ? AND tp.datum = date('now','localtime') AND tp.erledigt = 0 AND COALESCE(tp.geloescht, 0) = 0
            ORDER BY tp.reihenfolge, tp.id
        ''', (session['user_id'],))

    return render_template('neue_aktivitaet.html',
        min_datum=min_datum,
        verkaufsstellen=verkaufsstellen, biersorten=biersorten,
        displaysorte=displaysorte, vertretungs_gruppen=vertretungs_gruppen,
        heute=date.today().isoformat(), preselect_vs=preselect_vs,
        preselect_typ=preselect_typ,
        bestellung_id=bestellung_id, bestellung_info=bestellung_info,
        tagesplan_heute=tagesplan_heute, homeoffice_vs_ids=homeoffice_vs_ids, tanken_vs_ids=tanken_vs_ids)


@app.route('/aktivitaeten')
@login_required
def aktivitaeten_liste():
    is_admin   = session.get('rolle') == 'admin'
    is_manager = session.get('rolle') in ('admin', 'verkaufsleiter')
    jahr       = request.args.get('jahr', date.today().year, type=int)
    kw_filter  = request.args.get('kw',    '', type=str)
    mo_filter  = request.args.get('monat', '', type=str)
    mo_ids     = [x.strip().zfill(2) for x in mo_filter.split(',') if x.strip()] if mo_filter else []
    ma_filter  = request.args.get('ma',    '', type=str)
    ma_ids     = [x.strip() for x in ma_filter.split(',') if x.strip()] if ma_filter else []
    vs_filter  = request.args.get('vs',    '', type=str)
    vs_ids     = [x.strip() for x in vs_filter.split(',') if x.strip()] if vs_filter else []
    typ_filter = request.args.get('typ',   '', type=str)   # kommagetrennte Typen
    typ_ids    = [x.strip() for x in typ_filter.split(',') if x.strip()] if typ_filter else []

    sql = '''
        SELECT a.id, a.datum, m.name AS mitarbeiter, m.id AS mitarbeiter_id,
               v.name AS verkaufsstelle, v.id AS verkaufsstelle_id,
               v.ort, v.strasse, v.typ, a.anzahl_displays, a.notizen, a.erstellt_am,
               a.foto_pfad, a.foto_pfad_2, a.foto_pfad_3,
               COALESCE(a.aktionstyp, 'Aufbau') AS aktionstyp,
               COALESCE(SUM(b.kisten_anzahl), 0) AS kisten_gesamt
        FROM aktivitaet a
        JOIN mitarbeiter m ON m.id = a.mitarbeiter_id
        JOIN verkaufsstelle v ON v.id = a.verkaufsstelle_id
        LEFT JOIN bestellposition b ON b.aktivitaet_id = a.id
        WHERE 1=1
    '''
    params = []

    vs_history_mode = is_manager and bool(vs_ids)
    if not vs_history_mode:
        sql += " AND strftime('%Y', a.datum) = ?"
        params.append(str(jahr))

    if not is_manager:
        sql += " AND a.mitarbeiter_id = ?"
        params.append(session['user_id'])
    else:
        if ma_ids:
            _ph = ','.join('?' * len(ma_ids))
            sql += f" AND a.mitarbeiter_id IN ({_ph})"
            params.extend(ma_ids)
        # Bugreport 2026-07-21 (Hoch): fehlte hier komplett – ein VKL mit eigenem Team
        # sah ohne Filter firmenweit ALLE Aktivitäten aller Teams.
        _tc_sql, _tc_params = _team_ma_clause('a')
        sql += _tc_sql
        params.extend(_tc_params)

    if is_manager and vs_ids:
        _ph = ','.join('?' * len(vs_ids))
        sql += f" AND a.verkaufsstelle_id IN ({_ph})"
        params.extend(vs_ids)

    if mo_ids:
        _ph = ','.join('?' * len(mo_ids))
        sql += f" AND strftime('%m', a.datum) IN ({_ph})"
        params.extend(mo_ids)

    if kw_filter:
        # Bug: strftime('%W', datum) ist NICHT die ISO-Kalenderwoche (die überall sonst
        # in der App via date.isocalendar()[1] angezeigt wird, z.B. das "KW"-Badge auf
        # jeder Aktivitäts-Karte) – bei vielen Wochen weichen beide um 1 ab. Fix:
        # Wochenbereich (Montag–Sonntag) über date.fromisocalendar() statt strftime,
        # damit Filter und Badge dieselbe ISO-Woche meinen.
        try:
            kw_start = date.fromisocalendar(jahr, int(kw_filter), 1)
            kw_end   = kw_start + timedelta(days=7)
            sql += " AND a.datum >= ? AND a.datum < ?"
            params.append(kw_start.isoformat())
            params.append(kw_end.isoformat())
        except ValueError:
            pass  # ungültige KW/Jahr-Kombination (z.B. KW 53 in einem Jahr ohne 53. Woche)

    if typ_ids:
        _ph = ','.join('?' * len(typ_ids))
        sql += f" AND v.typ IN ({_ph})"
        params.extend(typ_ids)

    sql += " GROUP BY a.id ORDER BY a.datum DESC, a.erstellt_am DESC"

    aktivitaeten = query(sql, params)

    # Bestellpositionen für jede Aktivität
    detail = {}
    for a in aktivitaeten:
        positionen = query('''
            SELECT bs.name, bp.kisten_anzahl, bs.einheit
            FROM bestellposition bp JOIN biersorte bs ON bs.id = bp.biersorte_id
            WHERE bp.aktivitaet_id = ?
        ''', (a['id'],))
        detail[a['id']] = positionen

    # Displaypositionen für jede Aktivität
    disp_detail = {}
    for a in aktivitaeten:
        dp = query('''
            SELECT ds.name, dp.anzahl
            FROM displayposition dp JOIN displaysorte ds ON ds.id = dp.displaysorte_id
            WHERE dp.aktivitaet_id = ? AND dp.anzahl > 0
            ORDER BY ds.name
        ''', (a['id'],))
        if dp:
            disp_detail[a['id']] = dp

    _tm_sql, _tm_p = _team_m_clause('m')
    alle_ma = query(
        f"SELECT id, name FROM mitarbeiter m WHERE rolle IN ('rep','verkaufsleiter'){_tm_sql} ORDER BY name",
        _tm_p
    ) if is_manager else []
    # Alle VS für Dropdown (inkl. inaktive – für historische Suche). Gedeckelt wie
    # die Besuchsplanung im Dashboard – bei vielen VS würde das feste Einbetten
    # ALLER Checkboxen die Seite massiv aufblähen.
    # Auch für Reps geladen (nicht nur is_manager): Reps dürfen seit 2026-07-22 eigene
    # Aktivitäten vom selben Tag bearbeiten, das Bearbeiten-Modal braucht dafür dieselbe
    # VS-Auswahl wie VKL/Admin.
    alle_vs = query(
        "SELECT id, name, ort, aktiv FROM verkaufsstelle ORDER BY aktiv DESC, name LIMIT ?",
        (VS_DASHBOARD_SEITENGROESSE,)
    )
    alle_typen = [r[0] for r in query(
        "SELECT DISTINCT typ FROM verkaufsstelle WHERE typ IS NOT NULL AND typ != '' ORDER BY typ"
    )]
    jahre = [r[0] for r in query("SELECT DISTINCT CAST(strftime('%Y', datum) AS INTEGER) FROM aktivitaet ORDER BY 1 DESC")]
    if not jahre:
        jahre = [date.today().year]

    return render_template('aktivitaeten.html',
        aktivitaeten=aktivitaeten, detail=detail, disp_detail=disp_detail,
        jahr=jahr, jahre=jahre, kw_filter=kw_filter,
        mo_filter=mo_filter, mo_ids=mo_ids,
        ma_filter=ma_filter, ma_ids=ma_ids,
        vs_filter=vs_filter, vs_ids=vs_ids, vs_history_mode=vs_history_mode,
        typ_filter=typ_filter, typ_ids=typ_ids, alle_typen=alle_typen,
        alle_ma=alle_ma, alle_vs=alle_vs,
        is_admin=is_admin, is_manager=is_manager)


@app.route('/aktivitaet/<int:akt_id>/bearbeiten', methods=['POST'])
@login_required
def aktivitaet_bearbeiten(akt_id):
    if session.get('rolle') not in ('admin', 'verkaufsleiter', 'rep'):
        flash('Keine Berechtigung.', 'danger')
        return redirect(url_for('aktivitaeten_liste'))

    a = query("SELECT * FROM aktivitaet WHERE id=?", (akt_id,), one=True)
    if not a:
        flash('Aktivität nicht gefunden.', 'danger')
        return redirect(url_for('aktivitaeten_liste'))

    # Reps dürfen seit 2026-07-22 eigene Aktivitäten komplett bearbeiten – aber nur am
    # gleichen Tag, an dem die Aktivität stattfand (nicht rückwirkend, nicht im Voraus).
    if session.get('rolle') == 'rep':
        if a['mitarbeiter_id'] != session.get('user_id') or a['datum'] != date.today().isoformat():
            flash('Sie können nur eigene Aktivitäten vom heutigen Tag bearbeiten.', 'danger')
            return redirect(url_for('aktivitaeten_liste'))

    # VKL darf nur Aktivitäten des eigenen Teams bearbeiten (Bugreport 2026-07-21: bislang
    # ohne Team-Check, konnte also fremde Teams inkl. deren Fotos verändern/löschen).
    if session.get('rolle') == 'verkaufsleiter' and session.get('team_id'):
        _besitzer = query("SELECT team_id FROM mitarbeiter WHERE id=?", (a['mitarbeiter_id'],), one=True)
        if not _besitzer or _besitzer['team_id'] != session.get('team_id'):
            flash('Keine Berechtigung für diese Aktivität.', 'danger')
            return redirect(url_for('aktivitaeten_liste'))

    datum          = request.form.get('datum', a['datum'])
    vs_id          = request.form.get('verkaufsstelle_id', a['verkaufsstelle_id'], type=int)
    aktionstyp     = request.form.get('aktionstyp', a['aktionstyp'] or 'Aufbau')
    anzahl_displays = request.form.get('anzahl_displays', a['anzahl_displays'] or 0, type=int)
    notizen        = request.form.get('notizen', a['notizen'] or '')
    bestell_status = request.form.get('bestell_status') if aktionstyp == 'Bestellung' else a['bestell_status']

    if aktionstyp not in ('Aufbau', 'Bestellung', 'Besuch'):
        flash('Ungültiger Aktivitätstyp.', 'danger')
        return redirect(url_for('aktivitaeten_liste'))

    execute(
        "UPDATE aktivitaet SET datum=?, verkaufsstelle_id=?, aktionstyp=?, anzahl_displays=?, notizen=?, bestell_status=? WHERE id=?",
        (datum, vs_id, aktionstyp, anzahl_displays, notizen, bestell_status, akt_id)
    )
    flash('Aktivität aktualisiert.', 'success')
    return redirect(request.referrer or url_for('aktivitaeten_liste'))


@app.route('/aktivitaet/<int:akt_id>/loeschen', methods=['POST'])
@login_required
def aktivitaet_loeschen(akt_id):
    a = query("SELECT * FROM aktivitaet WHERE id=?", (akt_id,), one=True)
    if not a:
        flash('Aktivität nicht gefunden.', 'danger')
        return redirect(url_for('aktivitaeten_liste'))

    if session.get('rolle') != 'admin':
        flash('Keine Berechtigung. Aktivitäten können nur vom Admin gelöscht werden.', 'danger')
        return redirect(url_for('aktivitaeten_liste'))

    # Foto-Dateien mitlöschen (bis zu 3)
    for spalte in ('foto_pfad', 'foto_pfad_2', 'foto_pfad_3'):
        if a[spalte]:
            foto_path = os.path.join(UPLOAD_FOLDER, a[spalte])
            if os.path.exists(foto_path):
                os.remove(foto_path)

    execute("DELETE FROM bestellposition  WHERE aktivitaet_id=?", (akt_id,))
    execute("DELETE FROM displayposition  WHERE aktivitaet_id=?", (akt_id,))
    execute("DELETE FROM aktivitaet       WHERE id=?",            (akt_id,))
    flash('Aktivität gelöscht.', 'success')
    return redirect(url_for('aktivitaeten_liste'))


# ─── Freigabe-Workflow für Aufbauten ───────────────────────────────────────────

@app.route('/aufbauten/freigabe')
@manager_required
def aufbauten_freigabe():
    """Freigabe-Workflow: zielrelevante Aufbauten müssen vom VKL geprüft (Details/Fotos
    einsehen) und freigegeben oder mit Begründung abgelehnt werden, bevor sie in die
    Zielzahlen des Mitarbeiters einfließen."""
    _tc, _tp = _freigabe_scope_clause('a')
    pending = query(f'''
        SELECT dp.id AS dp_id, dp.anzahl, dp.displaysorte_id,
               ds.name AS ds_name,
               a.id AS aktivitaet_id, a.datum, a.notizen, a.von_uhrzeit, a.bis_uhrzeit,
               a.foto_pfad, a.foto_pfad_2, a.foto_pfad_3,
               v.name AS vs_name, v.ort AS vs_ort, v.strasse AS vs_strasse,
               m.name AS mitarbeiter_name
        FROM displayposition dp
        JOIN displaysorte ds ON ds.id = dp.displaysorte_id
        JOIN aktivitaet a ON a.id = dp.aktivitaet_id
        JOIN verkaufsstelle v ON v.id = a.verkaufsstelle_id
        JOIN mitarbeiter m ON m.id = a.mitarbeiter_id
        WHERE dp.status = 'offen'{_tc}
        ORDER BY a.datum DESC, dp.id DESC
    ''', _tp)
    return render_template('aufbauten_freigabe.html', pending=pending)


@app.route('/aufbau/<int:dp_id>/freigeben', methods=['POST'])
@manager_required
def aufbau_freigeben(dp_id):
    _tc, _tp = _freigabe_scope_clause('a')
    dp = query(f'''
        SELECT dp.id, dp.anzahl, dp.aktivitaet_id FROM displayposition dp
        JOIN aktivitaet a ON a.id = dp.aktivitaet_id
        WHERE dp.id = ?{_tc}
    ''', (dp_id,) + _tp, one=True)
    if not dp:
        flash('Aufbau nicht gefunden oder keine Berechtigung.', 'danger')
        return redirect(url_for('aufbauten_freigabe'))
    execute(
        "UPDATE displayposition SET status='freigegeben', entschieden_von=?, entschieden_am=? WHERE id=?",
        (session['user_id'], datetime.now().isoformat(), dp_id)
    )
    # Erst jetzt fließt die Menge in die Zielzahlen ein (anzahl_displays ist die einzige
    # Quelle für Zielzahlen/Reports, s. Kommentar bei neue_aktivitaet()).
    execute(
        "UPDATE aktivitaet SET anzahl_displays = anzahl_displays + ? WHERE id = ?",
        (dp['anzahl'], dp['aktivitaet_id'])
    )
    flash('Aufbau freigegeben.', 'success')
    return redirect(url_for('aufbauten_freigabe'))


@app.route('/aufbau/<int:dp_id>/ablehnen', methods=['POST'])
@manager_required
def aufbau_ablehnen(dp_id):
    grund = request.form.get('grund', '').strip()
    if not grund:
        flash('Bitte einen Ablehnungsgrund angeben.', 'danger')
        return redirect(url_for('aufbauten_freigabe'))
    _tc, _tp = _freigabe_scope_clause('a')
    dp = query(f'''
        SELECT dp.id FROM displayposition dp JOIN aktivitaet a ON a.id = dp.aktivitaet_id
        WHERE dp.id = ?{_tc}
    ''', (dp_id,) + _tp, one=True)
    if not dp:
        flash('Aufbau nicht gefunden oder keine Berechtigung.', 'danger')
        return redirect(url_for('aufbauten_freigabe'))
    execute(
        "UPDATE displayposition SET status='abgelehnt', ablehnungsgrund=?, entschieden_von=?, "
        "entschieden_am=?, mitarbeiter_gesehen=0 WHERE id=?",
        (grund, session['user_id'], datetime.now().isoformat(), dp_id)
    )
    flash('Aufbau abgelehnt.', 'success')
    return redirect(url_for('aufbauten_freigabe'))


@app.route('/aufbau/<int:dp_id>/gesehen', methods=['POST'])
@login_required
def aufbau_gesehen(dp_id):
    """Mitarbeiter bestätigt, die Ablehnungs-Notiz gesehen zu haben (Dashboard-Hinweis)."""
    execute(
        "UPDATE displayposition SET mitarbeiter_gesehen=1 WHERE id=? AND aktivitaet_id IN "
        "(SELECT id FROM aktivitaet WHERE mitarbeiter_id=?)",
        (dp_id, session['user_id'])
    )
    return redirect(request.referrer or url_for('dashboard'))


# ─── Excel-Hilfsfunktion (für Route + Auto-Export) ────────────────────────────

_EXCEL_FORMEL_PRAEFIXE = ('=', '+', '-', '@', '\t', '\r')

def _excel_formel_sicher(wert):
    """Verhindert Formel-Injection (Bugreport 2026-07-21): Freitext wie das Notizen-Feld
    landet ungeprüft in Excel-Zellen – ein Rep könnte z.B. '=HYPERLINK(...)' eintragen,
    das beim Öffnen durch einen Admin als Formel/Link ausgeführt wird statt als Text
    angezeigt zu werden. Ein führendes Apostroph erzwingt in Excel Text-Interpretation."""
    if isinstance(wert, str) and wert.startswith(_EXCEL_FORMEL_PRAEFIXE):
        return "'" + wert
    return wert


def _build_excel_bytes(jahr: int, is_admin: bool = True, mitarbeiter_id=None) -> bytes:
    """Erstellt die Excel-Auswertung und gibt sie als Bytes zurück. mitarbeiter_id (nur
    relevant wenn is_admin=False) akzeptiert sowohl eine einzelne ID (Rep-Eigenexport) als
    auch eine Liste von IDs (VKL-Team-Export, Bugreport 2026-07-21: VKL bekam bislang trotz
    is_admin=False Zugriff auf den kompletten unternehmensweiten Export, da export_excel()
    'admin' und 'verkaufsleiter' beide als is_admin=True behandelte)."""
    wb = openpyxl.Workbook()
    _ma_ids = tuple(mitarbeiter_id) if isinstance(mitarbeiter_id, (list, tuple, set)) else (mitarbeiter_id,)
    _ma_ph  = ','.join('?' * len(_ma_ids))

    HEADER_FILL = PatternFill("solid", fgColor="1a3a5c")
    SUB_FILL    = PatternFill("solid", fgColor="2e6da4")
    ALT_FILL    = PatternFill("solid", fgColor="eef4fb")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    SUB_FONT    = Font(bold=True, color="FFFFFF", size=10)
    TITLE_FONT  = Font(bold=True, size=14, color="1a3a5c")
    BOLD        = Font(bold=True)
    BORDER      = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC'),
    )
    CENTER = Alignment(horizontal='center', vertical='center')
    LEFT   = Alignment(horizontal='left', vertical='center')

    def style_header(ws, row, cols):
        for c in range(1, cols+1):
            cell = ws.cell(row=row, column=c)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = CENTER
            cell.border = BORDER

    def style_subheader(ws, row, cols):
        for c in range(1, cols+1):
            cell = ws.cell(row=row, column=c)
            cell.fill = SUB_FILL
            cell.font = SUB_FONT
            cell.alignment = CENTER
            cell.border = BORDER

    # ── Sheet 1: KW-Übersicht ──────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = f"KW-Übersicht {jahr}"
    ws1.column_dimensions['A'].width = 10
    ws1.column_dimensions['B'].width = 18
    ws1.column_dimensions['C'].width = 18
    ws1.column_dimensions['D'].width = 15

    ws1.merge_cells('A1:D1')
    ws1['A1'] = f"Wochenübersicht {jahr} – Displays & {UNIT_LABEL}"
    ws1['A1'].font = TITLE_FONT
    ws1['A1'].alignment = CENTER

    headers = ['Kalenderwoche', 'Anzahl Displays', f'{UNIT_LABEL} gesamt', 'Besuche']
    for col, h in enumerate(headers, 1):
        ws1.cell(row=2, column=col, value=h)
    style_header(ws1, 2, 4)

    _BP = "(SELECT aktivitaet_id, SUM(kisten_anzahl) AS kisten_total FROM bestellposition GROUP BY aktivitaet_id)"
    if is_admin:
        kw_data = query(f'''
            SELECT CAST(strftime('%W', a.datum) AS INTEGER) AS kw,
                   SUM(a.anzahl_displays) AS displays,
                   COALESCE(SUM(b.kisten_total), 0) AS kisten,
                   COUNT(a.id) AS besuche
            FROM aktivitaet a
            LEFT JOIN {_BP} b ON b.aktivitaet_id = a.id
            WHERE strftime('%Y', a.datum) = ?
            GROUP BY kw ORDER BY kw
        ''', (str(jahr),))
    else:
        kw_data = query(f'''
            SELECT CAST(strftime('%W', a.datum) AS INTEGER) AS kw,
                   SUM(a.anzahl_displays) AS displays,
                   COALESCE(SUM(b.kisten_total), 0) AS kisten,
                   COUNT(a.id) AS besuche
            FROM aktivitaet a
            LEFT JOIN {_BP} b ON b.aktivitaet_id = a.id
            WHERE strftime('%Y', a.datum) = ? AND a.mitarbeiter_id IN ({_ma_ph})
            GROUP BY kw ORDER BY kw
        ''', (str(jahr),) + _ma_ids)

    total_disp = total_kist = total_bes = 0
    for i, row in enumerate(kw_data):
        r = i + 3
        ws1.cell(r, 1, f"KW {row['kw']:02d}")
        ws1.cell(r, 2, row['displays'] or 0)
        ws1.cell(r, 3, row['kisten'] or 0)
        ws1.cell(r, 4, row['besuche'] or 0)
        total_disp += row['displays'] or 0
        total_kist += row['kisten'] or 0
        total_bes  += row['besuche'] or 0
        fill = ALT_FILL if i % 2 == 0 else None
        for c in range(1, 5):
            cell = ws1.cell(r, c)
            cell.alignment = CENTER
            cell.border = BORDER
            if fill:
                cell.fill = fill

    # Summenzeile
    sr = len(kw_data) + 3
    ws1.cell(sr, 1, 'GESAMT').font = BOLD
    ws1.cell(sr, 2, total_disp).font = BOLD
    ws1.cell(sr, 3, total_kist).font = BOLD
    ws1.cell(sr, 4, total_bes).font = BOLD
    for c in range(1, 5):
        ws1.cell(sr, c).fill = PatternFill("solid", fgColor="FFD700")
        ws1.cell(sr, c).alignment = CENTER
        ws1.cell(sr, c).border = BORDER

    # ── Sheet 2: Jahreswerte nach Mitarbeiter (nur Admin) ─────────────────
    if is_admin:
        ws2 = wb.create_sheet(f"Mitarbeiter {jahr}")
        ws2.column_dimensions['A'].width = 22
        ws2.column_dimensions['B'].width = 10
        ws2.column_dimensions['C'].width = 18
        ws2.column_dimensions['D'].width = 16
        ws2.column_dimensions['E'].width = 12

        ws2.merge_cells('A1:E1')
        ws2['A1'] = f"Jahresübersicht nach Mitarbeiter {jahr}"
        ws2['A1'].font = TITLE_FONT
        ws2['A1'].alignment = CENTER

        h2 = ['Mitarbeiter', 'Kürzel', 'Displays gesamt', f'{UNIT_LABEL} gesamt', 'Besuche']
        for col, h in enumerate(h2, 1):
            ws2.cell(2, col, h)
        style_header(ws2, 2, 5)

        ma_data = query(f'''
            SELECT m.name, m.kuerzel,
                   SUM(a.anzahl_displays) AS displays,
                   COALESCE(SUM(b.kisten_total), 0) AS kisten,
                   COUNT(a.id) AS besuche
            FROM mitarbeiter m
            JOIN aktivitaet a ON a.mitarbeiter_id = m.id
            LEFT JOIN {_BP} b ON b.aktivitaet_id = a.id
            WHERE strftime('%Y', a.datum) = ?
            GROUP BY m.id ORDER BY kisten DESC
        ''', (str(jahr),))

        for i, row in enumerate(ma_data):
            r = i + 3
            ws2.cell(r, 1, row['name'])
            ws2.cell(r, 2, row['kuerzel'])
            ws2.cell(r, 3, row['displays'] or 0)
            ws2.cell(r, 4, row['kisten'] or 0)
            ws2.cell(r, 5, row['besuche'] or 0)
            fill = ALT_FILL if i % 2 == 0 else None
            for c in range(1, 6):
                cell = ws2.cell(r, c)
                cell.border = BORDER
                cell.alignment = CENTER if c > 1 else LEFT
                if fill:
                    cell.fill = fill

    # ── Sheet 3: Alle Aktivitäten ─────────────────────────────────────────
    ws3 = wb.create_sheet("Aktivitäten-Detail")
    cols3 = ['Datum', 'KW', 'Mitarbeiter', 'Verkaufsstelle', 'Ort', 'Typ',
             'Displays', 'Produkt', UNIT_LABEL, 'Notizen']
    widths = [14, 6, 20, 28, 16, 16, 10, 20, 10, 35]
    for i, w in enumerate(widths, 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    ws3.merge_cells(f'A1:{get_column_letter(len(cols3))}1')
    ws3['A1'] = f"Aktivitäten-Detail {jahr}"
    ws3['A1'].font = TITLE_FONT
    ws3['A1'].alignment = CENTER

    for col, h in enumerate(cols3, 1):
        ws3.cell(2, col, h)
    style_header(ws3, 2, len(cols3))

    if is_admin:
        aktivitaeten = query('''
            SELECT a.id, a.datum, m.name AS mitarbeiter,
                   v.name AS verkaufsstelle, v.ort, v.typ,
                   a.anzahl_displays, a.notizen
            FROM aktivitaet a
            JOIN mitarbeiter m ON m.id = a.mitarbeiter_id
            JOIN verkaufsstelle v ON v.id = a.verkaufsstelle_id
            WHERE strftime('%Y', a.datum) = ?
            ORDER BY a.datum DESC
        ''', (str(jahr),))
    else:
        aktivitaeten = query(f'''
            SELECT a.id, a.datum, m.name AS mitarbeiter,
                   v.name AS verkaufsstelle, v.ort, v.typ,
                   a.anzahl_displays, a.notizen
            FROM aktivitaet a
            JOIN mitarbeiter m ON m.id = a.mitarbeiter_id
            JOIN verkaufsstelle v ON v.id = a.verkaufsstelle_id
            WHERE strftime('%Y', a.datum) = ? AND a.mitarbeiter_id IN ({_ma_ph})
            ORDER BY a.datum DESC
        ''', (str(jahr),) + _ma_ids)

    r = 3
    for i, a in enumerate(aktivitaeten):
        positionen = query('''
            SELECT bs.name, bp.kisten_anzahl
            FROM bestellposition bp JOIN biersorte bs ON bs.id = bp.biersorte_id
            WHERE bp.aktivitaet_id = ?
        ''', (a['id'],))

        import datetime as dt
        d = dt.date.fromisoformat(a['datum'])
        kw = int(d.strftime('%W'))
        fill = ALT_FILL if i % 2 == 0 else None

        if not positionen:
            ws3.cell(r, 1, a['datum'])
            ws3.cell(r, 2, f"KW {kw:02d}")
            ws3.cell(r, 3, a['mitarbeiter'])
            ws3.cell(r, 4, _excel_formel_sicher(a['verkaufsstelle']))
            ws3.cell(r, 5, _excel_formel_sicher(a['ort']))
            ws3.cell(r, 6, a['typ'])
            ws3.cell(r, 7, a['anzahl_displays'])
            ws3.cell(r, 8, '–')
            ws3.cell(r, 9, 0)
            ws3.cell(r, 10, _excel_formel_sicher(a['notizen'] or ''))
            for c in range(1, 11):
                ws3.cell(r, c).border = BORDER
                if fill:
                    ws3.cell(r, c).fill = fill
            r += 1
        else:
            for j, pos in enumerate(positionen):
                ws3.cell(r, 1, a['datum'] if j == 0 else '')
                ws3.cell(r, 2, f"KW {kw:02d}" if j == 0 else '')
                ws3.cell(r, 3, a['mitarbeiter'] if j == 0 else '')
                ws3.cell(r, 4, _excel_formel_sicher(a['verkaufsstelle']) if j == 0 else '')
                ws3.cell(r, 5, _excel_formel_sicher(a['ort']) if j == 0 else '')
                ws3.cell(r, 6, a['typ'] if j == 0 else '')
                ws3.cell(r, 7, a['anzahl_displays'] if j == 0 else '')
                ws3.cell(r, 8, pos['name'])
                ws3.cell(r, 9, pos['kisten_anzahl'])
                ws3.cell(r, 10, _excel_formel_sicher(a['notizen']) if j == 0 else '')
                for c in range(1, 11):
                    ws3.cell(r, c).border = BORDER
                    if fill:
                        ws3.cell(r, c).fill = fill
                r += 1

    # ── Sheet 4: Produkt-Übersicht ─────────────────────────────────────────
    ws4 = wb.create_sheet(f"Produkte {jahr}")
    ws4.column_dimensions['A'].width = 22
    ws4.column_dimensions['B'].width = 20
    ws4.column_dimensions['C'].width = 16

    ws4.merge_cells('A1:C1')
    ws4['A1'] = f"Bestellte Produkte {jahr}"
    ws4['A1'].font = TITLE_FONT
    ws4['A1'].alignment = CENTER

    for col, h in enumerate(['Produkt', 'Einheit', f'{UNIT_LABEL} gesamt'], 1):
        ws4.cell(2, col, h)
    style_header(ws4, 2, 3)

    if is_admin:
        bier_data = query('''
            SELECT bs.name, bs.einheit, SUM(bp.kisten_anzahl) AS kisten
            FROM bestellposition bp
            JOIN biersorte bs ON bs.id = bp.biersorte_id
            JOIN aktivitaet a ON a.id = bp.aktivitaet_id
            WHERE strftime('%Y', a.datum) = ?
            GROUP BY bs.id ORDER BY kisten DESC
        ''', (str(jahr),))
    else:
        bier_data = query(f'''
            SELECT bs.name, bs.einheit, SUM(bp.kisten_anzahl) AS kisten
            FROM bestellposition bp
            JOIN biersorte bs ON bs.id = bp.biersorte_id
            JOIN aktivitaet a ON a.id = bp.aktivitaet_id
            WHERE strftime('%Y', a.datum) = ? AND a.mitarbeiter_id IN ({_ma_ph})
            GROUP BY bs.id ORDER BY kisten DESC
        ''', (str(jahr),) + _ma_ids)

    for i, row in enumerate(bier_data):
        r2 = i + 3
        ws4.cell(r2, 1, row['name'])
        ws4.cell(r2, 2, row['einheit'])
        ws4.cell(r2, 3, row['kisten'])
        fill = ALT_FILL if i % 2 == 0 else None
        for c in range(1, 4):
            ws4.cell(r2, c).border = BORDER
            ws4.cell(r2, c).alignment = CENTER if c > 1 else LEFT
            if fill:
                ws4.cell(r2, c).fill = fill

    # Output
    output = io.BytesIO()
    # Output als Bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


# ─── Route: Excel Export ──────────────────────────────────────────────────────

def _build_tanken_excel_bytes(von: str, bis: str, monat_label: str) -> bytes:
    """Excel-Liste aller Tankungen im Zeitraum: Tankdatum, Kennzeichen, Kraftstoffsorte,
    Menge, Kilometerstand (aus dem Tankbeleg) + Mitarbeiter, Notiz, Beleg-Anzahl."""
    rows = query(
        "SELECT a.datum, m.name AS mitarbeiter, m.kuerzel, a.notizen, "
        "       a.tank_datum, a.tank_kennzeichen, a.tank_kraftstoffsorte, "
        "       a.tank_liter, a.tank_km_stand, "
        "       (CASE WHEN a.foto_pfad IS NOT NULL AND a.foto_pfad != '' THEN 1 ELSE 0 END "
        "        + CASE WHEN a.foto_pfad_2 IS NOT NULL AND a.foto_pfad_2 != '' THEN 1 ELSE 0 END "
        "        + CASE WHEN a.foto_pfad_3 IS NOT NULL AND a.foto_pfad_3 != '' THEN 1 ELSE 0 END) AS anzahl_belege "
        "FROM aktivitaet a "
        "JOIN mitarbeiter m ON m.id = a.mitarbeiter_id "
        "JOIN verkaufsstelle v ON v.id = a.verkaufsstelle_id "
        "WHERE v.typ = 'Firmenwagen-Tanken' AND a.datum BETWEEN ? AND ? "
        "ORDER BY a.datum, m.name",
        (von, bis)
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tankungen"

    HEADER_FILL = PatternFill("solid", fgColor="c8860a")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    TITLE_FONT  = Font(bold=True, size=14, color="1a3a5c")
    CENTER      = Alignment(horizontal='center', vertical='center')
    BORDER      = Border(
        left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'), bottom=Side(style='thin', color='CCCCCC'),
    )

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 40
    ws.column_dimensions['H'].width = 12

    ws.merge_cells('A1:H1')
    ws['A1'] = f"Firmenwagen-Tankungen – {monat_label}"
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = CENTER

    headers = ['Tankdatum', 'Mitarbeiter', 'Kennzeichen', 'Kraftstoffsorte', 'Menge (Ltr.)', 'Kilometerstand', 'Notiz', 'Belege (Fotos)']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER

    for i, row in enumerate(rows):
        r = i + 3
        ws.cell(r, 1, row['tank_datum'] or row['datum'])
        ws.cell(r, 2, f"{row['mitarbeiter']} ({row['kuerzel']})")
        ws.cell(r, 3, _excel_formel_sicher(row['tank_kennzeichen'] or ''))
        ws.cell(r, 4, row['tank_kraftstoffsorte'] or '')
        ws.cell(r, 5, row['tank_liter'] or '')
        ws.cell(r, 6, row['tank_km_stand'] or '')
        ws.cell(r, 7, _excel_formel_sicher(row['notizen'] or ''))
        ws.cell(r, 8, row['anzahl_belege'])
        for c in range(1, 9):
            ws.cell(r, c).border = BORDER

    if not rows:
        ws.cell(3, 1, 'Keine Tankungen in diesem Zeitraum.')

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


@app.route('/tanken-report')
@manager_required
def tanken_report():
    """Firmenwagen-Tanken-Report (Kernfunktion): Excel-Download der Tankungen für einen
    wählbaren Monat. Nur relevant, wenn TANKEN_MODUS aktiviert ist."""
    if not TANKEN_MODUS:
        flash('Firmenwagen-Tanken ist nicht aktiviert.', 'danger')
        return redirect(url_for('dashboard'))
    jahr  = request.args.get('jahr', date.today().year, type=int)
    monat = request.args.get('monat', date.today().month, type=int)
    von   = date(jahr, monat, 1)
    bis   = date(jahr + 1, 1, 1) - timedelta(days=1) if monat == 12 else date(jahr, monat + 1, 1) - timedelta(days=1)
    _monat_namen = ['Januar','Februar','März','April','Mai','Juni',
                    'Juli','August','September','Oktober','November','Dezember']
    monat_label = f"{_monat_namen[monat - 1]} {jahr}"
    data  = _build_tanken_excel_bytes(von.isoformat(), bis.isoformat(), monat_label)
    fname = f"Tankungen_{jahr}-{monat:02d}.xlsx"
    return send_file(io.BytesIO(data), as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


def _build_gratisware_report_excel(von_str: str, bis_str: str) -> bytes:
    """Eine Zeile je Besuch mit Gratisware Verleger und/oder Kofferraum im gegebenen
    Zeitraum."""
    rows = query('''
        SELECT a.datum, m.name AS mitarbeiter,
               v.name AS verkaufsstelle, v.strasse, v.plz, v.ort, v.lieferant, v.kundennummer,
               COALESCE(SUM(CASE WHEN bs.name='Gratisware Verleger'   THEN bp.kisten_anzahl ELSE 0 END), 0) AS verleger,
               COALESCE(SUM(CASE WHEN bs.name='Gratisware Kofferraum' THEN bp.kisten_anzahl ELSE 0 END), 0) AS kofferraum,
               GROUP_CONCAT(
                   CASE WHEN COALESCE(bs.ist_gratisware,0)=0 THEN bs.name || ' × ' || bp.kisten_anzahl END,
                   ', '
               ) AS bestellung
        FROM aktivitaet a
        JOIN mitarbeiter m ON m.id = a.mitarbeiter_id
        JOIN verkaufsstelle v ON v.id = a.verkaufsstelle_id
        JOIN bestellposition bp ON bp.aktivitaet_id = a.id
        JOIN biersorte bs ON bs.id = bp.biersorte_id
        WHERE a.aktionstyp='Bestellung' AND a.datum BETWEEN ? AND ?
        GROUP BY a.id
        HAVING verleger > 0 OR kofferraum > 0
        ORDER BY a.datum, m.name
    ''', (von_str, bis_str))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Gratisware"

    HEADER_FILL = PatternFill("solid", fgColor="1a3a5c")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    BORDER = Border(
        left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'), bottom=Side(style='thin', color='CCCCCC'),
    )
    CENTER = Alignment(horizontal='center', vertical='center')

    headers = ['Besuchstag', 'Mitarbeiter', 'Verkaufsstelle', 'Straße', 'PLZ', 'Ort',
                'Lieferant', 'Kundennummer',
                'Gratisware Verleger', 'Gratisware Kofferraum', 'Bestellung (Produkt × Kisten)']
    widths  = [12, 20, 28, 26, 8, 18, 18, 14, 16, 18, 45]
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col)].width = w

    for r, row in enumerate(rows, start=2):
        werte = [
            datetime.strptime(row['datum'], '%Y-%m-%d').strftime('%d.%m.%Y'),
            row['mitarbeiter'], row['verkaufsstelle'], row['strasse'] or '', row['plz'] or '',
            row['ort'] or '', row['lieferant'] or '', row['kundennummer'] or '',
            row['verleger'], row['kofferraum'], row['bestellung'] or '',
        ]
        for col, wert in enumerate(werte, 1):
            cell = ws.cell(row=r, column=col, value=wert)
            cell.border = BORDER
            if col in (9, 10):
                cell.alignment = CENTER

    if not rows:
        ws.cell(row=2, column=1, value="Keine Besuche mit Gratisware im gewählten Zeitraum.")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


@app.route('/gratisware-report')
@manager_required
def gratisware_report():
    """Gratisware-Report (Kernfunktion): Excel-Download aller Besuche mit Gratisware
    Verleger/Kofferraum für einen wählbaren Monat. Nur relevant, wenn GRATISWARE_MODUS
    aktiviert ist."""
    if not GRATISWARE_MODUS:
        flash('Gratisware-Report ist nicht aktiviert.', 'danger')
        return redirect(url_for('dashboard'))
    jahr  = request.args.get('jahr', date.today().year, type=int)
    monat = request.args.get('monat', date.today().month, type=int)
    von   = date(jahr, monat, 1)
    bis   = date(jahr + 1, 1, 1) - timedelta(days=1) if monat == 12 else date(jahr, monat + 1, 1) - timedelta(days=1)
    data  = _build_gratisware_report_excel(von.isoformat(), bis.isoformat())
    fname = f"Gratisware_Report_{jahr}-{monat:02d}.xlsx"
    return send_file(io.BytesIO(data), as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/export/excel')
@manager_required
def export_excel():
    jahr     = request.args.get('jahr', date.today().year, type=int)
    # Bugreport 2026-07-21: VKL bekam bislang denselben unternehmensweiten Export wie
    # Admin (is_admin wurde für beide Rollen gleich True gesetzt). VKL exportiert jetzt
    # nur sein eigenes Team, Admin weiterhin uneingeschränkt.
    is_admin = session.get('rolle') == 'admin'
    if is_admin:
        ma_id = None
    else:
        _t_m_sql, _t_m_p = _team_m_clause('m')
        ma_id = [r['id'] for r in query(f"SELECT id FROM mitarbeiter m WHERE 1=1{_t_m_sql}", _t_m_p)] or [session.get('user_id')]
    data     = _build_excel_bytes(jahr, is_admin=is_admin, mitarbeiter_id=ma_id)
    fname    = f"Aktions_Tracker_{jahr}.xlsx"
    app.logger.info(f"Audit: Excel-Export {jahr} heruntergeladen von {session.get('kuerzel')} (Mitarbeiter-ID {session.get('user_id')}, Rolle {session.get('rolle')})")
    return send_file(io.BytesIO(data), as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ─── Routes: Admin ────────────────────────────────────────────────────────────

@app.route('/admin/demo-cleanup', methods=['POST'])
@admin_required
def admin_demo_cleanup():
    """Einmaliger Trigger: Demo-Pipeline bereinigen (alte Bestellungen schließen, vergangene Vertretungen löschen, neue anlegen)."""
    try:
        _demo_pipeline_cleanup()
        _demo_tagesplan_fortschritt()
        flash('Demo-Cleanup erfolgreich: Pipeline bereinigt, Vertretungen aktualisiert, Tagesplan aktualisiert.', 'success')
    except Exception as e:
        flash(f'Fehler beim Cleanup: {e}', 'danger')
    return redirect(url_for('admin'))


def _geo_ausreisser_finden(min_gruppe=3, schwelle_km=50):
    """Findet Verkaufsstellen, deren Koordinaten weit von den übrigen VS desselben
    Landkreis-Werts entfernt liegen – fängt sowohl echte Geocoding-Fehler als auch
    falsch getippte landkreis-Werte ab. Referenzpunkt je Landkreis ist der Median
    der eigenen Gruppe (robust gegen einzelne Ausreißer), nur für Gruppen mit
    mindestens min_gruppe Mitgliedern."""
    import statistics
    rows = query("""
        SELECT id, name, ort, strasse, landkreis, lat, lng
        FROM verkaufsstelle
        WHERE aktiv=1 AND lat IS NOT NULL AND lng IS NOT NULL
          AND landkreis IS NOT NULL AND landkreis != ''
    """)
    gruppen = {}
    for r in rows:
        gruppen.setdefault(r['landkreis'], []).append(r)

    def _distanz_km(lat1, lng1, lat2, lng2):
        dlat = math.radians(lat2 - lat1)
        dlng = math.radians(lng2 - lng1)
        a = math.sin(dlat/2)**2 + math.cos(math.radians(lat1)) * math.cos(math.radians(lat2)) * math.sin(dlng/2)**2
        return 6371 * 2 * math.asin(math.sqrt(a))

    ausreisser = []
    for landkreis, mitglieder in gruppen.items():
        if len(mitglieder) < min_gruppe:
            continue
        zentrum_lat = statistics.median(m['lat'] for m in mitglieder)
        zentrum_lng = statistics.median(m['lng'] for m in mitglieder)
        for m in mitglieder:
            km = _distanz_km(zentrum_lat, zentrum_lng, m['lat'], m['lng'])
            if km > schwelle_km:
                ausreisser.append({
                    'id': m['id'], 'name': m['name'], 'ort': m['ort'], 'strasse': m['strasse'],
                    'landkreis': landkreis, 'lat': m['lat'], 'lng': m['lng'], 'km_abweichung': round(km, 1),
                })
    ausreisser.sort(key=lambda x: -x['km_abweichung'])
    return ausreisser


@app.route('/admin/geo-ausreisser')
@admin_required
def admin_geo_ausreisser():
    """Geo-Ausreißer-Prüfung (Kernfunktion): auf Abruf statt per E-Mail-Scheduler –
    listet Verkaufsstellen, deren Koordinaten weit außerhalb ihrer Landkreis-Gruppe
    liegen (Geocoding-Fehler oder falscher Landkreis-Wert)."""
    ausreisser = _geo_ausreisser_finden()
    return render_template('geo_ausreisser.html', ausreisser=ausreisser)


@app.route('/admin')
@admin_required
def admin():
    mitarbeiter     = query("SELECT m.*, t.name AS team_name FROM mitarbeiter m LEFT JOIN team t ON t.id = m.team_id WHERE m.kuerzel != 'ADMIN' ORDER BY m.rolle, m.name")
    # Bei vielen Verkaufsstellen würde das ungebremste Einbetten ALLER Zeilen die
    # Admin-Seite lahmlegen – initial nur die ersten laden, für den Rest steht die
    # Suche (api_admin_verkaufsstellen_suche) zur Verfügung.
    # Homeoffice-Einträge werden über den Wohnort-Dialog im Mitarbeiter-Panel gepflegt,
    # nicht über diese allgemeine VS-Tabelle – dort ausgeblendet, um sie nicht mit
    # echten Kunden-Verkaufsstellen zu vermischen.
    vs_admin_gesamt = query("SELECT COUNT(*) AS n FROM verkaufsstelle WHERE homeoffice_mitarbeiter_id IS NULL", one=True)['n']
    verkaufsstellen = query(
        "SELECT * FROM verkaufsstelle WHERE homeoffice_mitarbeiter_id IS NULL ORDER BY aktiv DESC, name LIMIT ?",
        (VS_ADMIN_SEITENGROESSE,)
    )
    biersorten      = query("SELECT * FROM biersorte ORDER BY name")
    displaysorte    = query("SELECT * FROM displaysorte ORDER BY name")
    teams           = query("SELECT t.*, COUNT(m.id) AS mitglieder FROM team t LEFT JOIN mitarbeiter m ON m.team_id = t.id GROUP BY t.id ORDER BY t.name")
    mail_konfiguriert = bool(MAIL_SERVER and MAIL_USERNAME)

    # Zuordnungen: {mitarbeiter_id: set(verkaufsstelle_id, ...)}
    zuordnungen_raw = query("SELECT mitarbeiter_id, verkaufsstelle_id FROM mitarbeiter_verkaufsstelle")
    zuordnungen = {}
    for z in zuordnungen_raw:
        zuordnungen.setdefault(z['mitarbeiter_id'], set()).add(z['verkaufsstelle_id'])

    # Besitzer je VS: {vs_id: mitarbeiter_name} – für Anzeige im Zuordnungs-Modal
    ma_namen = {m['id']: m['name'] for m in mitarbeiter}
    vs_besitzer = {}
    for ma_id_loop, vs_set in zuordnungen.items():
        for vs_id in vs_set:
            vs_besitzer[vs_id] = ma_namen.get(ma_id_loop, '')

    # Vertretungsregelungen
    vertretungen = query('''
        SELECT v.id, v.von, v.bis, v.status,
               a.name AS abwesender, r.name AS vertreter
        FROM vertretung v
        JOIN mitarbeiter a ON a.id = v.abwesender_id
        LEFT JOIN mitarbeiter r ON r.id = v.vertreter_id
        ORDER BY v.von DESC
    ''')
    # Alle Außendienst-Mitarbeiter für Dropdowns
    alle_ad = query("SELECT id, name FROM mitarbeiter WHERE rolle IN ('rep','verkaufsleiter') ORDER BY name")

    # Offene Verkaufsstellen-Hinweise (z.B. Adress-/Namensänderung, aus "Neue Aktivität" gemeldet)
    vs_hinweise_offen = query('''
        SELECT h.id, h.text, h.erstellt_am, v.name AS verkaufsstelle, v.id AS verkaufsstelle_id, m.name AS mitarbeiter
        FROM vs_hinweis_meldung h
        JOIN verkaufsstelle v ON v.id = h.verkaufsstelle_id
        JOIN mitarbeiter m ON m.id = h.mitarbeiter_id
        WHERE h.status = 'offen'
        ORDER BY h.erstellt_am DESC
    ''')

    cfg = query("SELECT urlaubsmail_empfaenger, neue_vs_empfaenger FROM wochenbericht_config WHERE id=1", one=True)
    urlaubsmail_empfaenger = cfg['urlaubsmail_empfaenger'] if cfg else ''
    neue_vs_empfaenger     = cfg['neue_vs_empfaenger']     if cfg else ''

    return render_template('admin.html',
        mitarbeiter=mitarbeiter,
        verkaufsstellen=verkaufsstellen,
        vs_admin_gesamt=vs_admin_gesamt,
        vs_admin_seitengroesse=VS_ADMIN_SEITENGROESSE,
        biersorten=biersorten,
        displaysorte=displaysorte,
        zuordnungen=zuordnungen,
        vs_besitzer=vs_besitzer,
        vertretungen=vertretungen,
        vs_hinweise_offen=vs_hinweise_offen,
        alle_ad=alle_ad,
        teams=teams,
        mail_konfiguriert=mail_konfiguriert,
        urlaubsmail_empfaenger=urlaubsmail_empfaenger,
        neue_vs_empfaenger=neue_vs_empfaenger,
        export_email=EXPORT_EMAIL,
        bundeslaender=_BUNDESLAENDER,
        urlaub_konten={m['id']: _urlaub_konto(m['id'], date.today().year) for m in mitarbeiter})


@app.route('/admin/mitarbeiter/neu', methods=['POST'])
@manager_required
def admin_mitarbeiter_neu():
    is_admin = session.get('rolle') == 'admin'
    is_vkl   = session.get('rolle') == 'verkaufsleiter'
    name     = request.form.get('name',    '').strip()
    kuerzel  = request.form.get('kuerzel', '').strip().upper()
    passwort = request.form.get('passwort', DEFAULT_PASSWORD).strip()
    email    = request.form.get('email',   '').strip().lower() or None
    rolle    = request.form.get('rolle',   'rep').strip()
    # VKL kann nur Reps anlegen, keine Rollenwahl
    if is_vkl:
        rolle = 'rep'
    if rolle not in ('rep', 'verkaufsleiter', 'admin'):
        rolle = 'rep'
    redirect_target = url_for('admin') if is_admin else url_for('team_verwaltung')
    if name and kuerzel:
        if rolle == 'rep' and MAX_MITARBEITER > 0:
            anzahl = query("SELECT COUNT(*) AS n FROM mitarbeiter WHERE rolle='rep'", one=True)['n']
            if anzahl >= MAX_MITARBEITER:
                flash(f'Ihr Plan erlaubt max. {MAX_MITARBEITER} Mitarbeiter. Bitte kontaktieren Sie uns für ein Upgrade.', 'danger')
                return redirect(redirect_target)
        new_id = execute(
            "INSERT OR IGNORE INTO mitarbeiter (name, kuerzel, passwort, email, rolle, muss_passwort_aendern) VALUES (?,?,?,?,?,1)",
            (name, kuerzel, generate_password_hash(passwort), email, rolle)
        )
        # VKL: neuen Rep direkt dem eigenen Team zuordnen
        if is_vkl and session.get('team_id') and new_id:
            execute("UPDATE mitarbeiter SET team_id=? WHERE id=?", (session['team_id'], new_id))
        rollen_label = {'rep': 'Mitarbeiter', 'verkaufsleiter': 'Verkaufsleiter', 'admin': 'Leitung'}.get(rolle, rolle)
        flash(f'{rollen_label} „{name}" angelegt.', 'success')
    return redirect(redirect_target)


@app.route('/admin/mitarbeiter/<int:ma_id>/email', methods=['POST'])
@admin_required
def admin_mitarbeiter_email(ma_id):
    ma    = query("SELECT * FROM mitarbeiter WHERE id=?", (ma_id,), one=True)
    email = request.form.get('email', '').strip().lower() or None
    if not ma:
        flash('Mitarbeiter nicht gefunden.', 'danger')
        return redirect(url_for('admin'))
    # Doppelte E-Mail prüfen
    if email:
        existing = query(
            "SELECT id FROM mitarbeiter WHERE LOWER(email)=? AND id!=?",
            (email, ma_id), one=True
        )
        if existing:
            flash(f'Die E-Mail „{email}" ist bereits einem anderen Mitarbeiter zugeordnet.', 'danger')
            return redirect(url_for('admin'))
    execute("UPDATE mitarbeiter SET email=? WHERE id=?", (email, ma_id))
    if email:
        flash(f'E-Mail für „{ma["name"]}" auf {email} gesetzt.', 'success')
    else:
        flash(f'E-Mail für „{ma["name"]}" entfernt.', 'info')
    return redirect(url_for('admin'))


@app.route('/admin/mitarbeiter/<int:ma_id>/name', methods=['POST'])
@admin_required
def admin_mitarbeiter_name(ma_id):
    ma = query("SELECT * FROM mitarbeiter WHERE id=?", (ma_id,), one=True)
    if not ma:
        flash('Mitarbeiter nicht gefunden.', 'danger')
        return redirect(url_for('admin'))
    name    = request.form.get('name', '').strip()
    kuerzel = request.form.get('kuerzel', '').strip().upper()
    if not name or not kuerzel:
        flash('Name und Kürzel sind Pflichtfelder.', 'danger')
        return redirect(url_for('admin'))
    existing = query(
        "SELECT id FROM mitarbeiter WHERE UPPER(kuerzel)=? AND id!=?",
        (kuerzel, ma_id), one=True
    )
    if existing:
        flash(f'Das Kürzel „{kuerzel}" ist bereits vergeben.', 'danger')
        return redirect(url_for('admin'))
    execute("UPDATE mitarbeiter SET name=?, kuerzel=? WHERE id=?", (name, kuerzel, ma_id))
    flash(f'Name und Kürzel aktualisiert: „{name}" ({kuerzel}).', 'success')
    return redirect(url_for('admin'))


@app.route('/api/admin/mitarbeiter/<int:ma_id>/vs-liste')
@admin_required
def api_admin_mitarbeiter_vs_liste(ma_id):
    """Liste aller aktiven Verkaufsstellen fürs Zuordnungs-Modal – lazy per Fetch geladen,
    damit die Admin-Seite nicht für jeden Mitarbeiter eine komplette Kopie der Liste
    vorrendern muss (Performance-Fix 2026-07-09)."""
    zugeordnet_ids = {
        r['verkaufsstelle_id'] for r in query(
            "SELECT verkaufsstelle_id FROM mitarbeiter_verkaufsstelle WHERE mitarbeiter_id=?", (ma_id,)
        )
    }
    besitzer_rows = query('''
        SELECT mv.verkaufsstelle_id, m.name AS besitzer
        FROM mitarbeiter_verkaufsstelle mv
        JOIN mitarbeiter m ON m.id = mv.mitarbeiter_id
        WHERE mv.mitarbeiter_id != ?
    ''', (ma_id,))
    besitzer_map = {r['verkaufsstelle_id']: r['besitzer'] for r in besitzer_rows}
    vs_rows = query(
        "SELECT id, name, ort, typ FROM verkaufsstelle "
        "WHERE aktiv=1 AND homeoffice_mitarbeiter_id IS NULL ORDER BY name"
    )
    return jsonify([{
        'id': v['id'], 'name': v['name'], 'ort': v['ort'], 'typ': v['typ'],
        'zugeordnet': v['id'] in zugeordnet_ids,
        'besitzer': besitzer_map.get(v['id']),
    } for v in vs_rows])


@app.route('/admin/mitarbeiter/<int:ma_id>/zuordnung', methods=['POST'])
@admin_required
def admin_mitarbeiter_zuordnung(ma_id):
    ma = query("SELECT * FROM mitarbeiter WHERE id=?", (ma_id,), one=True)
    if not ma:
        flash('Mitarbeiter nicht gefunden.', 'danger')
        return redirect(url_for('admin'))
    vs_ids = [int(x) for x in request.form.getlist('vs_ids')]
    # Exklusive Zuordnung: jede VS gehört nur einer Person
    # → zuerst diese VS bei allen anderen entfernen
    for vs_id in vs_ids:
        execute(
            "DELETE FROM mitarbeiter_verkaufsstelle WHERE verkaufsstelle_id=? AND mitarbeiter_id!=?",
            (vs_id, ma_id)
        )
    execute("DELETE FROM mitarbeiter_verkaufsstelle WHERE mitarbeiter_id=?", (ma_id,))
    for vs_id in vs_ids:
        execute(
            "INSERT OR IGNORE INTO mitarbeiter_verkaufsstelle (mitarbeiter_id, verkaufsstelle_id) VALUES (?,?)",
            (ma_id, vs_id)
        )
    anzahl = len(vs_ids)
    flash(f'Zuordnung für „{ma["name"]}" gespeichert: {anzahl} Verkaufsstelle(n).', 'success')
    return redirect(url_for('admin'))


@app.route('/admin/mitarbeiter/<int:ma_id>/rolle', methods=['POST'])
@admin_required
def admin_mitarbeiter_rolle(ma_id):
    if ma_id == session.get('user_id'):
        flash('Sie können Ihre eigene Rolle nicht ändern.', 'danger')
        return redirect(url_for('admin'))
    ma    = query("SELECT * FROM mitarbeiter WHERE id=?", (ma_id,), one=True)
    rolle = request.form.get('rolle', 'rep').strip()
    if not ma or rolle not in ('rep', 'verkaufsleiter', 'admin'):
        flash('Ungültige Anfrage.', 'danger')
        return redirect(url_for('admin'))
    execute("UPDATE mitarbeiter SET rolle=? WHERE id=?", (rolle, ma_id))
    rollen_label = {'rep': 'Mitarbeiter', 'verkaufsleiter': 'Verkaufsleiter', 'admin': 'Leitung'}.get(rolle, rolle)
    flash(f'Rolle von „{ma["name"]}" auf {rollen_label} geändert.', 'success')
    return redirect(url_for('admin'))


@app.route('/admin/mitarbeiter/<int:ma_id>/loeschen', methods=['POST'])
@admin_required
def admin_mitarbeiter_loeschen(ma_id):
    if ma_id == session.get('user_id'):
        flash('Sie können sich nicht selbst löschen.', 'danger')
        return redirect(url_for('admin'))
    ma = query("SELECT * FROM mitarbeiter WHERE id=?", (ma_id,), one=True)
    if not ma:
        flash('Mitarbeiter nicht gefunden.', 'danger')
        return redirect(url_for('admin'))
    if ma['rolle'] == 'admin':
        flash('Admin-Konten können nicht gelöscht werden.', 'danger')
        return redirect(url_for('admin'))
    count = query("SELECT COUNT(*) AS c FROM aktivitaet WHERE mitarbeiter_id=?", (ma_id,), one=True)['c']
    if count > 0:
        flash(f'„{ma["name"]}" hat noch {count} Aktivität(en) und kann nicht gelöscht werden, '
              f'da sonst Geschäftsdaten (Bestellungen/Umsatzhistorie) verloren gehen. '
              f'Für eine Löschanfrage nach DSGVO stattdessen „Anonymisieren" verwenden.', 'danger')
        return redirect(url_for('admin'))
    execute("DELETE FROM mitarbeiter WHERE id=?", (ma_id,))
    flash(f'Mitarbeiter „{ma["name"]}" wurde gelöscht.', 'success')
    return redirect(url_for('admin'))


@app.route('/admin/mitarbeiter/<int:ma_id>/anonymisieren', methods=['POST'])
@admin_required
def admin_mitarbeiter_anonymisieren(ma_id):
    """Löschung nach Art. 17 DSGVO ist bei Mitarbeitern mit Aktivitäten technisch nicht
    als Hard-Delete möglich (Geschäftsdaten wie Bestellungen/Umsatzhistorie hängen daran,
    siehe admin_mitarbeiter_loeschen). Diese Route anonymisiert stattdessen alle
    personenbezogenen Felder (Name, Kürzel, E-Mail, Passwort, Wohnort) – die
    Geschäftsdaten (Aktivitäten, Bestellungen, Tagespläne) bleiben unter derselben
    mitarbeiter_id erhalten, sind aber keiner Person mehr zuordenbar. Bugreport
    2026-07-21: vorher gab es dafür überhaupt keinen Weg, nur den nie greifenden
    Hard-Delete-Pfad."""
    ma = query("SELECT * FROM mitarbeiter WHERE id=?", (ma_id,), one=True)
    if not ma:
        flash('Mitarbeiter nicht gefunden.', 'danger')
        return redirect(url_for('admin'))
    if ma['rolle'] == 'admin':
        flash('Admin-Konten können nicht anonymisiert werden.', 'danger')
        return redirect(url_for('admin'))
    if ma_id == session.get('user_id'):
        flash('Sie können sich nicht selbst anonymisieren.', 'danger')
        return redirect(url_for('admin'))
    platzhalter_name = f'Ehemaliger Mitarbeiter #{ma_id}'
    platzhalter_kuerzel = f'EX{ma_id}'
    execute('''
        UPDATE mitarbeiter SET
            name=?, kuerzel=?, email=NULL, passwort=?,
            wohnort_strasse=NULL, wohnort_plz=NULL, wohnort_ort=NULL,
            reset_token=NULL, reset_token_ablauf=NULL, karte_benachrichtigung=NULL,
            aktiv=0, konto_gesperrt=1, muss_passwort_aendern=0
        WHERE id=?
    ''', (platzhalter_name, platzhalter_kuerzel, generate_password_hash(secrets.token_urlsafe(32)), ma_id))
    # Bugreport 2026-07-21 (Kritisch): die Anonymisierung vergaß bisher die synthetische
    # Homeoffice-Verkaufsstelle (admin_mitarbeiter_wohnort) – die enthielt weiterhin
    # "Homeoffice – <Klarname>" plus volle Wohnadresse, sichtbar in Aktivitäten-Historie,
    # Karte und Excel-Exports. Jetzt mit-bereinigt.
    if ma['homeoffice_vs_id']:
        execute(
            "UPDATE verkaufsstelle SET name=?, strasse=NULL, plz=NULL, ort=NULL, lat=NULL, lng=NULL, aktiv=0 WHERE id=?",
            (f'Homeoffice – {platzhalter_name}', ma['homeoffice_vs_id'])
        )
    flash(f'„{ma["name"]}" wurde anonymisiert (Name/E-Mail/Wohnort inkl. Homeoffice-Adresse entfernt, Konto deaktiviert). '
          f'Geschäftsdaten (Aktivitäten, Bestellungen) bleiben zur Auswertung erhalten, sind aber keiner Person mehr zuordenbar.', 'success')
    return redirect(url_for('admin'))


@app.route('/admin/mitarbeiter/<int:ma_id>/passwort', methods=['POST'])
@manager_required
def admin_mitarbeiter_passwort(ma_id):
    is_admin = session.get('rolle') == 'admin'
    is_vkl   = session.get('rolle') == 'verkaufsleiter'
    redirect_target = url_for('admin') if is_admin else url_for('team_verwaltung')
    ma = query("SELECT * FROM mitarbeiter WHERE id=?", (ma_id,), one=True)
    if not ma:
        flash('Mitarbeiter nicht gefunden.', 'danger')
        return redirect(redirect_target)
    # VKL: nur Reps im eigenen Team
    if is_vkl:
        if ma['rolle'] != 'rep' or ma['team_id'] != session.get('team_id'):
            flash('Keine Berechtigung für diesen Mitarbeiter.', 'danger')
            return redirect(redirect_target)
    neues_pw = request.form.get('passwort', '').strip()
    if len(neues_pw) < 8:
        flash('Passwort muss mindestens 8 Zeichen haben.', 'danger')
        return redirect(redirect_target)
    muss_aendern = 1 if request.form.get('muss_passwort_aendern') else 0
    execute(
        "UPDATE mitarbeiter SET passwort=?, muss_passwort_aendern=?, konto_gesperrt=0, fehlgeschlagene_logins=0 WHERE id=?",
        (generate_password_hash(neues_pw), muss_aendern, ma_id)
    )
    if muss_aendern:
        flash(f'Passwort für „{ma["name"]}" wurde geändert – er/sie muss bei der nächsten Anmeldung ein eigenes Passwort festlegen.', 'success')
    else:
        flash(f'Passwort für „{ma["name"]}" wurde geändert.', 'success')
    return redirect(redirect_target)


@app.route('/admin/mitarbeiter/passwort-alle-zuruecksetzen', methods=['POST'])
@admin_required
def admin_mitarbeiter_passwort_alle_zuruecksetzen():
    """Massen-Aktion (z.B. beim Kunden-Onboarding): Setzt bei allen Mitarbeitern das
    Passwort auf das Standard-Passwort zurück und erzwingt beim nächsten Login die
    Festlegung eines eigenen Passworts."""
    rows = query("SELECT id, name FROM mitarbeiter WHERE UPPER(kuerzel) NOT IN ('ADMIN', 'DEMO')")
    for r in rows:
        execute(
            "UPDATE mitarbeiter SET passwort=?, muss_passwort_aendern=1, konto_gesperrt=0, fehlgeschlagene_logins=0 WHERE id=?",
            (generate_password_hash(DEFAULT_PASSWORD), r['id'])
        )
    flash(f'Passwort für {len(rows)} Mitarbeiter auf das Standard-Passwort zurückgesetzt – Passwort-Wechsel beim nächsten Login erzwungen.', 'success')
    return redirect(url_for('admin'))


@app.route('/admin/team/neu', methods=['POST'])
@admin_required
def admin_team_neu():
    name = request.form.get('name', '').strip()
    if not name:
        flash('Bitte einen Teamnamen eingeben.', 'danger')
        return redirect(url_for('admin'))
    existing = query("SELECT id FROM team WHERE LOWER(name) = LOWER(?)", (name,), one=True)
    if existing:
        flash(f'Ein Team mit dem Namen „{name}" existiert bereits.', 'warning')
        return redirect(url_for('admin'))
    execute("INSERT INTO team (name) VALUES (?)", (name,))
    flash(f'Team „{name}" angelegt.', 'success')
    return redirect(url_for('admin'))


@app.route('/admin/team/<int:team_id>/loeschen', methods=['POST'])
@admin_required
def admin_team_loeschen(team_id):
    team = query("SELECT * FROM team WHERE id=?", (team_id,), one=True)
    if not team:
        flash('Team nicht gefunden.', 'danger')
        return redirect(url_for('admin'))
    # Mitglieder aus Team austragen (nicht löschen)
    execute("UPDATE mitarbeiter SET team_id = NULL WHERE team_id = ?", (team_id,))
    execute("DELETE FROM team WHERE id=?", (team_id,))
    flash(f'Team „{team["name"]}" gelöscht. Mitglieder wurden keinem Team zugeordnet.', 'success')
    return redirect(url_for('admin'))


@app.route('/admin/mitarbeiter/<int:ma_id>/team', methods=['POST'])
@admin_required
def admin_mitarbeiter_team(ma_id):
    ma = query("SELECT * FROM mitarbeiter WHERE id=?", (ma_id,), one=True)
    if not ma:
        flash('Mitarbeiter nicht gefunden.', 'danger')
        return redirect(url_for('admin'))
    team_id_raw = request.form.get('team_id', '').strip()
    team_id = int(team_id_raw) if team_id_raw and team_id_raw.isdigit() else None
    execute("UPDATE mitarbeiter SET team_id=? WHERE id=?", (team_id, ma_id))
    if team_id:
        team = query("SELECT name FROM team WHERE id=?", (team_id,), one=True)
        flash(f'„{ma["name"]}" dem Team „{team["name"]}" zugeordnet.', 'success')
    else:
        flash(f'„{ma["name"]}" aus allen Teams ausgetragen.', 'info')
    return redirect(url_for('admin'))


@app.route('/admin/mitarbeiter/<int:ma_id>/urlaubskonto', methods=['POST'])
@admin_required
def admin_mitarbeiter_urlaubskonto(ma_id):
    ma = query("SELECT * FROM mitarbeiter WHERE id=?", (ma_id,), one=True)
    if not ma:
        flash('Mitarbeiter nicht gefunden.', 'danger')
        return redirect(url_for('admin'))
    anspruch  = request.form.get('urlaubsanspruch_jahr', type=int)
    uebertrag = request.form.get('urlaub_uebertrag_vorjahr', type=int)
    manuell_genommen = request.form.get('urlaub_manuell_genommen', type=int)
    bundesland = request.form.get('bundesland', 'BY').strip()
    if anspruch is None or anspruch < 0:
        flash('Ungültiger Jahresanspruch.', 'danger')
        return redirect(url_for('admin'))
    if uebertrag is None or uebertrag < 0:
        uebertrag = 0
    if manuell_genommen is None or manuell_genommen < 0:
        manuell_genommen = 0
    if bundesland not in _BUNDESLAENDER:
        bundesland = 'BY'
    execute(
        "UPDATE mitarbeiter SET urlaubsanspruch_jahr=?, urlaub_uebertrag_vorjahr=?, bundesland=?, urlaub_manuell_genommen=? WHERE id=?",
        (anspruch, uebertrag, bundesland, manuell_genommen, ma_id)
    )
    flash(f'Urlaubskonto von „{ma["name"]}" aktualisiert.', 'success')
    return redirect(url_for('admin'))


@app.route('/admin/mitarbeiter/<int:ma_id>/wohnort', methods=['POST'])
@admin_required
def admin_mitarbeiter_wohnort(ma_id):
    """Speichert den Wohnort und pflegt dabei automatisch eine synthetische
    „Verkaufsstelle" vom Typ Homeoffice, die dem Mitarbeiter zugeordnet wird –
    dadurch funktioniert Homeoffice als ganz normaler Besuchsplanungs-Stopp,
    ohne dass Tagesplan/Aktivität an ihrer Grundstruktur geändert werden müssen."""
    ma = query("SELECT * FROM mitarbeiter WHERE id=?", (ma_id,), one=True)
    if not ma:
        flash('Mitarbeiter nicht gefunden.', 'danger')
        return redirect(url_for('admin'))
    strasse = request.form.get('wohnort_strasse', '').strip() or None
    plz     = request.form.get('wohnort_plz', '').strip() or None
    ort     = request.form.get('wohnort_ort', '').strip() or None
    execute(
        "UPDATE mitarbeiter SET wohnort_strasse=?, wohnort_plz=?, wohnort_ort=? WHERE id=?",
        (strasse, plz, ort, ma_id)
    )
    if ort:
        vs_name = f"Homeoffice – {ma['name']}"
        if ma['homeoffice_vs_id']:
            execute(
                "UPDATE verkaufsstelle SET name=?, strasse=?, plz=?, ort=?, lat=NULL, lng=NULL WHERE id=?",
                (vs_name, strasse, plz, ort, ma['homeoffice_vs_id'])
            )
            homeoffice_vs_id = ma['homeoffice_vs_id']
        else:
            homeoffice_vs_id = execute(
                "INSERT INTO verkaufsstelle (name, strasse, plz, ort, typ, aktiv, homeoffice_mitarbeiter_id) "
                "VALUES (?,?,?,?,'Homeoffice',1,?)",
                (vs_name, strasse, plz, ort, ma_id)
            )
            execute("UPDATE mitarbeiter SET homeoffice_vs_id=? WHERE id=?", (homeoffice_vs_id, ma_id))
        execute(
            "INSERT OR IGNORE INTO mitarbeiter_verkaufsstelle (mitarbeiter_id, verkaufsstelle_id) VALUES (?,?)",
            (ma_id, homeoffice_vs_id)
        )
        flash(f'Wohnort von „{ma["name"]}" gespeichert – Homeoffice ist jetzt in der Besuchsplanung wählbar.', 'success')
    else:
        flash(f'Wohnort von „{ma["name"]}" aktualisiert.', 'success')
    return redirect(url_for('admin'))


@app.route('/profil/meine-daten')
@login_required
def profil_meine_daten():
    """Selbstbedienungs-Auskunft nach Art. 15 DSGVO (Bugreport 2026-07-21, Mittel):
    bisher war eine strukturierte Auskunft über die eigenen gespeicherten Daten nur
    manuell per DB-Query durch einen Admin möglich. Passwort-Hash und interne Tokens
    werden bewusst nicht mit ausgegeben."""
    uid = session['user_id']
    ma = query(
        "SELECT name, kuerzel, email, rolle, wohnort_strasse, wohnort_plz, wohnort_ort, "
        "urlaubsanspruch_jahr, urlaub_uebertrag_vorjahr FROM mitarbeiter WHERE id=?",
        (uid,), one=True
    )
    aktivitaeten = query(
        "SELECT a.datum, v.name AS verkaufsstelle, a.notizen "
        "FROM aktivitaet a JOIN verkaufsstelle v ON v.id=a.verkaufsstelle_id "
        "WHERE a.mitarbeiter_id=? ORDER BY a.datum DESC", (uid,)
    )
    arbeitszeiten = query(
        "SELECT datum, beginn, ende, pause_minuten FROM arbeitszeit WHERE mitarbeiter_id=? ORDER BY datum DESC",
        (uid,)
    )
    vertretungen = query(
        "SELECT typ, von, bis, status, grund FROM vertretung WHERE vertreter_id=? OR abwesender_id=? ORDER BY von DESC",
        (uid, uid)
    )
    export = {
        'stammdaten': dict(ma) if ma else {},
        'aktivitaeten': [dict(r) for r in aktivitaeten],
        'arbeitszeiten': [dict(r) for r in arbeitszeiten],
        'abwesenheiten': [dict(r) for r in vertretungen],
        'hinweis': 'Auskunft nach Art. 15 DSGVO – automatisiert erzeugt am ' + date.today().isoformat(),
    }
    resp = jsonify(export)
    resp.headers['Content-Disposition'] = f'attachment; filename=meine_daten_{date.today().isoformat()}.json'
    app.logger.info(f"DSGVO-Selbstauskunft heruntergeladen: user_id={uid}")
    return resp


@app.route('/profil/passwort', methods=['POST'])
@login_required
def profil_passwort():
    altes_pw  = request.form.get('altes_passwort', '')
    neues_pw  = request.form.get('neues_passwort', '').strip()
    neues_pw2 = request.form.get('neues_passwort2', '').strip()
    user = query("SELECT * FROM mitarbeiter WHERE id=?", (session['user_id'],), one=True)
    if not _password_ok(user['passwort'], altes_pw):
        flash('Aktuelles Passwort ist falsch.', 'danger')
    elif len(neues_pw) < 8:
        flash('Neues Passwort muss mindestens 8 Zeichen haben.', 'danger')
    elif neues_pw != neues_pw2:
        flash('Passwörter stimmen nicht überein.', 'danger')
    else:
        execute("UPDATE mitarbeiter SET passwort=? WHERE id=?", (generate_password_hash(neues_pw), session['user_id']))
        flash('Passwort erfolgreich geändert.', 'success')
    return redirect(request.referrer or url_for('dashboard'))


@app.route('/profil/daten', methods=['POST'])
@login_required
def profil_daten():
    name    = request.form.get('name', '').strip()
    kuerzel = request.form.get('kuerzel', '').strip().upper()
    email   = request.form.get('email', '').strip().lower() or None
    if not name or not kuerzel:
        flash('Name und Kürzel sind Pflichtfelder.', 'danger')
        return redirect(request.referrer or url_for('dashboard'))
    conflict = query("SELECT id FROM mitarbeiter WHERE UPPER(kuerzel)=? AND id!=?",
                     (kuerzel, session['user_id']), one=True)
    if conflict:
        flash(f'Kürzel „{kuerzel}" wird bereits von jemand anderem verwendet.', 'danger')
        return redirect(request.referrer or url_for('dashboard'))
    execute("UPDATE mitarbeiter SET name=?, kuerzel=?, email=? WHERE id=?",
            (name, kuerzel, email, session['user_id']))
    session['name']    = name
    session['kuerzel'] = kuerzel
    flash('Ihre Daten wurden aktualisiert.', 'success')
    return redirect(request.referrer or url_for('dashboard'))


@app.route('/profil/signatur', methods=['POST'])
@login_required
def profil_signatur():
    """Speichert die einmalig hinterlegte Unterschrift (PNG-Data-URI aus dem Canvas-
    Signaturpad in base.html) am eigenen Mitarbeiter-Datensatz. Wird künftig automatisch
    in den Urlaubsantrag eingebettet (siehe _urlaubsantrag_pdf_bytes) – kein erneutes
    Unterschreiben je Antrag nötig. Grobe Plausibilitätsprüfung statt echter Bildvalidierung
    reicht hier aus, da das Canvas selbst nur PNG-Data-URIs erzeugen kann."""
    data = request.get_json(silent=True) or {}
    bild = (data.get('signatur') or '').strip()
    if not bild.startswith('data:image/png;base64,') or len(bild) > 200_000:
        return jsonify({'ok': False, 'error': 'Ungültige Signatur.'}), 400
    execute("UPDATE mitarbeiter SET signatur_bild=? WHERE id=?", (bild, session['user_id']))
    return jsonify({'ok': True})


@app.route('/profil/signatur/loeschen', methods=['POST'])
@login_required
def profil_signatur_loeschen():
    execute("UPDATE mitarbeiter SET signatur_bild=NULL WHERE id=?", (session['user_id'],))
    return jsonify({'ok': True})


@app.route('/faq')
@login_required
def faq():
    return render_template('faq.html')


@app.route('/admin/vertretung/neu', methods=['POST'])
@manager_required
def admin_vertretung_neu():
    is_admin      = session.get('rolle') == 'admin'
    _redir        = url_for('admin') if is_admin else url_for('team_verwaltung')
    abwesender_id = request.form.get('abwesender_id', type=int)
    vertreter_id  = request.form.get('vertreter_id',  type=int)
    von           = request.form.get('von', '').strip()
    bis           = request.form.get('bis', '').strip()
    typ           = request.form.get('typ', 'urlaub').strip()
    grund         = request.form.get('grund', '').strip()
    if typ not in ('urlaub', 'krankheit', 'sonderurlaub', 'frei_sonderurlaub', 'unbezahlt'):
        typ = 'urlaub'
    if not all([abwesender_id, von, bis]):
        flash('Abwesender Mitarbeiter, Von und Bis sind Pflichtfelder.', 'danger')
        return redirect(_redir)
    if typ == 'frei_sonderurlaub' and not grund:
        flash('Bitte einen Grund für Frei/Sonderurlaub angeben.', 'danger')
        return redirect(_redir)
    if vertreter_id and abwesender_id == vertreter_id:
        flash('Abwesender und Vertreter dürfen nicht dieselbe Person sein.', 'danger')
        return redirect(_redir)
    # VKL: nur Reps im eigenen Team als abwesend eintragen
    if not is_admin and session.get('team_id'):
        ma = query("SELECT team_id, rolle FROM mitarbeiter WHERE id=?", (abwesender_id,), one=True)
        if not ma or ma['team_id'] != session.get('team_id'):
            flash('Keine Berechtigung für diesen Mitarbeiter.', 'danger')
            return redirect(_redir)
    execute(
        "INSERT INTO vertretung (abwesender_id, vertreter_id, von, bis, status, typ, grund, erstellt_am) VALUES (?,?,?,?,'bestätigt',?,?,?)",
        (abwesender_id, vertreter_id, von, bis, typ, grund or None, date.today().isoformat())
    )
    flash('Urlaub / Vertretung gespeichert.', 'success')
    return redirect(_redir)


@app.route('/admin/vertretung/<int:vtr_id>/loeschen', methods=['POST'])
@manager_required
def admin_vertretung_loeschen(vtr_id):
    if not _vertretung_team_ok(vtr_id):
        flash('Keine Berechtigung für diesen Eintrag.', 'danger')
        return redirect(request.referrer or url_for('dashboard'))
    execute("DELETE FROM vertretung WHERE id=?", (vtr_id,))
    flash('Eintrag gelöscht.', 'success')
    return redirect(request.referrer or url_for('dashboard'))


def _vertretung_team_ok(vtr_id):
    """True wenn der aktuelle Manager diesen Vertretungs-/Urlaubseintrag bearbeiten darf.
    Admin: immer. VKL: nur wenn der Abwesende im eigenen Team ist – hat der VKL selbst
    kein Team zugeordnet, gilt wie überall sonst in der App keine Einschränkung (statt
    fälschlich jeden Antrag zu blockieren)."""
    if session.get('rolle') == 'admin':
        return True
    if session.get('rolle') == 'verkaufsleiter':
        tid = session.get('team_id')
        if not tid:
            return True
        row = query('''SELECT m.team_id FROM vertretung v
                       JOIN mitarbeiter m ON m.id = v.abwesender_id
                       WHERE v.id = ?''', (vtr_id,), one=True)
        return bool(row and row['team_id'] == tid)
    return False


def _send_vertretung_email(vtr_id: int, status: str):
    """Sendet E-Mail-Benachrichtigung nach Urlaubsentscheidung an Rep + VKL/Admin. Bei
    genehmigtem Urlaub (typ='urlaub') wird zusätzlich der digitale Urlaubsantrag als PDF
    angehängt (siehe _urlaubsantrag_pdf_bytes). Bei fehlgeschlagener PDF-Erzeugung wird
    die Mail trotzdem ohne Anhang verschickt (kein Blocker für die Benachrichtigung)."""
    vtr = query(
        """SELECT v.typ, v.von, v.bis, m.name AS abwesender_name, m.email AS abwesender_email,
                  vt.name AS vertreter_name
           FROM vertretung v
           JOIN mitarbeiter m ON m.id = v.abwesender_id
           LEFT JOIN mitarbeiter vt ON vt.id = v.vertreter_id
           WHERE v.id=?""",
        (vtr_id,), one=True
    )
    if not vtr:
        return
    farbe  = '#2cc4b0' if status == 'bestätigt' else '#e24b4a'
    icon   = '✓' if status == 'bestätigt' else '✗'
    subject = f"Urlaubsantrag {icon} {status} – {vtr['abwesender_name']}"

    attachments = []
    if status == 'bestätigt' and vtr['typ'] == 'urlaub':
        pdf = _urlaubsantrag_pdf_bytes(vtr_id)
        if pdf:
            dateiname = f"Urlaubsantrag_{re.sub(r'[^A-Za-z0-9]+', '_', vtr['abwesender_name']).strip('_')}_{vtr['von']}.pdf"
            attachments = [(dateiname, pdf, 'application/pdf')]

    vertreter_zeile = (
        f"<tr><td style='padding:.35rem .7rem;background:#f4f6fa;font-weight:600'>Vertretung</td>"
        f"<td style='padding:.35rem .7rem'>{vtr['vertreter_name']}</td></tr>"
    ) if vtr['vertreter_name'] else ''
    body = f"""
    <div style="font-family:Arial,sans-serif;max-width:560px;margin:0 auto">
      <div style="background:#1a3a5c;color:#fff;padding:1.4rem 1.6rem;border-radius:6px 6px 0 0">
        <h2 style="margin:0;font-size:1.2rem">Urlaubsantrag <span style="color:{farbe}">{icon} {status}</span></h2>
      </div>
      <div style="padding:1.4rem 1.6rem;border:1px solid #e0e0e0;border-top:none;border-radius:0 0 6px 6px">
        <p style="margin:0 0 1rem">Hallo {vtr['abwesender_name']},</p>
        <p style="margin:0 0 1rem">dein Urlaubsantrag wurde
          <strong style="color:{farbe}">{status}</strong>.</p>
        <table style="width:100%;border-collapse:collapse;margin:0 0 1rem">
          <tr><td style="padding:.35rem .7rem;background:#f4f6fa;font-weight:600">Zeitraum</td>
              <td style="padding:.35rem .7rem">{vtr['von']} bis {vtr['bis']}</td></tr>
          {vertreter_zeile}
        </table>
        {'<p style="color:#555;font-size:.85rem;margin:0 0 1rem"><i>Der ausgefüllte Urlaubsantrag ist als PDF angehängt.</i></p>' if attachments else ''}
        <hr style="border:none;border-top:1px solid #e8e8e8;margin:1rem 0">
        <p style="color:#777;font-size:.82rem;margin:0">
          Diese Nachricht wurde automatisch vom Aktionstracker gesendet.</p>
      </div>
    </div>"""

    def _versenden(to):
        if attachments:
            return send_email_with_attachments(to, subject, body, attachments)
        return send_email(to, subject, body)

    if vtr['abwesender_email']:
        _versenden(vtr['abwesender_email'])
    if status == 'bestätigt':
        manager_emails = query(
            "SELECT email FROM mitarbeiter WHERE rolle IN ('admin','verkaufsleiter') AND email IS NOT NULL AND email != ''",
            ()
        )
        gesendet = {vtr['abwesender_email']} if vtr['abwesender_email'] else set()
        for mgr in manager_emails:
            if mgr['email'] not in gesendet:
                _versenden(mgr['email'])
                gesendet.add(mgr['email'])
        cfg = query("SELECT urlaubsmail_empfaenger FROM wochenbericht_config WHERE id=1", one=True)
        if cfg and cfg['urlaubsmail_empfaenger']:
            for addr in [a.strip() for a in cfg['urlaubsmail_empfaenger'].split(',') if a.strip()]:
                if addr not in gesendet:
                    _versenden(addr)


@app.route('/admin/vertretung/<int:vtr_id>/bestaetigen', methods=['POST'])
@manager_required
def admin_vertretung_bestaetigen(vtr_id):
    if not _vertretung_team_ok(vtr_id):
        flash('Keine Berechtigung für diesen Antrag.', 'danger')
    else:
        execute("UPDATE vertretung SET status='bestätigt', bestaetigt_von_id=? WHERE id=?", (session['user_id'], vtr_id))
        flash('Urlaub bestätigt.', 'success')
        _send_vertretung_email(vtr_id, 'bestätigt')
    return redirect(request.referrer or url_for('dashboard'))


@app.route('/admin/vertretung/<int:vtr_id>/ablehnen', methods=['POST'])
@manager_required
def admin_vertretung_ablehnen(vtr_id):
    if not _vertretung_team_ok(vtr_id):
        flash('Keine Berechtigung für diesen Antrag.', 'danger')
    else:
        execute("UPDATE vertretung SET status='abgelehnt' WHERE id=?", (vtr_id,))
        flash('Urlaubsantrag abgelehnt.', 'warning')
        _send_vertretung_email(vtr_id, 'abgelehnt')
    return redirect(request.referrer or url_for('dashboard'))


@app.route('/admin/vs-hinweis/<int:hinweis_id>/erledigt', methods=['POST'])
@admin_required
def admin_vs_hinweis_erledigt(hinweis_id):
    execute(
        "UPDATE vs_hinweis_meldung SET status='erledigt', erledigt_am=datetime('now','localtime'), erledigt_von_id=? WHERE id=?",
        (session['user_id'], hinweis_id)
    )
    flash('Hinweis als erledigt markiert.', 'success')
    return redirect(url_for('admin') + '#vs-hinweise')


@app.route('/admin/urlaubsmail/empfaenger', methods=['POST'])
@admin_required
def admin_urlaubsmail_empfaenger():
    empfaenger = request.form.get('urlaubsmail_empfaenger', '').strip()
    execute("UPDATE wochenbericht_config SET urlaubsmail_empfaenger=? WHERE id=1", (empfaenger,))
    flash('Urlaubsmail-Empfänger gespeichert.', 'success')
    return redirect(url_for('admin') + '#vertretung')


@app.route('/admin/neue-vs/empfaenger', methods=['POST'])
@admin_required
def admin_neue_vs_empfaenger():
    empfaenger = request.form.get('neue_vs_empfaenger', '').strip()
    execute("UPDATE wochenbericht_config SET neue_vs_empfaenger=? WHERE id=1", (empfaenger,))
    flash('E-Mail-Adresse für neue Verkaufsstellen gespeichert.', 'success')
    return redirect(url_for('admin'))


def _notify_neue_vs(name, strasse, plz, ort, typ, ansprechpartner, erstellt_von_name):
    cfg = query("SELECT neue_vs_empfaenger FROM wochenbericht_config WHERE id=1", one=True)
    if not cfg or not cfg['neue_vs_empfaenger']:
        return
    adresse_teile = [t for t in [strasse, (f'{plz} {ort}').strip()] if t]
    adresse = ', '.join(adresse_teile) if adresse_teile else '–'
    jetzt = datetime.now().strftime('%d.%m.%Y %H:%M')
    html = f'''<!DOCTYPE html>
<html><head><meta charset="utf-8"></head>
<body style="margin:0;padding:0;background:#f0f4f8;font-family:Arial,Helvetica,sans-serif">
<div style="max-width:520px;margin:32px auto;background:#fff;border-radius:10px;overflow:hidden;box-shadow:0 2px 12px rgba(0,0,0,0.10)">
  <div style="background:#1a3a5c;padding:20px 28px">
    <div style="color:#fff;font-size:17px;font-weight:bold">Aktions Tracker – Neue Verkaufsstelle</div>
    <div style="color:#90b8d8;font-size:12px;margin-top:4px">Angelegt am {jetzt} von {erstellt_von_name}</div>
  </div>
  <div style="padding:24px 28px">
    <table width="100%" cellpadding="0" cellspacing="0">
      <tr><td style="padding:6px 0;font-size:13px;color:#666;width:130px">Name</td>
          <td style="padding:6px 0;font-size:13px;font-weight:bold;color:#1a3a5c">{name}</td></tr>
      <tr><td style="padding:6px 0;font-size:13px;color:#666">Adresse</td>
          <td style="padding:6px 0;font-size:13px;color:#333">{adresse}</td></tr>
      <tr><td style="padding:6px 0;font-size:13px;color:#666">Typ</td>
          <td style="padding:6px 0;font-size:13px;color:#333">{typ or '–'}</td></tr>
      <tr><td style="padding:6px 0;font-size:13px;color:#666">Ansprechpartner</td>
          <td style="padding:6px 0;font-size:13px;color:#333">{ansprechpartner or '–'}</td></tr>
      <tr><td style="padding:6px 0;font-size:13px;color:#666">Angelegt von</td>
          <td style="padding:6px 0;font-size:13px;color:#333">{erstellt_von_name}</td></tr>
    </table>
    <div style="margin-top:16px;padding:12px 16px;background:#fff8e1;border-left:4px solid #c8860a;border-radius:4px;font-size:12px;color:#8a5a00">
      Bitte Verkaufsstelle in den zentralen Stammdaten anlegen, damit Belieferung sichergestellt ist.
    </div>
  </div>
  <div style="padding:12px 28px;background:#f4f8fc;border-top:1px solid #e4eaf0;text-align:center">
    <div style="font-size:11px;color:#aaa">Aktions Tracker – automatische Benachrichtigung</div>
  </div>
</div>
</body></html>'''
    for addr in [a.strip() for a in cfg['neue_vs_empfaenger'].split(',') if a.strip()]:
        try:
            send_email(addr, f'Neue Verkaufsstelle: {name}, {ort}', html)
        except Exception as e:
            app.logger.error(f"Neue-VS-Benachrichtigung an {addr} fehlgeschlagen: {e}")


@app.route('/profil/vertretung/neu', methods=['POST'])
@login_required
def profil_vertretung_neu():
    vertreter_id = request.form.get('vertreter_id', type=int)
    von          = request.form.get('von', '').strip()
    bis          = request.form.get('bis', '').strip()
    typ          = request.form.get('typ', 'urlaub').strip()
    grund        = request.form.get('grund', '').strip()
    if typ not in ('urlaub', 'krankheit', 'sonderurlaub', 'frei_sonderurlaub', 'unbezahlt'):
        typ = 'urlaub'
    if not all([von, bis]):
        flash('Von und Bis sind Pflichtfelder.', 'danger')
        return redirect(request.referrer or url_for('dashboard'))
    if typ == 'frei_sonderurlaub' and not grund:
        flash('Bitte einen Grund für Frei/Sonderurlaub angeben.', 'danger')
        return redirect(request.referrer or url_for('dashboard'))
    if vertreter_id and vertreter_id == session['user_id']:
        flash('Sie können sich nicht selbst als Vertreter eintragen.', 'danger')
        return redirect(request.referrer or url_for('dashboard'))
    # VKL/GF tragen ihren eigenen Urlaub direkt bestätigt ein; Reps müssen anfragen.
    _status = 'bestätigt' if session.get('rolle') in ('admin', 'verkaufsleiter') else 'angefragt'
    execute(
        "INSERT INTO vertretung (abwesender_id, vertreter_id, von, bis, status, typ, grund, erstellt_am) VALUES (?,?,?,?,?,?,?,?)",
        (session['user_id'], vertreter_id, von, bis, _status, typ, grund or None, date.today().isoformat())
    )
    _typ_label = {'urlaub': 'Urlaub', 'krankheit': 'Krankheit', 'sonderurlaub': 'Sonderurlaub', 'frei_sonderurlaub': 'Frei/Sonderurlaub', 'unbezahlt': 'Unbezahlter Urlaub'}[typ]
    if typ == 'urlaub':
        _tage = _werktage_zaehlen(von, bis, query("SELECT bundesland FROM mitarbeiter WHERE id=?", (session['user_id'],), one=True)['bundesland'])
        _konto = _urlaub_konto(session['user_id'], date.fromisoformat(von).year)
        _rest_info = f' {_konto["verfuegbar"]} Tage verbleiben {date.fromisoformat(von).year}.' if _konto else ''
        if _status == 'angefragt':
            flash(f'{_typ_label} angefragt ({_tage} Werktage) – wartet auf Bestätigung durch Verkaufsleiter oder Leitung.{_rest_info}', 'success')
        else:
            flash(f'{_typ_label} eingetragen ({_tage} Werktage).{_rest_info}', 'success')
    else:
        if _status == 'angefragt':
            flash(f'{_typ_label} angefragt – wartet auf Bestätigung durch Verkaufsleiter oder Leitung.', 'success')
        else:
            flash(f'{_typ_label} eingetragen.', 'success')
    return redirect(request.referrer or url_for('dashboard'))


@app.route('/profil/vertretung/<int:vtr_id>/loeschen', methods=['POST'])
@login_required
def profil_vertretung_loeschen(vtr_id):
    vtr = query("SELECT * FROM vertretung WHERE id=?", (vtr_id,), one=True)
    if not vtr or vtr['abwesender_id'] != session['user_id']:
        flash('Nicht gefunden oder keine Berechtigung.', 'danger')
        return redirect(request.referrer or url_for('dashboard'))
    # Bestätigten Urlaub kann nur VKL/GF wieder entfernen.
    if vtr['status'] == 'bestätigt':
        flash('Bestätigter Urlaub kann nur von Verkaufsleiter oder Leitung gelöscht werden.', 'danger')
        return redirect(request.referrer or url_for('dashboard'))
    execute("DELETE FROM vertretung WHERE id=?", (vtr_id,))
    flash('Eintrag zurückgenommen.', 'success')
    return redirect(request.referrer or url_for('dashboard'))


@app.route('/admin/verkaufsstelle/neu', methods=['POST'])
@manager_required
def admin_vs_neu():
    name             = request.form.get('name',             '').strip()
    strasse          = request.form.get('strasse',          '').strip()
    plz              = request.form.get('plz',              '').strip()
    ort              = request.form.get('ort',              '').strip()
    landkreis        = request.form.get('landkreis',        '').strip()
    typ              = request.form.get('typ',              '').strip()
    ansprechpartner  = request.form.get('ansprechpartner',  '').strip()
    lieferant        = request.form.get('lieferant',        '').strip()
    kundennummer     = request.form.get('kundennummer',     '').strip()
    hinweis          = request.form.get('hinweis',          '').strip()
    if name:
        new_id = execute(
            "INSERT INTO verkaufsstelle (name, strasse, plz, ort, landkreis, typ, ansprechpartner, lieferant, kundennummer, hinweis) VALUES (?,?,?,?,?,?,?,?,?,?)",
            (name, strasse, plz or None, ort, landkreis or None, typ, ansprechpartner, lieferant or None, kundennummer or None, hinweis or None)
        )
        if KARTE_MODUS != 'aus' and (strasse or ort):
            lat, lng, quelle, kreis_aus_geo = _geocode_adresse(strasse, ort, plz=plz or None)
            if lat is not None:
                if kreis_aus_geo and not landkreis:
                    execute("UPDATE verkaufsstelle SET lat=?, lng=?, geocode_quelle=?, landkreis=? WHERE id=?",
                            (lat, lng, quelle, kreis_aus_geo, new_id))
                else:
                    execute("UPDATE verkaufsstelle SET lat=?, lng=?, geocode_quelle=? WHERE id=?",
                            (lat, lng, quelle, new_id))
                flash(f'Verkaufsstelle "{name}" angelegt und auf Karte verortet.', 'success')
            else:
                flash(f'Verkaufsstelle "{name}" angelegt. Koordinaten konnten nicht automatisch ermittelt werden – bitte "Koordinaten ermitteln" auf der Karte nutzen.', 'warning')
        else:
            flash(f'Verkaufsstelle "{name}" angelegt.', 'success')
        ersteller = query("SELECT name FROM mitarbeiter WHERE id=?", (session['user_id'],), one=True)
        _notify_neue_vs(name, strasse, plz, ort, typ, ansprechpartner,
                        ersteller['name'] if ersteller else session.get('user_name', 'Admin'))
    is_admin = session.get('rolle') == 'admin'
    return redirect(url_for('admin') if is_admin else url_for('team_verwaltung'))


@app.route('/admin/verkaufsstelle/<int:vs_id>/bearbeiten', methods=['POST'])
@admin_required
def admin_vs_bearbeiten(vs_id):
    vs = query("SELECT * FROM verkaufsstelle WHERE id=?", (vs_id,), one=True)
    if not vs:
        flash('Verkaufsstelle nicht gefunden.', 'danger')
        return redirect(url_for('admin'))
    name            = request.form.get('name',            '').strip()
    strasse         = request.form.get('strasse',         '').strip()
    plz             = request.form.get('plz',             '').strip()
    ort             = request.form.get('ort',             '').strip()
    landkreis       = request.form.get('landkreis',       '').strip()
    typ             = request.form.get('typ',             '').strip()
    ansprechpartner = request.form.get('ansprechpartner', '').strip()
    lieferant       = request.form.get('lieferant',       '').strip()
    kundennummer    = request.form.get('kundennummer',    '').strip()
    hinweis         = request.form.get('hinweis',         '').strip()
    if not name:
        flash('Name ist ein Pflichtfeld.', 'danger')
        return redirect(url_for('admin'))
    adresse_geaendert = strasse != (vs['strasse'] or '') or ort != (vs['ort'] or '')
    execute(
        "UPDATE verkaufsstelle SET name=?, strasse=?, plz=?, ort=?, landkreis=?, typ=?, ansprechpartner=?, "
        "lieferant=?, kundennummer=?, hinweis=? WHERE id=?",
        (name, strasse, plz or None, ort, landkreis or None, typ, ansprechpartner,
         lieferant or None, kundennummer or None, hinweis or None, vs_id)
    )
    if adresse_geaendert and KARTE_MODUS != 'aus' and (strasse or ort):
        lat, lng, quelle, kreis_aus_geo = _geocode_adresse(strasse, ort, plz=plz or None)
        if lat is not None:
            if kreis_aus_geo and not landkreis:
                execute("UPDATE verkaufsstelle SET lat=?, lng=?, geocode_quelle=?, landkreis=? WHERE id=?",
                        (lat, lng, quelle, kreis_aus_geo, vs_id))
            else:
                execute("UPDATE verkaufsstelle SET lat=?, lng=?, geocode_quelle=? WHERE id=?",
                        (lat, lng, quelle, vs_id))
            flash(f'„{name}" gespeichert und neu auf Karte verortet.', 'success')
        else:
            flash(f'„{name}" gespeichert. Koordinaten konnten nicht neu ermittelt werden.', 'warning')
    else:
        flash(f'„{name}" gespeichert.', 'success')
    return redirect(url_for('admin'))


@app.route('/admin/verkaufsstelle/<int:vs_id>/loeschen', methods=['POST'])
@admin_required
def admin_vs_loeschen(vs_id):
    vs = query("SELECT * FROM verkaufsstelle WHERE id=?", (vs_id,), one=True)
    if not vs:
        flash('Verkaufsstelle nicht gefunden.', 'danger')
        return redirect(url_for('admin'))
    count = query("SELECT COUNT(*) AS c FROM aktivitaet WHERE verkaufsstelle_id=?", (vs_id,), one=True)['c']
    if count > 0:
        execute("UPDATE verkaufsstelle SET aktiv=0 WHERE id=?", (vs_id,))
        flash(f'„{vs["name"]}" hat {count} verknüpfte Aktivität(en) und wurde deaktiviert '
              f'(erscheint nicht mehr in neuen Aktivitäten, historische Daten bleiben erhalten).', 'warning')
    else:
        execute("DELETE FROM verkaufsstelle WHERE id=?", (vs_id,))
        flash(f'Verkaufsstelle „{vs["name"]}" wurde gelöscht.', 'success')
    return redirect(url_for('admin'))


@app.route('/admin/verkaufsstelle/<int:vs_id>/reaktivieren', methods=['POST'])
@admin_required
def admin_vs_reaktivieren(vs_id):
    vs = query("SELECT * FROM verkaufsstelle WHERE id=?", (vs_id,), one=True)
    if vs:
        execute("UPDATE verkaufsstelle SET aktiv=1 WHERE id=?", (vs_id,))
        flash(f'Verkaufsstelle „{vs["name"]}" wurde reaktiviert.', 'success')
    return redirect(url_for('admin'))


@app.route('/admin/vs-ohne-koordinaten')
@admin_required
def admin_vs_ohne_koordinaten():
    stellen = query(
        "SELECT id, name, strasse, plz, ort, landkreis, typ, lat, lng FROM verkaufsstelle "
        "WHERE aktiv=1 AND (lat IS NULL OR lat=0) ORDER BY landkreis, ort, name"
    )
    return render_template('vs_ohne_koordinaten.html', stellen=stellen)


@app.route('/admin/vs-koordinaten-setzen/<int:vs_id>', methods=['POST'])
@admin_required
def admin_vs_koordinaten_setzen(vs_id):
    try:
        lat = float(request.form.get('lat', '').replace(',', '.'))
        lng = float(request.form.get('lng', '').replace(',', '.'))
        b = _DACH_BBOX
        if not (b['lat_min'] <= lat <= b['lat_max'] and b['lon_min'] <= lng <= b['lon_max']):
            return jsonify({'error': 'Koordinaten außerhalb des erlaubten Bereichs'}), 400
        execute("UPDATE verkaufsstelle SET lat=?, lng=? WHERE id=?", (lat, lng, vs_id))
        return jsonify({'ok': True, 'lat': lat, 'lng': lng})
    except (ValueError, TypeError) as e:
        return jsonify({'error': str(e)}), 400


@app.route('/verkaufsstelle/neu', methods=['POST'])
@login_required
def vs_neu_rep():
    name            = request.form.get('name',            '').strip()
    strasse         = request.form.get('strasse',         '').strip()
    ort             = request.form.get('ort',             '').strip()
    plz_rep         = request.form.get('plz',             '').strip()
    typ             = request.form.get('typ',             '').strip()
    ansprechpartner = request.form.get('ansprechpartner', '').strip()
    next_page       = request.form.get('next',            '').strip()
    vom_dashboard   = next_page == 'tagesplan'
    if not name or not strasse or not plz_rep or not ort:
        flash('Name, Straße, PLZ und Ort sind Pflichtfelder.', 'danger')
        return redirect(url_for('dashboard') if vom_dashboard else url_for('neue_aktivitaet'))
    if name:
        # Duplikat-Check: gleicher Name + Ort (Groß-/Kleinschreibung egal)
        vorhanden = query(
            "SELECT id, name, ort FROM verkaufsstelle WHERE LOWER(name)=LOWER(?) AND LOWER(COALESCE(ort,''))=LOWER(?) AND aktiv=1",
            (name, ort), one=True
        )
        if vorhanden:
            # Existiert bereits → direkt zuordnen und weiterleiten
            if session.get('rolle') in ('rep', 'verkaufsleiter'):
                execute(
                    "INSERT OR IGNORE INTO mitarbeiter_verkaufsstelle (mitarbeiter_id, verkaufsstelle_id) VALUES (?,?)",
                    (session['user_id'], vorhanden['id'])
                )
            flash(f'„{vorhanden["name"]}" in {vorhanden["ort"] or "unbekanntem Ort"} existiert bereits – direkt ausgewählt.', 'info')
            return redirect(url_for('dashboard') if vom_dashboard else url_for('neue_aktivitaet', vs_id=vorhanden['id']))

        new_id = execute(
            "INSERT INTO verkaufsstelle (name, strasse, plz, ort, typ, ansprechpartner) VALUES (?,?,?,?,?,?)",
            (name, strasse, plz_rep, ort, typ, ansprechpartner)
        )
        # Reps/VKL: neue Verkaufsstelle direkt dem Ersteller zuordnen
        if session.get('rolle') in ('rep', 'verkaufsleiter'):
            execute(
                "INSERT OR IGNORE INTO mitarbeiter_verkaufsstelle (mitarbeiter_id, verkaufsstelle_id) VALUES (?,?)",
                (session['user_id'], new_id)
            )
        flash(f'Verkaufsstelle "{name}" wurde angelegt und ausgewählt.', 'success')
        ersteller = query("SELECT name FROM mitarbeiter WHERE id=?", (session['user_id'],), one=True)
        _notify_neue_vs(name, strasse, plz_rep, ort, typ, ansprechpartner,
                        ersteller['name'] if ersteller else session.get('user_name', ''))
        return redirect(url_for('dashboard') if vom_dashboard else url_for('neue_aktivitaet', vs_id=new_id))
    flash('Name, Straße und Ort sind Pflichtfelder.', 'danger')
    return redirect(url_for('dashboard') if vom_dashboard else url_for('neue_aktivitaet'))


@app.route('/admin/biersorte/<int:b_id>/loeschen', methods=['POST'])
@admin_required
def admin_bier_loeschen(b_id):
    b = query("SELECT * FROM biersorte WHERE id=?", (b_id,), one=True)
    if not b:
        flash('Biersorte nicht gefunden.', 'danger')
        return redirect(url_for('admin'))
    count = query(
        "SELECT COUNT(*) AS c FROM bestellposition WHERE biersorte_id=?", (b_id,), one=True
    )['c']
    if count > 0:
        execute("UPDATE biersorte SET aktiv=0 WHERE id=?", (b_id,))
        flash(
            f'„{b["name"]}" hat {count} verknüpfte Bestellung(en) und wurde deaktiviert '
            f'(erscheint nicht mehr in neuen Aktivitäten, historische Daten bleiben erhalten).',
            'warning'
        )
    else:
        execute("DELETE FROM biersorte WHERE id=?", (b_id,))
        flash(f'Biersorte „{b["name"]}" wurde gelöscht.', 'success')
    return redirect(url_for('admin'))


@app.route('/admin/biersorte/<int:b_id>/reaktivieren', methods=['POST'])
@admin_required
def admin_bier_reaktivieren(b_id):
    b = query("SELECT * FROM biersorte WHERE id=?", (b_id,), one=True)
    if b:
        execute("UPDATE biersorte SET aktiv=1 WHERE id=?", (b_id,))
        flash(f'Biersorte „{b["name"]}" wurde reaktiviert.', 'success')
    return redirect(url_for('admin'))


@app.route('/admin/biersorte/neu', methods=['POST'])
@admin_required
def admin_bier_neu():
    name   = request.form.get('name', '').strip()
    einheit = request.form.get('einheit', 'Kiste (20x0.5L)').strip()
    if name:
        execute("INSERT OR IGNORE INTO biersorte (name, einheit) VALUES (?,?)", (name, einheit))
        flash(f'Biersorte "{name}" angelegt.', 'success')
    return redirect(url_for('admin'))


@app.route('/admin/biersorte/<int:b_id>/bearbeiten', methods=['POST'])
@admin_required
def admin_bier_bearbeiten(b_id):
    name    = request.form.get('name', '').strip()
    einheit = request.form.get('einheit', '').strip()
    if name:
        execute("UPDATE biersorte SET name=?, einheit=? WHERE id=?", (name, einheit, b_id))
        flash(f'Biersorte „{name}" aktualisiert.', 'success')
    return redirect(url_for('admin'))


@app.route('/admin/displaysorte/neu', methods=['POST'])
@admin_required
def admin_display_neu():
    name = request.form.get('name', '').strip()
    zaehlt = 1 if request.form.get('zaehlt_zur_zielerreichung') == '1' else 0
    if name:
        execute("INSERT OR IGNORE INTO displaysorte (name, zaehlt_zur_zielerreichung) VALUES (?,?)", (name, zaehlt))
        flash(f'Displaysorte "{name}" angelegt.', 'success')
    return redirect(url_for('admin'))


@app.route('/admin/displaysorte/<int:ds_id>/loeschen', methods=['POST'])
@admin_required
def admin_display_loeschen(ds_id):
    d = query("SELECT * FROM displaysorte WHERE id=?", (ds_id,), one=True)
    if d:
        execute("UPDATE displaysorte SET aktiv=0 WHERE id=?", (ds_id,))
        flash(f'Display-Typ „{d["name"]}" wurde deaktiviert.', 'warning')
    return redirect(url_for('admin'))


@app.route('/admin/displaysorte/<int:ds_id>/reaktivieren', methods=['POST'])
@admin_required
def admin_display_reaktivieren(ds_id):
    d = query("SELECT * FROM displaysorte WHERE id=?", (ds_id,), one=True)
    if d:
        execute("UPDATE displaysorte SET aktiv=1 WHERE id=?", (ds_id,))
        flash(f'Display-Typ „{d["name"]}" wurde reaktiviert.', 'success')
    return redirect(url_for('admin'))


@app.route('/admin/displaysorte/<int:ds_id>/bearbeiten', methods=['POST'])
@admin_required
def admin_display_bearbeiten(ds_id):
    name = request.form.get('name', '').strip()
    zaehlt = 1 if request.form.get('zaehlt_zur_zielerreichung') == '1' else 0
    if name:
        execute("UPDATE displaysorte SET name=?, zaehlt_zur_zielerreichung=? WHERE id=?", (name, zaehlt, ds_id))
        flash(f'Display-Typ „{name}" aktualisiert.', 'success')
    return redirect(url_for('admin'))


@app.route('/admin/displaysorte/<int:ds_id>/tier-umschalten', methods=['POST'])
@admin_required
def admin_display_tier_umschalten(ds_id):
    d = query("SELECT * FROM displaysorte WHERE id=?", (ds_id,), one=True)
    if d:
        neu = 0 if d['zaehlt_zur_zielerreichung'] else 1
        execute("UPDATE displaysorte SET zaehlt_zur_zielerreichung=? WHERE id=?", (neu, ds_id))
        label = 'zählt zur Zielerreichung' if neu else 'zählt nicht zur Zielerreichung'
        flash(f'Aufbautyp „{d["name"]}" {label}.', 'info')
    return redirect(url_for('admin'))


# ─── Routes: Excel-Import ────────────────────────────────────────────────────

@app.route('/admin/import-excel/vorlage')
@admin_required
def admin_import_vorlage():
    """Liefert die stammdaten_vorlage.xlsx zum Download."""
    vorlage = os.path.join(os.path.dirname(__file__), 'stammdaten_vorlage.xlsx')
    if not os.path.exists(vorlage):
        flash('Vorlagen-Datei nicht gefunden. Bitte erstelle_vorlage.py ausführen.', 'danger')
        return redirect(url_for('admin'))
    return send_file(vorlage, as_attachment=True,
                     download_name='stammdaten_vorlage.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/admin/import-excel', methods=['POST'])
@admin_required
def admin_import_excel():
    """Importiert Mitarbeiter und Verkaufsstellen aus einer hochgeladenen Excel-Datei."""
    if 'datei' not in request.files or request.files['datei'].filename == '':
        flash('Keine Datei ausgewählt.', 'danger')
        return redirect(url_for('admin'))

    datei = request.files['datei']
    if not datei.filename.lower().endswith(('.xlsx', '.xls')):
        flash('Nur Excel-Dateien (.xlsx) werden unterstützt.', 'danger')
        return redirect(url_for('admin'))

    try:
        import openpyxl
        wb = openpyxl.load_workbook(datei, data_only=True)
    except Exception as e:
        flash(f'Fehler beim Lesen der Excel-Datei: {e}', 'danger')
        return redirect(url_for('admin'))

    def _col(row, hmap, *namen):
        for name in namen:
            idx = hmap.get(name.lower())
            if idx is not None and idx < len(row):
                val = row[idx].value
                return str(val).strip() if val is not None else ''
        return ''

    def _hmap(sheet):
        headers = {}
        first_row = next(sheet.iter_rows(min_row=1, max_row=1))
        for i, cell in enumerate(first_row):
            if cell.value:
                headers[str(cell.value).strip().lower()] = i
        return headers

    stats = {'ma_neu': 0, 'ma_skip': 0, 'vs_neu': 0, 'vs_skip': 0, 'zuweis': 0, 'fehler': []}

    # ── Blatt Mitarbeiter ──────────────────────────────────────────────────────
    MA_NAMES = ['mitarbeiter', 'ma', 'reps']
    ma_sheet = next((wb[n] for n in wb.sheetnames if n.lower() in MA_NAMES), None)

    if ma_sheet:
        hmap = _hmap(ma_sheet)
        for row in ma_sheet.iter_rows(min_row=2):
            name     = _col(row, hmap, 'name', 'Name')
            kuerzel  = _col(row, hmap, 'kuerzel', 'Kürzel', 'kuerzel').upper()
            rolle    = _col(row, hmap, 'rolle', 'Rolle') or 'rep'
            passwort = _col(row, hmap, 'passwort', 'Passwort') or DEFAULT_PASSWORD
            email    = _col(row, hmap, 'email', 'Email', 'E-Mail').lower() or None

            if not name or not kuerzel:
                continue

            exists = query("SELECT id FROM mitarbeiter WHERE UPPER(kuerzel)=?", (kuerzel,), one=True)
            if exists:
                stats['ma_skip'] += 1
            else:
                execute("INSERT INTO mitarbeiter (name, kuerzel, rolle, passwort, email) VALUES (?,?,?,?,?)",
                        (name, kuerzel, rolle, generate_password_hash(passwort), email))
                stats['ma_neu'] += 1

    # ── Blatt Verkaufsstellen ──────────────────────────────────────────────────
    VS_NAMES = ['verkaufsstellen', 'kunden', 'vs']
    vs_sheet = next((wb[n] for n in wb.sheetnames if n.lower() in VS_NAMES), None)

    if vs_sheet:
        hmap = _hmap(vs_sheet)
        for row in vs_sheet.iter_rows(min_row=2):
            vs_name  = _col(row, hmap, 'name', 'Name')
            ort      = _col(row, hmap, 'ort', 'Ort', 'Stadt') or None
            strasse  = _col(row, hmap, 'strasse', 'Straße', 'Strasse', 'Adresse') or None
            typ      = _col(row, hmap, 'typ', 'Typ', 'Kategorie') or None
            ansprech = _col(row, hmap, 'ansprechpartner', 'Ansprechpartner', 'Kontakt') or None
            ma_raw   = _col(row, hmap, 'mitarbeiter', 'Mitarbeiter', 'rep', 'Rep')

            if not vs_name:
                continue

            existing = query("SELECT id FROM verkaufsstelle WHERE LOWER(name)=LOWER(?)", (vs_name,), one=True)
            if existing:
                vs_id = existing['id']
                stats['vs_skip'] += 1
            else:
                vs_id = execute(
                    "INSERT INTO verkaufsstelle (name, ort, strasse, typ, ansprechpartner, aktiv) VALUES (?,?,?,?,?,1)",
                    (vs_name, ort, strasse, typ, ansprech))
                stats['vs_neu'] += 1

            if ma_raw:
                for kuerzel in [k.strip().upper() for k in ma_raw.split(',') if k.strip()]:
                    ma = query("SELECT id FROM mitarbeiter WHERE UPPER(kuerzel)=?", (kuerzel,), one=True)
                    if not ma:
                        stats['fehler'].append(f'Kürzel „{kuerzel}" nicht gefunden (Zuweisung für „{vs_name}" übersprungen)')
                        continue
                    already = query("SELECT 1 FROM mitarbeiter_verkaufsstelle WHERE mitarbeiter_id=? AND verkaufsstelle_id=?",
                                    (ma['id'], vs_id), one=True)
                    if not already:
                        execute("INSERT INTO mitarbeiter_verkaufsstelle (mitarbeiter_id, verkaufsstelle_id) VALUES (?,?)",
                                (ma['id'], vs_id))
                        stats['zuweis'] += 1

    # ── Ergebnismeldung ────────────────────────────────────────────────────────
    teile = []
    if stats['ma_neu']:    teile.append(f"{stats['ma_neu']} Mitarbeiter neu")
    if stats['ma_skip']:   teile.append(f"{stats['ma_skip']} Mitarbeiter bereits vorhanden")
    if stats['vs_neu']:    teile.append(f"{stats['vs_neu']} Verkaufsstellen neu")
    if stats['vs_skip']:   teile.append(f"{stats['vs_skip']} Verkaufsstellen bereits vorhanden")
    if stats['zuweis']:    teile.append(f"{stats['zuweis']} Zuweisungen neu")

    if not teile:
        flash('Keine neuen Daten gefunden – alle Einträge bereits vorhanden oder Datei leer.', 'warning')
    else:
        flash('Import erfolgreich: ' + ' · '.join(teile) + '.', 'success')

    for fehler in stats['fehler']:
        flash(fehler, 'warning')

    return redirect(url_for('admin'))


@app.route('/admin/demo-seed', methods=['POST'])
@admin_required
def admin_demo_seed():
    """Manueller Trigger: Demo-Aktivitäten für die Vorwoche nachfüllen."""
    _do_demo_woche_nachfuellen(force=True)
    from datetime import date, timedelta
    today      = date.today()
    letzter_mo = today - timedelta(days=today.weekday() + 7)
    letzter_fr = letzter_mo + timedelta(days=4)
    kw         = letzter_mo.isocalendar()[1]
    flash(f'Demo-Daten für KW {kw} ({letzter_mo.strftime("%d.%m.")}–{letzter_fr.strftime("%d.%m.")}) wurden eingefügt.', 'success')
    return redirect(url_for('admin'))


# ─── Routes: Vergleich & Zielzahlen ──────────────────────────────────────────

@app.route('/vergleich')
@login_required
def vergleich():
    is_admin   = session.get('rolle') == 'admin'
    is_manager = session.get('rolle') in ('admin', 'verkaufsleiter')
    jahr       = request.args.get('jahr', date.today().year, type=int)

    # Reps sehen nur eigene Ziele; VKL/Admin können per Tab filtern
    if not is_manager:
        ma_filter = str(session['user_id'])
    else:
        ma_filter = request.args.get('ma', '', type=str)

    alle_jahre = [r[0] for r in query(
        "SELECT DISTINCT CAST(strftime('%Y', datum) AS INTEGER) FROM aktivitaet ORDER BY 1 DESC"
    )]
    if not alle_jahre:
        alle_jahre = [date.today().year]
    if date.today().year not in alle_jahre:
        alle_jahre.insert(0, date.today().year)

    # IST-Werte aller Reps – Subquery verhindert Doppelung von anzahl_displays
    _BP = "(SELECT aktivitaet_id, SUM(kisten_anzahl) AS kisten_total FROM bestellposition GROUP BY aktivitaet_id)"
    _vm_sql, _vm_p = _team_m_clause('m')
    ist = query(f'''
        SELECT m.id, m.name, m.kuerzel,
               COALESCE(SUM(a.anzahl_displays), 0) AS displays_ist,
               COALESCE(SUM(b.kisten_total), 0)    AS kisten_ist,
               COUNT(a.id)                          AS besuche
        FROM mitarbeiter m
        LEFT JOIN aktivitaet a
               ON a.mitarbeiter_id = m.id
              AND strftime('%Y', a.datum) = ?
        LEFT JOIN {_BP} b ON b.aktivitaet_id = a.id
        WHERE m.rolle IN ('rep','verkaufsleiter'){_vm_sql}
        GROUP BY m.id
        ORDER BY m.name
    ''', (str(jahr),) + _vm_p)

    # Zielzahlen
    ziele_raw = query(
        "SELECT mitarbeiter_id, displays_ziel, kisten_ziel FROM zielzahlen WHERE jahr = ?",
        (str(jahr),)
    )
    ziele = {r['mitarbeiter_id']: dict(r) for r in ziele_raw}

    # Teamziel (mitarbeiter_id IS NULL)
    teamziel_row = query(
        "SELECT displays_ziel, kisten_ziel FROM zielzahlen WHERE mitarbeiter_id IS NULL AND jahr = ?",
        (str(jahr),), one=True
    )
    teamziel = dict(teamziel_row) if teamziel_row else None

    # Daten für Charts aufbereiten (ohne saison_soll – zielkurs noch nicht befüllt)
    reps_namen  = [r['name'] for r in ist]
    kisten_ist  = [r['kisten_ist'] for r in ist]
    kisten_soll = [ziele.get(r['id'], {}).get('kisten_ziel', 0) or 0 for r in ist]
    disp_ist    = [r['displays_ist'] for r in ist]
    disp_soll   = [ziele.get(r['id'], {}).get('displays_ziel', 0) or 0 for r in ist]

    # KW-Verlauf je Mitarbeiter (für Liniendiagramm)
    kw_verlauf = query(f'''
        SELECT m.name,
               CAST(strftime('%W', a.datum) AS INTEGER) AS kw,
               COALESCE(SUM(bp.kisten_anzahl), 0) AS kisten
        FROM mitarbeiter m
        JOIN aktivitaet a ON a.mitarbeiter_id = m.id
        LEFT JOIN bestellposition bp ON bp.aktivitaet_id = a.id
        WHERE strftime('%Y', a.datum) = ? AND m.rolle IN ('rep','verkaufsleiter'){_vm_sql}
        GROUP BY m.id, kw
        ORDER BY m.name, kw
    ''', (str(jahr),) + _vm_p)

    # KW-Verlauf als Dict {rep_name: [(kw, kisten), ...]}
    from collections import defaultdict
    kw_by_rep = defaultdict(list)
    for row in kw_verlauf:
        kw_by_rep[row['name']].append({'kw': row['kw'], 'kisten': row['kisten']})

    # ── Saisonaler Zielkurs (Sonnenschlüssel) ────────────────────────────────
    aktueller_monat = date.today().month  # 1–12

    # Monatliche IST-Werte je Rep (mit korrekter BP-Subquery)
    _BPm = "(SELECT aktivitaet_id, SUM(kisten_anzahl) AS kisten_total FROM bestellposition GROUP BY aktivitaet_id)"
    monatlich_raw = query(f'''
        SELECT m.id AS rep_id,
               CAST(strftime('%m', a.datum) AS INTEGER) AS monat,
               COALESCE(SUM(a.anzahl_displays), 0) AS displays_ist,
               COALESCE(SUM(b.kisten_total),   0) AS kisten_ist
        FROM mitarbeiter m
        LEFT JOIN aktivitaet a
               ON a.mitarbeiter_id = m.id
              AND strftime('%Y', a.datum) = ?
        LEFT JOIN {_BPm} b ON b.aktivitaet_id = a.id
        WHERE m.rolle IN ('rep','verkaufsleiter'){_vm_sql}
        GROUP BY m.id, monat
    ''', (str(jahr),) + _vm_p)

    # {rep_id: {monat(1–12): {displays, kisten}}}
    monatlich_dict = {}
    for row in monatlich_raw:
        rid = row['rep_id']
        if rid not in monatlich_dict:
            monatlich_dict[rid] = {}
        if row['monat']:
            monatlich_dict[rid][row['monat']] = {
                'displays': row['displays_ist'],
                'kisten':   row['kisten_ist'],
            }

    monatsziele   = {}   # {rep_id: list[12 dicts]}
    zielkurs      = {}   # {rep_id: {disp_kurs, kist_kurs, kum_*}}
    kum_ist_data  = {}   # {kuerzel: [kum_kisten pro vergangenen Monat]}
    kum_ziel_data = {}   # {kuerzel: [kum_kisten_ziel pro vergangenen Monat]}

    for rep_row in ist:
        rid     = rep_row['id']
        kuerzel = rep_row['kuerzel']
        ziel    = ziele.get(rid, {})
        d_jz    = ziel.get('displays_ziel', 0) or 0
        k_jz    = ziel.get('kisten_ziel',   0) or 0

        monate = []
        kum_d_ziel = kum_k_ziel = kum_d_ist = kum_k_ist = 0
        kum_k_series = []
        kum_kz_series = []
        lauf_k = lauf_kz = 0

        for idx in range(12):
            m_num      = idx + 1
            mon_d_ziel = round(d_jz * SONNENSCHLUESSEL[idx])
            mon_k_ziel = round(k_jz * SONNENSCHLUESSEL[idx])
            mon_ist    = monatlich_dict.get(rid, {}).get(m_num, {'displays': 0, 'kisten': 0})
            vergangen  = (m_num <= aktueller_monat)

            d_pct_m = round(mon_ist['displays'] / mon_d_ziel * 100) if (mon_d_ziel and vergangen) else None
            k_pct_m = round(mon_ist['kisten']   / mon_k_ziel * 100) if (mon_k_ziel and vergangen) else None

            if vergangen:
                kum_d_ziel += mon_d_ziel;  kum_k_ziel += mon_k_ziel
                kum_d_ist  += mon_ist['displays']
                kum_k_ist  += mon_ist['kisten']
                lauf_k  += mon_ist['kisten'];  lauf_kz += mon_k_ziel
                kum_k_series.append(lauf_k)
                kum_kz_series.append(lauf_kz)

            monate.append({
                'num':           m_num,
                'name':          M_NAMEN[idx],
                'displays_ziel': mon_d_ziel,
                'kisten_ziel':   mon_k_ziel,
                'displays_ist':  mon_ist['displays'],
                'kisten_ist':    mon_ist['kisten'],
                'displays_pct':  d_pct_m,
                'kisten_pct':    k_pct_m,
                'vergangen':     vergangen,
            })

        monatsziele[rid] = monate
        zielkurs[rid] = {
            'disp_kurs': round(kum_d_ist / kum_d_ziel * 100) if kum_d_ziel else None,
            'kist_kurs': round(kum_k_ist / kum_k_ziel * 100) if kum_k_ziel else None,
            'kum_d_ist':  kum_d_ist,  'kum_d_ziel': kum_d_ziel,
            'kum_k_ist':  kum_k_ist,  'kum_k_ziel': kum_k_ziel,
        }
        kum_ist_data[kuerzel]  = kum_k_series
        kum_ziel_data[kuerzel] = kum_kz_series

    # Saisonaler Anteilszielkurs bis aktueller Monat – jetzt nach zielkurs-Loop
    kisten_saison_soll = [zielkurs.get(r['id'], {}).get('kum_k_ziel', 0) or 0 for r in ist]
    disp_saison_soll   = [zielkurs.get(r['id'], {}).get('kum_d_ziel', 0) or 0 for r in ist]

    sk_pct  = [int(x * 100) for x in SONNENSCHLUESSEL]
    alle_ma = query(
        f"SELECT id, name, kuerzel FROM mitarbeiter m WHERE rolle IN ('rep','verkaufsleiter'){_vm_sql} ORDER BY name",
        _vm_p
    ) if is_manager else []

    # Einzelner Rep für Detailansicht (Tabs)
    selected_rep = None
    if ma_filter:
        for rep_row in ist:
            if str(rep_row['id']) == ma_filter:
                selected_rep = dict(rep_row)
                break

    return render_template('vergleich.html',
        jahr=jahr, alle_jahre=alle_jahre,
        ist=ist, ziele=ziele, teamziel=teamziel,
        reps_namen=json.dumps(reps_namen),
        kisten_ist=json.dumps(kisten_ist),
        kisten_soll=json.dumps(kisten_soll),
        disp_ist=json.dumps(disp_ist),
        disp_soll=json.dumps(disp_soll),
        kw_by_rep=json.dumps(kw_by_rep),
        monatsziele=monatsziele,
        zielkurs=zielkurs,
        kum_ist_data=json.dumps(kum_ist_data),
        kum_ziel_data=json.dumps(kum_ziel_data),
        kisten_saison_soll=json.dumps(kisten_saison_soll),
        disp_saison_soll=json.dumps(disp_saison_soll),
        aktueller_monat=aktueller_monat,
        m_namen=M_NAMEN,
        sk_pct=sk_pct,
        is_manager=is_manager,
        is_admin=is_admin,
        ma_filter=ma_filter,
        alle_ma=alle_ma,
        selected_rep=selected_rep,
    )


@app.route('/zielzahlen', methods=['GET', 'POST'])
@manager_required
def zielzahlen():
    jahr = request.args.get('jahr', date.today().year, type=int)

    if request.method == 'POST':
        jar = request.form.get('jahr', date.today().year, type=int)
        _zz_sql, _zz_p = _team_m_clause('m')
        reps = query(
            f"SELECT id FROM mitarbeiter m WHERE rolle IN ('rep','verkaufsleiter'){_zz_sql}",
            _zz_p
        )

        for rep in reps:
            d_ziel = request.form.get(f'disp_{rep["id"]}', 0) or 0
            k_ziel = request.form.get(f'kist_{rep["id"]}', 0) or 0
            execute('''
                INSERT INTO zielzahlen (mitarbeiter_id, jahr, displays_ziel, kisten_ziel)
                VALUES (?, ?, ?, ?)
                ON CONFLICT(mitarbeiter_id, jahr) DO UPDATE SET
                    displays_ziel = excluded.displays_ziel,
                    kisten_ziel   = excluded.kisten_ziel
            ''', (rep['id'], jar, int(d_ziel), int(k_ziel)))

        # Teamziel
        td = request.form.get('team_disp', 0) or 0
        tk = request.form.get('team_kist', 0) or 0
        execute('''
            INSERT INTO zielzahlen (mitarbeiter_id, jahr, displays_ziel, kisten_ziel)
            VALUES (NULL, ?, ?, ?)
            ON CONFLICT(mitarbeiter_id, jahr) DO UPDATE SET
                displays_ziel = excluded.displays_ziel,
                kisten_ziel   = excluded.kisten_ziel
        ''', (jar, int(td), int(tk)))

        flash(f'Zielzahlen für {jar} gespeichert.', 'success')
        if request.form.get('redirect_to') == 'team_verwaltung':
            return redirect(url_for('team_verwaltung', jahr=jar))
        return redirect(url_for('zielzahlen', jahr=jar))

    _zz_sql, _zz_p = _team_m_clause('m')
    reps = query(
        f"SELECT id, name, kuerzel FROM mitarbeiter m WHERE rolle IN ('rep','verkaufsleiter'){_zz_sql} ORDER BY name",
        _zz_p
    )
    ziele_raw = query(
        "SELECT mitarbeiter_id, displays_ziel, kisten_ziel FROM zielzahlen WHERE jahr = ?",
        (str(jahr),)
    )
    ziele = {r['mitarbeiter_id']: dict(r) for r in ziele_raw}
    teamziel = ziele.get(None)

    alle_jahre = list(range(date.today().year, date.today().year + 3))

    return render_template('zielzahlen.html',
        reps=reps, ziele=ziele, teamziel=teamziel,
        jahr=jahr, alle_jahre=alle_jahre)


# ─── Team-Verwaltung (VKL+) ──────────────────────────────────────────────────

@app.route('/team-verwaltung')
@manager_required
def team_verwaltung():
    is_admin = session.get('rolle') == 'admin'
    is_vkl   = session.get('rolle') == 'verkaufsleiter'
    jahr     = request.args.get('jahr', date.today().year, type=int)

    _t_sql, _t_p = _team_m_clause('m')
    reps = query(
        f"SELECT m.id, m.name, m.kuerzel, m.email, m.rolle FROM mitarbeiter m "
        f"WHERE m.rolle IN ('rep','verkaufsleiter'){_t_sql} ORDER BY m.name",
        _t_p
    )

    # Vertretungen des eigenen Teams
    if is_vkl and session.get('team_id'):
        vertretungen = query("""
            SELECT v.id, v.von, v.bis, v.status, v.typ, ab.name AS abwesender, vtr.name AS vertreter
            FROM vertretung v
            JOIN mitarbeiter ab  ON ab.id  = v.abwesender_id
            LEFT JOIN mitarbeiter vtr ON vtr.id = v.vertreter_id
            WHERE ab.team_id = ?
            ORDER BY v.bis DESC
        """, (session['team_id'],))
    else:
        vertretungen = query("""
            SELECT v.id, v.von, v.bis, v.status, v.typ, ab.name AS abwesender, vtr.name AS vertreter
            FROM vertretung v
            JOIN mitarbeiter ab  ON ab.id  = v.abwesender_id
            LEFT JOIN mitarbeiter vtr ON vtr.id = v.vertreter_id
            ORDER BY v.bis DESC
        """)

    urlaub_konten = {rep['id']: _urlaub_konto(rep['id'], jahr) for rep in reps}

    # Zielzahlen-Daten für eingebettetes Formular
    ziele_raw = query(
        "SELECT mitarbeiter_id, displays_ziel, kisten_ziel FROM zielzahlen WHERE jahr = ?",
        (str(jahr),)
    )
    ziele    = {r['mitarbeiter_id']: dict(r) for r in ziele_raw}
    teamziel = ziele.get(None)
    alle_jahre = list(range(date.today().year, date.today().year + 3))

    return render_template('team_verwaltung.html',
        reps=reps, vertretungen=vertretungen,
        is_admin=is_admin, is_vkl=is_vkl,
        ziele=ziele, teamziel=teamziel, jahr=jahr, alle_jahre=alle_jahre,
        urlaub_konten=urlaub_konten)


# ─── Team-Vergleich (Admin) ───────────────────────────────────────────────────

@app.route('/team-vergleich')
@admin_required
def team_vergleich():
    periode = request.args.get('periode', 'woche')
    heute   = date.today()

    if periode == 'monat':
        start = heute.replace(day=1)
        end   = heute
        label = heute.strftime('%B %Y')
    elif periode == 'jahr':
        start = heute.replace(month=1, day=1)
        end   = heute
        label = str(heute.year)
    else:  # woche (default)
        start  = heute - timedelta(days=heute.weekday())
        end    = heute
        label  = f'KW {heute.isocalendar()[1]:02d} · {heute.year}'
        periode = 'woche'

    BP       = "(SELECT aktivitaet_id, SUM(kisten_anzahl) AS kisten_total FROM bestellposition GROUP BY aktivitaet_id)"
    DISP_IST = "SUM(CASE WHEN COALESCE(a.aktionstyp,'Aufbau')='Aufbau' THEN a.anzahl_displays ELSE 0 END)"
    KIST_IST = "COALESCE(SUM(CASE WHEN a.aktionstyp='Bestellung' THEN b.kisten_total ELSE 0 END), 0)"

    teams = query("SELECT id, name FROM team ORDER BY name")

    team_kpis    = {}
    prev_kpis    = {}
    rep_ranking  = {}
    jahres_ziele = {}
    jahres_ist   = {}
    inaktiv_per  = {}
    vkl_per_team = {}
    rep_count    = {}

    mo_kw = heute - timedelta(days=heute.weekday())

    for team in teams:
        tid = team['id']

        kpi = query(f"""
            SELECT {DISP_IST} AS displays, {KIST_IST} AS kisten, COUNT(a.id) AS besuche
            FROM aktivitaet a
            LEFT JOIN {BP} b ON b.aktivitaet_id = a.id
            WHERE a.datum BETWEEN ? AND ?
              AND a.mitarbeiter_id IN (SELECT id FROM mitarbeiter WHERE team_id=? AND rolle='rep')
        """, (start.isoformat(), end.isoformat(), tid), one=True)

        bestell = query("""
            SELECT COUNT(*) AS n FROM aktivitaet
            WHERE aktionstyp='Bestellung' AND datum BETWEEN ? AND ?
              AND mitarbeiter_id IN (SELECT id FROM mitarbeiter WHERE team_id=? AND rolle='rep')
        """, (start.isoformat(), end.isoformat(), tid), one=True)

        team_kpis[tid] = {
            'displays':    (kpi['displays']  or 0) if kpi else 0,
            'kisten':      (kpi['kisten']    or 0) if kpi else 0,
            'besuche':     (kpi['besuche']   or 0) if kpi else 0,
            'bestellungen': bestell['n'] if bestell else 0,
        }

        if periode == 'woche':
            p_start = start - timedelta(days=7)
            p_end   = end   - timedelta(days=7)
            prev = query(f"""
                SELECT {DISP_IST} AS displays, {KIST_IST} AS kisten, COUNT(a.id) AS besuche
                FROM aktivitaet a LEFT JOIN {BP} b ON b.aktivitaet_id = a.id
                WHERE a.datum BETWEEN ? AND ?
                  AND a.mitarbeiter_id IN (SELECT id FROM mitarbeiter WHERE team_id=? AND rolle='rep')
            """, (p_start.isoformat(), p_end.isoformat(), tid), one=True)
            prev_kpis[tid] = {'displays': (prev['displays'] or 0) if prev else 0,
                              'kisten':   (prev['kisten']   or 0) if prev else 0,
                              'besuche':  (prev['besuche']  or 0) if prev else 0}

        ranking = query(f"""
            SELECT m.id, m.name, m.kuerzel,
                   {DISP_IST} AS displays, {KIST_IST} AS kisten, COUNT(a.id) AS besuche
            FROM mitarbeiter m
            LEFT JOIN aktivitaet a ON a.mitarbeiter_id = m.id AND a.datum BETWEEN ? AND ?
            LEFT JOIN {BP} b ON b.aktivitaet_id = a.id
            WHERE m.team_id=? AND m.rolle='rep'
            GROUP BY m.id ORDER BY besuche DESC, kisten DESC
        """, (start.isoformat(), end.isoformat(), tid))
        rep_ranking[tid] = ranking
        rep_count[tid]   = len(ranking)

        vkl = query("SELECT name FROM mitarbeiter WHERE team_id=? AND rolle='verkaufsleiter' LIMIT 1", (tid,), one=True)
        vkl_per_team[tid] = vkl['name'] if vkl else '—'

        ziel = query("""
            SELECT SUM(z.displays_ziel) AS d, SUM(z.kisten_ziel) AS k
            FROM zielzahlen z JOIN mitarbeiter m ON m.id=z.mitarbeiter_id
            WHERE m.team_id=? AND z.jahr=?
        """, (tid, str(heute.year)), one=True)
        jahres_ziele[tid] = {'displays_ziel': (ziel['d'] or 0) if ziel else 0,
                             'kisten_ziel':   (ziel['k'] or 0) if ziel else 0}

        ist = query(f"""
            SELECT {DISP_IST} AS displays, {KIST_IST} AS kisten
            FROM aktivitaet a LEFT JOIN {BP} b ON b.aktivitaet_id = a.id
            WHERE strftime('%Y', a.datum)=?
              AND a.mitarbeiter_id IN (SELECT id FROM mitarbeiter WHERE team_id=? AND rolle='rep')
        """, (str(heute.year), tid), one=True)
        jahres_ist[tid] = {'displays': (ist['displays'] or 0) if ist else 0,
                           'kisten':   (ist['kisten']   or 0) if ist else 0}

        inaktiv = query("""
            SELECT m.id, m.name, m.kuerzel FROM mitarbeiter m
            WHERE m.team_id=? AND m.rolle='rep'
              AND m.id NOT IN (SELECT DISTINCT mitarbeiter_id FROM aktivitaet WHERE datum>=?)
              AND m.id NOT IN (SELECT abwesender_id FROM vertretung WHERE von<=? AND bis>=? AND status='bestätigt')
            ORDER BY m.name
        """, (tid, mo_kw.isoformat(), heute.isoformat(), heute.isoformat()))
        inaktiv_per[tid] = inaktiv

    return render_template('team_vergleich.html',
        teams=teams, team_kpis=team_kpis, prev_kpis=prev_kpis,
        rep_ranking=rep_ranking, rep_count=rep_count,
        vkl_per_team=vkl_per_team,
        jahres_ziele=jahres_ziele, jahres_ist=jahres_ist,
        inaktiv_per=inaktiv_per,
        periode=periode, label=label, heute=heute)


# ─── PWA ─────────────────────────────────────────────────────────────────────

@app.route('/sw.js')
def service_worker():
    from flask import send_from_directory
    return send_from_directory('static', 'sw.js', mimetype='application/javascript')


@app.route('/uploads/<path:filename>')
@login_required
def serve_upload(filename):
    """Fotos sind ausschließlich an aktivitaet.foto_pfad(_2/_3) verknüpft – Zugriff daher
    auf die zugehörige Aktivität geprüft: Rep nur eigene, VKL nur Team, Admin alles. Vorher
    ohne jede Prüfung abrufbar, sobald der (vorhersehbare) Dateiname bekannt war
    (Bugreport 2026-07-20, von Arco portiert)."""
    from flask import send_from_directory, abort
    akt = query(
        "SELECT mitarbeiter_id FROM aktivitaet WHERE foto_pfad=? OR foto_pfad_2=? OR foto_pfad_3=?",
        (filename, filename, filename), one=True
    )
    if not akt:
        abort(404)
    rolle = session.get('rolle')
    if rolle != 'admin':
        if rolle == 'verkaufsleiter' and session.get('team_id'):
            erlaubt = query(
                "SELECT 1 FROM mitarbeiter WHERE id=? AND team_id=?",
                (akt['mitarbeiter_id'], session['team_id']), one=True
            )
            if not erlaubt:
                abort(403)
        elif rolle != 'verkaufsleiter':
            if akt['mitarbeiter_id'] != session.get('user_id'):
                abort(403)
    return send_from_directory(UPLOAD_FOLDER, filename)


# ─── API: Autocomplete ────────────────────────────────────────────────────────

@app.route('/api/verkaufsstellen')
@login_required
def api_verkaufsstellen():
    """Freitext-Suche für die Verkaufsstellen-Auswahl in Neue Aktivität. Nutzt denselben
    Gebiets-Scope wie die eigene Verkaufsstellen-Liste (_verkaufsstellen_liste_sql) – Rep:
    eigene Zuordnung, VKL: Team bzw. alle ohne eigenes Team, Admin: alle. Vorher ohne jede
    Einschränkung für Reps (Bugreport 2026-07-20, von Arco portiert: lieferte den
    kompletten firmenweiten Kundenstamm statt nur das eigene Gebiet)."""
    q = request.args.get('q', '')
    from_sql, where_sql, params, distinct = _verkaufsstellen_liste_sql(suche=q)
    sel = "SELECT DISTINCT" if distinct else "SELECT"
    rows = query(
        f"{sel} v.id, v.name, v.ort, v.typ {from_sql} {where_sql} ORDER BY v.name LIMIT 20",
        params
    )
    return jsonify([dict(r) for r in rows])


# ─── Auto-Export ──────────────────────────────────────────────────────────────

def erstelle_fotos_zip_bytes(wochen: int = None, von: str = None, bis: str = None):
    """Erstellt ZIP-Archiv aller Fotos in einem Zeitraum.
    Entweder von/bis (YYYY-MM-DD) oder wochen (Anzahl Wochen zurück). Gibt (zip_bytes, anzahl) zurück."""
    _sql_basis = (
        "SELECT a.datum, m.kuerzel, a.foto_pfad, a.foto_pfad_2, a.foto_pfad_3, "
        "       COALESCE(a.aktionstyp, 'Aufbau') AS aktionstyp, "
        "       COALESCE(v.name, '') AS vs_name "
        "FROM aktivitaet a "
        "JOIN mitarbeiter m ON m.id = a.mitarbeiter_id "
        "LEFT JOIN verkaufsstelle v ON v.id = a.verkaufsstelle_id "
    )
    _foto_bedingung = ("((a.foto_pfad IS NOT NULL AND a.foto_pfad != '') "
                        " OR (a.foto_pfad_2 IS NOT NULL AND a.foto_pfad_2 != '') "
                        " OR (a.foto_pfad_3 IS NOT NULL AND a.foto_pfad_3 != ''))")
    if von and bis:
        fotos = query(_sql_basis + f"WHERE {_foto_bedingung} AND a.datum BETWEEN ? AND ? ORDER BY a.datum", (von, bis))
    else:
        w = wochen or 4
        grenzwert = (date.today() - timedelta(weeks=w)).isoformat()
        fotos = query(_sql_basis + f"WHERE {_foto_bedingung} AND a.datum >= ? ORDER BY a.datum", (grenzwert,))

    def _safe(s):
        import re
        return re.sub(r'[^\w\-]', '-', s).strip('-')

    buf   = io.BytesIO()
    count = 0
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        for f in fotos:
            for foto_pfad in (f['foto_pfad'], f['foto_pfad_2'], f['foto_pfad_3']):
                if not foto_pfad:
                    continue
                pfad = os.path.join(UPLOAD_FOLDER, foto_pfad)
                if os.path.exists(pfad):
                    ext     = os.path.splitext(foto_pfad)[1]
                    arcname = f"{f['datum']}_{f['kuerzel']}_{_safe(f['vs_name'])}_{_safe(f['aktionstyp'])}_{count+1:03d}{ext}"
                    zf.write(pfad, arcname)
                    count += 1
    buf.seek(0)
    return buf.read(), count


def auto_export_job():
    """Monatlicher Export am 1. jeden Monats: Excel-Jahresauswertung + Foto-ZIP des Vormonats."""
    if not EXPORT_EMAIL:
        app.logger.info("AUTO_EXPORT: EXPORT_EMAIL nicht gesetzt – übersprungen.")
        return
    with app.app_context():
        try:
            heute             = date.today()
            # Von "heute" aus über den 1. des laufenden Monats rechnen statt einfach
            # "gestern" zu nehmen – sonst liefert ein manueller Trigger (Button
            # "Export jetzt senden") an einem beliebigen Tag nur den bisherigen,
            # unvollständigen laufenden Monat statt des kompletten Vormonats.
            letzter_vormonat  = heute.replace(day=1) - timedelta(days=1)
            erster_vormonat   = letzter_vormonat.replace(day=1)
            _monat_namen = ['Januar','Februar','März','April','Mai','Juni',
                            'Juli','August','September','Oktober','November','Dezember']
            monat_label  = f"{_monat_namen[erster_vormonat.month - 1]} {erster_vormonat.year}"
            von_str      = erster_vormonat.isoformat()
            bis_str      = letzter_vormonat.isoformat()

            jahr        = heute.year
            heute_str   = heute.strftime('%d.%m.%Y')

            excel_bytes           = _build_excel_bytes(jahr, is_admin=True)
            zip_bytes, foto_count = erstelle_fotos_zip_bytes(von=von_str, bis=bis_str)

            body = f"""
            <div style="font-family:sans-serif;max-width:600px;color:#222">
              <h2 style="color:#1a3a5c">Aktions Tracker – Monatlicher Export</h2>
              <p>Sehr geehrte Damen und Herren,</p>
              <p>anbei erhalten Sie den automatischen Datenexport für <strong>{monat_label}</strong>.</p>
              <ul>
                <li><strong>Excel-Auswertung:</strong> Jahresübersicht {jahr}
                    (KW-Übersicht, Mitarbeiter-Ranking, Aktivitäten-Detail, Produktübersicht)</li>
                <li><strong>Foto-Archiv:</strong> {foto_count} Foto(s) aus {monat_label} als ZIP</li>
              </ul>
              <p style="color:#666;font-size:.9em">
                Die Fotos werden im System nach {FOTO_AUFBEWAHRUNG_WOCHEN} Wochen automatisch gelöscht –
                dieses Archiv enthält alle Aufnahmen des Vormonats ({von_str} bis {bis_str}).<br>
                Die Excel-Datei enthält den vollständigen Jahresstand zum Exportzeitpunkt.
              </p>
              <hr style="border:none;border-top:1px solid #eee;margin:1.5rem 0">
              <p style="font-size:.85em;color:#888">
                Aktions Tracker · Jan Anschütz · info@aktionstracker.de<br>
                Automatisch generiert – bitte nicht auf diese E-Mail antworten.
              </p>
            </div>
            """

            attachments = [
                (f"Aktions_Tracker_{jahr}.xlsx", excel_bytes,
                 "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
            ]
            if foto_count > 0:
                zip_name = f"Fotos_{erster_vormonat.strftime('%Y-%m')}.zip"
                attachments.append((zip_name, zip_bytes, "application/zip"))

            ok = send_email_with_attachments(
                EXPORT_EMAIL,
                f"Aktions Tracker – Export {monat_label}",
                body,
                attachments
            )
            if ok:
                app.logger.info(f"AUTO_EXPORT: Gesendet an {EXPORT_EMAIL} ({foto_count} Foto(s), {monat_label})")
            else:
                app.logger.error("AUTO_EXPORT: E-Mail-Versand fehlgeschlagen")
        except Exception as e:
            app.logger.error(f"AUTO_EXPORT Fehler: {e}", exc_info=True)


# ─── Wochenbericht ───────────────────────────────────────────────────────────

FIRMA_NAME = os.environ.get('FIRMA_NAME', '')

def _do_send_wochenbericht(force=False):
    """Kern-Logik – muss innerhalb eines aktiven App-Contexts aufgerufen werden."""
    try:
            config = query("SELECT * FROM wochenbericht_config WHERE id=1", one=True)
            if not config:
                return False, "Keine Konfiguration gefunden."
            if not force and not config['aktiv']:
                return False, "Wochenbericht ist deaktiviert."

            # Nicht zwei Mal in derselben Woche senden (außer bei force)
            kw_key = date.today().strftime('%Y-W%V')
            if not force and config['zuletzt_gesendet'] == kw_key:
                app.logger.info("WOCHENBERICHT: Diese Woche bereits gesendet – übersprungen.")
                return False, "Diese Woche bereits gesendet."

            # Admin-Empfaenger (empfaenger_2/3 – immer kumuliert)
            empfaenger_admin = []
            if config['empfaenger_2']:
                empfaenger_admin.append(config['empfaenger_2'])
            if config['empfaenger_3']:
                empfaenger_admin.append(config['empfaenger_3'])

            # VKLs mit E-Mail
            vkls = query(
                "SELECT email, name, team_id FROM mitarbeiter "
                "WHERE rolle='verkaufsleiter' AND email IS NOT NULL AND email != ''",
            ) or []

            # Zeiträume — berichtete Woche ist immer die abgeschlossene Vorwoche
            heute          = date.today()
            montag_diese   = heute - timedelta(days=heute.weekday() + 7)
            sonntag_diese  = montag_diese + timedelta(days=6)
            montag_letzte  = montag_diese - timedelta(days=7)
            sonntag_letzte = montag_letzte + timedelta(days=6)
            kw_nr          = montag_diese.strftime('%V')
            datum_von      = montag_diese.strftime('%d.%m.')
            datum_bis      = sonntag_diese.strftime('%d.%m.%Y')

            # Teams-Map für Multi-Team-Modus
            teams    = query("SELECT id, name FROM team ORDER BY name") or []
            team_map = {t['id']: t['name'] for t in teams}

            def trend_str(neu, alt):
                diff = neu - alt
                if diff > 0: return f'+{diff}'
                if diff < 0: return str(diff)
                return '±0'

            def trend_col(neu, alt):
                if neu > alt: return '#2d8a4e'
                if neu < alt: return '#c0392b'
                return '#888'

            def _offen_col(n):
                if n > 0:
                    return f'<span style="color:#c8860a;font-weight:bold">{n}</span>'
                return f'<span style="color:#aaa">0</span>'

            def build_html(team_id=None, team_name=None):
                t_p = [team_id] if team_id else []
                tf  = ' AND m.team_id=?' if team_id else ''

                def stats(von, bis):
                    return query(f'''
                        SELECT COUNT(DISTINCT a.id) AS besuche,
                               COUNT(DISTINCT CASE WHEN COALESCE(a.aktionstyp,\'Aufbau\')=\'Aufbau\'
                                                   THEN a.id END) AS aufbauten,
                               COUNT(DISTINCT CASE WHEN a.aktionstyp=\'Bestellung\'
                                                   THEN a.id END) AS bestellungen,
                               COALESCE(SUM(CASE WHEN a.aktionstyp=\'Bestellung\'
                                                 THEN bp.kisten_anzahl END), 0) AS kisten,
                               COALESCE(SUM(CASE WHEN COALESCE(a.aktionstyp,\'Aufbau\')=\'Aufbau\'
                                                 THEN a.anzahl_displays END), 0) AS displays
                        FROM aktivitaet a
                        JOIN mitarbeiter m ON m.id=a.mitarbeiter_id
                        LEFT JOIN bestellposition bp ON bp.aktivitaet_id = a.id
                        WHERE a.datum BETWEEN ? AND ?{tf}
                    ''', [von.isoformat(), bis.isoformat()] + t_p, one=True)

                diese  = stats(montag_diese,  sonntag_diese)
                letzte = stats(montag_letzte, sonntag_letzte)

                rs = query(f'''
                    SELECT m.id AS mitarbeiter_id, m.name,
                           COUNT(DISTINCT a.id) AS besuche,
                           COUNT(DISTINCT CASE WHEN a.aktionstyp=\'Bestellung\'
                                               THEN a.id END) AS bestellungen,
                           COUNT(DISTINCT CASE WHEN COALESCE(a.aktionstyp,\'Aufbau\')=\'Aufbau\'
                                               THEN a.id END) AS aufbauten,
                           COALESCE(SUM(CASE WHEN a.aktionstyp=\'Bestellung\'
                                             THEN bp.kisten_anzahl END), 0) AS kisten
                    FROM mitarbeiter m
                    LEFT JOIN aktivitaet a ON a.mitarbeiter_id = m.id AND a.datum BETWEEN ? AND ?
                    LEFT JOIN bestellposition bp ON bp.aktivitaet_id = a.id
                    WHERE (m.rolle='rep' OR (m.rolle='verkaufsleiter' AND EXISTS(SELECT 1 FROM mitarbeiter_verkaufsstelle mv WHERE mv.mitarbeiter_id=m.id))){tf}
                    GROUP BY m.id, m.name ORDER BY kisten DESC, m.name
                ''', [montag_diese.isoformat(), sonntag_diese.isoformat()] + t_p)

                _rs_vw = query(f'''
                    SELECT m.id AS mitarbeiter_id,
                           COUNT(DISTINCT a.id) AS besuche,
                           COALESCE(SUM(CASE WHEN a.aktionstyp=\'Bestellung\'
                                             THEN bp.kisten_anzahl END), 0) AS kisten
                    FROM aktivitaet a
                    JOIN mitarbeiter m ON m.id=a.mitarbeiter_id
                    LEFT JOIN bestellposition bp ON bp.aktivitaet_id=a.id
                    WHERE a.datum BETWEEN ? AND ? AND (m.rolle='rep' OR (m.rolle='verkaufsleiter' AND EXISTS(SELECT 1 FROM mitarbeiter_verkaufsstelle mv WHERE mv.mitarbeiter_id=m.id))){tf}
                    GROUP BY m.id
                ''', [montag_letzte.isoformat(), sonntag_letzte.isoformat()] + t_p) or []
                letzte_map = {r['mitarbeiter_id']: dict(r) for r in _rs_vw}

                def _delta_w(val, mid, key):
                    prev = letzte_map.get(mid, {}).get(key, 0)
                    col = trend_col(val, prev)
                    ts  = trend_str(val, prev)
                    return f'<div style="font-size:9px;font-weight:bold;color:{col};margin-top:1px">{ts}</div>'

                offen_map = {r['mitarbeiter_id']: r['n'] for r in query(
                    "SELECT a.mitarbeiter_id, COUNT(*) AS n FROM aktivitaet a "
                    "JOIN mitarbeiter m ON m.id=a.mitarbeiter_id "
                    "WHERE a.aktionstyp='Bestellung' AND COALESCE(a.bestell_status,'offen')='offen' "
                    f"AND (m.rolle='rep' OR (m.rolle='verkaufsleiter' AND EXISTS(SELECT 1 FROM mitarbeiter_verkaufsstelle mv WHERE mv.mitarbeiter_id=m.id))){tf} GROUP BY a.mitarbeiter_id",
                    t_p
                )}

                if team_id:
                    pipeline = query(
                        "SELECT COALESCE(SUM(CASE WHEN COALESCE(a.bestell_status,'offen')='offen' THEN 1 END),0) AS offen,"
                        "       COALESCE(SUM(CASE WHEN a.bestell_status='aufgebaut' THEN 1 END),0) AS aufgebaut,"
                        "       COALESCE(SUM(CASE WHEN a.bestell_status='storniert' THEN 1 END),0) AS storniert "
                        "FROM aktivitaet a JOIN mitarbeiter m ON m.id=a.mitarbeiter_id "
                        "WHERE a.aktionstyp='Bestellung' AND m.team_id=?",
                        (team_id,), one=True)
                    ue_rows = query(
                        "SELECT v.name AS station, m.name AS rep, a.datum, "
                        "CAST(julianday('now') - julianday(a.datum) AS INTEGER) AS tage "
                        "FROM aktivitaet a "
                        "JOIN mitarbeiter m ON m.id=a.mitarbeiter_id "
                        "JOIN verkaufsstelle v ON v.id=a.verkaufsstelle_id "
                        "WHERE a.aktionstyp='Bestellung' AND COALESCE(a.bestell_status,'offen')='offen' "
                        "AND julianday('now') - julianday(a.datum) > 28 AND m.team_id=? "
                        "ORDER BY tage DESC LIMIT 10",
                        (team_id,))
                else:
                    pipeline = query(
                        "SELECT COALESCE(SUM(CASE WHEN COALESCE(bestell_status,'offen')='offen' THEN 1 END),0) AS offen,"
                        "       COALESCE(SUM(CASE WHEN bestell_status='aufgebaut' THEN 1 END),0) AS aufgebaut,"
                        "       COALESCE(SUM(CASE WHEN bestell_status='storniert' THEN 1 END),0) AS storniert "
                        "FROM aktivitaet WHERE aktionstyp='Bestellung'",
                        one=True
                    )
                    ue_rows = query(
                        "SELECT v.name AS station, m.name AS rep, a.datum, "
                        "CAST(julianday('now') - julianday(a.datum) AS INTEGER) AS tage "
                        "FROM aktivitaet a "
                        "JOIN mitarbeiter m ON m.id=a.mitarbeiter_id "
                        "JOIN verkaufsstelle v ON v.id=a.verkaufsstelle_id "
                        "WHERE a.aktionstyp='Bestellung' AND COALESCE(a.bestell_status,'offen')='offen' "
                        "AND julianday('now') - julianday(a.datum) > 28 "
                        "ORDER BY tage DESC LIMIT 10"
                    )

                if ue_rows:
                    ue_trs = ''.join(f'''
                  <tr>
                    <td style="padding:7px 16px;border-bottom:1px solid #f0e8d0;font-size:12px;font-weight:600">{u["station"]}</td>
                    <td style="padding:7px 8px;border-bottom:1px solid #f0e8d0;font-size:12px;color:#666">{u["rep"]}</td>
                    <td style="padding:7px 16px;border-bottom:1px solid #f0e8d0;font-size:12px;text-align:right">
                      <span style="background:#fdecc8;color:#8a5a00;padding:2px 8px;border-radius:4px">{u["tage"]} Tage</span>
                    </td>
                  </tr>''' for u in ue_rows)
                    ueberfaellig_html = f'''
  <div style="padding:0 32px 20px">
    <div style="background:#fff8f0;border:1px solid #f0c674;border-radius:8px;overflow:hidden">
      <div style="background:#fdecc8;padding:10px 16px;font-size:13px;font-weight:bold;color:#8a5a00">
        &#9888; &Uuml;berf&auml;llig &ndash; Bestellungen offen seit &uuml;ber 4 Wochen ({len(ue_rows)})
      </div>
      <table width="100%" cellpadding="0" cellspacing="0">{ue_trs}
      </table>
    </div>
  </div>'''
                else:
                    ueberfaellig_html = ''

                # Tagesplan-Erfüllung berichtete Woche (= montag_diese, gleiche Periode wie Haupt-KPIs)
                _tp_team_row = query(f'''
                    SELECT COUNT(*) AS geplant,
                           COALESCE(SUM(tp.erledigt), 0) AS erledigt
                    FROM tagesplan tp JOIN mitarbeiter m ON m.id=tp.mitarbeiter_id
                    WHERE tp.datum BETWEEN ? AND ?
                      AND COALESCE(tp.geloescht,0)=0 AND (m.rolle='rep' OR (m.rolle='verkaufsleiter' AND EXISTS(SELECT 1 FROM mitarbeiter_verkaufsstelle mv WHERE mv.mitarbeiter_id=m.id))){tf}
                ''', [montag_diese.isoformat(), sonntag_diese.isoformat()] + t_p, one=True)
                _tp_team = dict(_tp_team_row) if _tp_team_row else {}
                _tp_reps = query(f'''
                    SELECT tp.mitarbeiter_id, COUNT(*) AS geplant,
                           COALESCE(SUM(tp.erledigt),0) AS erledigt
                    FROM tagesplan tp JOIN mitarbeiter m ON m.id=tp.mitarbeiter_id
                    WHERE tp.datum BETWEEN ? AND ?
                      AND COALESCE(tp.geloescht,0)=0 AND (m.rolle='rep' OR (m.rolle='verkaufsleiter' AND EXISTS(SELECT 1 FROM mitarbeiter_verkaufsstelle mv WHERE mv.mitarbeiter_id=m.id))){tf}
                    GROUP BY tp.mitarbeiter_id
                ''', [montag_diese.isoformat(), sonntag_diese.isoformat()] + t_p) or []
                tp_map = {r['mitarbeiter_id']: dict(r) for r in _tp_reps}

                def _plan_badge(geplant, erledigt):
                    if not geplant:
                        return '<span style="color:#aaa;font-size:11px">–</span>'
                    pct = round(erledigt / geplant * 100)
                    col = '#2d8a4e' if pct >= 80 else '#c8860a' if pct >= 60 else '#c0392b'
                    return (f'<span style="font-size:12px">{erledigt} erl. / {geplant} ges.</span>'
                            f'<div style="font-size:9px;font-weight:bold;color:{col};margin-top:1px">{pct}%</div>')

                tp_g = _tp_team.get('geplant', 0)
                tp_e = _tp_team.get('erledigt', 0)
                tp_o = tp_g - tp_e
                tp_pct = round(tp_e / tp_g * 100) if tp_g else 0
                tp_col = '#2d8a4e' if tp_pct >= 80 else '#c8860a' if tp_pct >= 60 else ('#c0392b' if tp_g else '#aaa')

                rep_rows = ''.join(f'''
                <tr>
                  <td style="padding:7px 10px;border-bottom:1px solid #f0f0f0;font-size:13px">{r["name"]}</td>
                  <td style="padding:7px 10px;border-bottom:1px solid #f0f0f0;text-align:center;font-size:13px">{r["besuche"]}{_delta_w(r["besuche"], r["mitarbeiter_id"], "besuche")}</td>
                  <td style="padding:7px 10px;border-bottom:1px solid #f0f0f0;text-align:center;font-size:13px;color:#2e6da4">{r["bestellungen"]}</td>
                  <td style="padding:7px 10px;border-bottom:1px solid #f0f0f0;text-align:center;font-size:13px;color:#27ae60">{r["aufbauten"]}</td>
                  <td style="padding:7px 10px;border-bottom:1px solid #f0f0f0;text-align:center;font-size:13px;font-weight:600;color:#c8860a">{r["kisten"]}{_delta_w(r["kisten"], r["mitarbeiter_id"], "kisten")}</td>
                  <td style="padding:7px 10px;border-bottom:1px solid #f0f0f0;text-align:center;font-size:13px">{_offen_col(offen_map.get(r["mitarbeiter_id"], 0))}</td>
                  <td style="padding:7px 10px;border-bottom:1px solid #f0f0f0;text-align:center">{_plan_badge(tp_map.get(r["mitarbeiter_id"],{}).get("geplant",0), tp_map.get(r["mitarbeiter_id"],{}).get("erledigt",0))}</td>
                </tr>''' for r in rs) or \
                '<tr><td colspan="7" style="padding:12px 14px;color:#999;text-align:center">Keine Aktivitäten diese Woche</td></tr>'

                tl = f' &ndash; {team_name}' if team_name else ''
                dl = APP_BASE_URL or '#'
                return f'''<!DOCTYPE html>
<html><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1.0">
<style>
@media only screen and (max-width:480px){{
  .outer{{margin:0!important;border-radius:0!important}}
  .hd{{padding:18px 16px!important}}
  .pad{{padding:16px 12px!important}}
  .kpi-cell{{display:block!important;width:100%!important;box-sizing:border-box!important;margin-bottom:8px!important}}
  .kpi-spc{{display:none!important}}
  .rep-wrap{{overflow-x:auto;-webkit-overflow-scrolling:touch}}
  .pipeline{{padding:12px 16px!important;font-size:12px!important}}
  .plan-bar{{padding:10px 16px!important;font-size:12px!important}}
}}
</style>
</head>
<body style="margin:0;padding:0;background:#f0f4f8;font-family:Arial,Helvetica,sans-serif">
<div class="outer" style="max-width:600px;margin:32px auto;background:#fff;border-radius:10px;overflow:hidden;box-shadow:0 2px 12px rgba(0,0,0,0.10)">

  <div class="hd" style="background:#1a3a5c;padding:26px 32px">
    <div style="color:#fff;font-size:20px;font-weight:bold;letter-spacing:.3px">Aktions Tracker{tl}</div>
    <div style="color:#90b8d8;font-size:13px;margin-top:5px">Wochenbericht KW {kw_nr} &nbsp;&middot;&nbsp; {datum_von} – {datum_bis}</div>
  </div>

  <div class="pad" style="padding:28px 32px 8px">
    <div style="font-size:15px;font-weight:bold;color:#1a3a5c;margin-bottom:16px">Gesamtübersicht</div>
    <table width="100%" cellpadding="0" cellspacing="0">
      <tr>
        <td class="kpi-cell" style="text-align:center;padding:18px 10px;background:#f4f8fc;border-radius:8px">
          <div style="font-size:30px;font-weight:bold;color:#1a3a5c">{diese["besuche"]}</div>
          <div style="font-size:12px;color:#666;margin-top:3px">Besuche</div>
          <div style="font-size:11px;font-weight:bold;color:{trend_col(diese["besuche"],letzte["besuche"])};margin-top:5px">{trend_str(diese["besuche"],letzte["besuche"])} ggü. Vorwoche</div>
        </td>
        <td class="kpi-spc" width="12"></td>
        <td class="kpi-cell" style="text-align:center;padding:18px 10px;background:#f4f8fc;border-radius:8px">
          <div style="font-size:30px;font-weight:bold;color:#c8860a">{diese["kisten"]}</div>
          <div style="font-size:12px;color:#666;margin-top:3px">{UNIT_LABEL}</div>
          <div style="font-size:11px;font-weight:bold;color:{trend_col(diese["kisten"],letzte["kisten"])};margin-top:5px">{trend_str(diese["kisten"],letzte["kisten"])} ggü. Vorwoche</div>
        </td>
        <td class="kpi-spc" width="12"></td>
        <td class="kpi-cell" style="text-align:center;padding:18px 10px;background:#f4f8fc;border-radius:8px">
          <div style="font-size:30px;font-weight:bold;color:#2e6da4">{diese["displays"]}</div>
          <div style="font-size:12px;color:#666;margin-top:3px">Aufbauten</div>
          <div style="font-size:11px;font-weight:bold;color:{trend_col(diese["displays"],letzte["displays"])};margin-top:5px">{trend_str(diese["displays"],letzte["displays"])} ggü. Vorwoche</div>
        </td>
      </tr>
    </table>
  </div>

  <div class="pipeline" style="padding:16px 32px;background:#fffbf0;border-top:1px solid #f0c674">
    <span style="font-size:13px;font-weight:bold;color:#1a3a5c">Bestellungen Pipeline:</span>
    <span style="margin-left:14px;font-size:13px">
      <span style="color:#c8860a;font-weight:bold">{pipeline["offen"]}</span><span style="color:#777"> offen</span>
      &nbsp;&nbsp;&middot;&nbsp;&nbsp;
      <span style="color:#27ae60;font-weight:bold">{pipeline["aufgebaut"]}</span><span style="color:#777"> aufgebaut</span>
      &nbsp;&nbsp;&middot;&nbsp;&nbsp;
      <span style="color:#6c757d;font-weight:bold">{pipeline["storniert"]}</span><span style="color:#777"> storniert</span>
    </span>
  </div>

  {ueberfaellig_html}

  <div class="plan-bar" style="padding:14px 32px;background:#f0f4f8;border-top:1px solid #e4eaf0">
    <span style="font-size:13px;font-weight:bold;color:#1a3a5c">&#128203; Besuchsplanung diese Woche:</span>
    <span style="margin-left:10px;font-size:13px">
      <span style="color:#555">{tp_g} geplant</span>
      &nbsp;&middot;&nbsp;
      <span style="color:#2d8a4e;font-weight:bold">{tp_e} erledigt</span>
      &nbsp;&middot;&nbsp;
      <span style="color:#c8860a;font-weight:bold">{tp_o} nicht erledigt</span>
      &nbsp;&middot;&nbsp;
      <span style="font-weight:bold;color:{tp_col}">{tp_pct}%</span>
    </span>
  </div>

  <div class="pad" style="padding:24px 32px">
    <div style="font-size:15px;font-weight:bold;color:#1a3a5c;margin-bottom:12px">Mitarbeiter diese Woche</div>
    <div class="rep-wrap" style="overflow-x:auto;-webkit-overflow-scrolling:touch">
    <table width="100%" cellpadding="0" cellspacing="0" style="border:1px solid #e4eaf0;border-radius:8px;overflow:hidden;min-width:420px">
      <thead>
        <tr style="background:#edf2f7">
          <th style="padding:8px 10px;text-align:left;font-size:10px;color:#666;font-weight:600;letter-spacing:.5px">MITARBEITER</th>
          <th style="padding:8px 10px;text-align:center;font-size:10px;color:#666;font-weight:600;letter-spacing:.5px">BESUCHE</th>
          <th style="padding:8px 10px;text-align:center;font-size:10px;color:#2e6da4;font-weight:600;letter-spacing:.5px">BESTELL.</th>
          <th style="padding:8px 10px;text-align:center;font-size:10px;color:#27ae60;font-weight:600;letter-spacing:.5px">AUFBAUT.</th>
          <th style="padding:8px 10px;text-align:center;font-size:10px;color:#c8860a;font-weight:600;letter-spacing:.5px">{UNIT_LABEL.upper()[:7]}</th>
          <th style="padding:8px 10px;text-align:center;font-size:10px;color:#c8860a;font-weight:600;letter-spacing:.5px">OFFEN</th>
          <th style="padding:8px 10px;text-align:center;font-size:10px;color:#5a3e9e;font-weight:600;letter-spacing:.5px">BESUCHSPL.</th>
        </tr>
      </thead>
      <tbody>{rep_rows}</tbody>
    </table>
    </div>
  </div>

  <div style="padding:16px 32px 24px;text-align:center">
    <a href="{dl}" style="display:inline-block;background:#1a3a5c;color:#fff;text-decoration:none;padding:10px 24px;border-radius:6px;font-size:13px;font-weight:bold">→ Zum Dashboard</a>
  </div>

  <div style="padding:14px 32px;background:#f4f8fc;border-top:1px solid #e4eaf0;text-align:center">
    <div style="font-size:11px;color:#aaa">Aktions Tracker · Automatischer Wochenbericht jeden Montag<br>
    Einstellungen unter <em>Einstellungen → Wochenbericht</em></div>
  </div>

</div>
</body></html>'''

            firma_teil = f' – {FIRMA_NAME}' if FIRMA_NAME else ''
            ok_count = 0
            _safe_co   = re.sub(r'[^\w]', '_', COMPANY_SHORT).strip('_') or 'AktionsTracker'

            def _send_wb(to, betreff, html, tag=''):
                pdf      = _html_to_pdf(html)
                fname    = f'Wochenbericht_{_safe_co}_KW{kw_nr}.pdf'
                atts     = [(fname, pdf, 'application/pdf')] if pdf else []
                fn       = send_email_with_attachments if atts else send_email
                args     = (to, betreff, html, atts) if atts else (to, betreff, html)
                if fn(*args):
                    app.logger.info(f"WOCHENBERICHT KW {kw_nr}{tag}: Gesendet an {to}")
                    return True
                app.logger.error(f"WOCHENBERICHT KW {kw_nr}{tag}: Versand an {to} fehlgeschlagen")
                return False

            # Multi-Team: VKLs in 2+ verschiedenen Teams -> separate Berichte
            vkl_teams = list(dict.fromkeys(v['team_id'] for v in vkls if v['team_id']))
            if len(vkl_teams) >= 2:
                for v in vkls:
                    if not v['team_id']:
                        continue
                    tname   = team_map.get(v['team_id'], f'Team {v["team_id"]}')
                    html    = build_html(team_id=v['team_id'], team_name=tname)
                    betreff = f'Wochenbericht{firma_teil} – {tname} – KW {kw_nr}'
                    if _send_wb(v['email'], betreff, html, f' [{tname}]'):
                        ok_count += 1
                if empfaenger_admin:
                    html_g  = build_html(team_id=None, team_name='Alle Teams')
                    betreff = f'Wochenbericht{firma_teil} – Alle Teams – KW {kw_nr}'
                    for mail in empfaenger_admin:
                        if _send_wb(mail, betreff, html_g, ' [Gesamt]'):
                            ok_count += 1
            else:
                # Einzel-Modus (ein Team oder keine Teams)
                empfaenger = []
                if vkls:
                    empfaenger.append(vkls[0]['email'])
                empfaenger.extend(empfaenger_admin)
                if not empfaenger:
                    app.logger.warning("WOCHENBERICHT: Keine Empfänger konfiguriert – übersprungen.")
                    return False, "Keine Empfänger konfiguriert. Bitte E-Mail-Adresse des Verkaufsleiters im Admin-Panel hinterlegen oder einen zusätzlichen Empfänger eintragen."
                html    = build_html(team_id=None)
                betreff = f'Wochenbericht Aktionstracker{firma_teil} – KW {kw_nr}'
                for mail in empfaenger:
                    if _send_wb(mail, betreff, html):
                        ok_count += 1

            if ok_count > 0:
                execute("UPDATE wochenbericht_config SET zuletzt_gesendet=? WHERE id=1", (kw_key,))
                return True, f"Gesendet an {ok_count} Empfänger"
            else:
                detail = f': {_smtp_last_error}' if _smtp_last_error else ''
                return False, f"E-Mail-Versand fehlgeschlagen{detail}"

    except Exception as e:
        app.logger.error(f"WOCHENBERICHT Fehler: {e}", exc_info=True)
        return False, f"Fehler: {e}"


def send_wochenbericht(force=False):
    """Wrapper für APScheduler – erstellt eigenen App-Context."""
    with app.app_context():
        return _do_send_wochenbericht(force=force)


# ─── Monatsbericht ───────────────────────────────────────────────────────────

def _do_send_monatsbericht(force=False):
    """Automatischer Monatsbericht – immer am 1. eines Monats für den abgeschlossenen Vormonat."""
    try:
        config = query("SELECT * FROM wochenbericht_config WHERE id=1", one=True)
        if not config:
            return False, "Keine Konfiguration gefunden."
        if not force and not config['aktiv']:
            return False, "Berichte sind deaktiviert."

        heute            = date.today()
        letzter_vormonat = heute - timedelta(days=1)
        erster_vormonat  = letzter_vormonat.replace(day=1)
        letzter_vorvorm  = erster_vormonat - timedelta(days=1)
        erster_vorvorm   = letzter_vorvorm.replace(day=1)

        monat_key = erster_vormonat.strftime('%Y-%m')
        if not force and config['zuletzt_gesendet_monat'] == monat_key:
            app.logger.info("MONATSBERICHT: Dieser Monat bereits gesendet – übersprungen.")
            return False, "Dieser Monat bereits gesendet."

        _monat_namen = ['Januar','Februar','März','April','Mai','Juni',
                        'Juli','August','September','Oktober','November','Dezember']
        monat_name  = _monat_namen[erster_vormonat.month - 1]
        vmonat_name = _monat_namen[erster_vorvorm.month - 1]
        monat_label = f"{monat_name} {erster_vormonat.year}"

        empfaenger_admin = [e for e in [config['empfaenger_2'], config['empfaenger_3']] if e]
        vkls     = query("SELECT email, name, team_id FROM mitarbeiter "
                         "WHERE rolle='verkaufsleiter' AND email IS NOT NULL AND email != ''") or []
        teams    = query("SELECT id, name FROM team ORDER BY name") or []
        team_map = {t['id']: t['name'] for t in teams}

        def trend_str(neu, alt):
            d = neu - alt
            return f'+{d}' if d > 0 else str(d) if d < 0 else '±0'

        def trend_col(neu, alt):
            return '#2d8a4e' if neu > alt else '#c0392b' if neu < alt else '#888'

        def build_html(team_id=None, team_name=None):
            t_p = [team_id] if team_id else []
            tf  = ' AND m.team_id=?' if team_id else ''

            def stats(von, bis):
                return query(f'''
                    SELECT COUNT(DISTINCT a.id) AS besuche,
                           COUNT(DISTINCT CASE WHEN COALESCE(a.aktionstyp,'Aufbau')='Aufbau'
                                               THEN a.id END) AS aufbauten,
                           COUNT(DISTINCT CASE WHEN a.aktionstyp='Bestellung'
                                               THEN a.id END) AS bestellungen,
                           COALESCE(SUM(CASE WHEN a.aktionstyp='Bestellung'
                                             THEN bp.kisten_anzahl END), 0) AS kisten,
                           COALESCE(SUM(CASE WHEN COALESCE(a.aktionstyp,'Aufbau')='Aufbau'
                                             THEN a.anzahl_displays END), 0) AS displays
                    FROM aktivitaet a
                    JOIN mitarbeiter m ON m.id=a.mitarbeiter_id
                    LEFT JOIN bestellposition bp ON bp.aktivitaet_id=a.id
                    WHERE a.datum BETWEEN ? AND ?{tf}
                ''', [von.isoformat(), bis.isoformat()] + t_p, one=True)

            dieser = stats(erster_vormonat, letzter_vormonat)
            vorher = stats(erster_vorvorm,  letzter_vorvorm)

            rs = query(f'''
                SELECT m.id AS mitarbeiter_id, m.name,
                       COUNT(DISTINCT a.id) AS besuche,
                       COUNT(DISTINCT CASE WHEN a.aktionstyp='Bestellung' THEN a.id END) AS bestellungen,
                       COUNT(DISTINCT CASE WHEN COALESCE(a.aktionstyp,'Aufbau')='Aufbau' THEN a.id END) AS aufbauten,
                       COALESCE(SUM(CASE WHEN a.aktionstyp='Bestellung'
                                         THEN bp.kisten_anzahl END), 0) AS kisten,
                       COALESCE(SUM(CASE WHEN COALESCE(a.aktionstyp,'Aufbau')='Aufbau'
                                         THEN a.anzahl_displays END), 0) AS displays
                FROM mitarbeiter m
                LEFT JOIN aktivitaet a ON a.mitarbeiter_id = m.id AND a.datum BETWEEN ? AND ?
                LEFT JOIN bestellposition bp ON bp.aktivitaet_id = a.id
                WHERE (m.rolle='rep' OR (m.rolle='verkaufsleiter' AND EXISTS(SELECT 1 FROM mitarbeiter_verkaufsstelle mv WHERE mv.mitarbeiter_id=m.id))){tf}
                GROUP BY m.id, m.name ORDER BY kisten DESC, m.name
            ''', [erster_vormonat.isoformat(), letzter_vormonat.isoformat()] + t_p)

            _rs_vm = query(f'''
                SELECT m.id AS mitarbeiter_id,
                       COUNT(DISTINCT a.id) AS besuche,
                       COALESCE(SUM(CASE WHEN a.aktionstyp='Bestellung'
                                         THEN bp.kisten_anzahl END), 0) AS kisten,
                       COALESCE(SUM(CASE WHEN COALESCE(a.aktionstyp,'Aufbau')='Aufbau'
                                         THEN a.anzahl_displays END), 0) AS displays
                FROM aktivitaet a
                JOIN mitarbeiter m ON m.id=a.mitarbeiter_id
                LEFT JOIN bestellposition bp ON bp.aktivitaet_id=a.id
                WHERE a.datum BETWEEN ? AND ? AND (m.rolle='rep' OR (m.rolle='verkaufsleiter' AND EXISTS(SELECT 1 FROM mitarbeiter_verkaufsstelle mv WHERE mv.mitarbeiter_id=m.id))){tf}
                GROUP BY m.id
            ''', [erster_vorvorm.isoformat(), letzter_vorvorm.isoformat()] + t_p) or []
            vorvorm_map = {r['mitarbeiter_id']: dict(r) for r in _rs_vm}

            def _delta_m(val, mid, key):
                prev = vorvorm_map.get(mid, {}).get(key, 0)
                col = trend_col(val, prev)
                ts  = trend_str(val, prev)
                return f'<div style="font-size:9px;font-weight:bold;color:{col};margin-top:1px">{ts}</div>'

            # Tagesplan-Erfüllung Vormonat (Team + pro Rep)
            _mtp_team_row = query(f'''
                SELECT COUNT(*) AS geplant,
                       COALESCE(SUM(tp.erledigt), 0) AS erledigt
                FROM tagesplan tp JOIN mitarbeiter m ON m.id=tp.mitarbeiter_id
                WHERE tp.datum BETWEEN ? AND ?
                  AND COALESCE(tp.geloescht,0)=0 AND (m.rolle='rep' OR (m.rolle='verkaufsleiter' AND EXISTS(SELECT 1 FROM mitarbeiter_verkaufsstelle mv WHERE mv.mitarbeiter_id=m.id))){tf}
            ''', [erster_vormonat.isoformat(), letzter_vormonat.isoformat()] + t_p, one=True)
            _mtp_team = dict(_mtp_team_row) if _mtp_team_row else {}
            _mtp_reps = query(f'''
                SELECT tp.mitarbeiter_id, COUNT(*) AS geplant,
                       COALESCE(SUM(tp.erledigt),0) AS erledigt
                FROM tagesplan tp JOIN mitarbeiter m ON m.id=tp.mitarbeiter_id
                WHERE tp.datum BETWEEN ? AND ?
                  AND COALESCE(tp.geloescht,0)=0 AND (m.rolle='rep' OR (m.rolle='verkaufsleiter' AND EXISTS(SELECT 1 FROM mitarbeiter_verkaufsstelle mv WHERE mv.mitarbeiter_id=m.id))){tf}
                GROUP BY tp.mitarbeiter_id
            ''', [erster_vormonat.isoformat(), letzter_vormonat.isoformat()] + t_p) or []
            mtp_map = {r['mitarbeiter_id']: dict(r) for r in _mtp_reps}

            def _mplan_badge(geplant, erledigt):
                if not geplant:
                    return '<span style="color:#aaa;font-size:11px">–</span>'
                pct = round(erledigt / geplant * 100)
                col = '#2d8a4e' if pct >= 80 else '#c8860a' if pct >= 60 else '#c0392b'
                return (f'<span style="font-size:12px">{erledigt} erl. / {geplant} ges.</span>'
                        f'<div style="font-size:9px;font-weight:bold;color:{col};margin-top:1px">{pct}%</div>')

            if team_id:
                pipeline = query(
                    "SELECT COALESCE(SUM(CASE WHEN COALESCE(a.bestell_status,'offen')='offen' THEN 1 END),0) AS offen,"
                    "       COALESCE(SUM(CASE WHEN a.bestell_status='aufgebaut' THEN 1 END),0) AS aufgebaut,"
                    "       COALESCE(SUM(CASE WHEN a.bestell_status='storniert' THEN 1 END),0) AS storniert "
                    "FROM aktivitaet a JOIN mitarbeiter m ON m.id=a.mitarbeiter_id "
                    "WHERE a.aktionstyp='Bestellung' AND m.team_id=?", (team_id,), one=True)
            else:
                pipeline = query(
                    "SELECT COALESCE(SUM(CASE WHEN COALESCE(bestell_status,'offen')='offen' THEN 1 END),0) AS offen,"
                    "       COALESCE(SUM(CASE WHEN bestell_status='aufgebaut' THEN 1 END),0) AS aufgebaut,"
                    "       COALESCE(SUM(CASE WHEN bestell_status='storniert' THEN 1 END),0) AS storniert "
                    "FROM aktivitaet WHERE aktionstyp='Bestellung'", one=True)

            rep_rows = ''.join(f'''
              <tr>
                <td style="padding:7px 10px;border-bottom:1px solid #f0f0f0;font-size:13px">{r["name"]}</td>
                <td style="padding:7px 10px;border-bottom:1px solid #f0f0f0;text-align:center;font-size:13px">{r["besuche"]}{_delta_m(r["besuche"], r["mitarbeiter_id"], "besuche")}</td>
                <td style="padding:7px 10px;border-bottom:1px solid #f0f0f0;text-align:center;font-size:13px;color:#2e6da4">{r["bestellungen"]}</td>
                <td style="padding:7px 10px;border-bottom:1px solid #f0f0f0;text-align:center;font-size:13px;color:#27ae60">{r["aufbauten"]}</td>
                <td style="padding:7px 10px;border-bottom:1px solid #f0f0f0;text-align:center;font-size:13px;font-weight:600;color:#c8860a">{r["kisten"]}{_delta_m(r["kisten"], r["mitarbeiter_id"], "kisten")}</td>
                <td style="padding:7px 10px;border-bottom:1px solid #f0f0f0;text-align:center;font-size:13px;color:#2e6da4">{r["displays"]}{_delta_m(r["displays"], r["mitarbeiter_id"], "displays")}</td>
                <td style="padding:7px 10px;border-bottom:1px solid #f0f0f0;text-align:center">{_mplan_badge(mtp_map.get(r["mitarbeiter_id"],{}).get("geplant",0), mtp_map.get(r["mitarbeiter_id"],{}).get("erledigt",0))}</td>
              </tr>''' for r in rs) or \
            '<tr><td colspan="7" style="padding:12px 14px;color:#999;text-align:center">Keine Aktivitäten im Vormonat</td></tr>'

            mtp_g = _mtp_team.get('geplant', 0)
            mtp_e = _mtp_team.get('erledigt', 0)
            mtp_o = mtp_g - mtp_e
            mtp_pct = round(mtp_e / mtp_g * 100) if mtp_g else 0
            mtp_col = '#2d8a4e' if mtp_pct >= 80 else '#c8860a' if mtp_pct >= 60 else ('#c0392b' if mtp_g else '#aaa')

            tl = f' &ndash; {team_name}' if team_name else ''
            dl = APP_BASE_URL or '#'
            return f'''<!DOCTYPE html>
<html><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1.0">
<style>
@media only screen and (max-width:480px){{
  .outer{{margin:0!important;border-radius:0!important}}
  .hd{{padding:18px 16px!important}}
  .pad{{padding:16px 12px!important}}
  .kpi-cell{{display:block!important;width:100%!important;box-sizing:border-box!important;margin-bottom:8px!important}}
  .kpi-spc{{display:none!important}}
  .rep-wrap{{overflow-x:auto;-webkit-overflow-scrolling:touch}}
  .pipeline{{padding:12px 16px!important;font-size:12px!important}}
  .plan-bar{{padding:10px 16px!important;font-size:12px!important}}
}}
</style>
</head>
<body style="margin:0;padding:0;background:#f0f4f8;font-family:Arial,Helvetica,sans-serif">
<div class="outer" style="max-width:600px;margin:32px auto;background:#fff;border-radius:10px;overflow:hidden;box-shadow:0 2px 12px rgba(0,0,0,0.10)">

  <div class="hd" style="background:#1a3a5c;padding:26px 32px">
    <div style="color:#fff;font-size:20px;font-weight:bold;letter-spacing:.3px">Aktions Tracker{tl}</div>
    <div style="color:#90b8d8;font-size:13px;margin-top:5px">Monatsbericht {monat_label} &nbsp;&middot;&nbsp; {erster_vormonat.strftime('%d.%m.')} &ndash; {letzter_vormonat.strftime('%d.%m.%Y')}</div>
  </div>

  <div class="pad" style="padding:28px 32px 8px">
    <div style="font-size:15px;font-weight:bold;color:#1a3a5c;margin-bottom:16px">Gesamtübersicht {monat_name}</div>
    <table width="100%" cellpadding="0" cellspacing="0">
      <tr>
        <td class="kpi-cell" style="text-align:center;padding:18px 10px;background:#f4f8fc;border-radius:8px">
          <div style="font-size:30px;font-weight:bold;color:#1a3a5c">{dieser["besuche"]}</div>
          <div style="font-size:12px;color:#666;margin-top:3px">Besuche</div>
          <div style="font-size:11px;font-weight:bold;color:{trend_col(dieser["besuche"],vorher["besuche"])};margin-top:5px">{trend_str(dieser["besuche"],vorher["besuche"])} ggü. {vmonat_name}</div>
        </td>
        <td class="kpi-spc" width="12"></td>
        <td class="kpi-cell" style="text-align:center;padding:18px 10px;background:#f4f8fc;border-radius:8px">
          <div style="font-size:30px;font-weight:bold;color:#c8860a">{dieser["kisten"]}</div>
          <div style="font-size:12px;color:#666;margin-top:3px">{UNIT_LABEL}</div>
          <div style="font-size:11px;font-weight:bold;color:{trend_col(dieser["kisten"],vorher["kisten"])};margin-top:5px">{trend_str(dieser["kisten"],vorher["kisten"])} ggü. {vmonat_name}</div>
        </td>
        <td class="kpi-spc" width="12"></td>
        <td class="kpi-cell" style="text-align:center;padding:18px 10px;background:#f4f8fc;border-radius:8px">
          <div style="font-size:30px;font-weight:bold;color:#2e6da4">{dieser["displays"]}</div>
          <div style="font-size:12px;color:#666;margin-top:3px">Aufbauten</div>
          <div style="font-size:11px;font-weight:bold;color:{trend_col(dieser["displays"],vorher["displays"])};margin-top:5px">{trend_str(dieser["displays"],vorher["displays"])} ggü. {vmonat_name}</div>
        </td>
      </tr>
    </table>
  </div>

  <div class="pipeline" style="padding:16px 32px;background:#fffbf0;border-top:1px solid #f0c674">
    <span style="font-size:13px;font-weight:bold;color:#1a3a5c">Bestellungen Pipeline:</span>
    <span style="margin-left:14px;font-size:13px">
      <span style="color:#c8860a;font-weight:bold">{pipeline["offen"]}</span><span style="color:#777"> offen</span>
      &nbsp;&nbsp;&middot;&nbsp;&nbsp;
      <span style="color:#27ae60;font-weight:bold">{pipeline["aufgebaut"]}</span><span style="color:#777"> aufgebaut</span>
      &nbsp;&nbsp;&middot;&nbsp;&nbsp;
      <span style="color:#6c757d;font-weight:bold">{pipeline["storniert"]}</span><span style="color:#777"> storniert</span>
    </span>
  </div>

  <div class="plan-bar" style="padding:14px 32px;background:#f0f4f8;border-top:1px solid #e4eaf0">
    <span style="font-size:13px;font-weight:bold;color:#1a3a5c">&#128203; Besuchsplanung {monat_name}:</span>
    <span style="margin-left:10px;font-size:13px">
      <span style="color:#555">{mtp_g} geplant</span>
      &nbsp;&middot;&nbsp;
      <span style="color:#2d8a4e;font-weight:bold">{mtp_e} erledigt</span>
      &nbsp;&middot;&nbsp;
      <span style="color:#c8860a;font-weight:bold">{mtp_o} nicht erledigt</span>
      &nbsp;&middot;&nbsp;
      <span style="font-weight:bold;color:{mtp_col}">{mtp_pct}%</span>
    </span>
  </div>

  <div class="pad" style="padding:24px 32px">
    <div style="font-size:15px;font-weight:bold;color:#1a3a5c;margin-bottom:12px">Mitarbeiter &ndash; {monat_name}</div>
    <div class="rep-wrap" style="overflow-x:auto;-webkit-overflow-scrolling:touch">
    <table width="100%" cellpadding="0" cellspacing="0" style="border:1px solid #e4eaf0;border-radius:8px;overflow:hidden;min-width:420px">
      <thead>
        <tr style="background:#edf2f7">
          <th style="padding:8px 10px;text-align:left;font-size:10px;color:#666;font-weight:600;letter-spacing:.5px">MITARBEITER</th>
          <th style="padding:8px 10px;text-align:center;font-size:10px;color:#666;font-weight:600;letter-spacing:.5px">BESUCHE</th>
          <th style="padding:8px 10px;text-align:center;font-size:10px;color:#2e6da4;font-weight:600;letter-spacing:.5px">BESTELL.</th>
          <th style="padding:8px 10px;text-align:center;font-size:10px;color:#27ae60;font-weight:600;letter-spacing:.5px">AUFBAUT.</th>
          <th style="padding:8px 10px;text-align:center;font-size:10px;color:#c8860a;font-weight:600;letter-spacing:.5px">{UNIT_LABEL.upper()[:7]}</th>
          <th style="padding:8px 10px;text-align:center;font-size:10px;color:#2e6da4;font-weight:600;letter-spacing:.5px">AUFBAUT.GES.</th>
          <th style="padding:8px 10px;text-align:center;font-size:10px;color:#5a3e9e;font-weight:600;letter-spacing:.5px">BESUCHSPL.</th>
        </tr>
      </thead>
      <tbody>{rep_rows}</tbody>
    </table>
    </div>
  </div>

  <div style="padding:16px 32px 24px;text-align:center">
    <a href="{dl}" style="display:inline-block;background:#1a3a5c;color:#fff;text-decoration:none;padding:10px 24px;border-radius:6px;font-size:13px;font-weight:bold">&rarr; Zum Dashboard</a>
  </div>

  <div style="padding:14px 32px;background:#f4f8fc;border-top:1px solid #e4eaf0;text-align:center">
    <div style="font-size:11px;color:#aaa">Aktions Tracker &middot; Automatischer Monatsbericht am 1. des Monats<br>
    Empfänger identisch zum Wochenbericht &ndash; Einstellungen unter <em>Einstellungen &rarr; Wochen-/Monatsbericht</em></div>
  </div>

</div>
</body></html>'''

        firma_teil = f' – {FIRMA_NAME}' if FIRMA_NAME else ''
        ok_count   = 0

        # Foto-ZIP für Vormonat erstellen
        von_str           = erster_vormonat.isoformat()
        bis_str           = letzter_vormonat.isoformat()
        zip_bytes, foto_count = erstelle_fotos_zip_bytes(von=von_str, bis=bis_str)
        base_atts = []
        if foto_count > 0:
            zip_name = f"Fotos_{erster_vormonat.strftime('%Y-%m')}.zip"
            base_atts.append((zip_name, zip_bytes, "application/zip"))

        _safe_co_m  = re.sub(r'[^\w]', '_', COMPANY_SHORT).strip('_') or 'AktionsTracker'
        _mb_label   = erster_vormonat.strftime('%Y-%m')

        def _atts_with_pdf(html):
            pdf = _html_to_pdf(html)
            if pdf:
                fname = f'Monatsbericht_{_safe_co_m}_{_mb_label}.pdf'
                return base_atts + [(fname, pdf, 'application/pdf')]
            return base_atts

        vkl_teams = list(dict.fromkeys(v['team_id'] for v in vkls if v['team_id']))
        if len(vkl_teams) >= 2:
            for v in vkls:
                if not v['team_id']:
                    continue
                tname   = team_map.get(v['team_id'], f'Team {v["team_id"]}')
                html    = build_html(team_id=v['team_id'], team_name=tname)
                betreff = f'Monatsbericht{firma_teil} – {tname} – {monat_label}'
                if send_email_with_attachments(v['email'], betreff, html, _atts_with_pdf(html)):
                    app.logger.info(f"MONATSBERICHT {monat_label} [{tname}]: Gesendet an {v['email']} ({foto_count} Fotos)")
                    ok_count += 1
                else:
                    app.logger.error(f"MONATSBERICHT {monat_label} [{tname}]: Fehler bei {v['email']}")
            if empfaenger_admin:
                html_g  = build_html(team_id=None, team_name='Alle Teams')
                betreff = f'Monatsbericht{firma_teil} – Alle Teams – {monat_label}'
                for mail in empfaenger_admin:
                    if send_email_with_attachments(mail, betreff, html_g, _atts_with_pdf(html_g)):
                        app.logger.info(f"MONATSBERICHT {monat_label} [Gesamt]: Gesendet an {mail} ({foto_count} Fotos)")
                        ok_count += 1
        else:
            empfaenger = []
            if vkls:
                empfaenger.append(vkls[0]['email'])
            empfaenger.extend(empfaenger_admin)
            if not empfaenger:
                app.logger.warning("MONATSBERICHT: Keine Empfänger konfiguriert – übersprungen.")
                return False, "Keine Empfänger konfiguriert."
            html    = build_html(team_id=None)
            betreff = f'Monatsbericht Aktionstracker{firma_teil} – {monat_label}'
            atts    = _atts_with_pdf(html)
            for mail in empfaenger:
                if send_email_with_attachments(mail, betreff, html, atts):
                    app.logger.info(f"MONATSBERICHT {monat_label}: Gesendet an {mail} ({foto_count} Fotos)")
                    ok_count += 1

        if ok_count > 0:
            execute("UPDATE wochenbericht_config SET zuletzt_gesendet_monat=? WHERE id=1", (monat_key,))
            return True, f"Gesendet an {ok_count} Empfänger"
        else:
            detail = f': {_smtp_last_error}' if _smtp_last_error else ''
            return False, f"E-Mail-Versand fehlgeschlagen{detail}"

    except Exception as e:
        app.logger.error(f"MONATSBERICHT Fehler: {e}", exc_info=True)
        return False, f"Fehler: {e}"


def send_monatsbericht(force=False):
    """Wrapper für APScheduler – erstellt eigenen App-Context."""
    with app.app_context():
        return _do_send_monatsbericht(force=force)


# ─── Demo-Frischhaltung: jede Woche neue Aktivitäten ─────────────────────────

def _demo_pipeline_cleanup():
    """Bereinigt Demo-Daten: alte offene Bestellungen schließen, vergangene Vertretungen löschen,
    zukünftige Vertretungen pro Rep anlegen. Läuft automatisch im Sunday-Job und per Admin-Trigger."""
    import random as rnd
    from datetime import date, timedelta

    today = date.today()
    db    = get_db()

    reps = db.execute("SELECT id, name FROM mitarbeiter WHERE rolle='rep' AND aktiv=1").fetchall()

    # 1. Alte offene Bestellungen schließen: pro Rep max 2 offen lassen, Rest → geliefert
    for rep in reps:
        offene = db.execute(
            "SELECT id FROM aktivitaet WHERE aktionstyp='Bestellung' "
            "AND COALESCE(bestell_status,'offen')='offen' AND mitarbeiter_id=? ORDER BY datum DESC",
            (rep['id'],)
        ).fetchall()
        # Die 2 neuesten offen lassen, Rest schließen
        zu_schliessen = offene[2:]
        for row in zu_schliessen:
            db.execute(
                "UPDATE aktivitaet SET bestell_status='geliefert', realisiert_am=? WHERE id=?",
                (today.isoformat(), row['id'])
            )

    # 2. Tagesplan: vergangene Tage als erledigt markieren; heute anteilig nach Uhrzeit
    db.execute(
        "UPDATE tagesplan SET erledigt=1 WHERE datum < ? AND erledigt=0 AND COALESCE(geloescht,0)=0",
        (today.isoformat(),)
    )
    from datetime import datetime
    stunde = datetime.now().hour
    if stunde >= 15:
        anteil = 0.8   # ab 15 Uhr: ~80% des Tages erledigt
    elif stunde >= 12:
        anteil = 0.5   # ab 12 Uhr: ~50%
    elif stunde >= 10:
        anteil = 0.3   # ab 10 Uhr: ~30%
    else:
        anteil = 0.0   # vor 10 Uhr: nichts
    if anteil > 0:
        heute_eintraege = db.execute(
            "SELECT id FROM tagesplan WHERE datum=? AND erledigt=0 AND COALESCE(geloescht,0)=0 ORDER BY reihenfolge",
            (today.isoformat(),)
        ).fetchall()
        n_erledigt = int(len(heute_eintraege) * anteil)
        for row in heute_eintraege[:n_erledigt]:
            db.execute("UPDATE tagesplan SET erledigt=1 WHERE id=?", (row['id'],))

    # 3. Vergangene Vertretungen löschen; falsch gesetzte 'offen'-Einträge auf 'angefragt' korrigieren
    db.execute("DELETE FROM vertretung WHERE bis < ?", (today.isoformat(),))
    db.execute("UPDATE vertretung SET status='angefragt' WHERE status='offen' AND von > ?", (today.isoformat(),))

    # 3. Zukünftige Vertretungen pro Rep anlegen falls keine vorhanden:
    #    - Erste 3 Reps → status='angefragt' (ausstehend, sichtbar für VKL zum Genehmigen)
    #    - Restliche    → status='bestätigt' (bereits genehmigt, sichtbar in Übersicht)
    for i, rep in enumerate(reps):
        existing = db.execute(
            "SELECT COUNT(*) FROM vertretung WHERE abwesender_id=? AND von > ?",
            (rep['id'], today.isoformat())
        ).fetchone()[0]
        if existing == 0:
            start_offset = rnd.randint(14, 45)
            dauer        = rnd.randint(3, 5)
            von    = (today + timedelta(days=start_offset)).isoformat()
            bis    = (today + timedelta(days=start_offset + dauer)).isoformat()
            status = 'angefragt' if i < 3 else 'bestätigt'
            db.execute(
                "INSERT INTO vertretung (abwesender_id, vertreter_id, von, bis, status) VALUES (?,NULL,?,?,?)",
                (rep['id'], von, bis, status)
            )

    db.commit()
    app.logger.info("Demo-Pipeline-Cleanup abgeschlossen.")


def _do_demo_woche_nachfuellen(force=False):
    """Fügt der vergangenen Woche je 5 Aktivitäten pro Rep hinzu (Mix: Aufbau/Bestellung/Besuch).
    Läuft jeden Sonntag 23:30 – Daten sind bereit für den Montags-Wochenbericht.
    Idempotent: Falls die Woche bereits Einträge hat, wird nichts eingefügt (außer force=True)."""
    import random as rnd
    from datetime import date, timedelta

    _demo_pipeline_cleanup()  # Erst aufräumen, dann neue Daten einfügen

    today       = date.today()
    letzter_mo  = today - timedelta(days=today.weekday() + 7)
    letzter_fr  = letzter_mo + timedelta(days=4)
    kw          = letzter_mo.isocalendar()[1]
    rnd.seed(kw * 1000 + letzter_mo.year)  # deterministisch pro KW

    db      = get_db()
    reps    = db.execute("SELECT id FROM mitarbeiter WHERE rolle='rep'").fetchall()
    stellen = db.execute("SELECT id, typ FROM verkaufsstelle WHERE aktiv=1").fetchall()
    biere   = db.execute("SELECT id FROM biersorte WHERE aktiv=1 AND COALESCE(ist_gratisware,0)=0").fetchall()
    bier_ids = [b['id'] for b in biere]

    # Bereits Daten für diese Woche? → überspringen (außer manueller Force)
    existing = db.execute(
        "SELECT COUNT(*) FROM aktivitaet WHERE datum BETWEEN ? AND ?",
        (letzter_mo.isoformat(), letzter_fr.isoformat())
    ).fetchone()[0]
    if existing > 0 and not force:
        app.logger.info(f"Demo-Seed KW {kw}/{letzter_mo.year}: {existing} Einträge vorhanden, übersprungen.")
        return

    NOTIZEN_AUFBAU = [
        '', '', '',
        'Sonderaktion vereinbart', 'Kunde sehr zufrieden',
        'Neues Kühlregal besprochen', 'Probierpaket mitgenommen',
        'Konkurrenzprodukte gesichtet', 'Rückgabe 3 leere Displays',
        'Termin für Herbstaktion vereinbart', 'Stammkunde, läuft sehr gut',
        'Aufbau problemlos, neues Regal eingerichtet',
    ]
    NOTIZEN_BESTELLUNG = [
        '', '',
        'Bestellung für nächste Lieferung',
        'Nachbestellung – läuft sehr gut',
        'Kunde bestellt für Herbst-Event',
        'Sonderbestellung Weihnachtsmarkt',
        'Erste Bestellung, neuer Gaststättenkunde',
        'Bestellung telefonisch bestätigt',
    ]
    NOTIZEN_BESUCH = [
        '', '',
        'Allgemeines Verkaufsgespräch',
        'Feedback eingeholt – positiv',
        'Konkurrenzprodukte gesichtet, gut positioniert',
        'Termin für nächsten Aufbau vereinbart',
        'Kein Bedarf aktuell, Wiedervorlage in 2 Wochen',
        'Neues Sortiment vorgestellt',
    ]

    gesamt = 0
    for rep in reps:
        # Nur dem Rep zugewiesene Stationen nutzen (geografisch korrekt)
        zugewiesen = db.execute("""
            SELECT v.id, v.typ FROM verkaufsstelle v
            JOIN mitarbeiter_verkaufsstelle mv ON mv.verkaufsstelle_id = v.id
            WHERE mv.mitarbeiter_id = ? AND v.aktiv = 1
        """, (rep['id'],)).fetchall()
        rep_stellen = list(zugewiesen) if len(zugewiesen) >= 2 else list(stellen)

        # Zusammensetzung pro Rep und Woche: ~10 Aktivitäten
        n_aufbau     = rnd.randint(4, 5)
        n_bestellung = rnd.randint(3, 4)
        n_besuch     = rnd.randint(1, 2)
        typen = ['Aufbau'] * n_aufbau + ['Bestellung'] * n_bestellung + ['Besuch'] * n_besuch
        rnd.shuffle(typen)
        n_total = len(typen)

        # Tage Mo–Fr mit Wiederholung (je ~2 Aktivitäten pro Tag)
        tage = sorted(rnd.choices(range(5), k=n_total))

        # Verkaufsstellen mit Wiederholung falls nötig
        if len(rep_stellen) >= n_total:
            vs_woche = rnd.sample(rep_stellen, k=n_total)
        else:
            vs_woche = [rnd.choice(rep_stellen) for _ in range(n_total)]

        for i, (tag, typ) in enumerate(zip(tage, typen)):
            datum          = (letzter_mo + timedelta(days=tag)).isoformat()
            vs             = vs_woche[i]
            displays       = rnd.choices([0,1,2,3,4,5], weights=[30,25,20,12,8,5])[0] if typ == 'Aufbau' else 0
            bestell_status = 'offen' if typ == 'Bestellung' else None
            if typ == 'Aufbau':
                notiz = rnd.choice(NOTIZEN_AUFBAU)
            elif typ == 'Bestellung':
                notiz = rnd.choice(NOTIZEN_BESTELLUNG)
            else:
                notiz = rnd.choice(NOTIZEN_BESUCH)

            cur = db.execute(
                "INSERT INTO aktivitaet "
                "(datum,mitarbeiter_id,verkaufsstelle_id,anzahl_displays,notizen,aktionstyp,bestell_status) "
                "VALUES (?,?,?,?,?,?,?)",
                (datum, rep['id'], vs['id'], displays, notiz, typ, bestell_status)
            )
            aid = cur.lastrowid

            # Aufbau: älteste offene Bestellung des Reps schließen (~75% der Aufbauten = 3-4 pro Woche)
            if typ == 'Aufbau' and rnd.random() < 0.75:
                offene = db.execute(
                    "SELECT id FROM aktivitaet "
                    "WHERE aktionstyp='Bestellung' AND COALESCE(bestell_status,'offen')='offen' "
                    "AND mitarbeiter_id=? ORDER BY datum ASC LIMIT 1",
                    (rep['id'],)
                ).fetchone()
                if offene:
                    db.execute(
                        "UPDATE aktivitaet SET bestell_status='aufgebaut', realisiert_am=? WHERE id=?",
                        (datum, offene['id'])
                    )

            # Bestellpositionen: nur bei Bestellung (nicht bei Aufbau oder Besuch)
            if typ == 'Bestellung':
                for bier_id in rnd.sample(bier_ids, k=rnd.randint(2, min(4, len(bier_ids)))):
                    db.execute(
                        "INSERT INTO bestellposition (aktivitaet_id,biersorte_id,kisten_anzahl) VALUES (?,?,?)",
                        (aid, bier_id, rnd.randint(3, 50))
                    )
            gesamt += 1

    db.commit()
    app.logger.info(f"Demo-Seed KW {kw}/{letzter_mo.year}: {gesamt} neue Aktivitäten eingefügt (Aufbau/Bestellung/Besuch).")


def demo_woche_nachfuellen():
    """Wrapper für APScheduler."""
    with app.app_context():
        _do_demo_woche_nachfuellen()


@app.route('/einstellungen/wochenbericht', methods=['GET', 'POST'])
@login_required
def einstellungen_wochenbericht():
    if session.get('rolle') not in ('admin', 'verkaufsleiter'):
        return redirect(url_for('dashboard'))

    # Sicherheits-Migration: Tabelle + Zeile anlegen falls DB älter als dieses Feature
    try:
        execute('''CREATE TABLE IF NOT EXISTS wochenbericht_config (
            id             INTEGER PRIMARY KEY CHECK (id = 1),
            aktiv          INTEGER DEFAULT 0,
            empfaenger_2   TEXT    DEFAULT '',
            empfaenger_3   TEXT    DEFAULT '',
            zuletzt_gesendet TEXT  DEFAULT ''
        )''')
        execute("INSERT OR IGNORE INTO wochenbericht_config (id) VALUES (1)")
    except Exception as _e:
        app.logger.warning(f"wochenbericht_config setup: {_e}")

    vkl = query(
        "SELECT email, name FROM mitarbeiter WHERE rolle IN ('verkaufsleiter','admin') "
        "AND email IS NOT NULL AND email != '' ORDER BY rolle='verkaufsleiter' DESC LIMIT 1",
        one=True
    )
    config = query("SELECT * FROM wochenbericht_config WHERE id=1", one=True)

    if request.method == 'POST':
        aktiv        = 1 if request.form.get('aktiv') else 0
        empfaenger_2 = request.form.get('empfaenger_2', '').strip()
        empfaenger_3 = request.form.get('empfaenger_3', '').strip()
        # Bugreport 2026-07-21: empfaenger_2/3 waren frei eingebbar, ohne Prüfung, dass es
        # sich um eine firmeninterne Adresse handelt – ein VKL (die Einstellung ist global,
        # nicht pro Team) konnte sich damit dauerhaft Team-KPI-Berichte an eine beliebige
        # externe/private Adresse schicken lassen. Nur noch E-Mails zulassen, die zu einem
        # bestehenden Mitarbeiter-Account gehören (Admin/VKL/Rep).
        _bekannte_mails = {r['email'].lower() for r in query(
            "SELECT email FROM mitarbeiter WHERE email IS NOT NULL AND email != ''"
        )}
        for _feld_name, _wert in (('Empfänger 2', empfaenger_2), ('Empfänger 3', empfaenger_3)):
            if _wert and _wert.lower() not in _bekannte_mails:
                flash(f'{_feld_name} „{_wert}" ist keine bekannte Mitarbeiter-E-Mail-Adresse im System und wurde nicht übernommen.', 'danger')
                return redirect(url_for('einstellungen_wochenbericht'))
        execute(
            "UPDATE wochenbericht_config SET aktiv=?, empfaenger_2=?, empfaenger_3=? WHERE id=1",
            (aktiv, empfaenger_2, empfaenger_3)
        )
        if request.form.get('jetzt_senden'):
            try:
                result = _do_send_wochenbericht(force=True)
                ok, msg = result if isinstance(result, tuple) and len(result) == 2 else (False, f'Unerwartetes Ergebnis: {result!r}')
            except BaseException as _bex:
                import traceback as _tb
                app.logger.error(f"WOCHENBERICHT UNCAUGHT:\n{_tb.format_exc()}")
                ok, msg = False, f'Fehler ({type(_bex).__name__}): {_bex}'
            flash(msg, 'success' if ok else 'danger')
        elif request.form.get('jetzt_monatsbericht_senden'):
            try:
                result = _do_send_monatsbericht(force=True)
                ok, msg = result if isinstance(result, tuple) and len(result) == 2 else (False, f'Unerwartetes Ergebnis: {result!r}')
            except BaseException as _bex:
                import traceback as _tb
                app.logger.error(f"MONATSBERICHT UNCAUGHT:\n{_tb.format_exc()}")
                ok, msg = False, f'Fehler ({type(_bex).__name__}): {_bex}'
            flash(msg, 'success' if ok else 'danger')
        else:
            flash('Einstellungen gespeichert.', 'success')
        return redirect(url_for('einstellungen_wochenbericht'))

    return render_template('einstellungen_wochenbericht.html',
                           config=config, vkl=vkl,
                           is_manager=True, is_admin=session.get('rolle')=='admin')


@app.route('/einstellungen/wochenbericht/vorschau')
@login_required
def wochenbericht_vorschau():
    """Rendert die Wochenbericht-E-Mail als HTML direkt im Browser (kein Versand)."""
    if session.get('rolle') not in ('admin', 'verkaufsleiter'):
        return redirect(url_for('dashboard'))

    heute          = date.today()
    montag_diese   = heute - timedelta(days=heute.weekday())
    sonntag_diese  = montag_diese + timedelta(days=6)
    montag_letzte  = montag_diese - timedelta(days=7)
    sonntag_letzte = montag_letzte + timedelta(days=6)
    kw_nr   = montag_diese.strftime('%V')
    datum_von = montag_diese.strftime('%d.%m.')
    datum_bis = sonntag_diese.strftime('%d.%m.%Y')

    def _stats(von, bis):
        return query('''
            SELECT COUNT(DISTINCT a.id) AS besuche,
                   COUNT(DISTINCT CASE WHEN COALESCE(a.aktionstyp,'Aufbau')='Aufbau'
                                       THEN a.id END) AS aufbauten,
                   COUNT(DISTINCT CASE WHEN a.aktionstyp='Bestellung'
                                       THEN a.id END) AS bestellungen,
                   COALESCE(SUM(CASE WHEN a.aktionstyp='Bestellung'
                                     THEN bp.kisten_anzahl END), 0) AS kisten,
                   COALESCE(SUM(CASE WHEN COALESCE(a.aktionstyp,'Aufbau')='Aufbau'
                                     THEN a.anzahl_displays END), 0) AS displays
            FROM aktivitaet a
            LEFT JOIN bestellposition bp ON bp.aktivitaet_id = a.id
            WHERE a.datum BETWEEN ? AND ?
        ''', (von.isoformat(), bis.isoformat()), one=True)

    diese  = _stats(montag_diese,  sonntag_diese)
    letzte = _stats(montag_letzte, sonntag_letzte)

    rep_stats = query('''
        SELECT m.id AS mitarbeiter_id, m.name,
               COUNT(DISTINCT a.id) AS besuche,
               COUNT(DISTINCT CASE WHEN a.aktionstyp='Bestellung' THEN a.id END) AS bestellungen,
               COUNT(DISTINCT CASE WHEN COALESCE(a.aktionstyp,'Aufbau')='Aufbau' THEN a.id END) AS aufbauten,
               COALESCE(SUM(CASE WHEN a.aktionstyp='Bestellung'
                                 THEN bp.kisten_anzahl END), 0) AS kisten
        FROM mitarbeiter m
        LEFT JOIN aktivitaet a ON a.mitarbeiter_id = m.id AND a.datum BETWEEN ? AND ?
        LEFT JOIN bestellposition bp ON bp.aktivitaet_id = a.id
        WHERE (m.rolle='rep' OR (m.rolle='verkaufsleiter' AND EXISTS(SELECT 1 FROM mitarbeiter_verkaufsstelle mv WHERE mv.mitarbeiter_id=m.id)))
        GROUP BY m.id, m.name ORDER BY kisten DESC, m.name
    ''', (montag_diese.isoformat(), sonntag_diese.isoformat()))

    rep_letzte_w = query('''
        SELECT m.id AS mitarbeiter_id,
               COUNT(DISTINCT a.id) AS besuche,
               COALESCE(SUM(CASE WHEN a.aktionstyp='Bestellung'
                                 THEN bp.kisten_anzahl END), 0) AS kisten
        FROM aktivitaet a
        JOIN mitarbeiter m ON m.id = a.mitarbeiter_id
        LEFT JOIN bestellposition bp ON bp.aktivitaet_id = a.id
        WHERE a.datum BETWEEN ? AND ? AND (m.rolle='rep' OR (m.rolle='verkaufsleiter' AND EXISTS(SELECT 1 FROM mitarbeiter_verkaufsstelle mv WHERE mv.mitarbeiter_id=m.id)))
        GROUP BY m.id
    ''', (montag_letzte.isoformat(), sonntag_letzte.isoformat()))
    letzte_map_w = {r['mitarbeiter_id']: r for r in rep_letzte_w}

    _tp_team_v_row = query(
        "SELECT COUNT(*) AS geplant, COALESCE(SUM(tp.erledigt),0) AS erledigt "
        "FROM tagesplan tp JOIN mitarbeiter m ON m.id=tp.mitarbeiter_id "
        "WHERE tp.datum BETWEEN ? AND ? AND COALESCE(tp.geloescht,0)=0 AND (m.rolle='rep' OR (m.rolle='verkaufsleiter' AND EXISTS(SELECT 1 FROM mitarbeiter_verkaufsstelle mv WHERE mv.mitarbeiter_id=m.id)))",
        (montag_diese.isoformat(), sonntag_diese.isoformat()), one=True)
    _tp_team_v = dict(_tp_team_v_row) if _tp_team_v_row else {}
    _tp_reps_v = query(
        "SELECT tp.mitarbeiter_id, COUNT(*) AS geplant, COALESCE(SUM(tp.erledigt),0) AS erledigt "
        "FROM tagesplan tp JOIN mitarbeiter m ON m.id=tp.mitarbeiter_id "
        "WHERE tp.datum BETWEEN ? AND ? AND COALESCE(tp.geloescht,0)=0 AND (m.rolle='rep' OR (m.rolle='verkaufsleiter' AND EXISTS(SELECT 1 FROM mitarbeiter_verkaufsstelle mv WHERE mv.mitarbeiter_id=m.id))) "
        "GROUP BY tp.mitarbeiter_id",
        (montag_diese.isoformat(), sonntag_diese.isoformat())) or []
    tp_map_v = {r['mitarbeiter_id']: dict(r) for r in _tp_reps_v}

    def _plan_badge_v(geplant, erledigt):
        if not geplant:
            return '<span style="color:#aaa;font-size:12px">–</span>'
        pct = round(erledigt / geplant * 100)
        col = '#2d8a4e' if pct >= 80 else '#c8860a' if pct >= 60 else '#c0392b'
        return (f'<span style="font-size:12px">{erledigt} erl. / {geplant} ges.</span>'
                f'<br><span style="font-size:11px;font-weight:bold;color:{col}">{pct}%</span>')

    tp_g_v = _tp_team_v.get('geplant', 0)
    tp_e_v = _tp_team_v.get('erledigt', 0)
    tp_o_v = tp_g_v - tp_e_v
    tp_pct_v = round(tp_e_v / tp_g_v * 100) if tp_g_v else 0
    tp_col_v = '#2d8a4e' if tp_pct_v >= 80 else '#c8860a' if tp_pct_v >= 60 else ('#c0392b' if tp_g_v else '#aaa')
    tp_summary_v = (
        f'<div style="padding:12px 32px 0">'
        f'<div style="background:#f3f0fa;border:1px solid #d5cdf0;border-radius:8px;padding:10px 16px;font-size:13px">'
        f'<span style="font-weight:bold;color:#5a3e9e">Besuchsplanung diese Woche:</span>'
        f'&nbsp;&nbsp;{tp_g_v} geplant &nbsp;·&nbsp; {tp_e_v} erledigt &nbsp;·&nbsp; {tp_o_v} offen'
        f'&nbsp;&nbsp;<span style="font-weight:bold;color:{tp_col_v}">{tp_pct_v}%</span>'
        f'</div></div>'
    ) if tp_g_v else ''

    offene_map = {r['mitarbeiter_id']: r['n'] for r in query(
        "SELECT a.mitarbeiter_id, COUNT(*) AS n FROM aktivitaet a "
        "JOIN mitarbeiter m ON m.id=a.mitarbeiter_id "
        "WHERE a.aktionstyp='Bestellung' AND COALESCE(a.bestell_status,'offen')='offen' "
        "AND (m.rolle='rep' OR (m.rolle='verkaufsleiter' AND EXISTS(SELECT 1 FROM mitarbeiter_verkaufsstelle mv WHERE mv.mitarbeiter_id=m.id))) GROUP BY a.mitarbeiter_id"
    )}
    pipeline = query(
        "SELECT COALESCE(SUM(CASE WHEN COALESCE(bestell_status,'offen')='offen' THEN 1 END),0) AS offen,"
        "       COALESCE(SUM(CASE WHEN bestell_status='aufgebaut' THEN 1 END),0) AS aufgebaut,"
        "       COALESCE(SUM(CASE WHEN bestell_status='storniert' THEN 1 END),0) AS storniert "
        "FROM aktivitaet WHERE aktionstyp='Bestellung'",
        one=True
    )

    ue_rows_v = query(
        "SELECT v.name AS station, m.name AS rep, a.datum, "
        "CAST(julianday('now') - julianday(a.datum) AS INTEGER) AS tage "
        "FROM aktivitaet a "
        "JOIN mitarbeiter m ON m.id=a.mitarbeiter_id "
        "JOIN verkaufsstelle v ON v.id=a.verkaufsstelle_id "
        "WHERE a.aktionstyp='Bestellung' AND COALESCE(a.bestell_status,'offen')='offen' "
        "AND julianday('now') - julianday(a.datum) > 28 "
        "ORDER BY tage DESC LIMIT 10"
    )
    if ue_rows_v:
        ue_trs_v = ''.join(f'''
          <tr>
            <td style="padding:7px 16px;border-bottom:1px solid #f0e8d0;font-size:12px;font-weight:600">{u["station"]}</td>
            <td style="padding:7px 8px;border-bottom:1px solid #f0e8d0;font-size:12px;color:#666">{u["rep"]}</td>
            <td style="padding:7px 16px;border-bottom:1px solid #f0e8d0;font-size:12px;text-align:right">
              <span style="background:#fdecc8;color:#8a5a00;padding:2px 8px;border-radius:4px">{u["tage"]} Tage</span>
            </td>
          </tr>''' for u in ue_rows_v)
        ueberfaellig_html_v = f'''
  <div style="padding:0 32px 20px">
    <div style="background:#fff8f0;border:1px solid #f0c674;border-radius:8px;overflow:hidden">
      <div style="background:#fdecc8;padding:10px 16px;font-size:13px;font-weight:bold;color:#8a5a00">
        &#9888; &Uuml;berf&auml;llig &ndash; Bestellungen offen seit &uuml;ber 4 Wochen ({len(ue_rows_v)})
      </div>
      <table width="100%" cellpadding="0" cellspacing="0">{ue_trs_v}
      </table>
    </div>
  </div>'''
    else:
        ueberfaellig_html_v = ''

    def trend_str(neu, alt):
        d = neu - alt
        return f'+{d}' if d > 0 else (str(d) if d < 0 else '±0')
    def trend_col(neu, alt):
        return '#2d8a4e' if neu > alt else ('#c0392b' if neu < alt else '#888')

    def _offen_col(n):
        return (f'<span style="color:#c8860a;font-weight:bold">{n}</span>'
                if n > 0 else f'<span style="color:#aaa">0</span>')

    def _trend_cell_w(neu, alt):
        d = neu - alt
        if d > 0:   return f'<span style="color:#2d8a4e;font-size:11px">&#x2191;+{d}</span>'
        if d < 0:   return f'<span style="color:#c0392b;font-size:11px">&#x2193;{d}</span>'
        return '<span style="color:#888;font-size:11px">±0</span>'

    _rep_0 = {'besuche': 0, 'kisten': 0}
    rep_rows = ''.join(f'''
        <tr>
          <td style="padding:7px 10px;border-bottom:1px solid #f0f0f0;font-size:13px">{r["name"]}</td>
          <td style="padding:7px 10px;border-bottom:1px solid #f0f0f0;text-align:center;font-size:13px">{r["besuche"]}<br>{_trend_cell_w(r["besuche"], letzte_map_w.get(r["mitarbeiter_id"], _rep_0)["besuche"])}</td>
          <td style="padding:7px 10px;border-bottom:1px solid #f0f0f0;text-align:center;font-size:13px;color:#2e6da4">{r["bestellungen"]}</td>
          <td style="padding:7px 10px;border-bottom:1px solid #f0f0f0;text-align:center;font-size:13px;color:#27ae60">{r["aufbauten"]}</td>
          <td style="padding:7px 10px;border-bottom:1px solid #f0f0f0;text-align:center;font-size:13px;font-weight:600;color:#c8860a">{r["kisten"]}<br>{_trend_cell_w(r["kisten"], letzte_map_w.get(r["mitarbeiter_id"], _rep_0)["kisten"])}</td>
          <td style="padding:7px 10px;border-bottom:1px solid #f0f0f0;text-align:center;font-size:13px">{_offen_col(offene_map.get(r["mitarbeiter_id"], 0))}</td>
          <td style="padding:7px 10px;border-bottom:1px solid #f0f0f0;text-align:center">{_plan_badge_v(tp_map_v.get(r["mitarbeiter_id"],{}).get("geplant",0), tp_map_v.get(r["mitarbeiter_id"],{}).get("erledigt",0))}</td>
        </tr>''' for r in rep_stats) or \
        '<tr><td colspan="7" style="padding:12px;color:#999;text-align:center">Keine Aktivitäten diese Woche</td></tr>'

    dashboard_link = APP_BASE_URL or 'http://localhost:5000'

    html = f'''<!DOCTYPE html>
<html><head><meta charset="utf-8">
<style>body{{margin:0;padding:0;background:#f0f4f8;font-family:Arial,Helvetica,sans-serif}}</style>
</head>
<body>
<div style="position:relative;background:#fffbf0;border:2px dashed #c8860a;padding:10px 24px;text-align:center;font-size:13px;color:#8a5a00">
  <a href="/einstellungen/wochenbericht" style="position:absolute;left:16px;top:50%;transform:translateY(-50%);color:#8a5a00;text-decoration:none;font-weight:bold">&larr; Zurück</a>
  <strong>Vorschau-Modus</strong> – Diese E-Mail wird nicht versendet &nbsp;·&nbsp;
  KW {kw_nr} ({datum_von} – {datum_bis})
</div>
<div style="max-width:600px;margin:24px auto;background:#fff;border-radius:10px;overflow:hidden;box-shadow:0 2px 12px rgba(0,0,0,0.10)">
  <div style="background:#1a3a5c;padding:26px 32px">
    <div style="color:#fff;font-size:20px;font-weight:bold;letter-spacing:.3px">Aktions Tracker</div>
    <div style="color:#90b8d8;font-size:13px;margin-top:5px">Wochenbericht KW {kw_nr} &nbsp;·&nbsp; {datum_von} – {datum_bis}</div>
  </div>
  <div style="padding:28px 32px 8px">
    <div style="font-size:15px;font-weight:bold;color:#1a3a5c;margin-bottom:16px">Gesamtübersicht</div>
    <table width="100%" cellpadding="0" cellspacing="0">
      <tr>
        <td style="text-align:center;padding:18px 10px;background:#f4f8fc;border-radius:8px">
          <div style="font-size:30px;font-weight:bold;color:#1a3a5c">{diese["besuche"]}</div>
          <div style="font-size:12px;color:#666;margin-top:3px">Besuche</div>
          <div style="font-size:11px;font-weight:bold;color:{trend_col(diese["besuche"],letzte["besuche"])};margin-top:5px">{trend_str(diese["besuche"],letzte["besuche"])} ggü. Vorwoche</div>
        </td>
        <td width="12"></td>
        <td style="text-align:center;padding:18px 10px;background:#f4f8fc;border-radius:8px">
          <div style="font-size:30px;font-weight:bold;color:#c8860a">{diese["kisten"]}</div>
          <div style="font-size:12px;color:#666;margin-top:3px">{UNIT_LABEL}</div>
          <div style="font-size:11px;font-weight:bold;color:{trend_col(diese["kisten"],letzte["kisten"])};margin-top:5px">{trend_str(diese["kisten"],letzte["kisten"])} ggü. Vorwoche</div>
        </td>
        <td width="12"></td>
        <td style="text-align:center;padding:18px 10px;background:#f4f8fc;border-radius:8px">
          <div style="font-size:30px;font-weight:bold;color:#2e6da4">{diese["displays"]}</div>
          <div style="font-size:12px;color:#666;margin-top:3px">Aufbauten</div>
          <div style="font-size:11px;font-weight:bold;color:{trend_col(diese["displays"],letzte["displays"])};margin-top:5px">{trend_str(diese["displays"],letzte["displays"])} ggü. Vorwoche</div>
        </td>
      </tr>
    </table>
  </div>
  <div style="padding:16px 32px;background:#fffbf0;border-top:1px solid #f0c674">
    <span style="font-size:13px;font-weight:bold;color:#1a3a5c">Bestellungen Pipeline:</span>
    <span style="margin-left:14px;font-size:13px">
      <span style="color:#c8860a;font-weight:bold">{pipeline["offen"]}</span><span style="color:#777"> offen</span>
      &nbsp;&nbsp;·&nbsp;&nbsp;
      <span style="color:#27ae60;font-weight:bold">{pipeline["aufgebaut"]}</span><span style="color:#777"> aufgebaut</span>
      &nbsp;&nbsp;·&nbsp;&nbsp;
      <span style="color:#6c757d;font-weight:bold">{pipeline["storniert"]}</span><span style="color:#777"> storniert</span>
    </span>
  </div>
  {ueberfaellig_html_v}
  {tp_summary_v}
  <div style="padding:24px 32px">
    <div style="font-size:15px;font-weight:bold;color:#1a3a5c;margin-bottom:12px">Mitarbeiter diese Woche</div>
    <table width="100%" cellpadding="0" cellspacing="0" style="border:1px solid #e4eaf0;border-radius:8px;overflow:hidden">
      <thead>
        <tr style="background:#edf2f7">
          <th style="padding:8px 10px;text-align:left;font-size:10px;color:#666;font-weight:600;letter-spacing:.5px">MITARBEITER</th>
          <th style="padding:8px 10px;text-align:center;font-size:10px;color:#666;font-weight:600;letter-spacing:.5px">BESUCHE</th>
          <th style="padding:8px 10px;text-align:center;font-size:10px;color:#2e6da4;font-weight:600;letter-spacing:.5px">BESTELL.</th>
          <th style="padding:8px 10px;text-align:center;font-size:10px;color:#27ae60;font-weight:600;letter-spacing:.5px">AUFBAUT.</th>
          <th style="padding:8px 10px;text-align:center;font-size:10px;color:#c8860a;font-weight:600;letter-spacing:.5px">{UNIT_LABEL[:7].upper()}</th>
          <th style="padding:8px 10px;text-align:center;font-size:10px;color:#c8860a;font-weight:600;letter-spacing:.5px">OFFEN</th>
          <th style="padding:8px 10px;text-align:center;font-size:10px;color:#5a3e9e;font-weight:600;letter-spacing:.5px">BESUCHSPL.</th>
        </tr>
      </thead>
      <tbody>{rep_rows}</tbody>
    </table>
  </div>
  <div style="padding:16px 32px 24px;text-align:center">
    <a href="{dashboard_link}" style="display:inline-block;background:#1a3a5c;color:#fff;text-decoration:none;padding:10px 24px;border-radius:6px;font-size:13px;font-weight:bold">→ Zum Dashboard</a>
  </div>
  <div style="padding:14px 32px;background:#f4f8fc;border-top:1px solid #e4eaf0;text-align:center">
    <div style="font-size:11px;color:#aaa">Aktions Tracker · Automatischer Wochenbericht jeden Montag</div>
  </div>
</div>
</body></html>'''

    return html, 200, {'Content-Type': 'text/html; charset=utf-8'}


@app.route('/einstellungen/monatsbericht/vorschau')
@login_required
def monatsbericht_vorschau():
    """Rendert den Monatsbericht für den laufenden Monat (kumuliert bis heute) im Browser."""
    if session.get('rolle') not in ('admin', 'verkaufsleiter'):
        return redirect(url_for('dashboard'))

    heute            = date.today()
    erster_dieses    = heute.replace(day=1)
    letzter_vorvorm  = erster_dieses - timedelta(days=1)
    erster_vorvorm   = letzter_vorvorm.replace(day=1)

    _monat_namen = ['Januar','Februar','März','April','Mai','Juni',
                    'Juli','August','September','Oktober','November','Dezember']
    monat_name  = _monat_namen[heute.month - 1]
    vmonat_name = _monat_namen[erster_vorvorm.month - 1]

    def _stats(von, bis):
        return query('''
            SELECT COUNT(DISTINCT a.id) AS besuche,
                   COUNT(DISTINCT CASE WHEN COALESCE(a.aktionstyp,'Aufbau')='Aufbau'
                                       THEN a.id END) AS aufbauten,
                   COUNT(DISTINCT CASE WHEN a.aktionstyp='Bestellung'
                                       THEN a.id END) AS bestellungen,
                   COALESCE(SUM(CASE WHEN a.aktionstyp='Bestellung'
                                     THEN bp.kisten_anzahl END), 0) AS kisten,
                   COALESCE(SUM(CASE WHEN COALESCE(a.aktionstyp,'Aufbau')='Aufbau'
                                     THEN a.anzahl_displays END), 0) AS displays
            FROM aktivitaet a
            LEFT JOIN bestellposition bp ON bp.aktivitaet_id=a.id
            WHERE a.datum BETWEEN ? AND ?
        ''', (von.isoformat(), bis.isoformat()), one=True)

    dieser = _stats(erster_dieses, heute)
    vorher = _stats(erster_vorvorm, letzter_vorvorm)

    rep_stats = query('''
        SELECT m.id AS mitarbeiter_id, m.name,
               COUNT(DISTINCT a.id) AS besuche,
               COUNT(DISTINCT CASE WHEN a.aktionstyp='Bestellung' THEN a.id END) AS bestellungen,
               COUNT(DISTINCT CASE WHEN COALESCE(a.aktionstyp,'Aufbau')='Aufbau' THEN a.id END) AS aufbauten,
               COALESCE(SUM(CASE WHEN a.aktionstyp='Bestellung'
                                 THEN bp.kisten_anzahl END), 0) AS kisten,
               COALESCE(SUM(CASE WHEN COALESCE(a.aktionstyp,'Aufbau')='Aufbau'
                                 THEN a.anzahl_displays END), 0) AS displays
        FROM mitarbeiter m
        LEFT JOIN aktivitaet a ON a.mitarbeiter_id = m.id AND a.datum BETWEEN ? AND ?
        LEFT JOIN bestellposition bp ON bp.aktivitaet_id = a.id
        WHERE (m.rolle='rep' OR (m.rolle='verkaufsleiter' AND EXISTS(SELECT 1 FROM mitarbeiter_verkaufsstelle mv WHERE mv.mitarbeiter_id=m.id)))
        GROUP BY m.id, m.name ORDER BY kisten DESC, m.name
    ''', (erster_dieses.isoformat(), heute.isoformat()))

    rep_letzte_m = query('''
        SELECT m.id AS mitarbeiter_id,
               COUNT(DISTINCT a.id) AS besuche,
               COALESCE(SUM(CASE WHEN a.aktionstyp='Bestellung'
                                 THEN bp.kisten_anzahl END), 0) AS kisten
        FROM aktivitaet a
        JOIN mitarbeiter m ON m.id = a.mitarbeiter_id
        LEFT JOIN bestellposition bp ON bp.aktivitaet_id = a.id
        WHERE a.datum BETWEEN ? AND ? AND (m.rolle='rep' OR (m.rolle='verkaufsleiter' AND EXISTS(SELECT 1 FROM mitarbeiter_verkaufsstelle mv WHERE mv.mitarbeiter_id=m.id)))
        GROUP BY m.id
    ''', (erster_vorvorm.isoformat(), letzter_vorvorm.isoformat()))
    letzte_map_m = {r['mitarbeiter_id']: r for r in rep_letzte_m}

    _mtp_team_v_row = query(
        "SELECT COUNT(*) AS geplant, COALESCE(SUM(tp.erledigt),0) AS erledigt "
        "FROM tagesplan tp JOIN mitarbeiter m ON m.id=tp.mitarbeiter_id "
        "WHERE tp.datum BETWEEN ? AND ? AND COALESCE(tp.geloescht,0)=0 AND (m.rolle='rep' OR (m.rolle='verkaufsleiter' AND EXISTS(SELECT 1 FROM mitarbeiter_verkaufsstelle mv WHERE mv.mitarbeiter_id=m.id)))",
        (erster_dieses.isoformat(), heute.isoformat()), one=True)
    _mtp_team_v = dict(_mtp_team_v_row) if _mtp_team_v_row else {}
    _mtp_reps_v = query(
        "SELECT tp.mitarbeiter_id, COUNT(*) AS geplant, COALESCE(SUM(tp.erledigt),0) AS erledigt "
        "FROM tagesplan tp JOIN mitarbeiter m ON m.id=tp.mitarbeiter_id "
        "WHERE tp.datum BETWEEN ? AND ? AND COALESCE(tp.geloescht,0)=0 AND (m.rolle='rep' OR (m.rolle='verkaufsleiter' AND EXISTS(SELECT 1 FROM mitarbeiter_verkaufsstelle mv WHERE mv.mitarbeiter_id=m.id))) "
        "GROUP BY tp.mitarbeiter_id",
        (erster_dieses.isoformat(), heute.isoformat())) or []
    mtp_map_v = {r['mitarbeiter_id']: dict(r) for r in _mtp_reps_v}

    def _mplan_badge_v(geplant, erledigt):
        if not geplant:
            return '<span style="color:#aaa;font-size:12px">–</span>'
        pct = round(erledigt / geplant * 100)
        col = '#2d8a4e' if pct >= 80 else '#c8860a' if pct >= 60 else '#c0392b'
        return (f'<span style="font-size:12px">{erledigt} erl. / {geplant} ges.</span>'
                f'<br><span style="font-size:11px;font-weight:bold;color:{col}">{pct}%</span>')

    mtp_g_v = _mtp_team_v.get('geplant', 0)
    mtp_e_v = _mtp_team_v.get('erledigt', 0)
    mtp_o_v = mtp_g_v - mtp_e_v
    mtp_pct_v = round(mtp_e_v / mtp_g_v * 100) if mtp_g_v else 0
    mtp_col_v = '#2d8a4e' if mtp_pct_v >= 80 else '#c8860a' if mtp_pct_v >= 60 else ('#c0392b' if mtp_g_v else '#aaa')
    mtp_summary_v = (
        f'<div style="padding:12px 32px 0">'
        f'<div style="background:#f3f0fa;border:1px solid #d5cdf0;border-radius:8px;padding:10px 16px;font-size:13px">'
        f'<span style="font-weight:bold;color:#5a3e9e">Besuchsplanung {monat_name}:</span>'
        f'&nbsp;&nbsp;{mtp_g_v} geplant &nbsp;·&nbsp; {mtp_e_v} erledigt &nbsp;·&nbsp; {mtp_o_v} offen'
        f'&nbsp;&nbsp;<span style="font-weight:bold;color:{mtp_col_v}">{mtp_pct_v}%</span>'
        f'</div></div>'
    ) if mtp_g_v else ''

    pipeline = query(
        "SELECT COALESCE(SUM(CASE WHEN COALESCE(bestell_status,'offen')='offen' THEN 1 END),0) AS offen,"
        "       COALESCE(SUM(CASE WHEN bestell_status='aufgebaut' THEN 1 END),0) AS aufgebaut,"
        "       COALESCE(SUM(CASE WHEN bestell_status='storniert' THEN 1 END),0) AS storniert "
        "FROM aktivitaet WHERE aktionstyp='Bestellung'", one=True)

    def trend_str(neu, alt):
        d = neu - alt
        return f'+{d}' if d > 0 else str(d) if d < 0 else '±0'
    def trend_col(neu, alt):
        return '#2d8a4e' if neu > alt else '#c0392b' if neu < alt else '#888'

    def _trend_cell_m(neu, alt):
        d = neu - alt
        if d > 0:   return f'<span style="color:#2d8a4e;font-size:11px">&#x2191;+{d}</span>'
        if d < 0:   return f'<span style="color:#c0392b;font-size:11px">&#x2193;{d}</span>'
        return '<span style="color:#888;font-size:11px">±0</span>'

    _rep_0 = {'besuche': 0, 'kisten': 0}
    rep_rows = ''.join(f'''
        <tr>
          <td style="padding:7px 10px;border-bottom:1px solid #f0f0f0;font-size:13px">{r["name"]}</td>
          <td style="padding:7px 10px;border-bottom:1px solid #f0f0f0;text-align:center;font-size:13px">{r["besuche"]}<br>{_trend_cell_m(r["besuche"], letzte_map_m.get(r["mitarbeiter_id"], _rep_0)["besuche"])}</td>
          <td style="padding:7px 10px;border-bottom:1px solid #f0f0f0;text-align:center;font-size:13px;color:#2e6da4">{r["bestellungen"]}</td>
          <td style="padding:7px 10px;border-bottom:1px solid #f0f0f0;text-align:center;font-size:13px;color:#27ae60">{r["aufbauten"]}</td>
          <td style="padding:7px 10px;border-bottom:1px solid #f0f0f0;text-align:center;font-size:13px;font-weight:600;color:#c8860a">{r["kisten"]}<br>{_trend_cell_m(r["kisten"], letzte_map_m.get(r["mitarbeiter_id"], _rep_0)["kisten"])}</td>
          <td style="padding:7px 10px;border-bottom:1px solid #f0f0f0;text-align:center;font-size:13px;color:#2e6da4">{r["displays"]}</td>
          <td style="padding:7px 10px;border-bottom:1px solid #f0f0f0;text-align:center">{_mplan_badge_v(mtp_map_v.get(r["mitarbeiter_id"],{}).get("geplant",0), mtp_map_v.get(r["mitarbeiter_id"],{}).get("erledigt",0))}</td>
        </tr>''' for r in rep_stats) or \
        '<tr><td colspan="7" style="padding:12px;color:#999;text-align:center">Noch keine Aktivitäten diesen Monat</td></tr>'

    tage_aktuell  = (heute - erster_dieses).days + 1
    tage_vormonat = (letzter_vorvorm - erster_vorvorm).days + 1

    html = f'''<!DOCTYPE html>
<html><head><meta charset="utf-8">
<style>body{{margin:0;padding:0;background:#f0f4f8;font-family:Arial,Helvetica,sans-serif}}</style>
</head>
<body>
<div style="position:relative;background:#fffbf0;border:2px dashed #c8860a;padding:10px 24px;text-align:center;font-size:13px;color:#8a5a00">
  <a href="/einstellungen/wochenbericht" style="position:absolute;left:16px;top:50%;transform:translateY(-50%);color:#8a5a00;text-decoration:none;font-weight:bold">&larr; Zurück</a>
  <strong>Vorschau-Modus</strong> – Diese E-Mail wird nicht versendet &nbsp;·&nbsp; {monat_name} {heute.year} ({erster_dieses.strftime('%d.%m.')} – {heute.strftime('%d.%m.')}, {tage_aktuell} Tage)
</div>
<div style="max-width:600px;margin:24px auto;background:#fff;border-radius:10px;overflow:hidden;box-shadow:0 2px 12px rgba(0,0,0,0.10)">
  <div style="background:#1a3a5c;padding:26px 32px">
    <div style="color:#fff;font-size:20px;font-weight:bold;letter-spacing:.3px">Aktions Tracker</div>
    <div style="color:#90b8d8;font-size:13px;margin-top:5px">Monatsbericht {monat_name} {heute.year} &nbsp;&middot;&nbsp; {erster_dieses.strftime('%d.%m.')} &ndash; {heute.strftime('%d.%m.%Y')} (laufend)</div>
  </div>

  <div style="padding:28px 32px 8px">
    <div style="font-size:15px;font-weight:bold;color:#1a3a5c;margin-bottom:16px">Gesamtübersicht {monat_name} (bis heute)</div>
    <table width="100%" cellpadding="0" cellspacing="0">
      <tr>
        <td style="text-align:center;padding:18px 10px;background:#f4f8fc;border-radius:8px">
          <div style="font-size:30px;font-weight:bold;color:#1a3a5c">{dieser["besuche"]}</div>
          <div style="font-size:12px;color:#666;margin-top:3px">Besuche</div>
          <div style="font-size:11px;font-weight:bold;color:{trend_col(dieser["besuche"],vorher["besuche"])};margin-top:5px">{trend_str(dieser["besuche"],vorher["besuche"])} ggü. {vmonat_name}</div>
        </td>
        <td width="12"></td>
        <td style="text-align:center;padding:18px 10px;background:#f4f8fc;border-radius:8px">
          <div style="font-size:30px;font-weight:bold;color:#c8860a">{dieser["kisten"]}</div>
          <div style="font-size:12px;color:#666;margin-top:3px">{UNIT_LABEL}</div>
          <div style="font-size:11px;font-weight:bold;color:{trend_col(dieser["kisten"],vorher["kisten"])};margin-top:5px">{trend_str(dieser["kisten"],vorher["kisten"])} ggü. {vmonat_name}</div>
        </td>
        <td width="12"></td>
        <td style="text-align:center;padding:18px 10px;background:#f4f8fc;border-radius:8px">
          <div style="font-size:30px;font-weight:bold;color:#2e6da4">{dieser["displays"]}</div>
          <div style="font-size:12px;color:#666;margin-top:3px">Aufbauten</div>
          <div style="font-size:11px;font-weight:bold;color:{trend_col(dieser["displays"],vorher["displays"])};margin-top:5px">{trend_str(dieser["displays"],vorher["displays"])} ggü. {vmonat_name}</div>
        </td>
      </tr>
    </table>
  </div>

  <div style="padding:16px 32px;background:#fffbf0;border-top:1px solid #f0c674">
    <span style="font-size:13px;font-weight:bold;color:#1a3a5c">Bestellungen Pipeline:</span>
    <span style="margin-left:14px;font-size:13px">
      <span style="color:#c8860a;font-weight:bold">{pipeline["offen"]}</span><span style="color:#777"> offen</span>
      &nbsp;&nbsp;&middot;&nbsp;&nbsp;
      <span style="color:#27ae60;font-weight:bold">{pipeline["aufgebaut"]}</span><span style="color:#777"> aufgebaut</span>
      &nbsp;&nbsp;&middot;&nbsp;&nbsp;
      <span style="color:#6c757d;font-weight:bold">{pipeline["storniert"]}</span><span style="color:#777"> storniert</span>
    </span>
  </div>

  {mtp_summary_v}
  <div style="padding:24px 32px">
    <div style="font-size:15px;font-weight:bold;color:#1a3a5c;margin-bottom:12px">Mitarbeiter &ndash; {monat_name} (bis heute)</div>
    <table width="100%" cellpadding="0" cellspacing="0" style="border:1px solid #e4eaf0;border-radius:8px;overflow:hidden">
      <thead>
        <tr style="background:#edf2f7">
          <th style="padding:8px 10px;text-align:left;font-size:10px;color:#666;font-weight:600;letter-spacing:.5px">MITARBEITER</th>
          <th style="padding:8px 10px;text-align:center;font-size:10px;color:#666;font-weight:600;letter-spacing:.5px">BESUCHE</th>
          <th style="padding:8px 10px;text-align:center;font-size:10px;color:#2e6da4;font-weight:600;letter-spacing:.5px">BESTELL.</th>
          <th style="padding:8px 10px;text-align:center;font-size:10px;color:#27ae60;font-weight:600;letter-spacing:.5px">AUFBAUT.</th>
          <th style="padding:8px 10px;text-align:center;font-size:10px;color:#c8860a;font-weight:600;letter-spacing:.5px">{UNIT_LABEL.upper()[:7]}</th>
          <th style="padding:8px 10px;text-align:center;font-size:10px;color:#2e6da4;font-weight:600;letter-spacing:.5px">AUFBAUT.GES.</th>
          <th style="padding:8px 10px;text-align:center;font-size:10px;color:#5a3e9e;font-weight:600;letter-spacing:.5px">BESUCHSPL.</th>
        </tr>
      </thead>
      <tbody>{rep_rows}</tbody>
    </table>
  </div>

  <div style="padding:14px 32px;background:#f4f8fc;border-top:1px solid #e4eaf0;text-align:center">
    <div style="font-size:11px;color:#aaa">Aktions Tracker &middot; Vorschau laufender Monat &ndash; am 1. des Folgemonats wird der abgeschlossene Monat versendet</div>
  </div>

</div>
</body></html>'''

    return html, 200, {'Content-Type': 'text/html; charset=utf-8'}


# ─── Verkaufsstellen (Rep/VKL-Selbstservice für Lieferant/Ansprechpartner) ────

VS_FELDER = "v.id, v.name, v.strasse, v.plz, v.ort, v.typ, v.landkreis, v.lieferant, v.ansprechpartner, v.hinweis, v.kundennummer"
VS_LISTE_SEITENGROESSE = 150
VS_ADMIN_SEITENGROESSE = 150
VS_DASHBOARD_SEITENGROESSE = 1000


def _verkaufsstellen_liste_sql(suche=None):
    """Baut FROM/WHERE + Parameter für die eigene Verkaufsstellen-Liste
    (Rep: eigene Zuordnung, VKL: Team bzw. alle ohne eigenes Team, Admin: alle).
    Zentral, damit Seiten-Rendering und Such-API exakt denselben Gebiets-Scope
    verwenden."""
    # Homeoffice-Einträge (private Wohnadressen) sind zwar allgemein aus geteilten Listen
    # ausgeblendet (siehe Privacy-Nachbesserung 2026-07-17), aber der Eigentümer selbst und
    # VKL/Admin müssen sie trotzdem sehen können (Team-Übersicht, eigene Datenpflege) – reine
    # Blanko-Ausblendung für alle Rollen war ein Bug, kein bewusstes Verhalten (siehe Arco-Fix,
    # Commit `d26e6f6`). Hier reicht das schlichte Entfernen der Ausblendung: die JOINs scopen
    # ohnehin schon per mitarbeiter_verkaufsstelle auf die eigene Zuordnung bzw. das Team des
    # Homeoffice-Eigentümers.
    rolle = session.get('rolle')
    if rolle == 'rep':
        from_sql  = "FROM verkaufsstelle v JOIN mitarbeiter_verkaufsstelle mv ON mv.verkaufsstelle_id = v.id"
        where_sql = "WHERE mv.mitarbeiter_id = ? AND v.aktiv = 1"
        params    = [session['user_id']]
        distinct  = False
    elif rolle == 'verkaufsleiter' and session.get('team_id'):
        from_sql  = ("FROM verkaufsstelle v JOIN mitarbeiter_verkaufsstelle mv ON mv.verkaufsstelle_id = v.id "
                     "JOIN mitarbeiter m ON m.id = mv.mitarbeiter_id")
        where_sql = "WHERE m.team_id = ? AND v.aktiv = 1"
        params    = [session['team_id']]
        distinct  = True
    else:
        from_sql  = "FROM verkaufsstelle v"
        where_sql = "WHERE v.aktiv = 1"
        params    = []
        distinct  = False

    if suche:
        where_sql += " AND (v.name LIKE ? OR v.ort LIKE ? OR v.landkreis LIKE ?)"
        like = f'%{suche}%'
        params += [like, like, like]

    return from_sql, where_sql, params, distinct


@app.route('/verkaufsstellen')
@login_required
def verkaufsstellen_liste():
    """Liste der Verkaufsstellen im eigenen Gebiet (Rep: eigene Zuordnung, VKL:
    ganzes Team) mit Möglichkeit, Lieferant und Ansprechpartner selbst zu
    korrigieren – diese Stammdaten kennt vor Ort meist nur der Außendienst,
    nicht der Admin. Initial nur die ersten Treffer laden (Performance bei
    VKL ohne eigenes Team / Admin mit vielen Verkaufsstellen); für den Rest
    steht die Suche zur Verfügung (serverseitig, siehe api_verkaufsstellen_mein_gebiet)."""
    from_sql, where_sql, params, distinct = _verkaufsstellen_liste_sql()
    select = f"SELECT {'DISTINCT ' if distinct else ''}{VS_FELDER}"
    gesamt = query(f"SELECT COUNT(*) AS n FROM ({select} {from_sql} {where_sql})", params, one=True)['n']
    verkaufsstellen = query(
        f"{select} {from_sql} {where_sql} ORDER BY v.name LIMIT ?",
        params + [VS_LISTE_SEITENGROESSE]
    )
    return render_template('verkaufsstellen_liste.html',
        verkaufsstellen=verkaufsstellen, vs_gesamt=gesamt, vs_seitengroesse=VS_LISTE_SEITENGROESSE)


@app.route('/api/verkaufsstellen-mein-gebiet')
@login_required
def api_verkaufsstellen_mein_gebiet():
    """Serverseitige Suche für die Verkaufsstellen-Selbstservice-Seite (ersetzt
    reines Client-Filtering, das bei vielen Zeilen den Browser einfrieren
    ließe)."""
    suche = request.args.get('q', '').strip()
    from_sql, where_sql, params, distinct = _verkaufsstellen_liste_sql(suche)
    select = f"SELECT {'DISTINCT ' if distinct else ''}{VS_FELDER}"
    rows = query(
        f"{select} {from_sql} {where_sql} ORDER BY v.name LIMIT ?",
        params + [VS_LISTE_SEITENGROESSE]
    )
    return jsonify([dict(r) for r in rows])


def _verkaufsstelle_im_eigenen_gebiet(vs_id):
    """True wenn die Verkaufsstelle im sichtbaren Gebiet des aktuellen Nutzers liegt
    (Rep: eigene Zuordnung, VKL: Team-Zuordnung bzw. alle ohne eigenes Team, Admin: immer)."""
    rolle = session.get('rolle')
    if rolle == 'admin':
        return True
    if rolle == 'rep':
        row = query(
            "SELECT 1 FROM mitarbeiter_verkaufsstelle WHERE mitarbeiter_id=? AND verkaufsstelle_id=?",
            (session['user_id'], vs_id), one=True
        )
        return bool(row)
    if rolle == 'verkaufsleiter':
        tid = session.get('team_id')
        if not tid:
            return True
        row = query('''
            SELECT 1 FROM mitarbeiter_verkaufsstelle mv
            JOIN mitarbeiter m ON m.id = mv.mitarbeiter_id
            WHERE mv.verkaufsstelle_id=? AND m.team_id=?
        ''', (vs_id, tid), one=True)
        return bool(row)
    return False


def _mitarbeiter_im_eigenen_team(ma_id):
    """True wenn ma_id im sichtbaren Verantwortungsbereich des aktuellen Nutzers liegt
    (Rep: nur sich selbst, VKL: eigenes Team bzw. alle ohne eigenes Team, Admin: immer).
    Analog zu _verkaufsstelle_im_eigenen_gebiet, aber für Mitarbeiter-bezogene Prüfungen
    (Tourenplanung/Tagesplan) statt Verkaufsstellen."""
    rolle = session.get('rolle')
    if rolle == 'admin':
        return True
    if rolle == 'rep':
        return ma_id == session.get('user_id')
    if rolle == 'verkaufsleiter':
        tid = session.get('team_id')
        if not tid:
            return True
        row = query("SELECT 1 FROM mitarbeiter WHERE id=? AND team_id=?", (ma_id, tid), one=True)
        return bool(row)
    return False


@app.route('/verkaufsstelle/<int:vs_id>/kontakt-aktualisieren', methods=['POST'])
@login_required
def verkaufsstelle_kontakt_aktualisieren(vs_id):
    if not _verkaufsstelle_im_eigenen_gebiet(vs_id):
        return jsonify({'ok': False, 'error': 'Kein Zugriff'}), 403
    data = request.get_json(silent=True) or {}
    lieferant       = (data.get('lieferant') or '').strip() or None
    ansprechpartner = (data.get('ansprechpartner') or '').strip() or None
    hinweis         = (data.get('hinweis') or '').strip() or None

    # Admin darf hier zusätzlich alle Stammdaten pflegen (nicht nur Lieferant/
    # Ansprechpartner/Hinweis wie Rep/VKL) – dieselbe Seite dient beiden Zwecken.
    if session.get('rolle') == 'admin':
        name = (data.get('name') or '').strip()
        if not name:
            return jsonify({'ok': False, 'error': 'Name ist ein Pflichtfeld'}), 400
        strasse      = (data.get('strasse') or '').strip() or None
        plz          = (data.get('plz') or '').strip() or None
        ort          = (data.get('ort') or '').strip() or None
        landkreis    = (data.get('landkreis') or '').strip() or None
        typ          = (data.get('typ') or '').strip() or None
        kundennummer = (data.get('kundennummer') or '').strip() or None

        vs_alt = query("SELECT strasse, ort FROM verkaufsstelle WHERE id=?", (vs_id,), one=True)
        adresse_geaendert = strasse != (vs_alt['strasse'] if vs_alt else None) or ort != (vs_alt['ort'] if vs_alt else None)

        execute(
            "UPDATE verkaufsstelle SET name=?, strasse=?, plz=?, ort=?, landkreis=?, typ=?, "
            "kundennummer=?, lieferant=?, ansprechpartner=?, hinweis=? WHERE id=?",
            (name, strasse, plz, ort, landkreis, typ, kundennummer, lieferant, ansprechpartner, hinweis, vs_id)
        )
        if adresse_geaendert and KARTE_MODUS != 'aus' and (strasse or ort):
            lat, lng, quelle, kreis_aus_geo = _geocode_adresse(strasse, ort, plz=plz)
            if lat is not None:
                if kreis_aus_geo and not landkreis:
                    execute("UPDATE verkaufsstelle SET lat=?, lng=?, geocode_quelle=?, landkreis=? WHERE id=?", (lat, lng, quelle, kreis_aus_geo, vs_id))
                else:
                    execute("UPDATE verkaufsstelle SET lat=?, lng=?, geocode_quelle=? WHERE id=?", (lat, lng, quelle, vs_id))
    else:
        execute(
            "UPDATE verkaufsstelle SET lieferant=?, ansprechpartner=?, hinweis=? WHERE id=?",
            (lieferant, ansprechpartner, hinweis, vs_id)
        )
    return jsonify({'ok': True})


@app.route('/api/admin/verkaufsstellen-suche')
@admin_required
def api_admin_verkaufsstellen_suche():
    """Serverseitige Suche für die Admin-Verkaufsstellen-Verwaltung (ersetzt reines
    Client-Filtering, das bei vielen Zeilen den Browser einfrieren ließe)."""
    suche = request.args.get('q', '').strip()
    where_sql = " WHERE homeoffice_mitarbeiter_id IS NULL"
    params = []
    if suche:
        like = f'%{suche}%'
        where_sql += " AND (name LIKE ? OR ort LIKE ? OR landkreis LIKE ? OR typ LIKE ? OR lieferant LIKE ?)"
        params = [like, like, like, like, like]
    rows = query(
        f"SELECT * FROM verkaufsstelle{where_sql} ORDER BY aktiv DESC, name LIMIT ?",
        tuple(params) + (VS_ADMIN_SEITENGROESSE,)
    )
    return jsonify([dict(r) for r in rows])


# ─── Karte ────────────────────────────────────────────────────────────────────

@app.route('/karte')
@login_required
def karte():
    if KARTE_MODUS == 'aus':
        flash('Die Karten-Funktion ist in Ihrem aktuellen Paket nicht verfügbar.', 'warning')
        return redirect(url_for('dashboard'))
    is_manager = session.get('rolle') in ('admin', 'verkaufsleiter')
    _km_sql, _km_p = _team_m_clause('m')
    reps = query(
        f"SELECT id, name, kuerzel FROM mitarbeiter m WHERE rolle IN ('rep','verkaufsleiter'){_km_sql} ORDER BY name",
        _km_p
    )
    _today = date.today()
    _km_woche_str = request.args.get('km_woche', None)
    if _km_woche_str:
        try:
            _km_w = date.fromisoformat(_km_woche_str)
        except ValueError:
            _km_w = _today
        _woche_montag = _km_w - timedelta(days=_km_w.weekday())
    else:
        _woche_montag = _today - timedelta(days=_today.weekday())
    datum_woche_karte = [(_woche_montag + timedelta(days=i)).isoformat() for i in range(7)]
    km_kw          = _woche_montag.isocalendar()[1]
    km_prev_woche  = (_woche_montag - timedelta(days=7)).isoformat()
    km_next_woche  = (_woche_montag + timedelta(days=7)).isoformat()
    return render_template('karte.html', reps=reps, is_manager=is_manager, karte_modus=KARTE_MODUS,
        datum_woche_karte=datum_woche_karte, today_str=_today.isoformat(),
        tomorrow_str=(_today + timedelta(days=1)).isoformat(),
        km_kw=km_kw, km_prev_woche=km_prev_woche, km_next_woche=km_next_woche)


@app.route('/api/karte/daten')
@login_required
def api_karte_daten():
    if KARTE_MODUS == 'aus':
        return jsonify({'error': 'Nicht verfügbar'}), 403
    is_manager = session.get('rolle') in ('admin', 'verkaufsleiter')

    if is_manager:
        # Bugreport 2026-07-21 (Hoch): fehlte bisher – ein VKL mit eigenem Team bekam
        # firmenweite Verkaufsstellen-Daten. Unzugeordnete VS bleiben sichtbar (damit sie
        # zugeordnet werden können), VS die exklusiv einem anderen Team zugeordnet sind
        # werden ausgeblendet.
        _team_karte_sql = ""
        _team_karte_params = []
        if session.get('rolle') == 'verkaufsleiter' and session.get('team_id'):
            _team_karte_sql = """ AND (
                NOT EXISTS (SELECT 1 FROM mitarbeiter_verkaufsstelle mv2 WHERE mv2.verkaufsstelle_id = v.id)
                OR EXISTS (
                    SELECT 1 FROM mitarbeiter_verkaufsstelle mv3
                    JOIN mitarbeiter m3 ON m3.id = mv3.mitarbeiter_id
                    WHERE mv3.verkaufsstelle_id = v.id AND m3.team_id = ?
                )
            )"""
            _team_karte_params = [session['team_id']]
        stellen = query(f"""
            SELECT v.id, v.name, v.ort, v.plz, v.typ, v.strasse, v.ansprechpartner, v.landkreis, v.lat, v.lng,
                   v.homeoffice_mitarbeiter_id,
                   GROUP_CONCAT(m.id || ':' || m.name || ':' || m.kuerzel, '|') AS zuordnungen
            FROM verkaufsstelle v
            LEFT JOIN mitarbeiter_verkaufsstelle mv ON mv.verkaufsstelle_id = v.id
            LEFT JOIN mitarbeiter m ON m.id = mv.mitarbeiter_id AND m.rolle IN ('rep', 'verkaufsleiter')
            WHERE v.aktiv = 1{_team_karte_sql}
            GROUP BY v.id
            ORDER BY v.name
        """, _team_karte_params)
    else:
        stellen = query("""
            SELECT v.id, v.name, v.ort, v.plz, v.typ, v.strasse, v.ansprechpartner, v.landkreis, v.lat, v.lng,
                   v.homeoffice_mitarbeiter_id,
                   m.id || ':' || m.name || ':' || m.kuerzel AS zuordnungen
            FROM verkaufsstelle v
            JOIN mitarbeiter_verkaufsstelle mv ON mv.verkaufsstelle_id = v.id
            JOIN mitarbeiter m ON m.id = mv.mitarbeiter_id
            WHERE v.aktiv = 1 AND mv.mitarbeiter_id = ?
            ORDER BY v.name
        """, (session['user_id'],))

    result = []
    for s in stellen:
        zuordnung_list = []
        if s['zuordnungen']:
            for z in s['zuordnungen'].split('|'):
                parts = z.split(':', 2)
                if len(parts) == 3:
                    zuordnung_list.append({
                        'id': int(parts[0]), 'name': parts[1], 'kuerzel': parts[2]
                    })
        result.append({
            'id': s['id'],
            'name': s['name'],
            'ort':  s['ort'] or '',
            'plz':  s['plz'] or '',
            'typ':  s['typ'] or '',
            'strasse': s['strasse'] or '',
            'ansprechpartner': s['ansprechpartner'] or '',
            'landkreis': s['landkreis'] or '',
            'lat': s['lat'],
            'lng': s['lng'],
            'zuordnungen': zuordnung_list,
            'homeoffice_mitarbeiter_id': s['homeoffice_mitarbeiter_id'],
        })

    # Für die Farb-/Legenden-Zuordnung im Frontend (repFarbe()) muss "reps" auch für
    # Nicht-Manager mindestens den eigenen Account enthalten, sonst findet die Karte
    # keinen Eintrag für die eigene Zuordnung und färbt alle Stationen grau ("nicht
    # zugeordnet"), obwohl sie korrekt zugewiesen sind.
    reps = query(
        "SELECT id, name, kuerzel FROM mitarbeiter WHERE rolle IN ('rep','verkaufsleiter') ORDER BY name"
    ) if is_manager else query(
        "SELECT id, name, kuerzel FROM mitarbeiter WHERE id=?", (session['user_id'],)
    )

    return jsonify({
        'stellen': result,
        'reps': [{'id': r['id'], 'name': r['name'], 'kuerzel': r['kuerzel']} for r in reps],
    })


@app.route('/api/karte/zuordnung-aendern', methods=['POST'])
@manager_required
def api_karte_zuordnung_aendern():
    if KARTE_MODUS == 'aus':
        return jsonify({'error': 'Nicht verfügbar'}), 403
    data      = request.get_json() or {}
    stelle_id = data.get('stelle_id')
    neue_ids  = set(int(i) for i in data.get('rep_ids', []))

    if not stelle_id:
        return jsonify({'error': 'stelle_id fehlt'}), 400

    stelle = query("SELECT name FROM verkaufsstelle WHERE id=?", (stelle_id,), one=True)
    if not stelle:
        return jsonify({'error': 'Station nicht gefunden'}), 404
    stelle_name = stelle['name']

    alte_rows = query(
        "SELECT mitarbeiter_id FROM mitarbeiter_verkaufsstelle WHERE verkaufsstelle_id=?",
        (stelle_id,)
    )
    alte_ids = {r['mitarbeiter_id'] for r in alte_rows}

    # VKL darf nur innerhalb des eigenen Teams umverteilen (Bugreport 2026-07-21: bislang
    # konnte ein VKL jede Station – auch aus fremden Teams – jedem beliebigen Mitarbeiter
    # zuordnen). Admin bleibt uneingeschränkt.
    if session.get('rolle') == 'verkaufsleiter' and session.get('team_id'):
        _team_mitglieder = {r['id'] for r in query(
            "SELECT id FROM mitarbeiter WHERE team_id=?", (session['team_id'],)
        )}
        if not neue_ids <= _team_mitglieder:
            return jsonify({'error': 'Nur Mitarbeiter des eigenen Teams zulässig'}), 403
        if alte_ids and not (alte_ids & _team_mitglieder):
            return jsonify({'error': 'Diese Station gehört nicht zu Ihrem Gebiet'}), 403

    entfernt     = alte_ids - neue_ids
    hinzugefuegt = neue_ids - alte_ids

    db = get_db()
    db.execute("DELETE FROM mitarbeiter_verkaufsstelle WHERE verkaufsstelle_id=?", (stelle_id,))
    for rep_id in neue_ids:
        db.execute(
            "INSERT OR IGNORE INTO mitarbeiter_verkaufsstelle (mitarbeiter_id, verkaufsstelle_id) VALUES (?,?)",
            (rep_id, stelle_id)
        )

    heute = date.today().strftime('%d.%m.%Y')
    for rep_id in entfernt | hinzugefuegt:
        ma_rolle = db.execute("SELECT rolle FROM mitarbeiter WHERE id=?", (rep_id,)).fetchone()
        if not ma_rolle or ma_rolle['rolle'] != 'rep':
            continue
        if rep_id in entfernt:
            msg = f'{heute}: Station "{stelle_name}" wurde aus Ihrem Gebiet entfernt.'
        else:
            msg = f'{heute}: Station "{stelle_name}" wurde Ihrem Gebiet hinzugefügt.'
        bestehend = db.execute(
            "SELECT karte_benachrichtigung FROM mitarbeiter WHERE id=?", (rep_id,)
        ).fetchone()
        alt = bestehend['karte_benachrichtigung'] if bestehend and bestehend['karte_benachrichtigung'] else ''
        neu = (alt + '\n' + msg).strip()
        db.execute("UPDATE mitarbeiter SET karte_benachrichtigung=? WHERE id=?", (neu, rep_id))

    db.commit()
    return jsonify({'ok': True, 'stelle_name': stelle_name})


@app.route('/api/karte/zuordnung-bulk-aendern', methods=['POST'])
@manager_required
def api_karte_zuordnung_bulk_aendern():
    """Überträgt mehrere Verkaufsstellen auf einen Mitarbeiter (oder entfernt Zuordnung).
    Für den Fall, dass ein ganzer Landkreis von einem Mitarbeiter zu einem anderen wechselt."""
    if KARTE_MODUS == 'aus':
        return jsonify({'error': 'Nicht verfügbar'}), 403
    data       = request.get_json() or {}
    stelle_ids = [int(i) for i in data.get('stelle_ids', []) if str(i).isdigit()]
    rep_id_raw = data.get('rep_id')
    rep_id     = int(rep_id_raw) if rep_id_raw not in (None, '', 'none') else None

    if not stelle_ids:
        return jsonify({'error': 'Keine Stationen ausgewählt'}), 400

    db = get_db()
    if rep_id is not None:
        gueltig = db.execute(
            "SELECT 1 FROM mitarbeiter WHERE id=? AND rolle IN ('rep','verkaufsleiter')",
            (rep_id,)
        ).fetchone()
        if not gueltig:
            return jsonify({'error': 'Ungültiger Mitarbeiter'}), 400

    # VKL darf nur innerhalb des eigenen Teams umverteilen (Bugreport 2026-07-21, siehe
    # api_karte_zuordnung_aendern). Admin bleibt uneingeschränkt. team_mitglieder=None
    # bedeutet "kein Scoping nötig" (Admin bzw. VKL ohne Team).
    team_mitglieder = None
    if session.get('rolle') == 'verkaufsleiter' and session.get('team_id'):
        team_mitglieder = {r['id'] for r in db.execute(
            "SELECT id FROM mitarbeiter WHERE team_id=?", (session['team_id'],)
        ).fetchall()}
        if rep_id is not None and rep_id not in team_mitglieder:
            return jsonify({'error': 'Nur Mitarbeiter des eigenen Teams zulässig'}), 403

    heute   = date.today().strftime('%d.%m.%Y')
    ph      = ','.join('?' * len(stelle_ids))
    stellen = db.execute(f"SELECT id, name FROM verkaufsstelle WHERE id IN ({ph})", stelle_ids).fetchall()
    namen_map = {s['id']: s['name'] for s in stellen}

    for stelle_id in stelle_ids:
        if stelle_id not in namen_map:
            continue
        alte_rows = db.execute(
            "SELECT mitarbeiter_id FROM mitarbeiter_verkaufsstelle WHERE verkaufsstelle_id=?",
            (stelle_id,)
        ).fetchall()
        alte_ids = {r['mitarbeiter_id'] for r in alte_rows}
        if team_mitglieder is not None and alte_ids and not (alte_ids & team_mitglieder):
            continue  # Station gehört nicht zum eigenen Gebiet
        neue_ids = {rep_id} if rep_id else set()

        entfernt     = alte_ids - neue_ids
        hinzugefuegt = neue_ids - alte_ids
        if not entfernt and not hinzugefuegt:
            continue

        db.execute("DELETE FROM mitarbeiter_verkaufsstelle WHERE verkaufsstelle_id=?", (stelle_id,))
        for rid in neue_ids:
            db.execute(
                "INSERT OR IGNORE INTO mitarbeiter_verkaufsstelle (mitarbeiter_id, verkaufsstelle_id) VALUES (?,?)",
                (rid, stelle_id)
            )

        stelle_name = namen_map[stelle_id]
        for rid in entfernt | hinzugefuegt:
            ma_rolle = db.execute("SELECT rolle FROM mitarbeiter WHERE id=?", (rid,)).fetchone()
            if not ma_rolle or ma_rolle['rolle'] != 'rep':
                continue
            if rid in entfernt:
                msg = f'{heute}: Station "{stelle_name}" wurde aus Ihrem Gebiet entfernt.'
            else:
                msg = f'{heute}: Station "{stelle_name}" wurde Ihrem Gebiet hinzugefügt.'
            bestehend = db.execute(
                "SELECT karte_benachrichtigung FROM mitarbeiter WHERE id=?", (rid,)
            ).fetchone()
            alt = bestehend['karte_benachrichtigung'] if bestehend and bestehend['karte_benachrichtigung'] else ''
            neu = (alt + '\n' + msg).strip()
            db.execute("UPDATE mitarbeiter SET karte_benachrichtigung=? WHERE id=?", (neu, rid))

    db.commit()
    return jsonify({'ok': True, 'anzahl': len(stelle_ids), 'rep_id': rep_id})


_DACH_BBOX = {'lat_min': 45.8, 'lat_max': 55.2, 'lon_min': 5.8, 'lon_max': 17.2}

def _in_dach(lat, lon):
    return (_DACH_BBOX['lat_min'] <= lat <= _DACH_BBOX['lat_max'] and
            _DACH_BBOX['lon_min'] <= lon <= _DACH_BBOX['lon_max'])

def _plz_zentroid(plz, timeout=6):
    """PLZ-Mittelpunkt: erst lokaler Cache (plz_zentrum), dann Nominatim-PLZ-Lookup.
    Thread-sicher durch eigene sqlite3-Verbindung. Gibt (lat, lng) oder (None, None) zurück."""
    if not plz or len(plz.strip()) < 4:
        return None, None
    plz = plz.strip()
    import sqlite3 as _sq3
    try:
        with _sq3.connect(DATABASE) as _c:
            r = _c.execute("SELECT lat, lng FROM plz_zentrum WHERE plz=?", (plz,)).fetchone()
            if r:
                return r[0], r[1]
    except Exception:
        pass
    url = (f'https://nominatim.openstreetmap.org/search?format=json&limit=1'
           f'&countrycodes=de,at,ch&postalcode={urllib.parse.quote(plz)}')
    try:
        req = urllib.request.Request(url, headers={'User-Agent': 'AktionsTracker/1.0 (info@aktionstracker.de)'})
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            hits = json.loads(resp.read().decode())
        if hits:
            lat, lng = float(hits[0]['lat']), float(hits[0]['lon'])
            if _in_dach(lat, lng):
                try:
                    with _sq3.connect(DATABASE) as _c:
                        _c.execute("INSERT OR IGNORE INTO plz_zentrum (plz, lat, lng) VALUES (?,?,?)",
                                   (plz, lat, lng))
                except Exception:
                    pass
                return lat, lng
    except Exception as exc:
        app.logger.warning(f"PLZ-Zentroid '{plz}': {exc}")
    return None, None

def _geocode_adresse(strasse, ort, plz=None, timeout=8):
    """Koordinaten via Nominatim mit strukturierten Parametern und PLZ-Priorisierung.
    Fallback-Kette: Straße+PLZ+Ort → PLZ+Ort → PLZ allein → Ort (Freitext) → PLZ-Zentroid.
    Ergebnis wird gegen DACH-Bounding-Box validiert.
    Gibt (lat, lng, quelle) zurück; quelle ist 'nominatim', 'plz' oder None."""
    base = 'https://nominatim.openstreetmap.org/search?format=json&addressdetails=1&limit=1&countrycodes=de,at,ch'
    headers = {'User-Agent': 'AktionsTracker/1.0 (info@aktionstracker.de)'}

    def _strukturiert(**felder):
        params = '&'.join(
            f"{k}={urllib.parse.quote(str(v))}"
            for k, v in felder.items() if v
        )
        return f"{base}&{params}" if params else None

    kandidaten = []
    if strasse and plz and ort:
        kandidaten.append(_strukturiert(street=strasse, postalcode=plz, city=ort))
    if plz and ort:
        kandidaten.append(_strukturiert(postalcode=plz, city=ort))
    if plz:
        kandidaten.append(_strukturiert(postalcode=plz))
    if ort:
        kandidaten.append(f"{base}&q={urllib.parse.quote(ort + ', Deutschland')}")

    for url in kandidaten:
        if not url:
            continue
        try:
            req = urllib.request.Request(url, headers=headers)
            with urllib.request.urlopen(req, timeout=timeout) as resp:
                hits = json.loads(resp.read().decode())
            if hits:
                lat, lon = float(hits[0]['lat']), float(hits[0]['lon'])
                if _in_dach(lat, lon):
                    addr = hits[0].get('address', {})
                    county = addr.get('county') or addr.get('state_district') or None
                    return lat, lon, 'nominatim', county
                app.logger.warning(f"Geocode außerhalb DACH verworfen: {lat},{lon}")
        except Exception as exc:
            app.logger.warning(f"Geocode-Fehler: {exc}")
        _time.sleep(1.1)

    if plz:
        _time.sleep(1.1)
        lat, lng = _plz_zentroid(plz)
        if lat is not None:
            app.logger.info(f"PLZ-Zentroid verwendet für PLZ {plz}")
            return lat, lng, 'plz', None

    return None, None, None, None


@app.route('/api/karte/geocode', methods=['POST'])
@manager_required
@limiter.limit('10 per minute')
def api_karte_geocode():
    if KARTE_MODUS == 'aus':
        return jsonify({'error': 'Nicht verfügbar'}), 403

    stellen = query(
        "SELECT id, name, strasse, plz, ort, landkreis FROM verkaufsstelle "
        "WHERE aktiv=1 AND (lat IS NULL OR lng IS NULL OR lat=0 OR lng=0)"
    )
    if not stellen:
        return jsonify({'geocoded': 0, 'total': 0, 'msg': 'Alle Stationen haben bereits Koordinaten.'})

    stellen_list = [dict(s) for s in stellen]

    def geocode_worker(items):
        import sqlite3 as _sqlite3
        conn = _sqlite3.connect(DATABASE)
        ok = fail = 0
        try:
            for stelle in items:
                lat, lng, quelle, kreis_aus_geo = _geocode_adresse(
                    stelle.get('strasse', ''), stelle.get('ort', ''), plz=stelle.get('plz')
                )
                if lat is not None:
                    if kreis_aus_geo and not stelle.get('landkreis'):
                        conn.execute(
                            "UPDATE verkaufsstelle SET lat=?, lng=?, geocode_quelle=?, landkreis=? WHERE id=?",
                            (lat, lng, quelle, kreis_aus_geo, stelle['id'])
                        )
                    else:
                        conn.execute(
                            "UPDATE verkaufsstelle SET lat=?, lng=?, geocode_quelle=? WHERE id=?",
                            (lat, lng, quelle, stelle['id'])
                        )
                    conn.commit()
                    ok += 1
                else:
                    fail += 1
                    app.logger.warning(f"Geocode fehlgeschlagen: {stelle['name']} ({stelle.get('ort')})")
        finally:
            conn.close()
        app.logger.info(f"Geocodierung: {ok} erfolgreich, {fail} fehlgeschlagen.")

    t = threading.Thread(target=geocode_worker, args=(stellen_list,), daemon=True)
    t.start()

    warte = max(30, round(len(stellen_list) * 1.2))
    return jsonify({
        'total': len(stellen_list),
        'msg':   (f'Geocodierung für {len(stellen_list)} Stationen gestartet – '
                  f'läuft im Hintergrund (~{warte} Sek.). '
                  f'Karte wird danach automatisch aktualisiert.'),
        'warte': warte,
    })


def _ma_ids_gebietsscope(ma_ids_angefragt):
    """Erzwingt den Mitarbeiter-Gebiets-Scope serverseitig, statt dem Client-Parameter
    'ma' blind zu vertrauen (Bugreport 2026-07-20, von Arco portiert: /api/karte/heatmap
    und /api/karte/besuche-zeitraum lieferten bei leerem/weggelassenem ma-Parameter
    firmenweite Daten statt nur das eigene Gebiet). Rep: immer nur die eigene ID,
    unabhängig vom Client-Wert. VKL mit Team: Client-Auswahl auf Team-Mitglieder
    beschränkt (leer/ungültig → ganzes Team). Admin bzw. VKL ohne Team: unverändert."""
    rolle = session.get('rolle')
    if rolle == 'rep':
        return [str(session['user_id'])]
    if rolle == 'verkaufsleiter' and session.get('team_id'):
        team_ids = {str(r['id']) for r in query(
            "SELECT id FROM mitarbeiter WHERE team_id=?", (session['team_id'],))}
        gefiltert = [m for m in ma_ids_angefragt if m in team_ids]
        return gefiltert if gefiltert else sorted(team_ids)
    return ma_ids_angefragt


@app.route('/api/karte/heatmap')
@login_required
def api_karte_heatmap():
    if KARTE_MODUS != 'heatmap':
        return jsonify({'error': 'Heatmap nicht verfügbar'}), 403
    jahr       = request.args.get('jahr', date.today().year, type=int)
    ma_raw     = request.args.get('ma', '', type=str)
    ma_ids     = [x.strip() for x in ma_raw.split(',') if x.strip()] if ma_raw else []
    ma_ids     = _ma_ids_gebietsscope(ma_ids)
    monate_raw = request.args.get('monate', '', type=str)
    monate_ids = [int(x.strip()) for x in monate_raw.split(',') if x.strip()] if monate_raw else []
    # betreuung (alle Aktivitäten) | aufbauten | volumen (Einheiten) | offene_bestellungen
    ebene      = request.args.get('ebene', 'betreuung')

    where_conds  = ["v.aktiv = 1", "v.lat IS NOT NULL", "v.lng IS NOT NULL"]
    where_params = []
    if ma_ids:
        ph = ','.join('?' * len(ma_ids))
        where_conds.append(
            f"v.id IN (SELECT verkaufsstelle_id FROM mitarbeiter_verkaufsstelle "
            f"WHERE mitarbeiter_id IN ({ph}))")
        where_params.extend(ma_ids)

    if ebene == 'offene_bestellungen':
        # Zeigt aktuelle offene Bestellungen – kein Jahresfilter
        join_conds  = ["a.verkaufsstelle_id = v.id",
                       "a.aktionstyp = 'Bestellung'",
                       "COALESCE(a.bestell_status,'offen') = 'offen'"]
        join_params = []
        if ma_ids:
            ph = ','.join('?' * len(ma_ids))
            join_conds.append(f"a.mitarbeiter_id IN ({ph})")
            join_params.extend(ma_ids)
        metric     = "COUNT(a.id) AS anzahl"
        extra_join = ""
    else:
        # Range-Vergleich statt strftime(), damit der Covering-Index
        # idx_aktivitaet_vs_datum(verkaufsstelle_id, datum) genutzt werden kann.
        join_conds  = ["a.verkaufsstelle_id = v.id", "a.datum >= ?", "a.datum < ?"]
        join_params = [f"{jahr}-01-01", f"{jahr + 1}-01-01"]
        if monate_ids:
            ph = ','.join('?' * len(monate_ids))
            join_conds.append(f"CAST(strftime('%m', a.datum) AS INTEGER) IN ({ph})")
            join_params.extend(str(m) for m in monate_ids)
        if ma_ids:
            ph = ','.join('?' * len(ma_ids))
            join_conds.append(f"a.mitarbeiter_id IN ({ph})")
            join_params.extend(ma_ids)
        if ebene == 'aufbauten':
            metric     = "COUNT(CASE WHEN COALESCE(a.aktionstyp,'Aufbau')='Aufbau' THEN 1 END) AS anzahl"
            extra_join = ""
        elif ebene == 'volumen':
            metric     = ("COALESCE(SUM(CASE WHEN a.aktionstyp='Bestellung' "
                          "THEN bp.kisten_anzahl END), 0) AS anzahl")
            extra_join = "LEFT JOIN bestellposition bp ON bp.aktivitaet_id = a.id"
        else:  # betreuung
            metric     = "COUNT(a.id) AS anzahl"
            extra_join = ""

    stellen = query(
        f"SELECT v.id, v.name, v.ort, v.plz, v.strasse, v.lat, v.lng, {metric}, "
        f"(SELECT COUNT(*) FROM mitarbeiter_verkaufsstelle WHERE verkaufsstelle_id = v.id) > 0 AS zugeordnet "
        f"FROM verkaufsstelle v "
        f"LEFT JOIN aktivitaet a ON {' AND '.join(join_conds)} "
        f"{extra_join + ' ' if extra_join else ''}"
        f"WHERE {' AND '.join(where_conds)} "
        f"GROUP BY v.id ORDER BY anzahl DESC",
        tuple(join_params + where_params)
    )
    jahre_raw = query("SELECT DISTINCT strftime('%Y', datum) AS jahr FROM aktivitaet ORDER BY jahr DESC")
    return jsonify({
        'stellen': [{'id': s['id'], 'name': s['name'], 'ort': s['ort'] or '',
                     'plz': s['plz'] or '', 'strasse': s['strasse'] or '',
                     'lat': s['lat'], 'lng': s['lng'], 'anzahl': s['anzahl'],
                     'zugeordnet': bool(s['zugeordnet'])} for s in stellen],
        'jahre':   [j['jahr'] for j in jahre_raw],
        'jahr':    jahr,
        'ebene':   ebene,
    })


@app.route('/api/karte/besuche-zeitraum')
@login_required
def api_karte_besuche_zeitraum():
    """Welche Verkaufsstellen wurden im Zeitraum (Jahr + Monate + MA) besucht?"""
    jahr       = request.args.get('jahr', date.today().year, type=int)
    monate_raw = request.args.get('monate', '', type=str)
    monate_ids = [int(x.strip()) for x in monate_raw.split(',') if x.strip()] if monate_raw else []
    ma_raw     = request.args.get('ma', '', type=str)
    ma_ids     = [x.strip() for x in ma_raw.split(',') if x.strip()] if ma_raw else []
    ma_ids     = _ma_ids_gebietsscope(ma_ids)

    conds  = ["strftime('%Y', a.datum) = ?"]
    params = [str(jahr)]
    if monate_ids:
        ph = ','.join('?' * len(monate_ids))
        conds.append(f"CAST(strftime('%m', a.datum) AS INTEGER) IN ({ph})")
        params.extend(str(m) for m in monate_ids)
    if ma_ids:
        ph = ','.join('?' * len(ma_ids))
        conds.append(f"a.mitarbeiter_id IN ({ph})")
        params.extend(ma_ids)

    rows = query(
        f"SELECT a.verkaufsstelle_id AS vid, COUNT(a.id) AS anzahl "
        f"FROM aktivitaet a WHERE {' AND '.join(conds)} GROUP BY a.verkaufsstelle_id",
        tuple(params)
    )
    return jsonify({'besuche': {str(r['vid']): r['anzahl'] for r in rows}})


@app.route('/api/karte/benachrichtigung-quittieren', methods=['POST'])
@login_required
def api_karte_benachrichtigung_quittieren():
    execute("UPDATE mitarbeiter SET karte_benachrichtigung=NULL WHERE id=?", (session['user_id'],))
    session.pop('karte_benachrichtigung', None)
    return jsonify({'ok': True})


# ─── Main ─────────────────────────────────────────────────────────────────────

# Wird von gunicorn (Railway) beim Import ausgeführt
init_db()

# ── Hintergrund-Scheduler ─────────────────────────────────────────────────────
try:
    from apscheduler.schedulers.background import BackgroundScheduler
    _scheduler = BackgroundScheduler(daemon=True, timezone='Europe/Berlin')
    _scheduler.add_job(backup_db,              'interval', days=1, id='backup_db',    replace_existing=True)
    _scheduler.add_job(send_wochenbericht,     'cron', day_of_week='mon', hour=7, minute=0,
                       id='wochenbericht',     replace_existing=True, timezone='Europe/Berlin')
    _scheduler.add_job(send_monatsbericht,     'cron', day=1, hour=7, minute=0,
                       id='monatsbericht',     replace_existing=True, timezone='Europe/Berlin')
    _scheduler.add_job(auto_export_job,        'cron', day=1, hour=8, minute=0,
                       id='auto_export',       replace_existing=True, timezone='Europe/Berlin')
    _scheduler.add_job(cleanup_alte_fotos,     'cron', day=1, hour=9, minute=0,
                       id='cleanup_fotos',     replace_existing=True, timezone='Europe/Berlin')
    _scheduler.add_job(demo_daily_reset,           'cron', hour=3,  minute=0,
                       id='demo_reset',             replace_existing=True, timezone='Europe/Berlin')
    _scheduler.add_job(demo_tagesplan_fortschritt, 'cron', hour=10, minute=0,
                       id='demo_tp_10',             replace_existing=True, timezone='Europe/Berlin')
    _scheduler.add_job(demo_tagesplan_fortschritt, 'cron', hour=13, minute=0,
                       id='demo_tp_13',             replace_existing=True, timezone='Europe/Berlin')
    _scheduler.add_job(demo_tagesplan_fortschritt, 'cron', hour=16, minute=0,
                       id='demo_tp_16',             replace_existing=True, timezone='Europe/Berlin')
    _scheduler.start()
    app.logger.info("Scheduler gestartet (Backup täglich, Wochenbericht Mo 07:00, Monatsbericht 1. 07:00, Export 1. 08:00, Foto-Cleanup 1. 09:00, Demo-Reset täglich 03:00, Tagesplan-Fortschritt 10/13/16 Uhr)")
except ImportError:
    app.logger.warning("APScheduler nicht installiert – automatische Jobs deaktiviert.")

if __name__ == '__main__':
    init_db()
    port = int(os.environ.get('PORT', 5000))
    # Bugreport 2026-07-21 (Niedrig): fail-closed statt fail-open – Debug läuft jetzt nur
    # noch bei explizitem Opt-in. Nur relevant für diesen direkten `python app.py`-Pfad,
    # Produktion läuft über Gunicorn.
    debug = os.environ.get('FLASK_ENV', '').strip().lower() == 'development'
    print("\n" + "="*55)
    print("  Aktions Tracker gestartet!")
    print(f"  http://127.0.0.1:{port}")
    print("  Admin-Login: ADMIN / admin123")
    print("  Rep-Login:   z.B. MM / demo123")
    print("="*55 + "\n")
    app.run(debug=debug, host='0.0.0.0', port=port)
